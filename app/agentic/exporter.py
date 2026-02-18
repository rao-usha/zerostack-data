"""
Portfolio data exporter for CSV and Excel formats.

Provides:
- CSV export with all portfolio fields
- Excel export with formatting and multiple sheets
- Optional filtering by source type, date range, etc.
- Streaming support for large datasets
"""

import csv
import io
import logging
from dataclasses import dataclass
from datetime import date, datetime
from enum import Enum
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)


class ExportFormat(str, Enum):
    """Supported export formats."""

    CSV = "csv"
    EXCEL = "xlsx"


@dataclass
class ExportFilter:
    """Filters for portfolio export."""

    source_type: Optional[str] = None
    min_confidence: Optional[float] = None
    industry: Optional[str] = None
    start_date: Optional[date] = None
    end_date: Optional[date] = None
    min_market_value: Optional[float] = None


@dataclass
class PortfolioCompany:
    """Portfolio company data for export."""

    id: int
    company_name: str
    company_industry: Optional[str]
    company_stage: Optional[str]
    investment_type: Optional[str]
    investment_date: Optional[date]
    market_value_usd: Optional[float]
    shares_held: Optional[int]
    ownership_percentage: Optional[float]
    source_type: str
    confidence_level: float
    source_url: Optional[str]
    collected_date: datetime
    ticker_symbol: Optional[str] = None
    cusip: Optional[str] = None
    company_description: Optional[str] = None

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary for export."""
        return {
            "ID": self.id,
            "Company Name": self.company_name,
            "Industry": self.company_industry or "",
            "Stage": self.company_stage or "",
            "Investment Type": self.investment_type or "",
            "Investment Date": self.investment_date.isoformat()
            if self.investment_date
            else "",
            "Market Value (USD)": self.market_value_usd or "",
            "Shares Held": self.shares_held or "",
            "Ownership %": self.ownership_percentage or "",
            "Ticker": self.ticker_symbol or "",
            "CUSIP": self.cusip or "",
            "Source": self.source_type,
            "Confidence": f"{self.confidence_level:.0%}",
            "Source URL": self.source_url or "",
            "Collected Date": self.collected_date.isoformat()
            if self.collected_date
            else "",
            "Description": self.company_description or "",
        }


class PortfolioExporter:
    """
    Exports portfolio data to various formats.

    Supports:
    - CSV: Simple comma-separated format
    - Excel: Multi-sheet workbook with formatting

    Usage:
        exporter = PortfolioExporter(companies, investor_name="Sequoia Capital")
        csv_bytes = exporter.to_csv()
        xlsx_bytes = exporter.to_excel()
    """

    # Column configuration for exports
    COLUMNS = [
        ("ID", "id", 8),
        ("Company Name", "company_name", 30),
        ("Industry", "company_industry", 20),
        ("Stage", "company_stage", 15),
        ("Investment Type", "investment_type", 15),
        ("Investment Date", "investment_date", 12),
        ("Market Value (USD)", "market_value_usd", 18),
        ("Shares Held", "shares_held", 12),
        ("Ownership %", "ownership_percentage", 12),
        ("Ticker", "ticker_symbol", 10),
        ("CUSIP", "cusip", 12),
        ("Source", "source_type", 15),
        ("Confidence", "confidence_level", 10),
        ("Source URL", "source_url", 40),
        ("Collected Date", "collected_date", 12),
    ]

    def __init__(
        self,
        companies: List[PortfolioCompany],
        investor_name: str = "Unknown",
        investor_type: str = "unknown",
        investor_id: int = 0,
    ):
        self.companies = companies
        self.investor_name = investor_name
        self.investor_type = investor_type
        self.investor_id = investor_id
        self._export_time = datetime.utcnow()

    def to_csv(self) -> bytes:
        """
        Export portfolio to CSV format.

        Returns:
            CSV file contents as bytes (UTF-8 encoded with BOM for Excel compatibility)
        """
        output = io.StringIO()
        writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)

        # Write header row
        headers = [col[0] for col in self.COLUMNS]
        writer.writerow(headers)

        # Write data rows
        for company in self.companies:
            row = self._company_to_row(company)
            writer.writerow(row)

        # Return with UTF-8 BOM for Excel compatibility
        csv_content = output.getvalue()
        return ("\ufeff" + csv_content).encode("utf-8")

    def to_excel(self) -> bytes:
        """
        Export portfolio to Excel format with formatting.

        Returns:
            Excel file contents as bytes

        Note:
            Requires openpyxl library. Falls back to CSV if not available.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.warning("openpyxl not installed, falling back to CSV export")
            return self.to_csv()

        wb = Workbook()

        # Create main portfolio sheet
        ws_portfolio = wb.active
        ws_portfolio.title = "Portfolio"
        self._write_portfolio_sheet(ws_portfolio)

        # Create summary sheet
        ws_summary = wb.create_sheet("Summary")
        self._write_summary_sheet(ws_summary)

        # Create source breakdown sheet
        ws_sources = wb.create_sheet("By Source")
        self._write_source_breakdown_sheet(ws_sources)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    def _company_to_row(self, company: PortfolioCompany) -> List[Any]:
        """Convert a company to a row of values."""
        return [
            company.id,
            company.company_name,
            company.company_industry or "",
            company.company_stage or "",
            company.investment_type or "",
            company.investment_date.isoformat() if company.investment_date else "",
            company.market_value_usd if company.market_value_usd else "",
            company.shares_held if company.shares_held else "",
            company.ownership_percentage if company.ownership_percentage else "",
            company.ticker_symbol or "",
            company.cusip or "",
            company.source_type,
            f"{company.confidence_level:.0%}",
            company.source_url or "",
            company.collected_date.isoformat() if company.collected_date else "",
        ]

    def _write_portfolio_sheet(self, ws) -> None:
        """Write the main portfolio data sheet."""
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Write headers
        for col_idx, (header, _, width) in enumerate(self.COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Write data rows
        for row_idx, company in enumerate(self.companies, start=2):
            row_data = self._company_to_row(company)
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

                # Format specific columns
                if col_idx == 7 and value:  # Market Value
                    cell.number_format = '"$"#,##0.00'
                elif col_idx == 8 and value:  # Shares
                    cell.number_format = "#,##0"

        # Freeze header row
        ws.freeze_panes = "A2"

        # Add auto-filter
        if self.companies:
            ws.auto_filter.ref = (
                f"A1:{get_column_letter(len(self.COLUMNS))}{len(self.companies) + 1}"
            )

    def _write_summary_sheet(self, ws) -> None:
        """Write the summary statistics sheet."""
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
        )

        # Title
        ws.cell(
            row=1, column=1, value=f"Portfolio Summary: {self.investor_name}"
        ).font = title_font
        ws.cell(row=2, column=1, value=f"Investor Type: {self.investor_type}")
        ws.cell(
            row=3,
            column=1,
            value=f"Generated: {self._export_time.strftime('%Y-%m-%d %H:%M UTC')}",
        )

        # Statistics
        total_companies = len(self.companies)
        total_value = sum(c.market_value_usd or 0 for c in self.companies)
        avg_confidence = sum(c.confidence_level for c in self.companies) / max(
            1, total_companies
        )

        # Industry breakdown
        industries: Dict[str, int] = {}
        for c in self.companies:
            ind = c.company_industry or "Unknown"
            industries[ind] = industries.get(ind, 0) + 1

        # Source breakdown
        sources: Dict[str, int] = {}
        for c in self.companies:
            sources[c.source_type] = sources.get(c.source_type, 0) + 1

        row = 5
        ws.cell(row=row, column=1, value="Key Metrics").font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        ws.merge_cells(f"A{row}:B{row}")

        row += 1
        metrics = [
            ("Total Portfolio Companies", total_companies),
            ("Total Market Value", f"${total_value:,.2f}" if total_value else "N/A"),
            ("Average Confidence", f"{avg_confidence:.0%}"),
            ("Unique Industries", len(industries)),
            ("Data Sources Used", len(sources)),
        ]

        for label, value in metrics:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            row += 1

        # Top industries
        row += 1
        ws.cell(row=row, column=1, value="Top Industries").font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        ws.merge_cells(f"A{row}:B{row}")

        row += 1
        sorted_industries = sorted(
            industries.items(), key=lambda x: x[1], reverse=True
        )[:10]
        for industry, count in sorted_industries:
            ws.cell(row=row, column=1, value=industry)
            ws.cell(row=row, column=2, value=count)
            row += 1

        # Column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20

    def _write_source_breakdown_sheet(self, ws) -> None:
        """Write the source breakdown sheet."""
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="70AD47", end_color="70AD47", fill_type="solid"
        )
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Headers
        headers = [
            "Source Type",
            "Companies",
            "Avg Confidence",
            "Total Value",
            "% of Portfolio",
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        # Calculate source stats
        source_stats: Dict[str, Dict[str, Any]] = {}
        total_companies = len(self.companies)

        for company in self.companies:
            source = company.source_type
            if source not in source_stats:
                source_stats[source] = {
                    "count": 0,
                    "confidence_sum": 0.0,
                    "value_sum": 0.0,
                }
            source_stats[source]["count"] += 1
            source_stats[source]["confidence_sum"] += company.confidence_level
            source_stats[source]["value_sum"] += company.market_value_usd or 0

        # Write data
        row = 2
        for source, stats in sorted(
            source_stats.items(), key=lambda x: x[1]["count"], reverse=True
        ):
            count = stats["count"]
            avg_conf = stats["confidence_sum"] / count if count > 0 else 0
            total_val = stats["value_sum"]
            pct = (count / total_companies * 100) if total_companies > 0 else 0

            ws.cell(row=row, column=1, value=source).border = thin_border
            ws.cell(row=row, column=2, value=count).border = thin_border
            ws.cell(row=row, column=3, value=f"{avg_conf:.0%}").border = thin_border
            cell_value = ws.cell(
                row=row, column=4, value=total_val if total_val else "N/A"
            )
            cell_value.border = thin_border
            if total_val:
                cell_value.number_format = '"$"#,##0.00'
            ws.cell(row=row, column=5, value=f"{pct:.1f}%").border = thin_border
            row += 1

        # Column widths
        widths = [20, 12, 15, 18, 15]
        for col_idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width


def create_exporter_from_db_rows(
    rows: List[Tuple],
    investor_name: str,
    investor_type: str,
    investor_id: int,
) -> PortfolioExporter:
    """
    Create a PortfolioExporter from database rows.

    Expected row format (from portfolio_companies table):
    (id, company_name, company_industry, company_stage, investment_type,
     investment_date, market_value_usd, shares_held, ownership_percentage,
     source_type, confidence_level, source_url, collected_date,
     ticker_symbol, cusip, company_description)

    Args:
        rows: List of database rows
        investor_name: Name of the investor
        investor_type: Type of investor ('lp' or 'family_office')
        investor_id: Investor ID

    Returns:
        Configured PortfolioExporter instance
    """
    companies = []

    for row in rows:
        company = PortfolioCompany(
            id=row[0],
            company_name=row[1],
            company_industry=row[2],
            company_stage=row[3],
            investment_type=row[4],
            investment_date=row[5],
            market_value_usd=float(row[6]) if row[6] else None,
            shares_held=int(row[7]) if row[7] else None,
            ownership_percentage=float(row[8]) if row[8] else None,
            source_type=row[9],
            confidence_level=float(row[10]) if row[10] else 0.5,
            source_url=row[11],
            collected_date=row[12],
            ticker_symbol=row[13] if len(row) > 13 else None,
            cusip=row[14] if len(row) > 14 else None,
            company_description=row[15] if len(row) > 15 else None,
        )
        companies.append(company)

    return PortfolioExporter(
        companies=companies,
        investor_name=investor_name,
        investor_type=investor_type,
        investor_id=investor_id,
    )


def export_portfolio(
    rows: List[Tuple],
    investor_name: str,
    investor_type: str,
    investor_id: int,
    format: ExportFormat = ExportFormat.CSV,
) -> Tuple[bytes, str, str]:
    """
    Export portfolio data to specified format.

    Args:
        rows: Database rows from portfolio_companies
        investor_name: Name of investor
        investor_type: Type of investor
        investor_id: Investor ID
        format: Export format (csv or xlsx)

    Returns:
        Tuple of (file_bytes, filename, content_type)
    """
    exporter = create_exporter_from_db_rows(
        rows=rows,
        investor_name=investor_name,
        investor_type=investor_type,
        investor_id=investor_id,
    )

    # Generate safe filename
    safe_name = "".join(c if c.isalnum() or c in " -_" else "_" for c in investor_name)
    safe_name = safe_name.strip().replace(" ", "_")[:50]
    timestamp = datetime.utcnow().strftime("%Y%m%d")

    if format == ExportFormat.EXCEL:
        content = exporter.to_excel()
        filename = f"portfolio_{safe_name}_{timestamp}.xlsx"
        content_type = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        content = exporter.to_csv()
        filename = f"portfolio_{safe_name}_{timestamp}.csv"
        content_type = "text/csv; charset=utf-8"

    logger.info(
        f"Exported {len(exporter.companies)} companies for {investor_name} "
        f"to {format.value} ({len(content)} bytes)"
    )

    return content, filename, content_type
