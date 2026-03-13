"""
PE Portfolio Import Service.

Handles CSV/Excel upload for PE firm portfolio data with 4 templates:
portfolio companies, financial history, deal history, and leadership.
Supports validation, preview, execute, and rollback.
"""

import csv
import io
import logging
import uuid
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from difflib import get_close_matches
from typing import Any, Dict, List, Optional, Tuple

from sqlalchemy import select, and_, delete
from sqlalchemy.orm import Session

from app.core.pe_models import (
    PEPortfolioCompany,
    PECompanyFinancials,
    PECompanyLeadership,
    PEDeal,
    PEFirm,
    PEPerson,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Template definitions
# ---------------------------------------------------------------------------

TEMPLATES = {
    "portfolio_companies": {
        "description": "Import portfolio company master data",
        "required": ["company_name"],
        "optional": [
            "industry", "naics_code", "revenue", "ebitda", "employees",
            "hq_city", "hq_state", "website", "status", "entry_date",
            "entry_multiple", "sector", "description", "founded_year",
        ],
        "model": PEPortfolioCompany,
    },
    "financial_history": {
        "description": "Import financial time-series per company",
        "required": ["company_name", "fiscal_year"],
        "optional": [
            "revenue", "ebitda", "gross_margin_pct", "net_income",
            "total_debt", "employees", "data_source", "ebitda_margin_pct",
            "revenue_growth_pct", "free_cash_flow",
        ],
        "model": PECompanyFinancials,
    },
    "deal_history": {
        "description": "Import M&A deal history",
        "required": ["deal_name", "company_name", "deal_type"],
        "optional": [
            "counterparty", "enterprise_value", "ev_revenue_multiple",
            "ev_ebitda_multiple", "status", "close_date", "announced_date",
            "seller_name", "seller_type", "deal_sub_type",
        ],
        "model": PEDeal,
    },
    "leadership": {
        "description": "Import company leadership / executive team",
        "required": ["company_name", "person_name", "title"],
        "optional": [
            "department", "start_date", "is_board_member",
            "linkedin_url", "email", "role_category",
        ],
        "model": PECompanyLeadership,
    },
}

# Column aliases for auto-mapping
COLUMN_ALIASES = {
    "company": "company_name",
    "name": "company_name",
    "comp_name": "company_name",
    "sector": "industry",
    "city": "hq_city",
    "state": "hq_state",
    "ebitda_margin": "ebitda_margin_pct",
    "gross_margin": "gross_margin_pct",
    "year": "fiscal_year",
    "fy": "fiscal_year",
    "ev": "enterprise_value",
    "enterprise_value_usd": "enterprise_value",
    "ev_ebitda": "ev_ebitda_multiple",
    "ev_revenue": "ev_revenue_multiple",
    "type": "deal_type",
    "person": "person_name",
    "full_name": "person_name",
    "role": "title",
    "position": "title",
    "board_member": "is_board_member",
    "linkedin": "linkedin_url",
    "date_closed": "close_date",
    "closed_date": "close_date",
}

# Signature columns per template (for auto-detection)
TEMPLATE_SIGNATURES = {
    "portfolio_companies": {"entry_multiple", "hq_city", "hq_state", "ownership_status"},
    "financial_history": {"fiscal_year", "ebitda", "gross_margin_pct", "revenue_growth_pct"},
    "deal_history": {"deal_name", "deal_type", "enterprise_value", "ev_ebitda_multiple"},
    "leadership": {"person_name", "title", "is_board_member", "linkedin_url"},
}


# ---------------------------------------------------------------------------
# Import record (for tracking/rollback)
# ---------------------------------------------------------------------------

class ImportRecord:
    """Tracks a single import operation."""

    def __init__(self, import_id: str, template_type: str, firm_name: Optional[str] = None):
        self.import_id = import_id
        self.template_type = template_type
        self.firm_name = firm_name
        self.created_at = datetime.utcnow()
        self.status = "pending"  # pending → previewed → imported → rolled_back
        self.rows_parsed = 0
        self.rows_imported = 0
        self.errors: List[Dict] = []
        self.warnings: List[Dict] = []
        self.created_ids: Dict[str, List[int]] = {}  # table_name → [ids]
        self.raw_data: List[Dict] = []

    def to_dict(self) -> Dict[str, Any]:
        return {
            "import_id": self.import_id,
            "template_type": self.template_type,
            "firm_name": self.firm_name,
            "status": self.status,
            "rows_parsed": self.rows_parsed,
            "rows_imported": self.rows_imported,
            "errors": self.errors[:50],
            "warnings": self.warnings[:50],
            "error_count": len(self.errors),
            "warning_count": len(self.warnings),
            "created_ids": {k: len(v) for k, v in self.created_ids.items()},
            "created_at": self.created_at.isoformat(),
        }


# ---------------------------------------------------------------------------
# In-memory import store (for preview → confirm flow)
# ---------------------------------------------------------------------------

_import_store: Dict[str, ImportRecord] = {}


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

class PEPortfolioImporter:
    """Handles CSV/Excel import for PE portfolio data."""

    def __init__(self, db: Session):
        self.db = db

    def get_templates(self) -> List[Dict[str, Any]]:
        """List available import templates with column definitions."""
        result = []
        for name, tpl in TEMPLATES.items():
            result.append({
                "name": name,
                "description": tpl["description"],
                "required_columns": tpl["required"],
                "optional_columns": tpl["optional"],
                "all_columns": tpl["required"] + tpl["optional"],
            })
        return result

    def upload(
        self,
        file_content: bytes,
        filename: str,
        template_type: Optional[str] = None,
        firm_name: Optional[str] = None,
    ) -> ImportRecord:
        """
        Parse uploaded file and create an import record.

        Returns ImportRecord in 'pending' status ready for preview/confirm.
        """
        import_id = f"imp_{uuid.uuid4().hex[:12]}"

        # Parse file
        if filename.endswith((".xlsx", ".xls")):
            rows = self._parse_excel(file_content)
        else:
            rows = self._parse_csv(file_content)

        if not rows:
            record = ImportRecord(import_id, template_type or "unknown", firm_name)
            record.errors.append({"row": 0, "message": "No data rows found in file"})
            _import_store[import_id] = record
            return record

        # Normalize column names
        rows = self._normalize_columns(rows)

        # Auto-detect template if not specified
        if not template_type:
            template_type = self._detect_template(rows[0].keys())
            if not template_type:
                record = ImportRecord(import_id, "unknown", firm_name)
                record.errors.append({
                    "row": 0,
                    "message": "Could not auto-detect template type. "
                               "Specify template_type parameter.",
                })
                _import_store[import_id] = record
                return record

        record = ImportRecord(import_id, template_type, firm_name)
        record.rows_parsed = len(rows)
        record.raw_data = rows

        # Validate
        self._validate(record, rows)

        _import_store[import_id] = record
        return record

    def preview(self, import_id: str) -> Optional[Dict[str, Any]]:
        """Get preview of a pending import."""
        record = _import_store.get(import_id)
        if not record:
            return None

        record.status = "previewed"
        result = record.to_dict()
        result["sample_rows"] = record.raw_data[:5]
        result["column_mappings"] = self._suggest_mappings(
            record.raw_data[0].keys() if record.raw_data else [],
            record.template_type,
        )
        return result

    def execute(self, import_id: str) -> Optional[Dict[str, Any]]:
        """Execute a previewed import, inserting records into the database."""
        record = _import_store.get(import_id)
        if not record:
            return None

        if record.errors:
            return {"error": "Cannot execute import with validation errors", **record.to_dict()}

        template_type = record.template_type
        rows = record.raw_data

        try:
            if template_type == "portfolio_companies":
                self._import_portfolio_companies(record, rows)
            elif template_type == "financial_history":
                self._import_financial_history(record, rows)
            elif template_type == "deal_history":
                self._import_deal_history(record, rows)
            elif template_type == "leadership":
                self._import_leadership(record, rows)

            self.db.commit()
            record.status = "imported"
            logger.info(
                "Import %s complete: %d rows imported (%s)",
                import_id, record.rows_imported, template_type,
            )
        except Exception as e:
            self.db.rollback()
            record.status = "failed"
            record.errors.append({"row": 0, "message": f"Import failed: {e}"})
            logger.error("Import %s failed: %s", import_id, e)

        return record.to_dict()

    def rollback(self, import_id: str) -> Optional[Dict[str, Any]]:
        """Undo an import by deleting created records."""
        record = _import_store.get(import_id)
        if not record:
            return None

        if record.status != "imported":
            return {"error": f"Cannot rollback import in status '{record.status}'"}

        deleted_counts = {}
        try:
            for table_name, ids in record.created_ids.items():
                model = self._table_to_model(table_name)
                if model and ids:
                    self.db.execute(
                        delete(model).where(model.id.in_(ids))
                    )
                    deleted_counts[table_name] = len(ids)

            self.db.commit()
            record.status = "rolled_back"
            logger.info("Rollback %s: deleted %s", import_id, deleted_counts)
        except Exception as e:
            self.db.rollback()
            logger.error("Rollback %s failed: %s", import_id, e)
            return {"error": f"Rollback failed: {e}"}

        return {"status": "rolled_back", "deleted": deleted_counts}

    def get_import(self, import_id: str) -> Optional[Dict[str, Any]]:
        """Get import record status."""
        record = _import_store.get(import_id)
        return record.to_dict() if record else None

    def list_imports(self) -> List[Dict[str, Any]]:
        """List all import records."""
        return [r.to_dict() for r in _import_store.values()]

    # ------------------------------------------------------------------
    # Parsing
    # ------------------------------------------------------------------

    def _parse_csv(self, content: bytes) -> List[Dict[str, str]]:
        """Parse CSV content into list of dicts."""
        try:
            text = content.decode("utf-8")
        except UnicodeDecodeError:
            text = content.decode("latin-1")

        reader = csv.DictReader(io.StringIO(text))
        rows = []
        for row in reader:
            # Skip completely empty rows
            if any(v and v.strip() for v in row.values()):
                rows.append({k: (v.strip() if v else "") for k, v in row.items()})
        return rows

    def _parse_excel(self, content: bytes) -> List[Dict[str, str]]:
        """Parse Excel content into list of dicts."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ValueError("openpyxl required for Excel import: pip install openpyxl")

        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)

        # First row = headers
        headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(next(rows_iter))]

        rows = []
        for row_vals in rows_iter:
            if any(v is not None for v in row_vals):
                row_dict = {}
                for i, val in enumerate(row_vals):
                    if i < len(headers):
                        row_dict[headers[i]] = str(val).strip() if val is not None else ""
                rows.append(row_dict)
        return rows

    # ------------------------------------------------------------------
    # Normalization & Detection
    # ------------------------------------------------------------------

    def _normalize_columns(self, rows: List[Dict]) -> List[Dict]:
        """Normalize column names: lowercase, strip, underscores, apply aliases."""
        normalized = []
        for row in rows:
            new_row = {}
            for k, v in row.items():
                key = k.lower().strip().replace(" ", "_").replace("-", "_")
                key = COLUMN_ALIASES.get(key, key)
                new_row[key] = v
            normalized.append(new_row)
        return normalized

    def _detect_template(self, columns) -> Optional[str]:
        """Auto-detect template type from column headers."""
        col_set = set(columns)

        best_match = None
        best_score = 0

        for tpl_name, sig_cols in TEMPLATE_SIGNATURES.items():
            overlap = len(col_set & sig_cols)
            required = set(TEMPLATES[tpl_name]["required"])
            has_required = required.issubset(col_set)

            score = overlap + (5 if has_required else 0)
            if score > best_score:
                best_score = score
                best_match = tpl_name

        return best_match if best_score >= 2 else None

    def _suggest_mappings(self, columns, template_type: str) -> Dict[str, str]:
        """Suggest column mappings for close-but-not-exact names."""
        if template_type not in TEMPLATES:
            return {}

        tpl = TEMPLATES[template_type]
        expected = set(tpl["required"] + tpl["optional"])
        mappings = {}

        for col in columns:
            if col not in expected:
                matches = get_close_matches(col, expected, n=1, cutoff=0.6)
                if matches:
                    mappings[col] = matches[0]

        return mappings

    # ------------------------------------------------------------------
    # Validation
    # ------------------------------------------------------------------

    def _validate(self, record: ImportRecord, rows: List[Dict]) -> None:
        """Validate rows against template requirements."""
        tpl = TEMPLATES.get(record.template_type)
        if not tpl:
            record.errors.append({"row": 0, "message": f"Unknown template: {record.template_type}"})
            return

        # Check required columns
        if rows:
            cols = set(rows[0].keys())
            missing = set(tpl["required"]) - cols
            if missing:
                record.errors.append({
                    "row": 0,
                    "message": f"Missing required columns: {', '.join(sorted(missing))}",
                })
                return

        # Per-row validation
        for i, row in enumerate(rows, start=1):
            for req_col in tpl["required"]:
                val = row.get(req_col, "").strip()
                if not val:
                    record.errors.append({
                        "row": i,
                        "message": f"Missing required value: {req_col}",
                    })

            # Type-specific checks
            if record.template_type == "financial_history":
                fy = row.get("fiscal_year", "")
                if fy and not fy.isdigit():
                    record.errors.append({"row": i, "message": f"Invalid fiscal_year: {fy}"})

            # Warn on empty optional numeric fields
            for col in ["revenue", "ebitda", "enterprise_value"]:
                if col in row and row[col] and not self._is_numeric(row[col]):
                    record.warnings.append({
                        "row": i,
                        "message": f"Non-numeric value for {col}: {row[col][:30]}",
                    })

    # ------------------------------------------------------------------
    # Import executors
    # ------------------------------------------------------------------

    def _import_portfolio_companies(self, record: ImportRecord, rows: List[Dict]) -> None:
        """Import portfolio company records."""
        created_ids = []
        for row in rows:
            name = row.get("company_name", "").strip()
            if not name:
                continue

            # Fuzzy match existing
            existing = self._find_company(name)
            if existing:
                record.warnings.append({
                    "row": 0,
                    "message": f"Company '{name}' already exists (id={existing.id}), skipping",
                })
                continue

            company = PEPortfolioCompany(
                name=name,
                industry=row.get("industry") or row.get("sector") or None,
                naics_code=row.get("naics_code") or None,
                headquarters_city=row.get("hq_city") or None,
                headquarters_state=row.get("hq_state") or None,
                website=row.get("website") or None,
                employee_count=self._to_int(row.get("employees")),
                status=row.get("status", "Active"),
                description=row.get("description") or None,
                founded_year=self._to_int(row.get("founded_year")),
                data_source=f"import:{record.import_id}",
            )
            self.db.add(company)
            self.db.flush()
            created_ids.append(company.id)
            record.rows_imported += 1

        record.created_ids["pe_portfolio_companies"] = created_ids

    def _import_financial_history(self, record: ImportRecord, rows: List[Dict]) -> None:
        """Import financial history records."""
        created_ids = []
        for row in rows:
            company_name = row.get("company_name", "").strip()
            company = self._find_company(company_name)
            if not company:
                record.warnings.append({
                    "row": 0,
                    "message": f"Company '{company_name}' not found, skipping financials",
                })
                continue

            fin = PECompanyFinancials(
                company_id=company.id,
                fiscal_year=self._to_int(row.get("fiscal_year")),
                revenue_usd=self._to_decimal(row.get("revenue")),
                ebitda_usd=self._to_decimal(row.get("ebitda")),
                gross_margin_pct=self._to_decimal(row.get("gross_margin_pct")),
                ebitda_margin_pct=self._to_decimal(row.get("ebitda_margin_pct")),
                net_income_usd=self._to_decimal(row.get("net_income")),
                total_debt_usd=self._to_decimal(row.get("total_debt")),
                revenue_growth_pct=self._to_decimal(row.get("revenue_growth_pct")),
                free_cash_flow_usd=self._to_decimal(row.get("free_cash_flow")),
                data_source=row.get("data_source") or f"import:{record.import_id}",
            )
            self.db.add(fin)
            self.db.flush()
            created_ids.append(fin.id)
            record.rows_imported += 1

        record.created_ids["pe_company_financials"] = created_ids

    def _import_deal_history(self, record: ImportRecord, rows: List[Dict]) -> None:
        """Import deal history records."""
        created_ids = []
        for row in rows:
            company_name = row.get("company_name", "").strip()
            company = self._find_company(company_name)
            if not company:
                record.warnings.append({
                    "row": 0,
                    "message": f"Company '{company_name}' not found, skipping deal",
                })
                continue

            deal = PEDeal(
                company_id=company.id,
                deal_name=row.get("deal_name", ""),
                deal_type=row.get("deal_type", "LBO"),
                deal_sub_type=row.get("deal_sub_type") or None,
                enterprise_value_usd=self._to_decimal(row.get("enterprise_value")),
                ev_revenue_multiple=self._to_decimal(row.get("ev_revenue_multiple")),
                ev_ebitda_multiple=self._to_decimal(row.get("ev_ebitda_multiple")),
                buyer_name=row.get("counterparty") or None,
                seller_name=row.get("seller_name") or None,
                seller_type=row.get("seller_type") or None,
                status=row.get("status", "Closed"),
                closed_date=self._to_date(row.get("close_date")),
                announced_date=self._to_date(row.get("announced_date")),
                data_source=f"import:{record.import_id}",
            )
            self.db.add(deal)
            self.db.flush()
            created_ids.append(deal.id)
            record.rows_imported += 1

        record.created_ids["pe_deals"] = created_ids

    def _import_leadership(self, record: ImportRecord, rows: List[Dict]) -> None:
        """Import leadership / executive team records."""
        created_person_ids = []
        created_leadership_ids = []

        for row in rows:
            company_name = row.get("company_name", "").strip()
            company = self._find_company(company_name)
            if not company:
                record.warnings.append({
                    "row": 0,
                    "message": f"Company '{company_name}' not found, skipping leader",
                })
                continue

            person_name = row.get("person_name", "").strip()
            if not person_name:
                continue

            # Create or find person
            person = self.db.execute(
                select(PEPerson).where(PEPerson.full_name == person_name)
            ).scalar_one_or_none()

            if not person:
                person = PEPerson(
                    full_name=person_name,
                    linkedin_url=row.get("linkedin_url") or None,
                    data_source=f"import:{record.import_id}",
                )
                self.db.add(person)
                self.db.flush()
                created_person_ids.append(person.id)

            title = row.get("title", "").strip()
            is_board = row.get("is_board_member", "").lower() in ("true", "yes", "1", "y")

            leadership = PECompanyLeadership(
                company_id=company.id,
                person_id=person.id,
                title=title,
                role_category=row.get("department") or row.get("role_category") or None,
                is_board_member=is_board,
                start_date=self._to_date(row.get("start_date")),
                is_current=True,
                data_source=f"import:{record.import_id}",
            )
            self.db.add(leadership)
            self.db.flush()
            created_leadership_ids.append(leadership.id)
            record.rows_imported += 1

        record.created_ids["pe_people"] = created_person_ids
        record.created_ids["pe_company_leadership"] = created_leadership_ids

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _find_company(self, name: str) -> Optional[PEPortfolioCompany]:
        """Find existing company by exact or near-exact name match."""
        if not name:
            return None
        return self.db.execute(
            select(PEPortfolioCompany).where(
                PEPortfolioCompany.name.ilike(name.strip())
            )
        ).scalar_one_or_none()

    def _table_to_model(self, table_name: str):
        """Map table name to SQLAlchemy model."""
        mapping = {
            "pe_portfolio_companies": PEPortfolioCompany,
            "pe_company_financials": PECompanyFinancials,
            "pe_deals": PEDeal,
            "pe_company_leadership": PECompanyLeadership,
            "pe_people": PEPerson,
        }
        return mapping.get(table_name)

    @staticmethod
    def _to_decimal(val: Optional[str]) -> Optional[Decimal]:
        if not val or not val.strip():
            return None
        try:
            cleaned = val.strip().replace(",", "").replace("$", "").replace("%", "")
            return Decimal(cleaned)
        except (InvalidOperation, ValueError):
            return None

    @staticmethod
    def _to_int(val: Optional[str]) -> Optional[int]:
        if not val or not val.strip():
            return None
        try:
            return int(val.strip().replace(",", ""))
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _to_date(val: Optional[str]) -> Optional[date]:
        if not val or not val.strip():
            return None
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except ValueError:
                continue
        return None

    @staticmethod
    def _is_numeric(val: str) -> bool:
        try:
            cleaned = val.strip().replace(",", "").replace("$", "").replace("%", "")
            float(cleaned)
            return True
        except (ValueError, TypeError):
            return False
