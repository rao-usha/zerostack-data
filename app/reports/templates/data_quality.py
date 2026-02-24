"""
Data Quality Assessment Report Template.

Analyzes all database tables for completeness, accuracy, and validity.
Generates a comprehensive report with domain breakdowns, column-level
null analysis, freshness tracking, and actionable recommendations.
"""

import json
import logging
from datetime import datetime, timedelta, timezone
from typing import Dict, Any, List, Optional
from io import BytesIO

from sqlalchemy import text
from sqlalchemy.orm import Session

from app.reports.design_system import (
    html_document, page_header, kpi_strip, kpi_card,
    toc, section_start, section_end,
    data_table, pill_badge, callout,
    chart_container, chart_init_js, page_footer,
    build_doughnut_config, build_horizontal_bar_config,
    build_bar_fallback, build_chart_legend,
    GREEN, ORANGE, RED, GRAY, BLUE,
)
from app.core.domains import classify_table, DOMAIN_LABELS, DOMAIN_COLORS

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Scoring
# ---------------------------------------------------------------------------

GRADE_THRESHOLDS = [
    (97, "A+"), (93, "A"), (90, "A-"),
    (87, "B+"), (83, "B"), (80, "B-"),
    (77, "C+"), (73, "C"), (70, "C-"),
    (60, "D"), (0, "F"),
]


def _letter_grade(score: float) -> str:
    for threshold, grade in GRADE_THRESHOLDS:
        if score >= threshold:
            return grade
    return "F"


def _fmt_number(n) -> str:
    if n is None:
        return "0"
    return f"{int(n):,}"


def _completeness_badge(pct: float) -> str:
    if pct >= 90:
        color = "#166534"
        bg = "#f0fdf4"
    elif pct >= 70:
        color = "#92400e"
        bg = "#fef3c7"
    else:
        color = "#991b1b"
        bg = "#fef2f2"
    return (
        f'<span style="display:inline-block;padding:2px 8px;border-radius:10px;'
        f'font-size:11px;font-weight:600;background:{bg};color:{color}">'
        f'{pct:.1f}%</span>'
    )


def _status_pill(status: str) -> str:
    if status == "Good":
        return pill_badge("Good", "private")
    elif status == "Warning":
        return pill_badge("Warning", "pe")
    elif status == "Critical":
        return pill_badge("Critical", "default")
    else:
        return pill_badge("Empty", "sub")


# ---------------------------------------------------------------------------
# Template
# ---------------------------------------------------------------------------


class DataQualityTemplate:
    """Data quality assessment report template."""

    name = "data_quality"
    description = "Platform-wide data quality assessment across all tables"

    def gather_data(self, db: Session, params: Dict[str, Any]) -> Dict[str, Any]:
        """Analyze all public tables for completeness, freshness, and validity."""
        now = datetime.now(timezone.utc)

        # Step 1: Discover tables
        tables = self._get_tables(db)
        if not tables:
            return self._empty_result(now)

        # Step 2: Row counts (estimated first, exact for small tables)
        row_counts = self._get_row_counts(db, tables)

        # Step 3: Column metadata
        col_meta = self._get_column_metadata(db)

        # Step 4: Column-level null analysis for populated tables
        col_nulls = self._get_column_nulls(db, tables, row_counts, col_meta)

        # Step 5: Ingestion job freshness
        source_freshness = self._get_source_freshness(db)

        # Step 6: Table-level freshness
        table_freshness = self._get_table_freshness(db, tables, col_meta)

        # Build per-table analysis
        table_details = []
        column_details = []
        for tbl in tables:
            count = row_counts.get(tbl, 0)
            cols = col_meta.get(tbl, [])
            nulls = col_nulls.get(tbl, {})

            # Completeness
            col_completeness = []
            null_columns = 0
            for col_info in cols:
                col_name = col_info["column_name"]
                non_null = nulls.get(col_name)
                if count > 0 and non_null is not None:
                    null_pct = round((1 - non_null / count) * 100, 2)
                    col_completeness.append(100 - null_pct)
                    if null_pct > 80:
                        null_columns += 1
                    column_details.append({
                        "table": tbl,
                        "column": col_name,
                        "data_type": col_info["data_type"],
                        "null_pct": null_pct,
                        "non_null_count": non_null,
                        "total_count": count,
                    })
                elif count == 0:
                    column_details.append({
                        "table": tbl,
                        "column": col_name,
                        "data_type": col_info["data_type"],
                        "null_pct": 100.0,
                        "non_null_count": 0,
                        "total_count": 0,
                    })

            completeness_pct = (
                round(sum(col_completeness) / len(col_completeness), 2)
                if col_completeness else 0.0
            )

            # Freshness
            last_updated = table_freshness.get(tbl)
            freshness_days = None
            if last_updated:
                freshness_days = (now - last_updated).days

            # Status
            if count == 0:
                status = "Empty"
            elif completeness_pct < 70:
                status = "Critical"
            elif completeness_pct < 90:
                status = "Warning"
            else:
                status = "Good"

            domain_key = classify_table(tbl)
            table_details.append({
                "name": tbl,
                "domain": domain_key,
                "domain_label": DOMAIN_LABELS.get(domain_key, "Other"),
                "row_count": count,
                "column_count": len(cols),
                "completeness_pct": completeness_pct,
                "null_columns": null_columns,
                "last_updated": last_updated.isoformat() if last_updated else None,
                "freshness_days": freshness_days,
                "status": status,
            })

        # Sort: critical first, then warning, then good, then empty
        status_order = {"Critical": 0, "Warning": 1, "Good": 2, "Empty": 3}
        table_details.sort(key=lambda t: (status_order.get(t["status"], 4), t["name"]))

        # Domain aggregation
        domain_agg: Dict[str, Dict] = {}
        for t in table_details:
            dk = t["domain"]
            if dk not in domain_agg:
                domain_agg[dk] = {
                    "name": DOMAIN_LABELS.get(dk, "Other"),
                    "key": dk,
                    "tables_total": 0,
                    "tables_with_data": 0,
                    "tables_empty": 0,
                    "total_records": 0,
                    "completeness_values": [],
                }
            agg = domain_agg[dk]
            agg["tables_total"] += 1
            if t["row_count"] > 0:
                agg["tables_with_data"] += 1
                agg["completeness_values"].append(t["completeness_pct"])
            else:
                agg["tables_empty"] += 1
            agg["total_records"] += t["row_count"]

        domains = []
        for dk, agg in sorted(domain_agg.items(), key=lambda x: -x[1]["total_records"]):
            avg_c = (
                round(sum(agg["completeness_values"]) / len(agg["completeness_values"]), 1)
                if agg["completeness_values"] else 0.0
            )
            domains.append({
                "name": agg["name"],
                "key": agg["key"],
                "tables_total": agg["tables_total"],
                "tables_with_data": agg["tables_with_data"],
                "tables_empty": agg["tables_empty"],
                "total_records": agg["total_records"],
                "avg_completeness": avg_c,
            })

        # Summary
        total_tables = len(table_details)
        tables_with_data = sum(1 for t in table_details if t["row_count"] > 0)
        tables_empty = total_tables - tables_with_data
        total_records = sum(t["row_count"] for t in table_details)
        tables_with_issues = sum(
            1 for t in table_details if t["status"] in ("Critical", "Warning")
        )

        # Overall completeness (avg of populated tables)
        populated_completeness = [
            t["completeness_pct"] for t in table_details if t["row_count"] > 0
        ]
        overall_completeness = (
            round(sum(populated_completeness) / len(populated_completeness), 1)
            if populated_completeness else 0.0
        )

        # Overall score: 60% completeness + 25% coverage + 15% freshness
        coverage_pct = (tables_with_data / total_tables * 100) if total_tables > 0 else 0
        fresh_tables = sum(
            1 for t in table_details
            if t["freshness_days"] is not None and t["freshness_days"] <= 30
        )
        freshness_pct = (fresh_tables / tables_with_data * 100) if tables_with_data > 0 else 0
        overall_score = round(
            0.60 * overall_completeness + 0.25 * coverage_pct + 0.15 * freshness_pct, 1
        )

        # Stale tables
        stale_tables = [
            {"name": t["name"], "days": t["freshness_days"]}
            for t in table_details
            if t["freshness_days"] is not None and t["freshness_days"] > 30
        ]

        # Recommendations
        recommendations = self._generate_recommendations(
            table_details, domains, overall_completeness, tables_empty,
            total_tables, stale_tables, source_freshness,
        )

        return {
            "generated_at": now.isoformat(),
            "summary": {
                "total_tables": total_tables,
                "tables_with_data": tables_with_data,
                "tables_empty": tables_empty,
                "total_records": total_records,
                "overall_completeness": overall_completeness,
                "overall_score": overall_score,
                "letter_grade": _letter_grade(overall_score),
                "tables_with_issues": tables_with_issues,
                "coverage_pct": round(coverage_pct, 1),
                "freshness_pct": round(freshness_pct, 1),
            },
            "domains": domains,
            "tables": table_details,
            "column_details": column_details,
            "freshness": {
                "sources": source_freshness,
                "stale_tables": stale_tables,
            },
            "recommendations": recommendations,
        }

    # ── SQL Queries ───────────────────────────────────────────────────

    def _get_tables(self, db: Session) -> List[str]:
        result = db.execute(text(
            "SELECT table_name FROM information_schema.tables "
            "WHERE table_schema = 'public' AND table_type = 'BASE TABLE' "
            "ORDER BY table_name"
        ))
        return [row[0] for row in result.fetchall()]

    def _get_row_counts(self, db: Session, tables: List[str]) -> Dict[str, int]:
        # Estimated counts first
        result = db.execute(text(
            "SELECT relname, n_live_tup FROM pg_stat_user_tables "
            "WHERE schemaname = 'public'"
        ))
        estimates = {row[0]: int(row[1]) for row in result.fetchall()}

        counts = {}
        for tbl in tables:
            est = estimates.get(tbl, 0)
            if est < 500000:
                try:
                    exact = db.execute(text(f'SELECT COUNT(*) FROM "{tbl}"'))
                    counts[tbl] = exact.scalar() or 0
                except Exception:
                    counts[tbl] = est
            else:
                counts[tbl] = est
        return counts

    def _get_column_metadata(self, db: Session) -> Dict[str, List[Dict]]:
        result = db.execute(text(
            "SELECT table_name, column_name, data_type, is_nullable "
            "FROM information_schema.columns "
            "WHERE table_schema = 'public' "
            "ORDER BY table_name, ordinal_position"
        ))
        meta: Dict[str, List[Dict]] = {}
        for row in result.fetchall():
            meta.setdefault(row[0], []).append({
                "column_name": row[1],
                "data_type": row[2],
                "is_nullable": row[3],
            })
        return meta

    def _get_column_nulls(
        self, db: Session, tables: List[str],
        row_counts: Dict[str, int], col_meta: Dict[str, List[Dict]],
    ) -> Dict[str, Dict[str, int]]:
        nulls: Dict[str, Dict[str, int]] = {}
        for tbl in tables:
            count = row_counts.get(tbl, 0)
            if count == 0:
                continue
            cols = col_meta.get(tbl, [])
            if not cols:
                continue
            # Build: SELECT COUNT(*), COUNT("col1"), COUNT("col2"), ... FROM "table"
            col_exprs = ", ".join(f'COUNT("{c["column_name"]}")' for c in cols)
            try:
                result = db.execute(text(f'SELECT {col_exprs} FROM "{tbl}"'))
                row = result.fetchone()
                if row:
                    tbl_nulls = {}
                    for i, c in enumerate(cols):
                        tbl_nulls[c["column_name"]] = int(row[i])
                    nulls[tbl] = tbl_nulls
            except Exception as e:
                logger.warning(f"Could not analyze nulls for {tbl}: {e}")
        return nulls

    def _get_source_freshness(self, db: Session) -> List[Dict]:
        try:
            result = db.execute(text("""
                SELECT source,
                       MAX(completed_at) AS last_completed,
                       COUNT(*) FILTER (WHERE status = 'success') AS successes,
                       COUNT(*) FILTER (WHERE status = 'failed') AS failures
                FROM ingestion_jobs
                GROUP BY source
                ORDER BY MAX(completed_at) DESC NULLS LAST
            """))
            now = datetime.now(timezone.utc)
            sources = []
            for row in result.fetchall():
                last = row[1]
                if last and last.tzinfo is None:
                    last = last.replace(tzinfo=timezone.utc)
                days_ago = (now - last).days if last else None
                sources.append({
                    "source": row[0],
                    "last_completed": last.isoformat() if last else None,
                    "days_ago": days_ago,
                    "successes": row[2],
                    "failures": row[3],
                })
            return sources
        except Exception:
            return []

    def _get_table_freshness(
        self, db: Session, tables: List[str], col_meta: Dict[str, List[Dict]],
    ) -> Dict[str, datetime]:
        ts_cols = ["updated_at", "created_at", "collected_at", "completed_at", "ingested_at"]
        freshness = {}
        for tbl in tables:
            cols = col_meta.get(tbl, [])
            col_names = {c["column_name"] for c in cols}
            ts_col = None
            for candidate in ts_cols:
                if candidate in col_names:
                    ts_col = candidate
                    break
            if ts_col:
                try:
                    result = db.execute(
                        text(f'SELECT MAX("{ts_col}") FROM "{tbl}"')
                    )
                    val = result.scalar()
                    if val:
                        if hasattr(val, 'tzinfo') and val.tzinfo is None:
                            val = val.replace(tzinfo=timezone.utc)
                        freshness[tbl] = val
                except Exception:
                    pass
        return freshness

    def _generate_recommendations(
        self, tables, domains, overall_completeness,
        tables_empty, total_tables, stale_tables, source_freshness,
    ) -> List[Dict]:
        recs = []

        # Empty table coverage
        empty_pct = (tables_empty / total_tables * 100) if total_tables > 0 else 0
        if empty_pct > 30:
            recs.append({
                "severity": "warn",
                "title": "Low Table Coverage",
                "detail": (
                    f"{tables_empty} of {total_tables} tables ({empty_pct:.0f}%) are empty. "
                    "Run ingestion pipelines to populate data sources."
                ),
            })

        # Domain-level completeness
        for d in domains:
            if d["tables_with_data"] > 0 and d["avg_completeness"] < 70:
                recs.append({
                    "severity": "warn",
                    "title": f"{d['name']} Domain — Low Completeness",
                    "detail": (
                        f"The {d['name']} domain has {d['avg_completeness']:.1f}% "
                        f"average completeness across {d['tables_with_data']} populated tables. "
                        "Review enrichment pipelines."
                    ),
                })

        # Stale sources
        for src in source_freshness:
            if src["days_ago"] is not None and src["days_ago"] > 30:
                recs.append({
                    "severity": "info",
                    "title": f"Stale Source: {src['source']}",
                    "detail": (
                        f"No successful ingestion in {src['days_ago']} days. "
                        "Consider re-running or disabling the schedule."
                    ),
                })

        # High-null columns
        critical_tables = [t for t in tables if t["status"] == "Critical"]
        if critical_tables:
            names = ", ".join(t["name"] for t in critical_tables[:5])
            more = f" and {len(critical_tables) - 5} more" if len(critical_tables) > 5 else ""
            recs.append({
                "severity": "warn",
                "title": f"{len(critical_tables)} Tables with Critical Completeness",
                "detail": (
                    f"Tables with < 70% completeness: {names}{more}. "
                    "Many columns have > 80% null values."
                ),
            })

        # Overall health
        if overall_completeness >= 90:
            recs.append({
                "severity": "good",
                "title": "Overall Data Quality is Strong",
                "detail": (
                    f"Platform-wide completeness is {overall_completeness:.1f}%. "
                    "Continue monitoring to maintain quality."
                ),
            })
        elif not recs:
            recs.append({
                "severity": "info",
                "title": "Data Quality Assessment Complete",
                "detail": "No critical issues detected. Review individual tables for optimization opportunities.",
            })

        return recs

    def _empty_result(self, now: datetime) -> Dict[str, Any]:
        return {
            "generated_at": now.isoformat(),
            "summary": {
                "total_tables": 0, "tables_with_data": 0, "tables_empty": 0,
                "total_records": 0, "overall_completeness": 0, "overall_score": 0,
                "letter_grade": "F", "tables_with_issues": 0,
                "coverage_pct": 0, "freshness_pct": 0,
            },
            "domains": [],
            "tables": [],
            "column_details": [],
            "freshness": {"sources": [], "stale_tables": []},
            "recommendations": [{
                "severity": "warn",
                "title": "Empty Database",
                "detail": "No tables found. Run ingestion pipelines to populate the database.",
            }],
        }

    # ── HTML Rendering ────────────────────────────────────────────────

    def render_html(self, data: Dict[str, Any]) -> str:
        summary = data.get("summary", {})
        domains = data.get("domains", [])
        tables = data.get("tables", [])
        col_details = data.get("column_details", [])
        freshness = data.get("freshness", {})
        recommendations = data.get("recommendations", [])

        grade = summary.get("letter_grade", "F")
        charts_js = ""
        body = ""

        # ── Page Header ──────────────────────────────────────────
        body += page_header(
            title="Data Quality Assessment",
            subtitle=f"{summary.get('total_tables', 0)} tables analyzed across "
                     f"{len(domains)} domains",
            badge=f"Grade: {grade}",
        )

        # ── KPI Strip ────────────────────────────────────────────
        cards = ""
        cards += kpi_card(
            "Overall Score", grade,
            delta=f"{summary.get('overall_completeness', 0):.1f}% completeness",
            delta_dir="up" if summary.get("overall_completeness", 0) >= 80 else "down",
        )
        cards += kpi_card("Tables Analyzed", _fmt_number(summary.get("total_tables", 0)))
        cards += kpi_card("Total Records", _fmt_number(summary.get("total_records", 0)))
        cards += kpi_card(
            "Tables with Issues", str(summary.get("tables_with_issues", 0)),
            delta="needs attention" if summary.get("tables_with_issues", 0) > 0 else "all clear",
            delta_dir="down" if summary.get("tables_with_issues", 0) > 0 else "up",
        )
        cards += kpi_card("Tables Empty", str(summary.get("tables_empty", 0)))

        body += '\n<div class="container">'
        body += "\n" + kpi_strip(cards)

        # ── TOC ──────────────────────────────────────────────────
        toc_items = [
            {"number": 1, "id": "executive-summary", "title": "Executive Summary"},
            {"number": 2, "id": "domain-breakdown", "title": "Domain Breakdown"},
            {"number": 3, "id": "table-analysis", "title": "Table-by-Table Analysis"},
            {"number": 4, "id": "column-completeness", "title": "Column Completeness"},
            {"number": 5, "id": "freshness", "title": "Freshness Analysis"},
            {"number": 6, "id": "recommendations", "title": "Recommendations"},
        ]
        body += "\n" + toc(toc_items)

        # ── Section 1: Executive Summary ─────────────────────────
        body += "\n" + section_start(1, "Executive Summary", "executive-summary")

        # Chart row: doughnut (left) + horizontal bar (right)
        body += '<div class="chart-row">'

        # Left: Overall Health doughnut
        completeness = summary.get("overall_completeness", 0)
        incomplete = round(100 - completeness, 1)
        doughnut_config = build_doughnut_config(
            ["Complete", "Incomplete"],
            [completeness, incomplete],
            [GREEN, GRAY],
        )
        doughnut_json = json.dumps(doughnut_config)
        body += '<div>'
        body += chart_container(
            "healthDoughnut", doughnut_json,
            build_bar_fallback(["Complete", "Incomplete"], [completeness, incomplete], GREEN),
            size="medium", title="Overall Completeness",
        )
        charts_js += chart_init_js("healthDoughnut", doughnut_json)
        body += build_chart_legend(
            ["Complete", "Incomplete"],
            [completeness, incomplete],
            [GREEN, GRAY],
            value_suffix="%", show_pct=False,
        )
        body += '</div>'

        # Right: Domain Scores horizontal bar
        if domains:
            d_labels = [d["name"] for d in domains if d["tables_with_data"] > 0]
            d_values = [d["avg_completeness"] for d in domains if d["tables_with_data"] > 0]
            d_colors = []
            for v in d_values:
                if v >= 90:
                    d_colors.append(GREEN)
                elif v >= 70:
                    d_colors.append(ORANGE)
                else:
                    d_colors.append(RED)
            bar_config = build_horizontal_bar_config(
                d_labels, d_values, d_colors, dataset_label="Completeness %",
            )
            bar_config["options"]["scales"]["x"]["max"] = 100
            bar_json = json.dumps(bar_config)
            bar_height = f"{max(len(d_labels) * 44 + 40, 140)}px"
            body += '<div>'
            body += chart_container(
                "domainBar", bar_json,
                build_bar_fallback(d_labels, d_values, BLUE),
                title="Domain Completeness", height=bar_height,
            )
            charts_js += chart_init_js("domainBar", bar_json)
            body += '</div>'

        body += '</div>'  # close chart-row

        # Assessment callout
        score = summary.get("overall_score", 0)
        if score >= 90:
            assess_variant = "good"
            assess_text = (
                f"<strong>Excellent:</strong> Overall data quality score is "
                f"{score:.1f}% ({grade}). {summary.get('tables_with_data', 0)} of "
                f"{summary.get('total_tables', 0)} tables are populated with "
                f"{summary.get('overall_completeness', 0):.1f}% average completeness."
            )
        elif score >= 70:
            assess_variant = "info"
            assess_text = (
                f"<strong>Good:</strong> Overall data quality score is "
                f"{score:.1f}% ({grade}). {summary.get('tables_with_issues', 0)} tables "
                f"need attention. Coverage is {summary.get('coverage_pct', 0):.0f}%."
            )
        else:
            assess_variant = "warn"
            assess_text = (
                f"<strong>Needs Improvement:</strong> Overall data quality score is "
                f"{score:.1f}% ({grade}). {summary.get('tables_empty', 0)} tables are empty "
                f"and {summary.get('tables_with_issues', 0)} have completeness issues."
            )
        body += callout(assess_text, assess_variant)
        body += "\n" + section_end()

        # ── Section 2: Domain Breakdown ──────────────────────────
        body += "\n" + section_start(2, "Domain Breakdown", "domain-breakdown")
        body += f'<p><strong>{len(domains)}</strong> data domains across the platform.</p>'

        domain_rows = []
        for d in domains:
            domain_rows.append([
                f'<strong>{d["name"]}</strong>',
                str(d["tables_total"]),
                str(d["tables_with_data"]),
                str(d["tables_empty"]),
                _fmt_number(d["total_records"]),
                _completeness_badge(d["avg_completeness"]),
            ])
        body += data_table(
            headers=["Domain", "Tables", "With Data", "Empty", "Records", "Completeness"],
            rows=domain_rows,
            numeric_columns={1, 2, 3, 4},
        )

        # Records per domain bar chart
        if domains:
            rec_labels = [d["name"] for d in domains if d["total_records"] > 0]
            rec_values = [float(d["total_records"]) for d in domains if d["total_records"] > 0]
            rec_colors = [
                DOMAIN_COLORS.get(d["key"], GRAY) for d in domains if d["total_records"] > 0
            ]
            if rec_labels:
                rec_config = build_horizontal_bar_config(
                    rec_labels, rec_values, rec_colors, dataset_label="Records",
                )
                rec_json = json.dumps(rec_config)
                rec_height = f"{max(len(rec_labels) * 44 + 40, 140)}px"
                body += chart_container(
                    "domainRecords", rec_json,
                    build_bar_fallback(rec_labels, rec_values),
                    title="Records by Domain", height=rec_height,
                )
                charts_js += chart_init_js("domainRecords", rec_json)

        body += "\n" + section_end()

        # ── Section 3: Table-by-Table Analysis ───────────────────
        body += "\n" + section_start(3, "Table-by-Table Analysis", "table-analysis")

        good_count = sum(1 for t in tables if t["status"] == "Good")
        warn_count = sum(1 for t in tables if t["status"] == "Warning")
        crit_count = sum(1 for t in tables if t["status"] == "Critical")
        empty_count = sum(1 for t in tables if t["status"] == "Empty")

        body += callout(
            f"<strong>Summary:</strong> {good_count} healthy, "
            f"{warn_count} need attention, {crit_count} critical, "
            f"{empty_count} empty.",
            "info" if crit_count == 0 else "warn",
        )

        table_rows = []
        for t in tables:
            last_upd = t.get("last_updated")
            if last_upd:
                try:
                    dt = datetime.fromisoformat(last_upd)
                    last_upd_display = dt.strftime("%Y-%m-%d")
                except Exception:
                    last_upd_display = str(last_upd)[:10]
            else:
                last_upd_display = "-"

            table_rows.append([
                f'<strong>{t["name"]}</strong>',
                t["domain_label"],
                _fmt_number(t["row_count"]),
                str(t["column_count"]),
                _completeness_badge(t["completeness_pct"]),
                last_upd_display,
                _status_pill(t["status"]),
            ])

        body += data_table(
            headers=["Table", "Domain", "Rows", "Columns", "Completeness", "Last Updated", "Status"],
            rows=table_rows,
            numeric_columns={2, 3},
        )
        body += "\n" + section_end()

        # ── Section 4: Column Completeness ───────────────────────
        body += "\n" + section_start(4, "Column Completeness", "column-completeness")

        # Only show tables with completeness < 90%
        problem_tables = [
            t for t in tables
            if t["status"] in ("Critical", "Warning") and t["row_count"] > 0
        ]

        if problem_tables:
            body += (
                f'<p>Showing column-level detail for <strong>{len(problem_tables)}</strong> '
                f'tables with completeness below 90%.</p>'
            )
            for pt in problem_tables:
                tbl_cols = [
                    c for c in col_details
                    if c["table"] == pt["name"] and c["total_count"] > 0
                ]
                tbl_cols.sort(key=lambda c: -c["null_pct"])
                if not tbl_cols:
                    continue

                body += (
                    f'<h3 style="font-size:15px;font-weight:600;color:var(--gray-800);'
                    f'margin:20px 0 4px">{pt["name"]} '
                    f'{_completeness_badge(pt["completeness_pct"])}</h3>'
                )

                col_rows = []
                for c in tbl_cols:
                    null_pct = c["null_pct"]
                    if null_pct > 80:
                        pct_display = f'<span style="color:#991b1b;font-weight:600">{null_pct:.1f}%</span>'
                    elif null_pct > 50:
                        pct_display = f'<span style="color:#92400e;font-weight:600">{null_pct:.1f}%</span>'
                    else:
                        pct_display = f'{null_pct:.1f}%'
                    col_rows.append([
                        c["column"],
                        c["data_type"],
                        pct_display,
                        _fmt_number(c["non_null_count"]),
                        _fmt_number(c["total_count"]),
                    ])
                body += data_table(
                    headers=["Column", "Type", "Null %", "Non-Null", "Total"],
                    rows=col_rows,
                    numeric_columns={2, 3, 4},
                )
        else:
            body += callout(
                "<strong>All tables healthy!</strong> All populated tables have "
                "completeness above 90%. No column-level issues detected.",
                "good",
            )

        body += "\n" + section_end()

        # ── Section 5: Freshness Analysis ────────────────────────
        body += "\n" + section_start(5, "Freshness Analysis", "freshness")

        sources = freshness.get("sources", [])
        stale = freshness.get("stale_tables", [])

        if sources:
            # Bar chart: days since last update
            fresh_labels = [s["source"] for s in sources if s["days_ago"] is not None]
            fresh_values = [float(s["days_ago"]) for s in sources if s["days_ago"] is not None]
            fresh_colors = []
            for v in fresh_values:
                if v < 7:
                    fresh_colors.append(GREEN)
                elif v <= 30:
                    fresh_colors.append(ORANGE)
                else:
                    fresh_colors.append(RED)

            if fresh_labels:
                fresh_config = build_horizontal_bar_config(
                    fresh_labels, fresh_values, fresh_colors,
                    dataset_label="Days Since Last Update",
                )
                fresh_json = json.dumps(fresh_config)
                fresh_height = f"{max(len(fresh_labels) * 44 + 40, 140)}px"
                body += chart_container(
                    "freshnessBar", fresh_json,
                    build_bar_fallback(fresh_labels, fresh_values),
                    title="Days Since Last Successful Ingestion",
                    height=fresh_height,
                )
                charts_js += chart_init_js("freshnessBar", fresh_json)

            if stale:
                stale_names = ", ".join(s["name"] for s in stale[:10])
                more = f" and {len(stale) - 10} more" if len(stale) > 10 else ""
                body += callout(
                    f"<strong>Stale Data:</strong> {len(stale)} tables have not been "
                    f"updated in over 30 days: {stale_names}{more}.",
                    "warn",
                )

            # Source table
            source_rows = []
            for s in sources:
                days_display = f'{s["days_ago"]}d ago' if s["days_ago"] is not None else "-"
                last_display = s.get("last_completed", "-")
                if last_display and last_display != "-":
                    try:
                        last_display = datetime.fromisoformat(last_display).strftime("%Y-%m-%d %H:%M")
                    except Exception:
                        pass
                source_rows.append([
                    f'<strong>{s["source"]}</strong>',
                    str(last_display),
                    days_display,
                    str(s.get("successes", 0)),
                    str(s.get("failures", 0)),
                ])
            body += data_table(
                headers=["Source", "Last Completed", "Days Ago", "Successes", "Failures"],
                rows=source_rows,
                numeric_columns={2, 3, 4},
            )
        else:
            body += callout(
                "<strong>No ingestion history found.</strong> "
                "Run data ingestion pipelines to track freshness.",
                "info",
            )

        body += "\n" + section_end()

        # ── Section 6: Recommendations ───────────────────────────
        body += "\n" + section_start(6, "Recommendations", "recommendations")

        if recommendations:
            for rec in recommendations:
                variant = rec.get("severity", "info")
                body += callout(
                    f'<strong>{rec["title"]}</strong><br>{rec["detail"]}',
                    variant,
                )
        else:
            body += callout(
                "<strong>No recommendations.</strong> All checks passed.",
                "good",
            )

        body += "\n" + section_end()

        body += '\n</div>'  # close container

        # ── Footer ───────────────────────────────────────────────
        notes = [
            "Completeness = average non-null percentage across all columns in each table.",
            "Overall score = 60% completeness + 25% coverage (populated/total) + 15% freshness (updated within 30 days).",
            "Row counts use exact COUNT(*) for tables < 500K rows; pg_stat_user_tables estimates for larger tables.",
            "Freshness determined from timestamp columns (updated_at, created_at, collected_at).",
        ]
        body += "\n" + page_footer(
            notes=notes,
            generated_line=f"Report generated {data.get('generated_at', 'N/A')} | Nexdata Data Quality Assessment",
        )

        extra_css = """
.score-a { color: #166534; }
.score-b { color: #1e40af; }
.score-c { color: #92400e; }
.score-d { color: #991b1b; }
.score-f { color: #991b1b; }
"""

        return html_document(
            title="Data Quality Assessment",
            body_content=body,
            charts_js=charts_js,
            extra_css=extra_css,
        )

    # ── Excel Rendering ───────────────────────────────────────────────

    def render_excel(self, data: Dict[str, Any]) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()

        header_fill = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        title_font = Font(bold=True, size=14)

        # ── Sheet 1: Summary ─────────────────────────────────────
        ws = wb.active
        ws.title = "Summary"
        summary = data.get("summary", {})

        ws["A1"] = "Data Quality Assessment"
        ws["A1"].font = Font(bold=True, size=18)
        ws.merge_cells("A1:D1")
        ws["A2"] = f"Generated: {data.get('generated_at', 'N/A')}"

        ws["A4"] = "Overall Score"
        ws["B4"] = f"{summary.get('overall_score', 0):.1f}% ({summary.get('letter_grade', 'F')})"
        ws["A5"] = "Total Tables"
        ws["B5"] = summary.get("total_tables", 0)
        ws["A6"] = "Tables with Data"
        ws["B6"] = summary.get("tables_with_data", 0)
        ws["A7"] = "Tables Empty"
        ws["B7"] = summary.get("tables_empty", 0)
        ws["A8"] = "Total Records"
        ws["B8"] = summary.get("total_records", 0)
        ws["A9"] = "Overall Completeness"
        ws["B9"] = f"{summary.get('overall_completeness', 0):.1f}%"
        ws["A10"] = "Tables with Issues"
        ws["B10"] = summary.get("tables_with_issues", 0)

        for r in range(4, 11):
            ws[f"A{r}"].font = Font(bold=True)

        # Domain breakdown
        ws["A12"] = "Domain Breakdown"
        ws["A12"].font = title_font
        domain_headers = ["Domain", "Tables", "With Data", "Empty", "Records", "Completeness"]
        for col, h in enumerate(domain_headers, 1):
            cell = ws.cell(row=13, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        for i, d in enumerate(data.get("domains", []), 14):
            ws.cell(row=i, column=1, value=d["name"])
            ws.cell(row=i, column=2, value=d["tables_total"])
            ws.cell(row=i, column=3, value=d["tables_with_data"])
            ws.cell(row=i, column=4, value=d["tables_empty"])
            ws.cell(row=i, column=5, value=d["total_records"])
            ws.cell(row=i, column=6, value=f"{d['avg_completeness']:.1f}%")

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 15

        # ── Sheet 2: Tables ──────────────────────────────────────
        ws2 = wb.create_sheet("Tables")
        tbl_headers = ["Table", "Domain", "Rows", "Columns", "Completeness", "Last Updated", "Status"]
        for col, h in enumerate(tbl_headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        for i, t in enumerate(data.get("tables", []), 2):
            ws2.cell(row=i, column=1, value=t["name"])
            ws2.cell(row=i, column=2, value=t["domain_label"])
            ws2.cell(row=i, column=3, value=t["row_count"])
            ws2.cell(row=i, column=4, value=t["column_count"])
            ws2.cell(row=i, column=5, value=f"{t['completeness_pct']:.1f}%")
            ws2.cell(row=i, column=6, value=t.get("last_updated") or "-")
            ws2.cell(row=i, column=7, value=t["status"])

        ws2.column_dimensions["A"].width = 35
        ws2.column_dimensions["B"].width = 22
        ws2.column_dimensions["C"].width = 12
        ws2.column_dimensions["D"].width = 10
        ws2.column_dimensions["E"].width = 15
        ws2.column_dimensions["F"].width = 18
        ws2.column_dimensions["G"].width = 12

        # ── Sheet 3: Column Details ──────────────────────────────
        ws3 = wb.create_sheet("Column Details")
        col_headers = ["Table", "Column", "Data Type", "Null %", "Non-Null Count", "Total Count"]
        for col, h in enumerate(col_headers, 1):
            cell = ws3.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        for i, c in enumerate(data.get("column_details", []), 2):
            ws3.cell(row=i, column=1, value=c["table"])
            ws3.cell(row=i, column=2, value=c["column"])
            ws3.cell(row=i, column=3, value=c["data_type"])
            ws3.cell(row=i, column=4, value=f"{c['null_pct']:.1f}%")
            ws3.cell(row=i, column=5, value=c["non_null_count"])
            ws3.cell(row=i, column=6, value=c["total_count"])

        ws3.column_dimensions["A"].width = 35
        ws3.column_dimensions["B"].width = 30
        ws3.column_dimensions["C"].width = 20
        ws3.column_dimensions["D"].width = 10
        ws3.column_dimensions["E"].width = 15
        ws3.column_dimensions["F"].width = 15

        # ── Sheet 4: Freshness ───────────────────────────────────
        ws4 = wb.create_sheet("Freshness")
        fresh_headers = ["Source", "Last Completed", "Days Ago", "Successes", "Failures"]
        for col, h in enumerate(fresh_headers, 1):
            cell = ws4.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        for i, s in enumerate(data.get("freshness", {}).get("sources", []), 2):
            ws4.cell(row=i, column=1, value=s["source"])
            ws4.cell(row=i, column=2, value=s.get("last_completed") or "-")
            ws4.cell(row=i, column=3, value=s.get("days_ago"))
            ws4.cell(row=i, column=4, value=s.get("successes", 0))
            ws4.cell(row=i, column=5, value=s.get("failures", 0))

        ws4.column_dimensions["A"].width = 25
        ws4.column_dimensions["B"].width = 22
        ws4.column_dimensions["C"].width = 10
        ws4.column_dimensions["D"].width = 12
        ws4.column_dimensions["E"].width = 12

        # ── Sheet 5: Recommendations ─────────────────────────────
        ws5 = wb.create_sheet("Recommendations")
        rec_headers = ["Severity", "Title", "Detail"]
        for col, h in enumerate(rec_headers, 1):
            cell = ws5.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        for i, r in enumerate(data.get("recommendations", []), 2):
            ws5.cell(row=i, column=1, value=r["severity"])
            ws5.cell(row=i, column=2, value=r["title"])
            ws5.cell(row=i, column=3, value=r["detail"])

        ws5.column_dimensions["A"].width = 12
        ws5.column_dimensions["B"].width = 40
        ws5.column_dimensions["C"].width = 80

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        return output.getvalue()
