"""
Deep Data Quality Report Template — Transparency Rebuild.

15-section report with full scoring methodology disclosure, column-level
profiling, anomaly context (sigma values, baselines), cross-source orphan
examples, schema drift history, programmatic recommendations, and rules engine.
"""

import json as _json
import logging
from datetime import datetime, date, timedelta, timezone
from typing import Dict, Any, List, Optional
from io import BytesIO

from sqlalchemy import func, case
from sqlalchemy.orm import Session

from app.reports.design_system import (
    html_document, page_header, kpi_strip, kpi_card,
    toc, section_start, section_end,
    data_table, pill_badge, callout,
    chart_container, chart_init_js, page_footer,
    build_doughnut_config, build_horizontal_bar_config,
    build_bar_fallback, build_chart_legend,
    GREEN, ORANGE, RED, GRAY, BLUE,
)
from app.core.models import (
    DataProfileSnapshot, DataProfileColumn,
    DQAnomalyAlert, DQAnomalyThreshold,
    DQCrossSourceValidation, DQCrossSourceResult,
    DQQualitySnapshot, DQSLATarget,
    AnomalyAlertStatus, AnomalyAlertType,
    IngestionJob, JobStatus,
    DataQualityRule, DataQualityResult,
    DatasetRegistry,
)
from app.core.domains import classify_table, DOMAIN_LABELS, DOMAIN_COLORS

logger = logging.getLogger(__name__)

# ── Constants ────────────────────────────────────────────────────────────────

GRADE_THRESHOLDS = [
    (97, "A+"), (93, "A"), (90, "A-"),
    (87, "B+"), (83, "B"), (80, "B-"),
    (77, "C+"), (73, "C"), (70, "C-"),
    (60, "D"), (0, "F"),
]

DIMENSION_WEIGHTS = {
    "completeness": 0.30,
    "freshness": 0.20,
    "validity": 0.30,
    "consistency": 0.20,
}

SCHEMA_DRIFT_TYPES = {
    AnomalyAlertType.NEW_COLUMN,
    AnomalyAlertType.DROPPED_COLUMN,
    AnomalyAlertType.TYPE_CHANGE,
}

EXTRA_CSS = """
.score-bar { display:inline-block; height:8px; border-radius:4px; background:var(--primary-light); vertical-align:middle; margin-left:6px; }
.weight-row { display:flex; align-items:center; gap:12px; padding:6px 0; font-size:13px; }
.weight-label { min-width:160px; font-weight:500; color:var(--gray-700); }
.weight-bar-track { height:8px; width:200px; background:var(--gray-100); border-radius:4px; overflow:hidden; }
.weight-bar-fill { height:100%; background:var(--primary-light); border-radius:4px; }
.weight-pct { font-weight:600; color:var(--gray-900); min-width:40px; }
.sigma-badge { display:inline-block; padding:2px 8px; border-radius:10px; font-size:11px; font-weight:600; }
.sigma-high { background:#fef2f2; color:#991b1b; }
.sigma-medium { background:#fef3c7; color:#92400e; }
.sigma-low { background:#f0fdf4; color:#166534; }
"""

# ── Helpers ──────────────────────────────────────────────────────────────────


def _letter_grade(score: float) -> str:
    for threshold, grade in GRADE_THRESHOLDS:
        if score >= threshold:
            return grade
    return "F"


def _fmt_number(n) -> str:
    if n is None:
        return "0"
    return f"{int(n):,}"


def _fmt_float(n, decimals: int = 1) -> str:
    if n is None:
        return "-"
    return f"{n:.{decimals}f}"


def _score_color(score: Optional[float]) -> str:
    if score is None:
        return GRAY
    if score >= 90:
        return GREEN
    if score >= 70:
        return ORANGE
    return RED


def _score_html(score: Optional[float], show_bar: bool = True) -> str:
    """Score value with optional inline bar."""
    if score is None:
        return '<span style="color:#a0aec0">N/A</span>'
    color = _score_color(score)
    bar = ""
    if show_bar:
        width = min(score, 100)
        bar = f'<span class="score-bar" style="width:{width:.0f}px;background:{color}"></span>'
    return f'<span style="color:{color};font-weight:600">{score:.1f}</span>{bar}'


def _grade_pill(score: Optional[float]) -> str:
    if score is None:
        return pill_badge("N/A", "default")
    grade = _letter_grade(score)
    if score >= 90:
        variant = "private"  # green
    elif score >= 70:
        variant = "pe"  # amber
    else:
        variant = "default"  # gray
    return pill_badge(grade, variant)


def _status_pill(status: str) -> str:
    if status in ("Good", "passed", "PASSED"):
        return pill_badge("Good", "private")
    elif status in ("Warning", "warning"):
        return pill_badge("Warning", "pe")
    elif status in ("Critical", "failed", "FAILED"):
        return pill_badge("Critical", "default")
    return pill_badge(status, "default")


def _severity_pill(sev: str) -> str:
    sev_lower = sev.lower() if sev else ""
    if sev_lower == "info":
        return pill_badge("INFO", "public")
    elif sev_lower == "warning":
        return pill_badge("WARNING", "pe")
    elif sev_lower in ("critical", "error"):
        return pill_badge("CRITICAL", "default")
    return pill_badge(sev.upper() if sev else "?", "default")


def _sigma_badge(sigma: Optional[float]) -> str:
    if sigma is None:
        return "-"
    if sigma > 4:
        cls = "sigma-high"
    elif sigma > 2:
        cls = "sigma-medium"
    else:
        cls = "sigma-low"
    return f'<span class="sigma-badge {cls}">{sigma:.1f}&sigma;</span>'


def _days_ago_html(days: Optional[float]) -> str:
    if days is None:
        return '<span style="color:#a0aec0">Never</span>'
    d = int(days)
    if d <= 7:
        color = GREEN
    elif d <= 30:
        color = ORANGE
    else:
        color = RED
    return f'<span style="color:{color};font-weight:600">{d}d</span>'


def _delta_html(delta: Optional[float]) -> str:
    if delta is None:
        return "-"
    if delta > 0:
        return f'<span style="color:{GREEN};font-weight:600">+{delta:.1f}</span>'
    elif delta < 0:
        return f'<span style="color:{RED};font-weight:600">{delta:.1f}</span>'
    return f'<span style="color:{GRAY}">0.0</span>'


# ── Template Class ───────────────────────────────────────────────────────────


class DataQualityDeepTemplate:
    """Deep data quality report with profiling, anomalies, cross-source validation, and trending."""

    name = "data_quality_deep"
    description = "Deep data quality analysis with profiling, anomalies, cross-source validation, and trending"

    # ── gather_data ──────────────────────────────────────────────────────

    def gather_data(self, db: Session, params: Dict[str, Any]) -> Dict[str, Any]:
        """Gather all deep quality data."""
        now = datetime.now(timezone.utc)
        today = date.today()
        window_days = params.get("window_days", 30)
        cutoff = today - timedelta(days=window_days)

        data: Dict[str, Any] = {
            "generated_at": now.isoformat(),
            "window_days": window_days,
        }

        # ── 1. Quality snapshots (today or yesterday) ────────────────
        quality_snapshots = (
            db.query(DQQualitySnapshot)
            .filter(DQQualitySnapshot.snapshot_date == today)
            .all()
        )
        if not quality_snapshots:
            quality_snapshots = (
                db.query(DQQualitySnapshot)
                .filter(DQQualitySnapshot.snapshot_date == today - timedelta(days=1))
                .all()
            )
        data["quality_snapshots"] = quality_snapshots

        scores = [s.quality_score for s in quality_snapshots if s.quality_score is not None]
        platform_score = sum(scores) / len(scores) if scores else None
        data["platform_score"] = round(platform_score, 1) if platform_score else None
        data["platform_grade"] = _letter_grade(platform_score) if platform_score else "N/A"
        data["tables_tracked"] = len(quality_snapshots)

        # Dimension averages
        for dim in ["completeness", "freshness", "validity", "consistency"]:
            vals = [getattr(s, f"{dim}_score") for s in quality_snapshots
                    if getattr(s, f"{dim}_score") is not None]
            data[f"avg_{dim}"] = round(sum(vals) / len(vals), 1) if vals else None

        # Health buckets
        good = sum(1 for s in quality_snapshots if (s.quality_score or 0) >= 90)
        warning = sum(1 for s in quality_snapshots if 70 <= (s.quality_score or 0) < 90)
        critical = sum(1 for s in quality_snapshots if (s.quality_score or 0) < 70)
        data["health_good"] = good
        data["health_warning"] = warning
        data["health_critical"] = critical

        # Week-over-week delta
        prior_day = (
            db.query(func.avg(DQQualitySnapshot.quality_score))
            .filter(DQQualitySnapshot.snapshot_date == today - timedelta(days=7))
            .scalar()
        )
        data["wow_delta"] = round(platform_score - prior_day, 1) if platform_score and prior_day else None

        # ── 2. Domain breakdown ──────────────────────────────────────
        domains: Dict[str, List] = {}
        for s in quality_snapshots:
            domain_key = classify_table(s.table_name)
            domains.setdefault(domain_key, []).append(s)

        domain_data = []
        for domain_key, snaps in sorted(domains.items()):
            d_scores = [s.quality_score for s in snaps if s.quality_score is not None]
            c_scores = [s.completeness_score for s in snaps if s.completeness_score is not None]
            f_scores = [s.freshness_score for s in snaps if s.freshness_score is not None]
            v_scores = [s.validity_score for s in snaps if s.validity_score is not None]
            domain_data.append({
                "key": domain_key,
                "label": DOMAIN_LABELS.get(domain_key, domain_key.replace("_", " ").title()),
                "color": DOMAIN_COLORS.get(domain_key, GRAY),
                "tables": len(snaps),
                "avg_score": round(sum(d_scores) / len(d_scores), 1) if d_scores else None,
                "avg_completeness": round(sum(c_scores) / len(c_scores), 1) if c_scores else None,
                "avg_freshness": round(sum(f_scores) / len(f_scores), 1) if f_scores else None,
                "avg_validity": round(sum(v_scores) / len(v_scores), 1) if v_scores else None,
                "total_rows": sum(s.row_count or 0 for s in snaps),
            })
        data["domains"] = domain_data

        # ── 3. Column-level profiles ─────────────────────────────────
        recent_profiles = (
            db.query(DataProfileSnapshot)
            .order_by(DataProfileSnapshot.profiled_at.desc())
            .limit(20)
            .all()
        )
        data["recent_profiles"] = recent_profiles

        snapshot_ids = [p.id for p in recent_profiles]
        column_profiles = []
        if snapshot_ids:
            column_profiles = (
                db.query(DataProfileColumn)
                .filter(DataProfileColumn.snapshot_id.in_(snapshot_ids))
                .all()
            )
        data["column_profiles"] = column_profiles

        # Map snapshot_id → table_name
        snap_table_map = {p.id: p.table_name for p in recent_profiles}
        data["snap_table_map"] = snap_table_map

        # ── 4. Anomaly alerts (full fields, last 7 days) ─────────────
        open_anomalies = (
            db.query(DQAnomalyAlert)
            .filter(DQAnomalyAlert.status == AnomalyAlertStatus.OPEN)
            .count()
        )
        recent_anomalies = (
            db.query(DQAnomalyAlert)
            .filter(DQAnomalyAlert.detected_at >= now - timedelta(days=7))
            .order_by(DQAnomalyAlert.detected_at.desc())
            .limit(50)
            .all()
        )
        data["open_anomalies"] = open_anomalies
        data["recent_anomalies"] = recent_anomalies

        # Counts by severity
        data["anomaly_critical"] = sum(1 for a in recent_anomalies if a.severity and a.severity.value == "critical")
        data["anomaly_warning"] = sum(1 for a in recent_anomalies if a.severity and a.severity.value == "warning")
        data["anomaly_info"] = sum(1 for a in recent_anomalies if a.severity and a.severity.value == "info")

        # ── 5. Anomaly thresholds ────────────────────────────────────
        anomaly_thresholds = db.query(DQAnomalyThreshold).all()
        data["anomaly_thresholds"] = anomaly_thresholds

        # ── 6. Cross-source validations + results ────────────────────
        validations = db.query(DQCrossSourceValidation).filter(
            DQCrossSourceValidation.is_enabled == 1
        ).all()

        cross_source_data = []
        for v in validations:
            latest_result = (
                db.query(DQCrossSourceResult)
                .filter(DQCrossSourceResult.validation_id == v.id)
                .order_by(DQCrossSourceResult.evaluated_at.desc())
                .first()
            )
            cross_source_data.append({
                "validation": v,
                "result": latest_result,
            })
        data["cross_source"] = cross_source_data

        # ── 7. Source freshness (ingestion_jobs) ─────────────────────
        try:
            freshness_rows = (
                db.query(
                    IngestionJob.source,
                    func.max(IngestionJob.completed_at).label("last_completed"),
                    func.count(case(
                        (IngestionJob.status == JobStatus.SUCCESS, 1),
                    )).label("successes"),
                    func.count(case(
                        (IngestionJob.status == JobStatus.FAILED, 1),
                    )).label("failures"),
                )
                .filter(IngestionJob.completed_at.isnot(None))
                .group_by(IngestionJob.source)
                .order_by(func.max(IngestionJob.completed_at).desc())
                .all()
            )
        except Exception:
            freshness_rows = []
        data["source_freshness"] = freshness_rows

        # ── 8. Quality trending (daily) ──────────────────────────────
        trend_data = (
            db.query(
                DQQualitySnapshot.snapshot_date,
                func.avg(DQQualitySnapshot.quality_score).label("avg_quality"),
                func.avg(DQQualitySnapshot.completeness_score).label("avg_completeness"),
                func.avg(DQQualitySnapshot.freshness_score).label("avg_freshness"),
                func.avg(DQQualitySnapshot.validity_score).label("avg_validity"),
                func.avg(DQQualitySnapshot.consistency_score).label("avg_consistency"),
                func.count(DQQualitySnapshot.id).label("count"),
            )
            .filter(DQQualitySnapshot.snapshot_date >= cutoff)
            .group_by(DQQualitySnapshot.snapshot_date)
            .order_by(DQQualitySnapshot.snapshot_date.asc())
            .all()
        )
        data["trend"] = trend_data

        # ── 9. Weekly rollups ────────────────────────────────────────
        try:
            weekly_data = (
                db.query(
                    func.date_trunc("week", DQQualitySnapshot.snapshot_date).label("week"),
                    func.avg(DQQualitySnapshot.quality_score).label("avg_quality"),
                    func.avg(DQQualitySnapshot.completeness_score).label("avg_completeness"),
                    func.avg(DQQualitySnapshot.freshness_score).label("avg_freshness"),
                    func.avg(DQQualitySnapshot.validity_score).label("avg_validity"),
                    func.avg(DQQualitySnapshot.consistency_score).label("avg_consistency"),
                    func.count(DQQualitySnapshot.id).label("count"),
                )
                .filter(DQQualitySnapshot.snapshot_date >= cutoff)
                .group_by(func.date_trunc("week", DQQualitySnapshot.snapshot_date))
                .order_by(func.date_trunc("week", DQQualitySnapshot.snapshot_date).asc())
                .all()
            )
        except Exception:
            weekly_data = []
        data["weekly_rollup"] = weekly_data

        # ── 10. SLA targets ──────────────────────────────────────────
        targets = db.query(DQSLATarget).filter(DQSLATarget.is_enabled == 1).all()
        data["sla_targets"] = targets

        sla_results = []
        for target in targets:
            matching = [s for s in quality_snapshots if not target.source or s.source == target.source]
            for s in matching:
                for dim, attr, tgt_attr in [
                    ("Quality", "quality_score", "target_quality_score"),
                    ("Completeness", "completeness_score", "target_completeness"),
                    ("Freshness", "freshness_score", "target_freshness"),
                    ("Validity", "validity_score", "target_validity"),
                ]:
                    score_val = getattr(s, attr) or 0
                    target_val = getattr(target, tgt_attr) or 0
                    sla_results.append({
                        "source": s.source,
                        "table": s.table_name,
                        "dimension": dim,
                        "score": score_val,
                        "target": target_val,
                        "gap": round(score_val - target_val, 1),
                        "met": score_val >= target_val,
                    })
        data["sla_results"] = sla_results

        # ── 11. Schema drift alerts ──────────────────────────────────
        schema_alerts = (
            db.query(DQAnomalyAlert)
            .filter(DQAnomalyAlert.alert_type.in_([
                AnomalyAlertType.NEW_COLUMN,
                AnomalyAlertType.DROPPED_COLUMN,
                AnomalyAlertType.TYPE_CHANGE,
            ]))
            .order_by(DQAnomalyAlert.detected_at.desc())
            .limit(50)
            .all()
        )
        data["schema_alerts"] = schema_alerts

        # ── 12. Rules engine data ────────────────────────────────────
        all_rules = db.query(DataQualityRule).all()
        data["rules_all"] = all_rules
        data["rules_total"] = len(all_rules)
        data["rules_enabled"] = sum(1 for r in all_rules if r.is_enabled)
        data["rules_auto"] = sum(1 for r in all_rules if (r.name or "").startswith("auto_"))
        data["rules_manual"] = data["rules_total"] - data["rules_auto"]

        # Rules by type
        rules_by_type: Dict[str, int] = {}
        for r in all_rules:
            rt = r.rule_type.value if r.rule_type else "unknown"
            rules_by_type[rt] = rules_by_type.get(rt, 0) + 1
        data["rules_by_type"] = rules_by_type

        # Recent results (last 7 days)
        results_cutoff = now - timedelta(days=7)
        recent_results = (
            db.query(DataQualityResult)
            .filter(DataQualityResult.evaluated_at >= results_cutoff)
            .all()
        )
        data["rules_results_7d"] = recent_results
        data["rules_pass_count_7d"] = sum(1 for r in recent_results if r.passed)
        data["rules_fail_count_7d"] = sum(1 for r in recent_results if not r.passed)
        total_7d = len(recent_results)
        data["rules_pass_rate_7d"] = (
            round(data["rules_pass_count_7d"] / total_7d * 100, 1) if total_7d > 0 else None
        )

        # Top failing rules
        top_failing = sorted(
            [r for r in all_rules if r.times_failed > 0],
            key=lambda r: r.times_failed,
            reverse=True,
        )[:10]
        data["rules_top_failing"] = top_failing

        # Rule coverage: tables with rules vs without
        all_registered = db.query(DatasetRegistry.table_name).all()
        all_table_names_set = {t[0] for t in all_registered}
        tables_with_rules = set()
        for r in all_rules:
            if r.dataset_pattern:
                try:
                    import re as _re
                    pat = _re.compile(r.dataset_pattern)
                    for tn in all_table_names_set:
                        if pat.match(tn):
                            tables_with_rules.add(tn)
                except Exception:
                    pass
        data["rules_tables_covered"] = len(tables_with_rules)
        data["rules_tables_uncovered"] = len(all_table_names_set) - len(tables_with_rules)
        data["rules_coverage_pct"] = (
            round(len(tables_with_rules) / len(all_table_names_set) * 100, 1)
            if all_table_names_set else 0
        )

        return data

    # ── render_html ──────────────────────────────────────────────────────

    def render_html(self, data: Dict[str, Any]) -> str:
        """Render the deep quality report as HTML."""
        charts_js = ""
        body = ""

        # ── Page Header ──────────────────────────────────────────────
        body += page_header(
            title="Deep Data Quality Report",
            subtitle=f"{data.get('tables_tracked', 0)} tables tracked | "
                     f"{data.get('window_days', 30)}-day analysis window",
        )

        # ── TOC ──────────────────────────────────────────────────────
        toc_items = [
            {"number": "1", "id": "exec-summary", "title": "Executive Summary & Health Dashboard"},
            {"number": "2", "id": "methodology", "title": "Scoring Methodology"},
            {"number": "3", "id": "table-scoreboard", "title": "Table Health Scoreboard"},
            {"number": "4", "id": "domain-breakdown", "title": "Domain Quality Breakdown"},
            {"number": "5", "id": "column-profiling", "title": "Column-Level Profiling Deep Dive"},
            {"number": "6", "id": "distributions", "title": "Data Distribution Samples"},
            {"number": "7", "id": "anomaly-alerts", "title": "Anomaly Alerts with Full Context"},
            {"number": "8", "id": "anomaly-config", "title": "Anomaly Detection Configuration"},
            {"number": "9", "id": "cross-source", "title": "Cross-Source Validation Detail"},
            {"number": "10", "id": "freshness", "title": "Data Freshness by Source"},
            {"number": "11", "id": "trending", "title": "Quality Trending & Weekly Rollups"},
            {"number": "12", "id": "sla-compliance", "title": "SLA Compliance Detail"},
            {"number": "13", "id": "schema-drift", "title": "Schema Drift History"},
            {"number": "14", "id": "recommendations", "title": "Recommendations & Action Items"},
            {"number": "15", "id": "rules-engine", "title": "Data Quality Rules Engine"},
        ]
        body += toc(toc_items)

        # ── Section 1: Executive Summary & Health Dashboard ──────────
        body += self._render_section_1(data)
        charts_js += self._charts_section_1(data)

        # ── Section 2: Scoring Methodology ───────────────────────────
        body += self._render_section_2(data)

        # ── Section 3: Table Health Scoreboard ───────────────────────
        body += self._render_section_3(data)

        # ── Section 4: Domain Quality Breakdown ──────────────────────
        body += self._render_section_4(data)
        charts_js += self._charts_section_4(data)

        # ── Section 5: Column-Level Profiling ────────────────────────
        body += self._render_section_5(data)

        # ── Section 6: Data Distribution Samples ─────────────────────
        body += self._render_section_6(data)

        # ── Section 7: Anomaly Alerts with Full Context ──────────────
        body += self._render_section_7(data)
        charts_js += self._charts_section_7(data)

        # ── Section 8: Anomaly Detection Configuration ───────────────
        body += self._render_section_8(data)

        # ── Section 9: Cross-Source Validation Detail ────────────────
        body += self._render_section_9(data)

        # ── Section 10: Data Freshness by Source ─────────────────────
        body += self._render_section_10(data)
        charts_js += self._charts_section_10(data)

        # ── Section 11: Quality Trending & Weekly Rollups ────────────
        body += self._render_section_11(data)
        charts_js += self._charts_section_11(data)

        # ── Section 12: SLA Compliance Detail ────────────────────────
        body += self._render_section_12(data)

        # ── Section 13: Schema Drift History ─────────────────────────
        body += self._render_section_13(data)

        # ── Section 14: Recommendations & Action Items ───────────────
        body += self._render_section_14(data)

        # ── Section 15: Data Quality Rules Engine ─────────────────────
        body += self._render_section_15(data)
        charts_js += self._charts_section_15(data)

        body += page_footer(
            notes=[
                "Quality scores are computed daily from profiling snapshots.",
                "Anomalies are flagged when a metric deviates beyond the configured sigma threshold.",
                "Cross-source validations check record-level consistency between tables.",
                "All data sourced from publicly available APIs and datasets.",
            ],
            generated_line=f"Generated: {data.get('generated_at', '')} | Nexdata Deep Quality Engine",
        )

        return html_document(
            title="Deep Data Quality Report",
            body_content=body,
            charts_js=charts_js,
            extra_css=EXTRA_CSS,
        )

    # ── Section 1: Executive Summary & Health Dashboard ──────────────

    def _render_section_1(self, data: Dict[str, Any]) -> str:
        s = ""
        s += section_start(1, "Executive Summary & Health Dashboard", "exec-summary")

        platform_score = data.get("platform_score")
        platform_grade = data.get("platform_grade", "N/A")
        wow_delta = data.get("wow_delta")

        delta_str = f"Grade: {platform_grade}"
        delta_dir = "neutral"
        if wow_delta is not None:
            sign = "+" if wow_delta > 0 else ""
            delta_str = f"{platform_grade} | {sign}{wow_delta:.1f} vs last week"
            delta_dir = "up" if wow_delta > 0 else ("down" if wow_delta < 0 else "neutral")

        kpis = [
            kpi_card(
                "Quality Score",
                f"{platform_score:.1f}" if platform_score else "N/A",
                delta=delta_str,
                delta_dir=delta_dir,
            ),
            kpi_card("Completeness", _fmt_float(data.get("avg_completeness")),
                     delta="30% weight"),
            kpi_card("Freshness", _fmt_float(data.get("avg_freshness")),
                     delta="20% weight"),
            kpi_card("Validity", _fmt_float(data.get("avg_validity")),
                     delta="30% weight"),
            kpi_card("Consistency", _fmt_float(data.get("avg_consistency")),
                     delta="20% weight"),
        ]
        s += kpi_strip(kpis)

        # Health distribution doughnut + dimension bars (chart row)
        good = data.get("health_good", 0)
        warn = data.get("health_warning", 0)
        crit = data.get("health_critical", 0)

        health_config = build_doughnut_config(
            [f"Good ({good})", f"Warning ({warn})", f"Critical ({crit})"],
            [good, warn, crit],
            [GREEN, ORANGE, RED],
        )
        health_json = _json.dumps(health_config)

        dims = ["Completeness", "Freshness", "Validity", "Consistency"]
        dim_vals = [
            data.get("avg_completeness") or 0,
            data.get("avg_freshness") or 0,
            data.get("avg_validity") or 0,
            data.get("avg_consistency") or 0,
        ]
        dim_bar_config = build_horizontal_bar_config(
            dims, dim_vals,
            colors=[BLUE, ORANGE, "#8B5CF6", GREEN],
            dataset_label="Score",
        )
        dim_bar_config["options"]["scales"]["x"] = {"min": 0, "max": 100}
        dim_bar_json = _json.dumps(dim_bar_config)

        s += '<div class="chart-row">'
        s += '<div>'
        s += chart_container("healthDoughnut", health_json,
                             fallback_html=build_chart_legend(
                                 ["Good", "Warning", "Critical"],
                                 [good, warn, crit],
                                 [GREEN, ORANGE, RED]),
                             size="medium", title="Table Health Distribution")
        s += '</div><div>'
        s += chart_container("dimBarChart", dim_bar_json,
                             fallback_html=build_bar_fallback(dims, dim_vals),
                             size="medium", title="Dimension Scores")
        s += '</div></div>'

        # Assessment callout
        if platform_score and platform_score >= 90:
            s += callout(
                f"<strong>Overall Assessment: Good.</strong> Platform quality score is "
                f"{platform_score:.1f} ({platform_grade}). "
                f"{good} of {data.get('tables_tracked', 0)} tables are in healthy state.",
                "good",
            )
        elif platform_score and platform_score >= 70:
            s += callout(
                f"<strong>Overall Assessment: Needs Attention.</strong> Platform quality score is "
                f"{platform_score:.1f} ({platform_grade}). "
                f"{crit} table(s) are in critical state and need investigation.",
                "warn",
            )
        else:
            s += callout(
                f"<strong>Overall Assessment: Action Required.</strong> Platform quality score is "
                f"{_fmt_float(platform_score)} ({platform_grade}). "
                f"Multiple tables need immediate attention.",
                "warn",
            )

        s += section_end()
        return s

    def _charts_section_1(self, data: Dict[str, Any]) -> str:
        js = ""
        good = data.get("health_good", 0)
        warn = data.get("health_warning", 0)
        crit = data.get("health_critical", 0)
        health_config = build_doughnut_config(
            [f"Good ({good})", f"Warning ({warn})", f"Critical ({crit})"],
            [good, warn, crit], [GREEN, ORANGE, RED])
        js += chart_init_js("healthDoughnut", _json.dumps(health_config))

        dims = ["Completeness", "Freshness", "Validity", "Consistency"]
        dim_vals = [data.get("avg_completeness") or 0, data.get("avg_freshness") or 0,
                    data.get("avg_validity") or 0, data.get("avg_consistency") or 0]
        dim_bar_config = build_horizontal_bar_config(
            dims, dim_vals, colors=[BLUE, ORANGE, "#8B5CF6", GREEN], dataset_label="Score")
        dim_bar_config["options"]["scales"]["x"] = {"min": 0, "max": 100}
        js += chart_init_js("dimBarChart", _json.dumps(dim_bar_config))
        return js

    # ── Section 2: Scoring Methodology ───────────────────────────────

    def _render_section_2(self, data: Dict[str, Any]) -> str:
        s = ""
        s += section_start(2, "Scoring Methodology", "methodology")

        s += callout(
            "<strong>Transparency Notice:</strong> This section explains exactly how every "
            "score in this report is computed. All thresholds, weights, and detection "
            "parameters are disclosed below.",
            "info",
        )

        # Weight visualization
        s += '<h3 style="font-size:15px;font-weight:600;color:var(--gray-900);margin:16px 0 8px">Dimension Weights</h3>'
        for dim_name, weight in DIMENSION_WEIGHTS.items():
            pct = int(weight * 100)
            bar_width = int(weight * 100 * 2)  # scale to 200px max
            s += f"""<div class="weight-row">
    <span class="weight-label">{dim_name.title()}</span>
    <div class="weight-bar-track"><div class="weight-bar-fill" style="width:{bar_width}px"></div></div>
    <span class="weight-pct">{pct}%</span>
</div>"""

        s += '<p style="margin-top:8px;font-size:13px;color:var(--gray-500)">Quality Score = (Completeness &times; 0.30) + (Freshness &times; 0.20) + (Validity &times; 0.30) + (Consistency &times; 0.20)</p>'

        # Grade thresholds table
        s += '<h3 style="font-size:15px;font-weight:600;color:var(--gray-900);margin:20px 0 8px">Grade Thresholds</h3>'
        grade_headers = ["Grade", "Score Range"]
        grade_rows = [
            ["A+", "97 - 100"], ["A", "93 - 96.9"], ["A-", "90 - 92.9"],
            ["B+", "87 - 89.9"], ["B", "83 - 86.9"], ["B-", "80 - 82.9"],
            ["C+", "77 - 79.9"], ["C", "73 - 76.9"], ["C-", "70 - 72.9"],
            ["D", "60 - 69.9"], ["F", "0 - 59.9"],
        ]
        s += data_table(grade_headers, grade_rows)

        # Anomaly detection disclosure
        s += '<h3 style="font-size:15px;font-weight:600;color:var(--gray-900);margin:20px 0 8px">Anomaly Detection</h3>'
        s += '<p>Anomalies are flagged when a metric exceeds <strong>N sigma</strong> from its rolling 30-day mean. Higher sigma thresholds = fewer (but more significant) alerts.</p>'

        thresholds = data.get("anomaly_thresholds", [])
        global_thresholds = [t for t in thresholds if t.source is None and t.table_pattern is None]
        if global_thresholds:
            gt = global_thresholds[0]
            s += '<p style="font-size:13px;color:var(--gray-700)"><strong>Global defaults:</strong> '
            s += f'Row Count: {gt.row_count_sigma}&sigma; | '
            s += f'Null Rate: {gt.null_rate_sigma}&sigma; | '
            s += f'Distribution: {gt.distribution_sigma}&sigma;</p>'

        # Freshness decay
        s += '<h3 style="font-size:15px;font-weight:600;color:var(--gray-900);margin:20px 0 8px">Freshness Scoring</h3>'
        s += '<p>Freshness uses linear decay: <strong>100</strong> at 0 hours &rarr; <strong>0</strong> at 168 hours (1 week). Tables not refreshed in 7+ days score 0 on freshness.</p>'

        # SLA disclosure
        sla_targets = data.get("sla_targets", [])
        s += '<h3 style="font-size:15px;font-weight:600;color:var(--gray-900);margin:20px 0 8px">SLA Targets</h3>'
        s += f'<p><strong>{len(sla_targets)}</strong> SLA targets configured. '
        if sla_targets:
            s += f'Degradation alert triggers after <strong>{sla_targets[0].consecutive_drops_threshold}</strong> consecutive drops below target.</p>'
        else:
            s += 'No targets configured yet.</p>'

        s += section_end()
        return s

    # ── Section 3: Table Health Scoreboard ────────────────────────────

    def _render_section_3(self, data: Dict[str, Any]) -> str:
        s = ""
        s += section_start(3, "Table Health Scoreboard", "table-scoreboard")

        quality_snapshots = data.get("quality_snapshots", [])
        if not quality_snapshots:
            s += callout("No quality snapshots available. Run <code>/data-quality/trends/compute</code> first.", "info")
            s += section_end()
            return s

        # Sort worst-first
        sorted_snaps = sorted(quality_snapshots, key=lambda x: x.quality_score or 0)

        headers = ["Table", "Source", "Domain", "Quality", "Grade",
                    "Completeness", "Freshness", "Validity", "Consistency",
                    "Rows", "Anomalies"]
        rows = []

        # Count anomalies per table
        anomaly_counts: Dict[str, int] = {}
        for a in data.get("recent_anomalies", []):
            tbl = a.table_name if hasattr(a, "table_name") else ""
            anomaly_counts[tbl] = anomaly_counts.get(tbl, 0) + 1

        for snap in sorted_snaps:
            domain_key = classify_table(snap.table_name)
            domain_label = DOMAIN_LABELS.get(domain_key, domain_key)
            domain_color = DOMAIN_COLORS.get(domain_key, GRAY)
            anom_count = anomaly_counts.get(snap.table_name, 0)
            anom_str = f'<span style="color:{RED};font-weight:600">{anom_count}</span>' if anom_count > 0 else "0"

            rows.append([
                f"<strong>{snap.table_name}</strong>",
                snap.source or "-",
                f'<span style="color:{domain_color}">{domain_label}</span>',
                _score_html(snap.quality_score),
                _grade_pill(snap.quality_score),
                _score_html(snap.completeness_score, show_bar=False),
                _score_html(snap.freshness_score, show_bar=False),
                _score_html(snap.validity_score, show_bar=False),
                _score_html(snap.consistency_score, show_bar=False),
                _fmt_number(snap.row_count),
                anom_str,
            ])

        # Footer with averages
        avg_q = data.get("platform_score")
        avg_c = data.get("avg_completeness")
        avg_f = data.get("avg_freshness")
        avg_v = data.get("avg_validity")
        avg_con = data.get("avg_consistency")
        total_rows = sum(s.row_count or 0 for s in quality_snapshots)
        total_anom = sum(anomaly_counts.values())
        footer = [
            f"<strong>Average ({len(quality_snapshots)} tables)</strong>",
            "", "",
            _fmt_float(avg_q), "",
            _fmt_float(avg_c), _fmt_float(avg_f), _fmt_float(avg_v), _fmt_float(avg_con),
            _fmt_number(total_rows), str(total_anom),
        ]

        s += data_table(headers, rows, numeric_columns={5, 6, 7, 8, 9, 10}, footer_row=footer)
        s += section_end()
        return s

    # ── Section 4: Domain Quality Breakdown ──────────────────────────

    def _render_section_4(self, data: Dict[str, Any]) -> str:
        s = ""
        s += section_start(4, "Domain Quality Breakdown", "domain-breakdown")

        domains = data.get("domains", [])
        if not domains:
            s += callout("No domain data available.", "info")
            s += section_end()
            return s

        # Charts row
        d_labels = [d["label"] for d in domains]
        d_tables = [float(d["tables"]) for d in domains]
        d_colors = [d["color"] for d in domains]
        d_scores = [d["avg_score"] or 0 for d in domains]

        doughnut_config = build_doughnut_config(d_labels, d_tables, d_colors)
        doughnut_json = _json.dumps(doughnut_config)

        bar_config = build_horizontal_bar_config(
            d_labels, d_scores, d_colors, dataset_label="Avg Quality Score")
        bar_config["options"]["scales"]["x"] = {"min": 0, "max": 100}
        bar_json = _json.dumps(bar_config)

        bar_height = f"{max(len(d_labels) * 48 + 40, 200)}px"

        s += '<div class="chart-row">'
        s += '<div>'
        s += chart_container("domainDoughnut", doughnut_json,
                             fallback_html=build_chart_legend(d_labels, d_tables, d_colors),
                             size="medium", title="Tables per Domain")
        s += '</div><div>'
        s += chart_container("domainBarChart", bar_json,
                             fallback_html=build_bar_fallback(d_labels, d_scores),
                             title="Avg Quality Score by Domain", height=bar_height)
        s += '</div></div>'

        # Domain table
        headers = ["Domain", "Tables", "Avg Quality", "Completeness", "Freshness", "Validity", "Total Rows"]
        rows = []
        for d in sorted(domains, key=lambda x: -(x.get("avg_score") or 0)):
            color = d["color"]
            rows.append([
                f'<span style="display:inline-block;width:10px;height:10px;border-radius:3px;background:{color};margin-right:6px;vertical-align:middle"></span>'
                f'<strong>{d["label"]}</strong>',
                str(d["tables"]),
                _score_html(d["avg_score"]),
                _fmt_float(d.get("avg_completeness")),
                _fmt_float(d.get("avg_freshness")),
                _fmt_float(d.get("avg_validity")),
                _fmt_number(d.get("total_rows")),
            ])
        s += data_table(headers, rows, numeric_columns={1, 3, 4, 5, 6})

        s += section_end()
        return s

    def _charts_section_4(self, data: Dict[str, Any]) -> str:
        js = ""
        domains = data.get("domains", [])
        if not domains:
            return js
        d_labels = [d["label"] for d in domains]
        d_tables = [float(d["tables"]) for d in domains]
        d_colors = [d["color"] for d in domains]
        d_scores = [d["avg_score"] or 0 for d in domains]

        doughnut_config = build_doughnut_config(d_labels, d_tables, d_colors)
        js += chart_init_js("domainDoughnut", _json.dumps(doughnut_config))

        bar_config = build_horizontal_bar_config(d_labels, d_scores, d_colors, dataset_label="Avg Quality Score")
        bar_config["options"]["scales"]["x"] = {"min": 0, "max": 100}
        js += chart_init_js("domainBarChart", _json.dumps(bar_config))
        return js

    # ── Section 5: Column-Level Profiling Deep Dive ──────────────────

    def _render_section_5(self, data: Dict[str, Any]) -> str:
        s = ""
        s += section_start(5, "Column-Level Profiling Deep Dive", "column-profiling")

        column_profiles: List = data.get("column_profiles", [])
        snap_table_map: Dict = data.get("snap_table_map", {})

        if not column_profiles:
            s += callout("No column profiles available. Run <code>POST /data-quality/profile/all</code> to profile tables.", "info")
            s += section_end()
            return s

        # KPIs
        total_cols = len(column_profiles)
        high_null = sum(1 for c in column_profiles if (c.null_pct or 0) > 50)
        very_null = sum(1 for c in column_profiles if (c.null_pct or 0) > 90)
        card_vals = [c.cardinality_ratio for c in column_profiles if c.cardinality_ratio is not None]
        avg_card = sum(card_vals) / len(card_vals) if card_vals else None

        kpis = [
            kpi_card("Columns Profiled", _fmt_number(total_cols)),
            kpi_card("Columns >50% Null", str(high_null),
                     delta="Needs review" if high_null > 0 else "Clean",
                     delta_dir="down" if high_null > 0 else "up"),
            kpi_card("Columns >90% Null", str(very_null),
                     delta="Consider dropping" if very_null > 0 else "Clean",
                     delta_dir="down" if very_null > 0 else "up"),
            kpi_card("Avg Cardinality", f"{avg_card:.3f}" if avg_card else "N/A"),
        ]
        s += kpi_strip(kpis)

        s += callout(
            "<strong>Methodology:</strong> Per-column stats include null rates, distinct values, "
            "and type-specific distributions (numeric: percentiles; string: length + top values; "
            "temporal: date range). Cardinality ratio = distinct values / total rows.",
            "info",
        )

        # Worst columns by null_pct (top 20)
        sorted_by_null = sorted(column_profiles, key=lambda c: -(c.null_pct or 0))[:20]
        if sorted_by_null:
            s += '<h3 style="font-size:15px;font-weight:600;color:var(--gray-900);margin:16px 0 8px">Worst Columns by Null Rate (Top 20)</h3>'
            headers = ["Table", "Column", "Type", "Null %", "Distinct", "Cardinality Ratio"]
            rows = []
            for c in sorted_by_null:
                table_name = snap_table_map.get(c.snapshot_id, "?")
                null_pct = c.null_pct or 0
                null_color = RED if null_pct > 90 else (ORANGE if null_pct > 50 else GREEN)
                rows.append([
                    table_name,
                    f"<strong>{c.column_name}</strong>",
                    c.data_type or "-",
                    f'<span style="color:{null_color};font-weight:600">{null_pct:.1f}%</span>',
                    _fmt_number(c.distinct_count),
                    f"{c.cardinality_ratio:.4f}" if c.cardinality_ratio else "-",
                ])
            s += data_table(headers, rows, numeric_columns={3, 4, 5})

        # Low-cardinality flags
        low_card = [c for c in column_profiles
                    if c.cardinality_ratio is not None and c.cardinality_ratio < 0.01
                    and (c.distinct_count or 0) < 10]
        if low_card:
            s += '<h3 style="font-size:15px;font-weight:600;color:var(--gray-900);margin:20px 0 8px">Low-Cardinality Flags</h3>'
            s += '<p style="font-size:13px;color:var(--gray-500)">Columns with cardinality ratio &lt; 0.01 and &lt; 10 distinct values. May indicate enum fields or data quality issues.</p>'
            headers = ["Table", "Column", "Distinct", "Cardinality Ratio", "Null %"]
            rows = []
            for c in sorted(low_card, key=lambda x: x.cardinality_ratio or 0)[:20]:
                table_name = snap_table_map.get(c.snapshot_id, "?")
                rows.append([
                    table_name,
                    f"<strong>{c.column_name}</strong>",
                    str(c.distinct_count or 0),
                    f"{c.cardinality_ratio:.4f}" if c.cardinality_ratio else "-",
                    f"{c.null_pct:.1f}%" if c.null_pct else "0%",
                ])
            s += data_table(headers, rows, numeric_columns={2, 3, 4})

        s += section_end()
        return s

    # ── Section 6: Data Distribution Samples ─────────────────────────

    def _render_section_6(self, data: Dict[str, Any]) -> str:
        s = ""
        s += section_start(6, "Data Distribution Samples", "distributions")

        column_profiles: List = data.get("column_profiles", [])
        snap_table_map: Dict = data.get("snap_table_map", {})

        if not column_profiles:
            s += callout("No column profiles available for distribution analysis.", "info")
            s += section_end()
            return s

        s += callout(
            "<strong>Data Distributions:</strong> These show what your data looks like inside. "
            "Use them to verify values are reasonable and catch outliers.",
            "info",
        )

        # Numeric distributions
        numeric_cols = [c for c in column_profiles
                        if c.stats and isinstance(c.stats, dict) and "mean" in c.stats]
        if numeric_cols:
            # Sort by widest range
            def _range(c):
                st = c.stats or {}
                mn = st.get("min")
                mx = st.get("max")
                if mn is not None and mx is not None:
                    try:
                        return float(mx) - float(mn)
                    except (ValueError, TypeError):
                        pass
                return 0
            numeric_cols = sorted(numeric_cols, key=_range, reverse=True)[:10]

            s += '<h3 style="font-size:15px;font-weight:600;color:var(--gray-900);margin:16px 0 8px">Numeric Distributions</h3>'
            headers = ["Table", "Column", "Min", "P25", "Median", "Mean", "P75", "Max", "StdDev"]
            rows = []
            for c in numeric_cols:
                st = c.stats or {}
                table_name = snap_table_map.get(c.snapshot_id, "?")
                rows.append([
                    table_name,
                    f"<strong>{c.column_name}</strong>",
                    _fmt_float(st.get("min")),
                    _fmt_float(st.get("p25")),
                    _fmt_float(st.get("median")),
                    _fmt_float(st.get("mean")),
                    _fmt_float(st.get("p75")),
                    _fmt_float(st.get("max")),
                    _fmt_float(st.get("stddev"), 2),
                ])
            s += data_table(headers, rows, numeric_columns={2, 3, 4, 5, 6, 7, 8})

        # String length stats
        string_cols = [c for c in column_profiles
                       if c.stats and isinstance(c.stats, dict) and "avg_length" in c.stats]
        if string_cols:
            s += '<h3 style="font-size:15px;font-weight:600;color:var(--gray-900);margin:20px 0 8px">String Length Stats</h3>'
            headers = ["Table", "Column", "Min Length", "Avg Length", "Max Length"]
            rows = []
            for c in string_cols[:15]:
                st = c.stats or {}
                table_name = snap_table_map.get(c.snapshot_id, "?")
                rows.append([
                    table_name,
                    f"<strong>{c.column_name}</strong>",
                    str(st.get("min_length", "-")),
                    _fmt_float(st.get("avg_length")),
                    str(st.get("max_length", "-")),
                ])
            s += data_table(headers, rows, numeric_columns={2, 3, 4})

        # Top values for string columns
        top_val_cols = [c for c in column_profiles
                        if c.stats and isinstance(c.stats, dict)
                        and c.stats.get("top_values")][:5]
        if top_val_cols:
            s += '<h3 style="font-size:15px;font-weight:600;color:var(--gray-900);margin:20px 0 8px">Top Values (String Columns)</h3>'
            for c in top_val_cols:
                table_name = snap_table_map.get(c.snapshot_id, "?")
                s += f'<p style="font-size:13px;font-weight:600;color:var(--gray-700);margin:12px 0 4px">{table_name}.{c.column_name}</p>'
                top_vals = c.stats.get("top_values", [])[:5]
                labels = [str(tv.get("value", "?")) for tv in top_vals]
                counts = [float(tv.get("count", 0)) for tv in top_vals]
                if labels and counts:
                    s += build_bar_fallback(labels, counts)

        # Date ranges
        date_cols = [c for c in column_profiles
                     if c.stats and isinstance(c.stats, dict) and "min_date" in c.stats]
        if date_cols:
            s += '<h3 style="font-size:15px;font-weight:600;color:var(--gray-900);margin:20px 0 8px">Date Ranges</h3>'
            headers = ["Table", "Column", "Earliest", "Latest", "Span (days)"]
            rows = []
            for c in date_cols[:15]:
                st = c.stats or {}
                table_name = snap_table_map.get(c.snapshot_id, "?")
                rows.append([
                    table_name,
                    f"<strong>{c.column_name}</strong>",
                    str(st.get("min_date", "-")),
                    str(st.get("max_date", "-")),
                    str(st.get("date_range_days", "-")),
                ])
            s += data_table(headers, rows, numeric_columns={4})

        s += section_end()
        return s

    # ── Section 7: Anomaly Alerts with Full Context ──────────────────

    def _render_section_7(self, data: Dict[str, Any]) -> str:
        s = ""
        s += section_start(7, "Anomaly Alerts with Full Context", "anomaly-alerts")

        open_count = data.get("open_anomalies", 0)
        recent: List = data.get("recent_anomalies", [])
        crit_count = data.get("anomaly_critical", 0)
        warn_count = data.get("anomaly_warning", 0)
        info_count = data.get("anomaly_info", 0)

        kpis = [
            kpi_card("Open Alerts", str(open_count),
                     delta="Requires attention" if open_count > 0 else "All clear",
                     delta_dir="down" if open_count > 0 else "up"),
            kpi_card("Last 7 Days", str(len(recent))),
            kpi_card("Critical", str(crit_count),
                     delta_dir="down" if crit_count > 0 else "neutral"),
            kpi_card("Warning / Info", f"{warn_count} / {info_count}"),
        ]
        s += kpi_strip(kpis)

        if not recent:
            s += callout("No anomalies detected in the last 7 days.", "good")
            s += section_end()
            return s

        # Doughnut by type
        type_counts: Dict[str, int] = {}
        for a in recent:
            atype = a.alert_type.value if a.alert_type else "unknown"
            type_counts[atype] = type_counts.get(atype, 0) + 1

        if type_counts:
            t_labels = [t.replace("_", " ").title() for t in type_counts.keys()]
            t_values = list(type_counts.values())
            anom_doughnut = build_doughnut_config(t_labels, t_values)
            anom_doughnut_json = _json.dumps(anom_doughnut)
            s += '<div class="chart-row"><div>'
            s += chart_container("anomTypeDoughnut", anom_doughnut_json,
                                 fallback_html=build_chart_legend(t_labels, t_values),
                                 size="medium", title="Alerts by Type")
            s += '</div><div></div></div>'

        # Full detail table
        s += callout(
            "<strong>Deviation sigma</strong> measures how many standard deviations the current "
            "value is from the 30-day rolling mean. Higher sigma = more unusual.",
            "info",
        )

        headers = ["Table", "Column", "Type", "Severity", "Current", "Baseline",
                    "Deviation", "Detected", "Status"]
        rows = []
        for a in recent:
            atype = a.alert_type.value if a.alert_type else ""
            sev = a.severity.value if a.severity else ""
            rows.append([
                a.table_name or "",
                a.column_name or "<em>table-level</em>",
                atype.replace("_", " ").title(),
                _severity_pill(sev),
                str(a.current_value or "-"),
                str(a.baseline_value or "-"),
                _sigma_badge(a.deviation_sigma),
                a.detected_at.strftime("%Y-%m-%d %H:%M") if a.detected_at else "",
                a.status.value if a.status else "",
            ])
        s += data_table(headers, rows)

        s += section_end()
        return s

    def _charts_section_7(self, data: Dict[str, Any]) -> str:
        js = ""
        recent: List = data.get("recent_anomalies", [])
        if not recent:
            return js
        type_counts: Dict[str, int] = {}
        for a in recent:
            atype = a.alert_type.value if a.alert_type else "unknown"
            type_counts[atype] = type_counts.get(atype, 0) + 1
        if type_counts:
            t_labels = [t.replace("_", " ").title() for t in type_counts.keys()]
            t_values = list(type_counts.values())
            anom_doughnut = build_doughnut_config(t_labels, t_values)
            js += chart_init_js("anomTypeDoughnut", _json.dumps(anom_doughnut))
        return js

    # ── Section 8: Anomaly Detection Configuration ───────────────────

    def _render_section_8(self, data: Dict[str, Any]) -> str:
        s = ""
        s += section_start(8, "Anomaly Detection Configuration", "anomaly-config")

        thresholds: List = data.get("anomaly_thresholds", [])
        if not thresholds:
            s += callout("No anomaly thresholds configured. Using system defaults.", "info")
            s += section_end()
            return s

        s += callout(
            "<strong>Threshold Cascade:</strong> Table-specific &rarr; source-specific &rarr; "
            "global defaults. Lower sigma = more sensitive (more alerts).",
            "info",
        )

        headers = ["Source", "Table Pattern", "Row Count &sigma;", "Null Rate &sigma;",
                    "Distribution &sigma;", "Schema Drift", "Active"]
        rows = []
        for t in thresholds:
            rows.append([
                t.source or '<em>Global</em>',
                t.table_pattern or '<em>All</em>',
                f"{t.row_count_sigma:.1f}",
                f"{t.null_rate_sigma:.1f}",
                f"{t.distribution_sigma:.1f}",
                pill_badge("Yes", "private") if t.schema_drift_enabled else pill_badge("No", "default"),
                pill_badge("Active", "private") if t.is_enabled else pill_badge("Disabled", "default"),
            ])
        s += data_table(headers, rows, numeric_columns={2, 3, 4})

        s += section_end()
        return s

    # ── Section 9: Cross-Source Validation Detail ────────────────────

    def _render_section_9(self, data: Dict[str, Any]) -> str:
        s = ""
        s += section_start(9, "Cross-Source Validation Detail", "cross-source")

        cross_source: List[Dict] = data.get("cross_source", [])
        if not cross_source:
            s += callout("No cross-source validations configured. Run <code>/data-quality/cross-source/seed-defaults</code>.", "info")
            s += section_end()
            return s

        # KPIs
        total_validations = len(cross_source)
        rates = [cs["result"].match_rate * 100 for cs in cross_source
                 if cs["result"] and cs["result"].match_rate is not None]
        avg_rate = sum(rates) / len(rates) if rates else None
        passing = sum(1 for cs in cross_source if cs["result"] and cs["result"].passed)
        total_orphans = sum(
            (cs["result"].orphan_left or 0) + (cs["result"].orphan_right or 0)
            for cs in cross_source if cs["result"]
        )

        kpis = [
            kpi_card("Validations", str(total_validations)),
            kpi_card("Avg Match Rate", f"{avg_rate:.1f}%" if avg_rate else "N/A"),
            kpi_card("Passing", f"{passing}/{total_validations}",
                     delta_dir="up" if passing == total_validations else "down"),
            kpi_card("Total Orphans", _fmt_number(total_orphans),
                     delta="Records without match" if total_orphans > 0 else "All matched"),
        ]
        s += kpi_strip(kpis)

        s += callout(
            "<strong>Cross-source validation</strong> checks whether records in one table have "
            "matching counterparts in another. Orphans are records present in one side but not the other.",
            "info",
        )

        # Validation table
        headers = ["Name", "Type", "Tables", "Match Rate", "Status",
                    "Orphans L/R", "Evaluations", "Exec Time"]
        rows = []
        for cs in cross_source:
            v = cs["validation"]
            r = cs["result"]
            config = v.config or {}
            left_table = config.get("left_table", "?")
            right_table = config.get("right_table", "?")
            table_str = f"{left_table} &rarr; {right_table}"

            if r:
                rate = r.match_rate * 100 if r.match_rate is not None else None
                rate_str = f"{rate:.1f}%" if rate is not None else "N/A"
                rate_color = _score_color(rate)
                status = _status_pill("Good" if r.passed else "Critical")
                orphans = f"L:{r.orphan_left or 0} R:{r.orphan_right or 0}"
                exec_time = f"{r.execution_time_ms}ms" if r.execution_time_ms else "-"
            else:
                rate_str = "N/A"
                rate_color = GRAY
                status = pill_badge("Not Run", "default")
                orphans = "-"
                exec_time = "-"

            rows.append([
                f"<strong>{v.name}</strong>",
                (v.validation_type or "").replace("_", " ").title(),
                table_str,
                f'<span style="color:{rate_color};font-weight:600">{rate_str}</span>',
                status,
                orphans,
                str(v.times_evaluated or 0),
                exec_time,
            ])
        s += data_table(headers, rows, numeric_columns={6, 7})

        # Orphan examples
        has_orphans = False
        for cs in cross_source:
            r = cs["result"]
            if r and r.sample_orphans:
                if not has_orphans:
                    s += '<h3 style="font-size:15px;font-weight:600;color:var(--gray-900);margin:20px 0 8px">Orphan Examples</h3>'
                    has_orphans = True
                v = cs["validation"]
                orphan_data = r.sample_orphans
                if isinstance(orphan_data, dict):
                    for side, examples in orphan_data.items():
                        if examples:
                            sample_list = examples[:10] if isinstance(examples, list) else [str(examples)]
                            s += f'<p style="font-size:13px;margin:8px 0 4px"><strong>{v.name}</strong> ({side}):</p>'
                            s += '<p style="font-size:12px;color:var(--gray-500);font-style:italic;margin-left:16px">'
                            s += ", ".join(str(ex) for ex in sample_list)
                            s += '</p>'
                elif isinstance(orphan_data, list) and orphan_data:
                    s += f'<p style="font-size:13px;margin:8px 0 4px"><strong>{v.name}</strong>:</p>'
                    s += '<p style="font-size:12px;color:var(--gray-500);font-style:italic;margin-left:16px">'
                    s += ", ".join(str(ex) for ex in orphan_data[:10])
                    s += '</p>'

        s += section_end()
        return s

    # ── Section 10: Data Freshness by Source ─────────────────────────

    def _render_section_10(self, data: Dict[str, Any]) -> str:
        s = ""
        s += section_start(10, "Data Freshness by Source", "freshness")

        freshness_rows = data.get("source_freshness", [])
        if not freshness_rows:
            s += callout("No ingestion jobs found. Run an ingestion to populate freshness data.", "info")
            s += section_end()
            return s

        now = datetime.now(timezone.utc)
        stale_sources = []

        headers = ["Source", "Last Ingestion", "Days Ago", "Successes", "Failures", "Freshness Score"]
        rows = []
        source_labels = []
        days_values = []

        for fr in freshness_rows:
            source = fr.source or "?"
            last = fr.last_completed
            if last:
                if last.tzinfo is None:
                    from datetime import timezone as tz
                    last = last.replace(tzinfo=tz.utc)
                days_ago = (now - last).total_seconds() / 86400
                # Linear decay: 100 at 0h -> 0 at 168h (7 days)
                freshness_score = max(0, 100 - (days_ago * 100 / 7))
            else:
                days_ago = None
                freshness_score = 0

            if days_ago and days_ago > 30:
                stale_sources.append(source)

            source_labels.append(source)
            days_values.append(days_ago or 0)

            rows.append([
                f"<strong>{source}</strong>",
                last.strftime("%Y-%m-%d %H:%M") if last else "Never",
                _days_ago_html(days_ago),
                str(fr.successes or 0),
                str(fr.failures or 0),
                _score_html(freshness_score, show_bar=False),
            ])

        s += data_table(headers, rows, numeric_columns={3, 4, 5})

        # Bar chart: days since last ingestion (worst first)
        sorted_pairs = sorted(zip(source_labels, days_values), key=lambda x: -x[1])
        bar_labels = [p[0] for p in sorted_pairs[:15]]
        bar_vals = [p[1] for p in sorted_pairs[:15]]
        bar_colors = [RED if v > 30 else (ORANGE if v > 7 else GREEN) for v in bar_vals]

        if bar_labels:
            bar_config = build_horizontal_bar_config(bar_labels, bar_vals, bar_colors,
                                                     dataset_label="Days Since Last Ingestion")
            bar_json = _json.dumps(bar_config)
            bar_height = f"{max(len(bar_labels) * 48 + 40, 200)}px"
            s += chart_container("freshnessBar", bar_json,
                                 fallback_html=build_bar_fallback(bar_labels, bar_vals),
                                 title="Days Since Last Ingestion (Worst First)",
                                 height=bar_height)

        if stale_sources:
            s += callout(
                f"<strong>Stale Sources ({len(stale_sources)}):</strong> "
                + ", ".join(stale_sources)
                + " — not ingested in over 30 days.",
                "warn",
            )

        s += section_end()
        return s

    def _charts_section_10(self, data: Dict[str, Any]) -> str:
        js = ""
        freshness_rows = data.get("source_freshness", [])
        if not freshness_rows:
            return js

        now = datetime.now(timezone.utc)
        source_labels = []
        days_values = []
        for fr in freshness_rows:
            source = fr.source or "?"
            last = fr.last_completed
            if last:
                if last.tzinfo is None:
                    from datetime import timezone as tz
                    last = last.replace(tzinfo=tz.utc)
                days_ago = (now - last).total_seconds() / 86400
            else:
                days_ago = 0
            source_labels.append(source)
            days_values.append(days_ago)

        sorted_pairs = sorted(zip(source_labels, days_values), key=lambda x: -x[1])
        bar_labels = [p[0] for p in sorted_pairs[:15]]
        bar_vals = [p[1] for p in sorted_pairs[:15]]
        bar_colors = [RED if v > 30 else (ORANGE if v > 7 else GREEN) for v in bar_vals]

        if bar_labels:
            bar_config = build_horizontal_bar_config(bar_labels, bar_vals, bar_colors,
                                                     dataset_label="Days Since Last Ingestion")
            js += chart_init_js("freshnessBar", _json.dumps(bar_config))
        return js

    # ── Section 11: Quality Trending & Weekly Rollups ────────────────

    def _render_section_11(self, data: Dict[str, Any]) -> str:
        s = ""
        s += section_start(11, "Quality Trending & Weekly Rollups", "trending")

        trend = data.get("trend", [])
        if not trend:
            s += callout("No trend data available. Run <code>/data-quality/trends/compute</code> to generate snapshots.", "info")
            s += section_end()
            return s

        # Line chart
        labels = [str(t.snapshot_date) for t in trend]
        quality_vals = [round(t.avg_quality, 1) if t.avg_quality else None for t in trend]
        completeness_vals = [round(t.avg_completeness, 1) if t.avg_completeness else None for t in trend]
        freshness_vals = [round(t.avg_freshness, 1) if t.avg_freshness else None for t in trend]
        validity_vals = [round(t.avg_validity, 1) if t.avg_validity else None for t in trend]
        consistency_vals = [round(t.avg_consistency, 1) if t.avg_consistency else None for t in trend]

        window = data.get("window_days", 30)
        chart_config = _json.dumps({
            "type": "line",
            "data": {
                "labels": labels,
                "datasets": [
                    {"label": "Quality Score", "data": quality_vals, "borderColor": BLUE,
                     "backgroundColor": BLUE + "22", "fill": True, "tension": 0.3},
                    {"label": "Completeness", "data": completeness_vals, "borderColor": GREEN, "tension": 0.3},
                    {"label": "Freshness", "data": freshness_vals, "borderColor": ORANGE, "tension": 0.3},
                    {"label": "Validity", "data": validity_vals, "borderColor": "#8B5CF6", "tension": 0.3},
                    {"label": "Consistency", "data": consistency_vals, "borderColor": "#d53f8c", "tension": 0.3},
                ],
            },
            "options": {
                "responsive": True,
                "plugins": {"legend": {"position": "top"},
                            "title": {"display": True, "text": f"Quality Score Trends ({window} days)"}},
                "scales": {"y": {"min": 0, "max": 100, "title": {"display": True, "text": "Score"}}},
            },
        })
        s += chart_container("qualityTrendChart", chart_config, size="tall",
                             title="Quality Trends")

        # Weekly rollup table
        weekly = data.get("weekly_rollup", [])
        if weekly:
            s += '<h3 style="font-size:15px;font-weight:600;color:var(--gray-900);margin:20px 0 8px">Weekly Rollups</h3>'
            headers = ["Week", "Avg Quality", "Delta", "Completeness", "Freshness",
                        "Validity", "Consistency", "Tables"]
            rows = []
            prev_quality = None
            for w in weekly:
                week_str = str(w.week)[:10] if w.week else "-"
                avg_q = round(w.avg_quality, 1) if w.avg_quality else None
                delta = round(avg_q - prev_quality, 1) if avg_q is not None and prev_quality is not None else None

                rows.append([
                    week_str,
                    _fmt_float(avg_q),
                    _delta_html(delta),
                    _fmt_float(round(w.avg_completeness, 1) if w.avg_completeness else None),
                    _fmt_float(round(w.avg_freshness, 1) if w.avg_freshness else None),
                    _fmt_float(round(w.avg_validity, 1) if w.avg_validity else None),
                    _fmt_float(round(w.avg_consistency, 1) if w.avg_consistency else None),
                    str(w.count),
                ])
                prev_quality = avg_q

            s += data_table(headers, rows, numeric_columns={1, 3, 4, 5, 6, 7})

            # Trend assessment
            if len(weekly) >= 2:
                first_q = weekly[0].avg_quality if weekly[0].avg_quality else 0
                last_q = weekly[-1].avg_quality if weekly[-1].avg_quality else 0
                if last_q > first_q + 1:
                    s += callout(
                        f"<strong>Trending Up:</strong> Quality improved from {first_q:.1f} to {last_q:.1f} "
                        f"over the last {len(weekly)} weeks.", "good")
                elif last_q < first_q - 1:
                    s += callout(
                        f"<strong>Trending Down:</strong> Quality declined from {first_q:.1f} to {last_q:.1f} "
                        f"over the last {len(weekly)} weeks.", "warn")
                else:
                    s += callout(
                        f"<strong>Stable:</strong> Quality score holding steady around {last_q:.1f} "
                        f"over the last {len(weekly)} weeks.", "info")

        s += section_end()
        return s

    def _charts_section_11(self, data: Dict[str, Any]) -> str:
        js = ""
        trend = data.get("trend", [])
        if not trend:
            return js

        labels = [str(t.snapshot_date) for t in trend]
        quality_vals = [round(t.avg_quality, 1) if t.avg_quality else None for t in trend]
        completeness_vals = [round(t.avg_completeness, 1) if t.avg_completeness else None for t in trend]
        freshness_vals = [round(t.avg_freshness, 1) if t.avg_freshness else None for t in trend]
        validity_vals = [round(t.avg_validity, 1) if t.avg_validity else None for t in trend]
        consistency_vals = [round(t.avg_consistency, 1) if t.avg_consistency else None for t in trend]

        window = data.get("window_days", 30)
        chart_config = _json.dumps({
            "type": "line",
            "data": {
                "labels": labels,
                "datasets": [
                    {"label": "Quality Score", "data": quality_vals, "borderColor": BLUE,
                     "backgroundColor": BLUE + "22", "fill": True, "tension": 0.3},
                    {"label": "Completeness", "data": completeness_vals, "borderColor": GREEN, "tension": 0.3},
                    {"label": "Freshness", "data": freshness_vals, "borderColor": ORANGE, "tension": 0.3},
                    {"label": "Validity", "data": validity_vals, "borderColor": "#8B5CF6", "tension": 0.3},
                    {"label": "Consistency", "data": consistency_vals, "borderColor": "#d53f8c", "tension": 0.3},
                ],
            },
            "options": {
                "responsive": True,
                "plugins": {"legend": {"position": "top"},
                            "title": {"display": True, "text": f"Quality Score Trends ({window} days)"}},
                "scales": {"y": {"min": 0, "max": 100, "title": {"display": True, "text": "Score"}}},
            },
        })
        js += chart_init_js("qualityTrendChart", chart_config)
        return js

    # ── Section 12: SLA Compliance Detail ────────────────────────────

    def _render_section_12(self, data: Dict[str, Any]) -> str:
        s = ""
        s += section_start(12, "SLA Compliance Detail", "sla-compliance")

        sla_results: List[Dict] = data.get("sla_results", [])

        if not sla_results:
            s += callout("No SLA targets configured. Create targets via <code>/data-quality/trends/sla-targets</code>.", "info")
            s += section_end()
            return s

        # Per-dimension compliance
        dims = ["Quality", "Completeness", "Freshness", "Validity"]
        dim_kpis = []
        total_met = sum(1 for r in sla_results if r["met"])
        total_all = len(sla_results)
        overall_pct = (total_met / total_all * 100) if total_all > 0 else 0

        dim_kpis.append(kpi_card("Overall SLA", f"{overall_pct:.0f}%",
                                 delta=f"{total_met}/{total_all} passing"))

        for dim in dims:
            dim_results = [r for r in sla_results if r["dimension"] == dim]
            dim_met = sum(1 for r in dim_results if r["met"])
            dim_total = len(dim_results)
            dim_pct = (dim_met / dim_total * 100) if dim_total > 0 else 0
            dim_kpis.append(kpi_card(f"{dim} SLA", f"{dim_pct:.0f}%",
                                     delta=f"{dim_met}/{dim_total}"))

        s += kpi_strip(dim_kpis)

        # Detail table
        headers = ["Source", "Table", "Dimension", "Score", "Target", "Gap", "Status"]
        rows = []
        for r in sorted(sla_results, key=lambda x: x["gap"]):
            gap = r["gap"]
            gap_color = GREEN if gap >= 0 else RED
            rows.append([
                r["source"] or "-",
                r["table"],
                r["dimension"],
                _fmt_float(r["score"]),
                _fmt_float(r["target"]),
                f'<span style="color:{gap_color};font-weight:600">{gap:+.1f}</span>',
                _status_pill("Good" if r["met"] else "Critical"),
            ])
        s += data_table(headers, rows, numeric_columns={3, 4, 5})

        s += callout(
            "<strong>SLA targets</strong> define minimum acceptable scores per dimension. "
            "Negative gaps indicate the score is below the target threshold.",
            "info",
        )

        s += section_end()
        return s

    # ── Section 13: Schema Drift History ─────────────────────────────

    def _render_section_13(self, data: Dict[str, Any]) -> str:
        s = ""
        s += section_start(13, "Schema Drift History", "schema-drift")

        schema_alerts: List = data.get("schema_alerts", [])

        # KPIs
        total_changes = len(schema_alerts)
        new_cols = sum(1 for a in schema_alerts if a.alert_type == AnomalyAlertType.NEW_COLUMN)
        dropped_cols = sum(1 for a in schema_alerts if a.alert_type == AnomalyAlertType.DROPPED_COLUMN)
        type_changes = sum(1 for a in schema_alerts if a.alert_type == AnomalyAlertType.TYPE_CHANGE)

        kpis = [
            kpi_card("Schema Changes", str(total_changes)),
            kpi_card("New Columns", str(new_cols)),
            kpi_card("Dropped Columns", str(dropped_cols)),
            kpi_card("Type Changes", str(type_changes)),
        ]
        s += kpi_strip(kpis)

        if not schema_alerts:
            s += callout("No schema drift events detected.", "good")
            s += section_end()
            return s

        s += callout(
            "<strong>Schema drift</strong> is detected by comparing schema snapshots between "
            "consecutive profiling runs. New/dropped columns and type changes are tracked automatically.",
            "info",
        )

        headers = ["Table", "Change Type", "Column", "Details", "Detected"]
        rows = []
        type_pill_map = {
            AnomalyAlertType.NEW_COLUMN: ("NEW", "private"),
            AnomalyAlertType.DROPPED_COLUMN: ("DROPPED", "default"),
            AnomalyAlertType.TYPE_CHANGE: ("TYPE CHANGE", "pe"),
        }
        for a in schema_alerts:
            ptype, pvariant = type_pill_map.get(a.alert_type, (str(a.alert_type), "default"))
            details = a.message or ""
            if a.details and isinstance(a.details, dict):
                details = _json.dumps(a.details, default=str)[:100]
            rows.append([
                f"<strong>{a.table_name}</strong>",
                pill_badge(ptype, pvariant),
                a.column_name or "-",
                details[:80],
                a.detected_at.strftime("%Y-%m-%d %H:%M") if a.detected_at else "",
            ])
        s += data_table(headers, rows)

        s += section_end()
        return s

    # ── Section 14: Recommendations & Action Items ───────────────────

    def _render_section_14(self, data: Dict[str, Any]) -> str:
        s = ""
        s += section_start(14, "Recommendations & Action Items", "recommendations")

        quality_snapshots = data.get("quality_snapshots", [])
        column_profiles: List = data.get("column_profiles", [])
        snap_table_map: Dict = data.get("snap_table_map", {})
        cross_source: List[Dict] = data.get("cross_source", [])
        sla_results: List[Dict] = data.get("sla_results", [])
        freshness_rows = data.get("source_freshness", [])
        now = datetime.now(timezone.utc)
        has_issues = False

        # Critical tables
        critical_tables = [s for s in quality_snapshots if (s.quality_score or 0) < 70]
        if critical_tables:
            has_issues = True
            names = ", ".join(f"<strong>{t.table_name}</strong> ({t.quality_score:.1f})" for t in critical_tables[:10])
            s += callout(
                f"<strong>Critical Quality ({len(critical_tables)} tables):</strong> {names}. "
                "Investigate completeness, validity, and freshness dimensions.",
                "warn",
            )

        # Stale sources
        stale = []
        for fr in freshness_rows:
            last = fr.last_completed
            if last:
                if last.tzinfo is None:
                    from datetime import timezone as tz
                    last = last.replace(tzinfo=tz.utc)
                days = (now - last).total_seconds() / 86400
                if days > 30:
                    stale.append(fr.source)
        if stale:
            has_issues = True
            s += callout(
                f"<strong>Stale Sources ({len(stale)}):</strong> {', '.join(stale)} "
                "have not been ingested in over 30 days. Re-run ingestion or verify API keys.",
                "warn",
            )

        # High-null columns
        very_null = [c for c in column_profiles if (c.null_pct or 0) > 90]
        if very_null:
            has_issues = True
            top5 = very_null[:5]
            names = ", ".join(
                f"<strong>{snap_table_map.get(c.snapshot_id, '?')}.{c.column_name}</strong> ({c.null_pct:.0f}%)"
                for c in top5
            )
            s += callout(
                f"<strong>High-Null Columns ({len(very_null)} total):</strong> {names}"
                + (" and more..." if len(very_null) > 5 else "")
                + " Consider dropping or investigating these columns.",
                "warn",
            )

        # Failing cross-source validations
        failing_cs = [cs for cs in cross_source if cs["result"] and not cs["result"].passed]
        if failing_cs:
            has_issues = True
            names = ", ".join(f"<strong>{cs['validation'].name}</strong>" for cs in failing_cs[:5])
            s += callout(
                f"<strong>Failing Validations ({len(failing_cs)}):</strong> {names}. "
                "Check orphan records and data consistency between sources.",
                "warn",
            )

        # SLA breaches
        sla_failing = [r for r in sla_results if not r["met"]]
        if sla_failing:
            has_issues = True
            # Group by dimension
            dim_counts: Dict[str, int] = {}
            for r in sla_failing:
                dim_counts[r["dimension"]] = dim_counts.get(r["dimension"], 0) + 1
            dim_str = ", ".join(f"{dim}: {cnt}" for dim, cnt in dim_counts.items())
            s += callout(
                f"<strong>SLA Breaches ({len(sla_failing)}):</strong> {dim_str}. "
                "Review targets and improve data quality for affected dimensions.",
                "warn",
            )

        if not has_issues:
            s += callout(
                "<strong>All Clear!</strong> No critical issues detected. Quality scores are healthy, "
                "sources are fresh, and all validations are passing.",
                "good",
            )

        s += section_end()
        return s

    # ── Section 15: Data Quality Rules Engine ─────────────────────────

    def _render_section_15(self, data: Dict[str, Any]) -> str:
        s = ""
        s += section_start(15, "Data Quality Rules Engine", "rules-engine")

        total = data.get("rules_total", 0)
        enabled = data.get("rules_enabled", 0)
        pass_rate = data.get("rules_pass_rate_7d")
        coverage = data.get("rules_coverage_pct", 0)
        auto_count = data.get("rules_auto", 0)
        manual_count = data.get("rules_manual", 0)

        if total == 0:
            s += callout(
                "<strong>No Rules Defined.</strong> Use <code>POST /data-quality/rules/seed</code> "
                "to auto-generate rules from profiling data, or create manual rules via the API.",
                "info",
            )
            s += section_end()
            return s

        # KPIs
        pass_rate_str = f"{pass_rate:.1f}%" if pass_rate is not None else "N/A"
        s += kpi_strip([
            kpi_card("Total Rules", str(total)),
            kpi_card("Enabled", str(enabled)),
            kpi_card("Pass Rate (7d)", pass_rate_str),
            kpi_card("Coverage", f"{coverage:.0f}%"),
        ])

        # Auto vs manual callout
        s += callout(
            f"<strong>Rule Sources:</strong> {auto_count} auto-generated, {manual_count} manually created. "
            f"Coverage: {data.get('rules_tables_covered', 0)} tables have rules, "
            f"{data.get('rules_tables_uncovered', 0)} tables have none.",
            "info",
        )

        # Rules by type table
        rules_by_type = data.get("rules_by_type", {})
        if rules_by_type:
            # Calculate pass rate per type from recent results
            results_7d = data.get("rules_results_7d", [])
            all_rules = data.get("rules_all", [])
            rule_type_map = {r.id: r.rule_type.value for r in all_rules if r.rule_type}

            type_pass: Dict[str, int] = {}
            type_total: Dict[str, int] = {}
            for res in results_7d:
                rt = rule_type_map.get(res.rule_id, "unknown")
                type_total[rt] = type_total.get(rt, 0) + 1
                if res.passed:
                    type_pass[rt] = type_pass.get(rt, 0) + 1

            rows = []
            for rt, count in sorted(rules_by_type.items(), key=lambda x: -x[1]):
                t_total = type_total.get(rt, 0)
                t_pass = type_pass.get(rt, 0)
                rate = f"{t_pass / t_total * 100:.0f}%" if t_total > 0 else "—"
                rows.append([rt, str(count), str(t_total), str(t_pass), rate])

            s += data_table(
                headers=["Rule Type", "Count", "Evals (7d)", "Passed (7d)", "Pass Rate"],
                rows=rows,
            )

        # Doughnut chart
        if rules_by_type:
            labels = list(rules_by_type.keys())
            values = list(rules_by_type.values())
            colors = [BLUE, GREEN, ORANGE, RED, GRAY, "#9F7AEA", "#38B2AC", "#ED8936", "#E53E3E"]
            bg = colors[:len(labels)]
            doughnut_config = build_doughnut_config(labels, values, bg)
            doughnut_json = _json.dumps(doughnut_config)
            s += chart_container(
                "rulesTypeDoughnut", doughnut_json,
                fallback_html=build_chart_legend(labels, values, bg),
                size="medium", title="Rules by Type",
            )

        # Top failing rules table
        top_failing = data.get("rules_top_failing", [])
        if top_failing:
            rows = []
            for r in top_failing:
                total_evals = r.times_evaluated or 0
                failed = r.times_failed or 0
                fail_rate = f"{failed / total_evals * 100:.0f}%" if total_evals > 0 else "—"
                rows.append([
                    r.name,
                    r.rule_type.value if r.rule_type else "—",
                    r.severity.value if r.severity else "—",
                    str(failed),
                    str(total_evals),
                    fail_rate,
                ])

            s += data_table(
                headers=["Rule Name", "Type", "Severity", "Failures", "Evaluations", "Fail Rate"],
                rows=rows,
            )
        else:
            s += callout(
                "<strong>No failing rules.</strong> All evaluated rules have passed.",
                "good",
            )

        s += section_end()
        return s

    def _charts_section_15(self, data: Dict[str, Any]) -> str:
        rules_by_type = data.get("rules_by_type", {})
        if not rules_by_type:
            return ""

        labels = list(rules_by_type.keys())
        values = list(rules_by_type.values())
        colors = [BLUE, GREEN, ORANGE, RED, GRAY, "#9F7AEA", "#38B2AC", "#ED8936", "#E53E3E"]
        bg = colors[:len(labels)]

        config = build_doughnut_config(labels, values, bg)
        return chart_init_js("rulesTypeDoughnut", _json.dumps(config))

    # ── render_excel ─────────────────────────────────────────────────

    def render_excel(self, data: Dict[str, Any]) -> bytes:
        """Render the deep quality report as Excel with 11 sheets."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        header_fill = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        title_font = Font(bold=True, size=14)

        def _write_headers(ws, headers, row=1):
            for i, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=i, value=h)
                cell.fill = header_fill
                cell.font = header_font

        # ── Sheet 1: Summary ─────────────────────────────────────────
        ws = wb.active
        ws.title = "Summary"
        ws["A1"] = "Deep Data Quality Report"
        ws["A1"].font = title_font
        ws["A3"] = "Platform Score"
        ws["B3"] = data.get("platform_score")
        ws["A4"] = "Grade"
        ws["B4"] = data.get("platform_grade")
        ws["A5"] = "Tables Tracked"
        ws["B5"] = data.get("tables_tracked")
        ws["A7"] = "Avg Completeness"
        ws["B7"] = data.get("avg_completeness")
        ws["A8"] = "Avg Freshness"
        ws["B8"] = data.get("avg_freshness")
        ws["A9"] = "Avg Validity"
        ws["B9"] = data.get("avg_validity")
        ws["A10"] = "Avg Consistency"
        ws["B10"] = data.get("avg_consistency")
        ws["A12"] = "Open Anomalies"
        ws["B12"] = data.get("open_anomalies")
        ws["A13"] = "Week-over-Week Delta"
        ws["B13"] = data.get("wow_delta")

        # ── Sheet 2: Table Scoreboard ────────────────────────────────
        ws2 = wb.create_sheet("Table Scoreboard")
        headers = ["Table", "Source", "Domain", "Quality Score", "Grade",
                    "Completeness", "Freshness", "Validity", "Consistency", "Rows"]
        _write_headers(ws2, headers)
        for idx, snap in enumerate(data.get("quality_snapshots", []), 2):
            domain_key = classify_table(snap.table_name)
            ws2.cell(row=idx, column=1, value=snap.table_name)
            ws2.cell(row=idx, column=2, value=snap.source or "")
            ws2.cell(row=idx, column=3, value=DOMAIN_LABELS.get(domain_key, domain_key))
            ws2.cell(row=idx, column=4, value=snap.quality_score)
            ws2.cell(row=idx, column=5, value=_letter_grade(snap.quality_score) if snap.quality_score else "N/A")
            ws2.cell(row=idx, column=6, value=snap.completeness_score)
            ws2.cell(row=idx, column=7, value=snap.freshness_score)
            ws2.cell(row=idx, column=8, value=snap.validity_score)
            ws2.cell(row=idx, column=9, value=snap.consistency_score)
            ws2.cell(row=idx, column=10, value=snap.row_count)

        # ── Sheet 3: Domains ─────────────────────────────────────────
        ws3 = wb.create_sheet("Domains")
        headers = ["Domain", "Tables", "Avg Quality", "Avg Completeness",
                    "Avg Freshness", "Avg Validity", "Total Rows"]
        _write_headers(ws3, headers)
        for idx, d in enumerate(data.get("domains", []), 2):
            ws3.cell(row=idx, column=1, value=d.get("label", ""))
            ws3.cell(row=idx, column=2, value=d.get("tables", 0))
            ws3.cell(row=idx, column=3, value=d.get("avg_score"))
            ws3.cell(row=idx, column=4, value=d.get("avg_completeness"))
            ws3.cell(row=idx, column=5, value=d.get("avg_freshness"))
            ws3.cell(row=idx, column=6, value=d.get("avg_validity"))
            ws3.cell(row=idx, column=7, value=d.get("total_rows", 0))

        # ── Sheet 4: Column Profiles ─────────────────────────────────
        ws4 = wb.create_sheet("Column Profiles")
        headers = ["Table", "Column", "Type", "Null %", "Distinct Count",
                    "Cardinality Ratio", "Stats"]
        _write_headers(ws4, headers)
        snap_table_map = data.get("snap_table_map", {})
        for idx, c in enumerate(data.get("column_profiles", []), 2):
            table_name = snap_table_map.get(c.snapshot_id, "?")
            ws4.cell(row=idx, column=1, value=table_name)
            ws4.cell(row=idx, column=2, value=c.column_name)
            ws4.cell(row=idx, column=3, value=c.data_type or "")
            ws4.cell(row=idx, column=4, value=c.null_pct)
            ws4.cell(row=idx, column=5, value=c.distinct_count)
            ws4.cell(row=idx, column=6, value=c.cardinality_ratio)
            stats_str = _json.dumps(c.stats, default=str)[:500] if c.stats else ""
            ws4.cell(row=idx, column=7, value=stats_str)

        # ── Sheet 5: Anomalies ───────────────────────────────────────
        ws5 = wb.create_sheet("Anomalies")
        headers = ["Table", "Column", "Type", "Severity", "Current Value",
                    "Baseline Value", "Deviation Sigma", "Message", "Detected", "Status"]
        _write_headers(ws5, headers)
        for idx, a in enumerate(data.get("recent_anomalies", []), 2):
            ws5.cell(row=idx, column=1, value=a.table_name or "")
            ws5.cell(row=idx, column=2, value=a.column_name or "")
            ws5.cell(row=idx, column=3, value=a.alert_type.value if a.alert_type else "")
            ws5.cell(row=idx, column=4, value=a.severity.value if a.severity else "")
            ws5.cell(row=idx, column=5, value=str(a.current_value or ""))
            ws5.cell(row=idx, column=6, value=str(a.baseline_value or ""))
            ws5.cell(row=idx, column=7, value=a.deviation_sigma)
            ws5.cell(row=idx, column=8, value=(a.message or "")[:200])
            ws5.cell(row=idx, column=9, value=a.detected_at.strftime("%Y-%m-%d %H:%M") if a.detected_at else "")
            ws5.cell(row=idx, column=10, value=a.status.value if a.status else "")

        # ── Sheet 6: Cross-Source ────────────────────────────────────
        ws6 = wb.create_sheet("Cross-Source")
        headers = ["Validation", "Type", "Left Table", "Right Table",
                    "Match Rate %", "Passed", "Orphan Left", "Orphan Right",
                    "Exec Time (ms)", "Last Run"]
        _write_headers(ws6, headers)
        for idx, cs in enumerate(data.get("cross_source", []), 2):
            v = cs["validation"]
            r = cs["result"]
            config = v.config or {}
            ws6.cell(row=idx, column=1, value=v.name)
            ws6.cell(row=idx, column=2, value=v.validation_type or "")
            ws6.cell(row=idx, column=3, value=config.get("left_table", ""))
            ws6.cell(row=idx, column=4, value=config.get("right_table", ""))
            if r:
                ws6.cell(row=idx, column=5, value=round(r.match_rate * 100, 1) if r.match_rate else None)
                ws6.cell(row=idx, column=6, value="Yes" if r.passed else "No")
                ws6.cell(row=idx, column=7, value=r.orphan_left)
                ws6.cell(row=idx, column=8, value=r.orphan_right)
                ws6.cell(row=idx, column=9, value=r.execution_time_ms)
                ws6.cell(row=idx, column=10, value=r.evaluated_at.strftime("%Y-%m-%d %H:%M") if r.evaluated_at else "")

        # ── Sheet 7: Freshness ───────────────────────────────────────
        ws7 = wb.create_sheet("Freshness")
        headers = ["Source", "Last Ingestion", "Days Ago", "Successes", "Failures"]
        _write_headers(ws7, headers)
        now = datetime.now(timezone.utc)
        for idx, fr in enumerate(data.get("source_freshness", []), 2):
            last = fr.last_completed
            if last:
                if last.tzinfo is None:
                    from datetime import timezone as tz
                    last = last.replace(tzinfo=tz.utc)
                days_ago = round((now - last).total_seconds() / 86400, 1)
            else:
                days_ago = None
            ws7.cell(row=idx, column=1, value=fr.source or "")
            ws7.cell(row=idx, column=2, value=last.strftime("%Y-%m-%d %H:%M") if last else "Never")
            ws7.cell(row=idx, column=3, value=days_ago)
            ws7.cell(row=idx, column=4, value=fr.successes or 0)
            ws7.cell(row=idx, column=5, value=fr.failures or 0)

        # ── Sheet 8: Weekly Trending ─────────────────────────────────
        ws8 = wb.create_sheet("Weekly Trending")
        headers = ["Week", "Avg Quality", "Completeness", "Freshness",
                    "Validity", "Consistency", "Tables"]
        _write_headers(ws8, headers)
        for idx, w in enumerate(data.get("weekly_rollup", []), 2):
            ws8.cell(row=idx, column=1, value=str(w.week)[:10] if w.week else "")
            ws8.cell(row=idx, column=2, value=round(w.avg_quality, 1) if w.avg_quality else None)
            ws8.cell(row=idx, column=3, value=round(w.avg_completeness, 1) if w.avg_completeness else None)
            ws8.cell(row=idx, column=4, value=round(w.avg_freshness, 1) if w.avg_freshness else None)
            ws8.cell(row=idx, column=5, value=round(w.avg_validity, 1) if w.avg_validity else None)
            ws8.cell(row=idx, column=6, value=round(w.avg_consistency, 1) if w.avg_consistency else None)
            ws8.cell(row=idx, column=7, value=w.count)

        # ── Sheet 9: SLA Detail ──────────────────────────────────────
        ws9 = wb.create_sheet("SLA Detail")
        headers = ["Source", "Table", "Dimension", "Score", "Target", "Gap", "Met"]
        _write_headers(ws9, headers)
        for idx, r in enumerate(data.get("sla_results", []), 2):
            ws9.cell(row=idx, column=1, value=r.get("source", ""))
            ws9.cell(row=idx, column=2, value=r.get("table", ""))
            ws9.cell(row=idx, column=3, value=r.get("dimension", ""))
            ws9.cell(row=idx, column=4, value=r.get("score"))
            ws9.cell(row=idx, column=5, value=r.get("target"))
            ws9.cell(row=idx, column=6, value=r.get("gap"))
            ws9.cell(row=idx, column=7, value="Yes" if r.get("met") else "No")

        # ── Sheet 10: Schema Drift ───────────────────────────────────
        ws10 = wb.create_sheet("Schema Drift")
        headers = ["Table", "Change Type", "Column", "Message", "Detected"]
        _write_headers(ws10, headers)
        for idx, a in enumerate(data.get("schema_alerts", []), 2):
            ws10.cell(row=idx, column=1, value=a.table_name or "")
            ws10.cell(row=idx, column=2, value=a.alert_type.value if a.alert_type else "")
            ws10.cell(row=idx, column=3, value=a.column_name or "")
            ws10.cell(row=idx, column=4, value=(a.message or "")[:200])
            ws10.cell(row=idx, column=5, value=a.detected_at.strftime("%Y-%m-%d %H:%M") if a.detected_at else "")

        # ── Sheet 11: Rules Engine ────────────────────────────────────
        ws11 = wb.create_sheet("Rules Engine")
        headers = [
            "Rule Name", "Type", "Severity", "Source", "Column",
            "Enabled", "Priority", "Evaluated", "Passed", "Failed",
            "Last Evaluated",
        ]
        _write_headers(ws11, headers)
        for idx, r in enumerate(data.get("rules_all", []), 2):
            ws11.cell(row=idx, column=1, value=r.name or "")
            ws11.cell(row=idx, column=2, value=r.rule_type.value if r.rule_type else "")
            ws11.cell(row=idx, column=3, value=r.severity.value if r.severity else "")
            ws11.cell(row=idx, column=4, value=r.source or "all")
            ws11.cell(row=idx, column=5, value=r.column_name or "—")
            ws11.cell(row=idx, column=6, value="Yes" if r.is_enabled else "No")
            ws11.cell(row=idx, column=7, value=r.priority or 5)
            ws11.cell(row=idx, column=8, value=r.times_evaluated or 0)
            ws11.cell(row=idx, column=9, value=r.times_passed or 0)
            ws11.cell(row=idx, column=10, value=r.times_failed or 0)
            ws11.cell(
                row=idx, column=11,
                value=r.last_evaluated_at.strftime("%Y-%m-%d %H:%M") if r.last_evaluated_at else "",
            )

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()
