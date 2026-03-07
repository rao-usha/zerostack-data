"""
Datacenter Site Selection Report Template.

15-section investor-ready report for datacenter site selection.
Combines county suitability scores, power grid analysis, connectivity,
regulatory environment, incentives, and capital modeling.
"""

import json
import logging
from datetime import datetime
from typing import Any, Dict, List, Optional

from sqlalchemy import text
from sqlalchemy.orm import Session

from app.reports.design_system import (
    html_document,
    page_header,
    kpi_strip,
    kpi_card,
    toc,
    section_start,
    section_end,
    data_table,
    callout,
    chart_container,
    chart_init_js,
    page_footer,
    build_horizontal_bar_config,
    build_bar_fallback,
    BLUE,
    BLUE_LIGHT,
    GREEN,
)

logger = logging.getLogger(__name__)

DATACENTER_EXTRA_CSS = """
/* CEO Overview — narrative section */
.ceo-overview {
    background: var(--white);
    border-radius: 12px;
    padding: 32px;
    margin-bottom: 24px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.08);
}
.ceo-overview h3 {
    font-size: 16px;
    font-weight: 700;
    color: var(--primary);
    margin: 24px 0 12px;
}
.ceo-overview h3:first-child { margin-top: 0; }
.narrative-text {
    font-size: 14px;
    line-height: 1.7;
    color: var(--gray-700);
    margin-bottom: 12px;
}
.narrative-text strong { color: var(--gray-900); }

/* Metric grid for capital model & deal scenarios */
.metric-grid {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
    gap: 16px;
    margin: 16px 0;
}
.metric-item { text-align: center; padding: 12px; }
.metric-item .value { font-size: 28px; font-weight: 700; color: var(--primary-light); }
.metric-item .label { font-size: 13px; color: var(--gray-500); }

/* Muted helper */
.muted-text { color: var(--gray-500); font-size: 13px; }

/* Score bars */
.score-bar-container { display: flex; align-items: center; gap: 8px; }
.score-bar {
    height: 8px; border-radius: 4px; background: var(--gray-200);
    flex: 1; position: relative; overflow: hidden;
}
.score-bar-fill {
    height: 100%; border-radius: 4px;
    background: linear-gradient(90deg, var(--primary-light), var(--accent));
}

/* Deal cards */
.deal-scenario-grid {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(280px, 1fr));
    gap: 20px;
    margin: 20px 0;
}
.deal-card {
    background: var(--white);
    border: 1px solid var(--gray-200);
    border-radius: 12px;
    padding: 24px;
}
.deal-card h4 { color: var(--gray-900); }
.deal-card.recommended {
    border-color: var(--primary-light);
    box-shadow: 0 0 0 2px rgba(43,108,176,0.1);
}

/* Dark mode overrides */
[data-theme="dark"] .ceo-overview { box-shadow: 0 2px 8px rgba(0,0,0,0.3); }
[data-theme="dark"] .deal-card { border-color: var(--gray-200); }
[data-theme="dark"] .deal-card.recommended {
    border-color: var(--primary-light);
    box-shadow: 0 0 0 2px rgba(144,205,244,0.15);
}
"""


def _fmt(n, prefix="", suffix="", decimals=0):
    """Format a number with optional prefix/suffix."""
    if n is None:
        return "N/A"
    if isinstance(n, float):
        if decimals == 0:
            return f"{prefix}{n:,.0f}{suffix}"
        return f"{prefix}{n:,.{decimals}f}{suffix}"
    return f"{prefix}{n:,}{suffix}"


class DatacenterSiteTemplate:
    name = "datacenter_site"
    description = (
        "Investor-ready datacenter site selection report with county rankings, "
        "power/connectivity analysis, regulatory scoring, and capital modeling."
    )

    def _safe_db_call(self, db: Session, fn, default):
        """Call fn, rollback session on failure to keep it usable."""
        try:
            return fn()
        except Exception as e:
            logger.warning(f"Report query failed: {e}")
            try:
                db.rollback()
            except Exception:
                pass
            return default

    def gather_data(self, db: Session, params: Dict) -> Dict[str, Any]:
        """Gather all data for the report."""
        state = params.get("state")
        top_n = params.get("top_n", 20)
        target_mw = params.get("target_mw", 50)

        data = {}
        data["params"] = {"state": state, "top_n": top_n, "target_mw": target_mw}
        data["generated_at"] = datetime.utcnow().strftime("%B %d, %Y at %H:%M UTC")
        data["summary"] = self._safe_db_call(db, lambda: self._get_summary(db, state),
            {"total_counties": 0, "a_grade": 0, "states_covered": 0, "avg_score": 0})
        data["state_averages"] = self._safe_db_call(db, lambda: self._get_state_averages(db), [])
        data["top_counties"] = self._safe_db_call(db, lambda: self._get_top_counties(db, state, top_n), [])
        data["power_analysis"] = self._safe_db_call(db, lambda: self._get_power_analysis(db, state),
            {"fuel_mix": [], "prices": []})
        data["connectivity"] = self._safe_db_call(db, lambda: self._get_connectivity(db, state),
            {"ix_by_city": [], "dc_by_state": []})
        data["regulatory"] = self._safe_db_call(db, lambda: self._get_regulatory(db, state), [])
        data["real_estate"] = self._safe_db_call(db, lambda: self._get_real_estate(db, state),
            {"industrial_sites": [], "brownfields": []})
        data["incentives"] = self._safe_db_call(db, lambda: self._get_incentives(db, state),
            {"programs": [], "deals": []})
        data["environment"] = self._safe_db_call(db, lambda: self._get_environment(db, state),
            {"risk_by_county": [], "elevation": [], "wetlands": [], "flood_zones": []})
        data["workforce"] = self._safe_db_call(db, lambda: self._get_workforce(db, state), [])
        data["transmission"] = self._safe_db_call(db, lambda: self._get_transmission(db, state),
            {"summary": {}, "by_voltage": []})
        data["dc_clusters"] = self._safe_db_call(db, lambda: self._get_dc_clusters(db, state),
            {"peeringdb": [], "epoch": []})
        data["capital_model"] = self._build_capital_model(target_mw)
        data["deal_scenarios"] = self._build_deal_scenarios(data["top_counties"], target_mw)
        data["data_sources"] = self._safe_db_call(db, lambda: self._get_data_sources(db), [])
        data["ceo_overview"] = self._compute_ceo_overview(data)

        return data

    # ------------------------------------------------------------------
    # Data gathering helpers
    # ------------------------------------------------------------------

    def _get_summary(self, db: Session, state: Optional[str]) -> Dict:
        try:
            state_filter = "AND state = :state" if state else ""
            params: Dict[str, Any] = {}
            if state:
                params["state"] = state.upper()
            result = db.execute(text(f"""
                SELECT
                    COUNT(*) as total_counties,
                    COUNT(CASE WHEN grade = 'A' THEN 1 END) as a_grade,
                    COUNT(CASE WHEN grade = 'B' THEN 1 END) as b_grade,
                    COUNT(DISTINCT state) as states_covered,
                    ROUND(AVG(overall_score)::numeric, 1) as avg_score,
                    ROUND(MAX(overall_score)::numeric, 1) as max_score
                FROM datacenter_site_scores
                WHERE score_date = (SELECT MAX(score_date) FROM datacenter_site_scores)
                {state_filter}
            """), params)
            row = result.fetchone()
            if row:
                return dict(zip(result.keys(), row))
        except Exception as e:
            logger.warning(f"Summary fetch failed: {e}")
            db.rollback()
        return {"total_counties": 0, "a_grade": 0, "states_covered": 0, "avg_score": 0}

    def _get_state_averages(self, db: Session) -> List[Dict]:
        try:
            result = db.execute(text("""
                SELECT state,
                       ROUND(AVG(overall_score)::numeric, 1) as avg_score,
                       COUNT(*) as county_count,
                       COUNT(CASE WHEN grade IN ('A','B') THEN 1 END) as top_counties
                FROM datacenter_site_scores
                WHERE score_date = (SELECT MAX(score_date) FROM datacenter_site_scores)
                GROUP BY state
                ORDER BY avg_score DESC
                LIMIT 15
            """))
            return [dict(zip(result.keys(), row)) for row in result.fetchall()]
        except Exception:
            db.rollback()
            return []

    def _get_top_counties(self, db: Session, state: Optional[str], top_n: int) -> List[Dict]:
        try:
            state_filter = "AND state = :state" if state else ""
            params: Dict[str, Any] = {"n": top_n}
            if state:
                params["state"] = state.upper()
            result = db.execute(text(f"""
                SELECT county_fips, county_name, state, overall_score, grade,
                       power_score, connectivity_score, regulatory_score,
                       labor_score, risk_score, cost_incentive_score,
                       national_rank, state_rank,
                       electricity_price_cents_kwh, ix_count, dc_facility_count
                FROM datacenter_site_scores
                WHERE score_date = (SELECT MAX(score_date) FROM datacenter_site_scores)
                {state_filter}
                ORDER BY overall_score DESC
                LIMIT :n
            """), params)
            return [dict(zip(result.keys(), row)) for row in result.fetchall()]
        except Exception:
            db.rollback()
            return []

    def _get_power_analysis(self, db: Session, state: Optional[str]) -> Dict:
        try:
            state_filter = "WHERE state = :state" if state else ""
            params: Dict[str, Any] = {}
            if state:
                params["state"] = state.upper()
            result = db.execute(text(f"""
                SELECT state,
                       SUM(nameplate_capacity_mw) as total_mw,
                       COUNT(*) as plant_count,
                       primary_fuel,
                       SUM(nameplate_capacity_mw) as fuel_mw
                FROM power_plant
                {state_filter}
                GROUP BY state, primary_fuel
                ORDER BY total_mw DESC
                LIMIT 20
            """), params)
            rows = [dict(zip(result.keys(), row)) for row in result.fetchall()]

            # Price comparison — exclude census regions
            price_result = db.execute(text("""
                SELECT geography_name as state, AVG(avg_price_cents_kwh) as avg_rate
                FROM electricity_price
                WHERE sector = 'commercial'
                  AND geography_name NOT IN (
                      'New England', 'Middle Atlantic', 'East North Central',
                      'West North Central', 'South Atlantic', 'East South Central',
                      'West South Central', 'Mountain', 'Pacific', 'US')
                GROUP BY geography_name
                ORDER BY avg_rate
                LIMIT 15
            """))
            prices = [dict(zip(price_result.keys(), row)) for row in price_result.fetchall()]
            return {"fuel_mix": rows, "prices": prices}
        except Exception:
            db.rollback()
            return {"fuel_mix": [], "prices": []}

    def _get_connectivity(self, db: Session, state: Optional[str]) -> Dict:
        try:
            params: Dict[str, Any] = {}
            if state:
                params["state"] = state.upper()
                # IX records often lack state — match by city within the state's DC facilities
                ix_result = db.execute(text("""
                    SELECT ix.city, COUNT(*) as ix_count
                    FROM internet_exchange ix
                    WHERE ix.city IN (
                        SELECT DISTINCT city FROM data_center_facility WHERE state = :state
                    )
                    GROUP BY ix.city
                    ORDER BY ix_count DESC
                    LIMIT 10
                """), params)
            else:
                ix_result = db.execute(text("""
                    SELECT city, COUNT(*) as ix_count
                    FROM internet_exchange
                    GROUP BY city
                    ORDER BY ix_count DESC
                    LIMIT 10
                """))
            ix_data = [dict(zip(ix_result.keys(), row)) for row in ix_result.fetchall()]

            dc_filter = "WHERE state = :state" if state else ""
            dc_result = db.execute(text(f"""
                SELECT state, COUNT(*) as dc_count
                FROM data_center_facility
                {dc_filter}
                GROUP BY state
                ORDER BY dc_count DESC
                LIMIT 10
            """), params)
            dc_data = [dict(zip(dc_result.keys(), row)) for row in dc_result.fetchall()]
            return {"ix_by_city": ix_data, "dc_by_state": dc_data}
        except Exception:
            db.rollback()
            return {"ix_by_city": [], "dc_by_state": []}

    def _get_regulatory(self, db: Session, state: Optional[str]) -> List[Dict]:
        try:
            state_filter = "AND state = :state" if state else ""
            params: Dict[str, Any] = {}
            if state:
                params["state"] = state.upper()
            result = db.execute(text(f"""
                SELECT county_fips, county_name, state, overall_score, grade,
                       permit_velocity_score, jurisdictional_simplicity_score,
                       energy_siting_score, historical_dc_deals_score
                FROM county_regulatory_scores
                WHERE score_date = (SELECT MAX(score_date) FROM county_regulatory_scores)
                {state_filter}
                ORDER BY overall_score DESC
                LIMIT 15
            """), params)
            return [dict(zip(result.keys(), row)) for row in result.fetchall()]
        except Exception:
            db.rollback()
            return []

    def _get_real_estate(self, db: Session, state: Optional[str]) -> Dict:
        try:
            state_filter = "WHERE state = :state" if state else ""
            params: Dict[str, Any] = {}
            if state:
                params["state"] = state.upper()
            # Industrial sites
            sites_result = db.execute(text(f"""
                SELECT site_name, city, state, county, acreage, zoning,
                       utilities_available, rail_served, highway_access, edo_name
                FROM industrial_site
                {state_filter}
                ORDER BY acreage DESC NULLS LAST
                LIMIT 15
            """), params)
            sites = [dict(zip(sites_result.keys(), row)) for row in sites_result.fetchall()]

            # Brownfields
            bf_result = db.execute(text(f"""
                SELECT site_name, city, state, county, acreage,
                       cleanup_status, land_use_current
                FROM brownfield_site
                {state_filter}
                ORDER BY acreage DESC NULLS LAST
                LIMIT 10
            """), params)
            brownfields = [dict(zip(bf_result.keys(), row)) for row in bf_result.fetchall()]
            return {"industrial_sites": sites, "brownfields": brownfields}
        except Exception:
            db.rollback()
            return {"industrial_sites": [], "brownfields": []}

    def _get_incentives(self, db: Session, state: Optional[str]) -> Dict:
        try:
            state_filter = "WHERE state = :state" if state else ""
            params: Dict[str, Any] = {}
            if state:
                params["state"] = state.upper()

            programs = db.execute(text(f"""
                SELECT program_name, state, program_type, target_industries,
                       max_benefit, benefit_duration_years
                FROM incentive_program
                {state_filter}
                ORDER BY max_benefit DESC NULLS LAST
                LIMIT 15
            """), params)
            prog_data = [dict(zip(programs.keys(), row)) for row in programs.fetchall()]

            deal_filter = "WHERE state = :state" if state else ""
            deals = db.execute(text(f"""
                SELECT company_name, state, county, subsidy_value, year,
                       jobs_announced, industry
                FROM incentive_deal
                {deal_filter}
                ORDER BY subsidy_value DESC NULLS LAST
                LIMIT 15
            """), params)
            deal_data = [dict(zip(deals.keys(), row)) for row in deals.fetchall()]
            return {"programs": prog_data, "deals": deal_data}
        except Exception:
            db.rollback()
            return {"programs": [], "deals": []}

    def _get_environment(self, db: Session, state: Optional[str]) -> Dict:
        try:
            state_filter = "AND state = :state" if state else ""
            params: Dict[str, Any] = {}
            if state:
                params["state"] = state.upper()

            # FEMA NRI risk data
            result = db.execute(text(f"""
                SELECT county_name, state, risk_rating,
                       ROUND(risk_score::numeric, 1) as risk_score,
                       ROUND(tornado_score::numeric, 1) as tornado_score,
                       ROUND(hurricane_score::numeric, 1) as hurricane_score,
                       ROUND(wildfire_score::numeric, 1) as wildfire_score
                FROM national_risk_index
                WHERE state IS NOT NULL
                  AND risk_rating != 'Insufficient Data'
                {state_filter}
                ORDER BY risk_score DESC
                LIMIT 15
            """), params)
            risk_data = [dict(zip(result.keys(), row)) for row in result.fetchall()]

            # County elevation data (USGS 3DEP)
            elev_filter = "WHERE state = :state" if state else ""
            elev_result = db.execute(text(f"""
                SELECT state, county, fips_code,
                       ROUND(min_elevation_ft::numeric, 0) as min_ft,
                       ROUND(max_elevation_ft::numeric, 0) as max_ft,
                       ROUND(mean_elevation_ft::numeric, 0) as mean_ft,
                       ROUND(elevation_range_ft::numeric, 0) as range_ft,
                       sample_points
                FROM county_elevation
                {elev_filter}
                ORDER BY mean_elevation_ft
                LIMIT 15
            """), params)
            elevation = [dict(zip(elev_result.keys(), row)) for row in elev_result.fetchall()]

            # Elevation summary
            elev_summary_result = db.execute(text(f"""
                SELECT COUNT(*) as counties,
                       ROUND(AVG(mean_elevation_ft)::numeric, 0) as avg_elev,
                       ROUND(MIN(min_elevation_ft)::numeric, 0) as min_elev,
                       ROUND(MAX(max_elevation_ft)::numeric, 0) as max_elev
                FROM county_elevation
                {elev_filter}
            """), params)
            elev_summary_row = elev_summary_result.fetchone()
            elev_summary = dict(zip(elev_summary_result.keys(), elev_summary_row)) if elev_summary_row else {}

            # Wetlands (NWI)
            wet_filter = "WHERE state = :state" if state else ""
            wet_result = db.execute(text(f"""
                SELECT state,
                       COUNT(*) as wetland_count,
                       ROUND(SUM(acres)::numeric, 0) as total_acres,
                       wetland_type
                FROM wetland
                {wet_filter}
                GROUP BY state, wetland_type
                ORDER BY total_acres DESC NULLS LAST
                LIMIT 15
            """), params)
            wetlands = [dict(zip(wet_result.keys(), row)) for row in wet_result.fetchall()]

            # Flood zones (FEMA NFHL)
            fz_filter = "WHERE state = :state" if state else ""
            fz_result = db.execute(text(f"""
                SELECT state, zone_code, zone_description,
                       is_high_risk, COUNT(*) as zone_count
                FROM flood_zone
                {fz_filter}
                GROUP BY state, zone_code, zone_description, is_high_risk
                ORDER BY zone_count DESC
                LIMIT 15
            """), params)
            flood_zones = [dict(zip(fz_result.keys(), row)) for row in fz_result.fetchall()]

            return {
                "risk_by_county": risk_data,
                "elevation": elevation,
                "elev_summary": elev_summary,
                "wetlands": wetlands,
                "flood_zones": flood_zones,
            }
        except Exception:
            db.rollback()
            return {"risk_by_county": [], "elevation": [], "elev_summary": {},
                    "wetlands": [], "flood_zones": []}

    def _get_workforce(self, db: Session, state: Optional[str]) -> List[Dict]:
        try:
            state_filter = "AND ie.area_fips LIKE :state_fips || '%'" if state else ""
            params: Dict[str, Any] = {}
            if state:
                from app.sources.site_intel.labor.bls_qcew_collector import STATE_FIPS
                fips = STATE_FIPS.get(state.upper(), "")
                params["state_fips"] = fips
            result = db.execute(text(f"""
                SELECT ie.area_name, ie.industry_title,
                       ie.avg_monthly_employment, ie.avg_weekly_wage,
                       ie.establishments
                FROM industry_employment ie
                WHERE ie.industry_code IN ('1022', '1024', '1013')
                  AND ie.ownership = 'private'
                  AND ie.avg_monthly_employment > 0
                {state_filter}
                ORDER BY ie.avg_monthly_employment DESC
                LIMIT 15
            """), params)
            return [dict(zip(result.keys(), row)) for row in result.fetchall()]
        except Exception:
            db.rollback()
            return []

    def _get_dc_clusters(self, db: Session, state: Optional[str]) -> Dict:
        try:
            state_filter = "WHERE state = :state" if state else ""
            params: Dict[str, Any] = {}
            if state:
                params["state"] = state.upper()

            # PeeringDB / existing DC facilities
            dc_result = db.execute(text(f"""
                SELECT name, operator, city, state
                FROM data_center_facility
                {state_filter}
                ORDER BY city, name
                LIMIT 15
            """), params)
            facilities = [dict(zip(dc_result.keys(), row)) for row in dc_result.fetchall()]

            # Epoch AI datacenters
            epoch_result = db.execute(text(f"""
                SELECT company, facility_name, city, state,
                       power_capacity_mw, year_opened, status
                FROM epoch_datacenter
                {state_filter}
                ORDER BY power_capacity_mw DESC NULLS LAST
                LIMIT 15
            """), params)
            epoch = [dict(zip(epoch_result.keys(), row)) for row in epoch_result.fetchall()]
            return {"peeringdb": facilities, "epoch": epoch}
        except Exception:
            db.rollback()
            return {"peeringdb": [], "epoch": []}

    def _get_transmission(self, db: Session, state: Optional[str]) -> Dict:
        try:
            state_filter = "WHERE state = :state" if state else ""
            params: Dict[str, Any] = {}
            if state:
                params["state"] = state.upper()

            # Summary stats
            summary_result = db.execute(text(f"""
                SELECT COUNT(*) as total_lines,
                       ROUND(SUM(length_miles)::numeric, 0) as total_miles,
                       COUNT(DISTINCT owner) as unique_owners,
                       ROUND(AVG(voltage_kv)::numeric, 0) as avg_voltage_kv
                FROM transmission_line
                {state_filter}
            """), params)
            summary_row = summary_result.fetchone()
            summary = dict(zip(summary_result.keys(), summary_row)) if summary_row else {}

            # Breakdown by voltage class
            volt_result = db.execute(text(f"""
                SELECT voltage_class,
                       COUNT(*) as line_count,
                       ROUND(SUM(length_miles)::numeric, 0) as total_miles,
                       ROUND(AVG(voltage_kv)::numeric, 0) as avg_kv
                FROM transmission_line
                {state_filter}
                GROUP BY voltage_class
                ORDER BY avg_kv DESC NULLS LAST
            """), params)
            by_voltage = [dict(zip(volt_result.keys(), row)) for row in volt_result.fetchall()]

            # Top states by line count (national view)
            if not state:
                state_result = db.execute(text("""
                    SELECT state, COUNT(*) as line_count,
                           ROUND(SUM(length_miles)::numeric, 0) as total_miles
                    FROM transmission_line
                    WHERE state IS NOT NULL
                    GROUP BY state
                    ORDER BY line_count DESC
                    LIMIT 15
                """))
                by_state = [dict(zip(state_result.keys(), row)) for row in state_result.fetchall()]
            else:
                by_state = []

            return {"summary": summary, "by_voltage": by_voltage, "by_state": by_state}
        except Exception:
            db.rollback()
            return {"summary": {}, "by_voltage": [], "by_state": []}

    def _build_capital_model(self, target_mw: int) -> Dict:
        """Build a capital cost model per MW by tier."""
        tiers = {
            "Tier III": {"cost_per_mw": 8_000_000, "pue": 1.3, "uptime": 99.982},
            "Tier IV": {"cost_per_mw": 12_000_000, "pue": 1.2, "uptime": 99.995},
        }
        models = {}
        for tier_name, tier in tiers.items():
            total_cost = tier["cost_per_mw"] * target_mw
            models[tier_name] = {
                "cost_per_mw": tier["cost_per_mw"],
                "total_build_cost": total_cost,
                "pue": tier["pue"],
                "uptime_sla": tier["uptime"],
                "target_mw": target_mw,
                "annual_power_cost_5c": target_mw * 1000 * 8760 * 0.05 * tier["pue"],
                "annual_power_cost_8c": target_mw * 1000 * 8760 * 0.08 * tier["pue"],
                "breakdown": {
                    "Land & Site Prep": 0.08,
                    "Power Infrastructure": 0.30,
                    "Cooling Systems": 0.18,
                    "IT Infrastructure": 0.22,
                    "Fiber & Network": 0.07,
                    "Security & Fire": 0.05,
                    "Soft Costs": 0.10,
                },
            }
        return models

    def _build_deal_scenarios(
        self, top_counties: List[Dict], target_mw: int
    ) -> List[Dict]:
        """Build deal comparison for top 3 counties."""
        scenarios = []
        for county in top_counties[:3]:
            price = float(county.get("electricity_price_cents_kwh") or 7)
            annual_power = target_mw * 1000 * 8760 * (price / 100) * 1.25
            build_cost = target_mw * 10_000_000  # Mid-tier estimate
            annual_revenue = target_mw * 1000 * 8760 * 0.12 * 0.85  # $0.12/kWh, 85% util
            annual_opex = annual_power + build_cost * 0.03  # Power + 3% maintenance
            annual_noi = annual_revenue - annual_opex
            irr = (annual_noi / build_cost) * 100

            scenarios.append({
                "county_name": county.get("county_name", "Unknown"),
                "state": county.get("state", ""),
                "overall_score": county.get("overall_score", 0),
                "grade": county.get("grade", ""),
                "build_cost": build_cost,
                "annual_power_cost": annual_power,
                "annual_revenue": annual_revenue,
                "annual_noi": annual_noi,
                "estimated_irr": round(irr, 1),
                "electricity_rate": price,
            })
        return scenarios

    def _get_data_sources(self, db: Session) -> List[Dict]:
        """List data sources with freshness info."""
        sources = [
            {"name": "EIA Power Plants", "table": "power_plant", "weight": "30% (Power)"},
            {"name": "HIFLD Substations", "table": "substation", "weight": "30% (Power)"},
            {"name": "HIFLD Transmission Lines", "table": "transmission_line", "weight": "30% (Power)"},
            {"name": "PeeringDB", "table": "data_center_facility", "weight": "20% (Connectivity)"},
            {"name": "FCC Broadband", "table": "broadband_availability", "weight": "20% (Connectivity)"},
            {"name": "Census Building Permits", "table": "building_permit", "weight": "20% (Regulatory)"},
            {"name": "Census of Governments", "table": "government_unit", "weight": "20% (Regulatory)"},
            {"name": "BLS QCEW", "table": "industry_employment", "weight": "15% (Labor)"},
            {"name": "FEMA NRI", "table": "national_risk_index", "weight": "10% (Risk)"},
            {"name": "USGS 3DEP Elevation", "table": "county_elevation", "weight": "10% (Risk)"},
            {"name": "NWI Wetlands", "table": "wetland", "weight": "10% (Risk)"},
            {"name": "FEMA Flood Zones", "table": "flood_zone", "weight": "10% (Risk)"},
            {"name": "EPA Brownfields", "table": "brownfield_site", "weight": "10% (Risk)"},
            {"name": "EPA Envirofacts", "table": "environmental_facility", "weight": "10% (Risk)"},
            {"name": "Good Jobs First", "table": "incentive_deal", "weight": "5% (Cost)"},
            {"name": "NREL Solar/Wind", "table": "renewable_resource", "weight": "Power analysis"},
            {"name": "Epoch AI DCs", "table": "epoch_datacenter", "weight": "Connectivity analysis"},
            {"name": "State EDO Sites", "table": "industrial_site", "weight": "Real estate analysis"},
        ]
        for src in sources:
            try:
                result = db.execute(
                    text(f"SELECT COUNT(*), MAX(collected_at) FROM {src['table']}")
                )
                row = result.fetchone()
                src["row_count"] = row[0] if row else 0
                src["last_collected"] = row[1].strftime("%Y-%m-%d") if row and row[1] else "N/A"
            except Exception:
                db.rollback()
                src["row_count"] = 0
                src["last_collected"] = "N/A"
        return sources

    def _compute_ceo_overview(self, data: Dict) -> Dict:
        """Build data-driven narrative overview from all report data."""
        summary = data.get("summary", {})
        top_counties = data.get("top_counties", [])
        deal_scenarios = data.get("deal_scenarios", [])
        data_sources = data.get("data_sources", [])

        total = summary.get("total_counties", 0)
        a_grade = summary.get("a_grade", 0)
        b_grade = summary.get("b_grade", 0)
        avg_score = float(summary.get("avg_score") or 0)
        max_score = float(summary.get("max_score") or 0)
        states = summary.get("states_covered", 0)

        # --- Paragraph 1: Market scope ---
        a_pct = round(a_grade / max(total, 1) * 100)
        if a_pct >= 50:
            quality_note = "a strong majority scored A-grade, indicating broadly favorable conditions"
        elif a_pct >= 20:
            quality_note = "a meaningful cluster of A-grade counties emerged as clear frontrunners"
        elif a_grade > 0:
            quality_note = "a select few counties stand out with A-grade scores"
        else:
            quality_note = "no counties reached A-grade threshold, suggesting challenging conditions across the board"

        scope_para = (
            f"This analysis evaluated <strong>{total:,}</strong> counties across "
            f"<strong>{states}</strong> state{'s' if states != 1 else ''}. "
            f"Average suitability score is <strong>{avg_score:.1f}</strong> out of 100, "
            f"with the top-ranked county reaching <strong>{max_score:.1f}</strong>. "
            f"Of all counties scored, <strong>{a_grade:,}</strong> ({a_pct}%) achieved A-grade (80+) — "
            f"{quality_note}."
        )

        # --- Paragraph 2: Top opportunities ---
        if top_counties:
            top3 = top_counties[:3]
            top_names = ", ".join(
                f"<strong>{c.get('county_name', '?')}, {c.get('state', '?')}</strong> ({c.get('overall_score', 0):.1f})"
                for c in top3
            )
            # Find strongest and weakest domain across top counties
            domain_keys = [
                ("power_score", "Power"), ("connectivity_score", "Connectivity"),
                ("regulatory_score", "Regulatory"), ("labor_score", "Labor"),
                ("risk_score", "Risk"), ("cost_incentive_score", "Cost/Incentives"),
            ]
            domain_avgs = {}
            for key, label in domain_keys:
                vals = [float(c.get(key) or 0) for c in top3]
                domain_avgs[label] = sum(vals) / max(len(vals), 1)
            best_domain = max(domain_avgs, key=domain_avgs.get)
            worst_domain = min(domain_avgs, key=domain_avgs.get)

            opps_para = (
                f"The highest-scoring candidates are {top_names}. "
                f"Across the top counties, <strong>{best_domain}</strong> is the strongest dimension "
                f"(avg {domain_avgs[best_domain]:.0f}/100), while <strong>{worst_domain}</strong> "
                f"is the weakest (avg {domain_avgs[worst_domain]:.0f}/100) and warrants "
                f"closer due diligence."
            )
        else:
            opps_para = "No counties have been scored yet. Run the scoring pipeline to populate rankings."

        # --- Paragraph 3: Data quality & next steps ---
        populated = sum(1 for s in data_sources if s.get("row_count", 0) > 0)
        total_sources = len(data_sources)
        coverage_pct = round(populated / max(total_sources, 1) * 100)

        gaps = [s["name"] for s in data_sources if s.get("row_count", 0) == 0]
        gap_note = ""
        if gaps:
            shown = gaps[:5]
            gap_note = f" Missing data: {', '.join(shown)}{'.' if len(gaps) <= 5 else f', and {len(gaps) - 5} more.'}"

        worst_domain = None
        if top_counties:
            # already computed above in the opps_para block
            domain_keys_check = [
                ("power_score", "Power"), ("connectivity_score", "Connectivity"),
                ("regulatory_score", "Regulatory"), ("labor_score", "Labor"),
                ("risk_score", "Risk"), ("cost_incentive_score", "Cost/Incentives"),
            ]
            _davgs = {}
            for key, label in domain_keys_check:
                vals = [float(c.get(key) or 0) for c in top_counties[:3]]
                _davgs[label] = sum(vals) / max(len(vals), 1)
            worst_domain = min(_davgs, key=_davgs.get)

        next_steps = []
        if coverage_pct < 50:
            next_steps.append("Ingest missing data sources to improve score accuracy")
        if not deal_scenarios:
            next_steps.append("Run deal scenario modeling once data coverage exceeds 50%")
        if a_grade > 0:
            next_steps.append("Conduct site visits for top A-grade counties")
        if worst_domain:
            next_steps.append(f"Deep-dive {worst_domain} data for shortlisted counties")

        best_irr = max((s.get("estimated_irr", 0) for s in deal_scenarios), default=0)

        data_para = (
            f"Data coverage: <strong>{populated}</strong> of {total_sources} sources populated "
            f"(<strong>{coverage_pct}%</strong>).{gap_note}"
        )

        return {
            "scope_para": scope_para,
            "opps_para": opps_para,
            "data_para": data_para,
            "coverage_pct": coverage_pct,
            "best_irr": best_irr,
            "top_counties": top_counties[:3],
            "next_steps": next_steps,
            "data_gaps": gaps,
        }

    # ------------------------------------------------------------------
    # HTML Rendering
    # ------------------------------------------------------------------

    def render_html(self, data: Dict) -> str:
        """Render full HTML report."""
        summary = data.get("summary", {})
        state_averages = data.get("state_averages", [])
        top_counties = data.get("top_counties", [])
        power = data.get("power_analysis", {})
        connectivity = data.get("connectivity", {})
        regulatory = data.get("regulatory", [])
        real_estate = data.get("real_estate", {})
        incentives = data.get("incentives", {})
        environment = data.get("environment", {})
        transmission = data.get("transmission", {})
        workforce = data.get("workforce", [])
        dc_clusters = data.get("dc_clusters", {})
        capital_model = data.get("capital_model", {})
        deal_scenarios = data.get("deal_scenarios", [])
        data_sources = data.get("data_sources", [])
        ceo = data.get("ceo_overview", {})
        params = data.get("params", {})
        generated_at = data.get("generated_at", "")

        title = data.get("report_title", "Datacenter Site Selection Report")
        state_label = params.get("state", "National")
        charts_js = ""
        body = ""

        # Header
        body += page_header(
            title,
            f"County-level suitability analysis | {state_label}",
            f"Generated {generated_at}",
        )

        # KPI Strip
        body += kpi_strip(
            kpi_card("Counties Analyzed", _fmt(summary.get("total_counties")))
            + kpi_card("A-Grade", _fmt(summary.get("a_grade")))
            + kpi_card("States", _fmt(summary.get("states_covered")))
            + kpi_card("Top Score", _fmt(summary.get("max_score"), decimals=1))
            + kpi_card("Avg Score", _fmt(summary.get("avg_score"), decimals=1))
        )

        # TOC
        body += toc([
            {"number": "★", "id": "ceo", "title": "CEO Overview"},
            {"number": 1, "id": "exec-summary", "title": "Executive Summary"},
            {"number": 2, "id": "geo-heat", "title": "Geographic Heat Map"},
            {"number": 3, "id": "top-counties", "title": "Top Candidate Counties"},
            {"number": 4, "id": "power", "title": "Power Grid Analysis"},
            {"number": 5, "id": "connectivity", "title": "Connectivity & Fiber"},
            {"number": 6, "id": "regulatory", "title": "Regulatory & Permitting"},
            {"number": 7, "id": "real-estate", "title": "Real Estate & Land"},
            {"number": 8, "id": "incentives", "title": "Tax & Incentives"},
            {"number": 9, "id": "environment", "title": "Environmental & Risk"},
            {"number": 10, "id": "workforce", "title": "Workforce & Labor"},
            {"number": 11, "id": "dc-clusters", "title": "Existing DC Clusters"},
            {"number": 12, "id": "capital", "title": "Capital Model"},
            {"number": 13, "id": "deal-scenarios", "title": "Deal Scenarios"},
            {"number": 14, "id": "data-sources", "title": "Data Sources"},
        ])

        # ★ CEO Overview — data-driven narrative
        body += '<div class="container" id="ceo">'
        body += '<div class="ceo-overview">'

        body += '<h3>Market Scope</h3>'
        body += f'<p class="narrative-text">{ceo.get("scope_para", "")}</p>'

        body += '<h3>Top Opportunities</h3>'
        body += f'<p class="narrative-text">{ceo.get("opps_para", "")}</p>'

        # Top recommendations as table
        ceo_top = ceo.get("top_counties", [])
        if ceo_top:
            rec_headers = ["County", "State", "Score", "Grade"]
            rec_rows = [[
                c.get("county_name", ""),
                c.get("state", ""),
                _fmt(c.get("overall_score"), decimals=1),
                c.get("grade", ""),
            ] for c in ceo_top]
            body += data_table(rec_headers, rec_rows, numeric_columns=[2])

        body += '<h3>Data Coverage &amp; Next Steps</h3>'
        body += f'<p class="narrative-text">{ceo.get("data_para", "")}</p>'

        next_steps = ceo.get("next_steps", [])
        if next_steps:
            body += callout(
                "<strong>Recommended next steps:</strong> " + " · ".join(next_steps),
                "info",
            )

        if ceo.get("data_gaps"):
            body += callout(
                f"<strong>{len(ceo['data_gaps'])} data sources have no records.</strong> "
                "Scores are based on available data only — ingest additional sources "
                "to improve accuracy.",
                "warn",
            )

        body += "</div></div>"  # ceo-overview, container

        # Section 1: Executive Summary
        body += section_start(1, "Executive Summary", "exec-summary")
        body += callout(
            f"<strong>Investment Thesis:</strong> {summary.get('total_counties', 0)} counties analyzed "
            f"across {summary.get('states_covered', 0)} states. {summary.get('a_grade', 0)} counties "
            f"scored A-grade (80+) for datacenter suitability based on power, connectivity, "
            f"regulatory speed, workforce, risk, and incentives.",
            "info",
        )
        body += section_end()

        # Section 2: Geographic Heat Map (bar chart of avg score by state)
        body += section_start(2, "Geographic Heat Map", "geo-heat")
        if state_averages:
            labels = [s["state"] for s in state_averages[:15]]
            values = [float(s["avg_score"] or 0) for s in state_averages[:15]]
            bar_cfg = build_horizontal_bar_config(labels, values, dataset_label="Avg Score")
            bar_json = json.dumps(bar_cfg)
            fallback = build_bar_fallback(labels, values, BLUE)
            chart_h = f"{max(300, len(labels) * 36)}px"
            body += chart_container("chart_state_avg", bar_json, fallback, title="Average Site Score by State", height=chart_h)
            charts_js += chart_init_js("chart_state_avg", bar_json)
        else:
            body += callout("No state-level data available yet. Run scoring first.", "warn")
        body += section_end()

        # Section 3: Top Candidate Counties
        body += section_start(3, "Top Candidate Counties", "top-counties")
        if top_counties:
            headers = ["Rank", "County", "State", "Score", "Grade", "Power", "Conn.", "Reg.", "Labor", "Risk", "Cost"]
            rows = []
            for c in top_counties:
                rows.append([
                    c.get("national_rank", ""),
                    c.get("county_name", ""),
                    c.get("state", ""),
                    _fmt(c.get("overall_score"), decimals=1),
                    c.get("grade", ""),
                    _fmt(c.get("power_score"), decimals=0),
                    _fmt(c.get("connectivity_score"), decimals=0),
                    _fmt(c.get("regulatory_score"), decimals=0),
                    _fmt(c.get("labor_score"), decimals=0),
                    _fmt(c.get("risk_score"), decimals=0),
                    _fmt(c.get("cost_incentive_score"), decimals=0),
                ])
            body += data_table(headers, rows, numeric_columns=[0, 3, 5, 6, 7, 8, 9, 10])
        else:
            body += callout("No county scores available. Run scoring first.", "warn")
        body += section_end()

        # Section 4: Power Grid Analysis
        body += section_start(4, "Power Grid Analysis", "power")
        prices = power.get("prices", [])
        if prices:
            labels = [p["state"] for p in prices[:10]]
            values = [float(p["avg_rate"] or 0) for p in prices[:10]]
            cfg = build_horizontal_bar_config(labels, values, dataset_label="cents/kWh")
            cfg_json = json.dumps(cfg)
            body += chart_container("chart_elec_price", cfg_json, build_bar_fallback(labels, values, GREEN), title="Avg Commercial Electricity Rate by State")
            charts_js += chart_init_js("chart_elec_price", cfg_json)
        else:
            body += callout("No electricity price data available.", "warn")

        # Transmission lines
        tx_summary = transmission.get("summary", {})
        tx_voltage = transmission.get("by_voltage", [])
        tx_by_state = transmission.get("by_state", [])
        if tx_summary.get("total_lines"):
            body += "<h4 style='margin-top:24px'>Transmission Line Infrastructure</h4>"
            body += callout(
                f"<strong>{_fmt(tx_summary.get('total_lines'))} transmission lines</strong> "
                f"spanning <strong>{_fmt(tx_summary.get('total_miles'))} miles</strong> "
                f"across <strong>{_fmt(tx_summary.get('unique_owners'))} owners</strong>. "
                f"Average voltage: <strong>{_fmt(tx_summary.get('avg_voltage_kv'))} kV</strong>.",
                "info",
            )
        if tx_voltage:
            headers = ["Voltage Class", "Lines", "Total Miles", "Avg kV"]
            rows = [[
                v.get("voltage_class") or "Unknown",
                _fmt(v.get("line_count")),
                _fmt(v.get("total_miles")),
                _fmt(v.get("avg_kv")),
            ] for v in tx_voltage]
            body += data_table(headers, rows, numeric_columns=[1, 2, 3])
        if tx_by_state:
            labels = [s["state"] for s in tx_by_state[:10]]
            values = [s["line_count"] for s in tx_by_state[:10]]
            cfg = build_horizontal_bar_config(labels, values, dataset_label="Transmission Lines")
            cfg_json = json.dumps(cfg)
            body += chart_container("chart_tx_lines", cfg_json, build_bar_fallback(labels, values, BLUE), title="Transmission Lines by State")
            charts_js += chart_init_js("chart_tx_lines", cfg_json)
        body += section_end()

        # Section 5: Connectivity & Fiber
        body += section_start(5, "Connectivity & Fiber", "connectivity")
        ix_data = connectivity.get("ix_by_city", [])
        if ix_data:
            labels = [d["city"] for d in ix_data[:10]]
            values = [d["ix_count"] for d in ix_data[:10]]
            cfg = build_horizontal_bar_config(labels, values, dataset_label="Internet Exchanges")
            cfg_json = json.dumps(cfg)
            body += chart_container("chart_ix", cfg_json, build_bar_fallback(labels, values, BLUE_LIGHT), title="Internet Exchanges by City")
            charts_js += chart_init_js("chart_ix", cfg_json)
        else:
            body += callout("No IX data available.", "warn")
        body += section_end()

        # Section 6: Regulatory & Permitting
        body += section_start(6, "Regulatory & Permitting", "regulatory")
        if regulatory:
            headers = ["County", "State", "Score", "Grade", "Permits", "Simplicity", "Energy", "DC Deals"]
            rows = []
            for r in regulatory:
                rows.append([
                    r.get("county_name", ""),
                    r.get("state", ""),
                    _fmt(r.get("overall_score"), decimals=1),
                    r.get("grade", ""),
                    _fmt(r.get("permit_velocity_score"), decimals=0),
                    _fmt(r.get("jurisdictional_simplicity_score"), decimals=0),
                    _fmt(r.get("energy_siting_score"), decimals=0),
                    _fmt(r.get("historical_dc_deals_score"), decimals=0),
                ])
            body += data_table(headers, rows, numeric_columns=[2, 4, 5, 6, 7])
        else:
            body += callout("No regulatory scores available.", "warn")
        body += section_end()

        # Section 7: Real Estate & Land
        body += section_start(7, "Real Estate & Land", "real-estate")
        sites = real_estate.get("industrial_sites", [])
        if sites:
            headers = ["Site Name", "City", "State", "Acres", "Zoning", "Rail", "Highway"]
            rows = [[
                s.get("site_name", ""),
                s.get("city", ""),
                s.get("state", ""),
                _fmt(s.get("acreage"), decimals=0),
                s.get("zoning", ""),
                "Yes" if s.get("rail_served") else "No",
                s.get("highway_access", ""),
            ] for s in sites]
            body += data_table(headers, rows, numeric_columns=[3])
        brownfields = real_estate.get("brownfields", [])
        if brownfields:
            body += "<h4 style='margin-top:20px'>Brownfield Sites (EPA ACRES)</h4>"
            headers = ["Site", "City", "State", "Acres", "Status", "Current Use"]
            rows = [[
                b.get("site_name", ""),
                b.get("city", ""),
                b.get("state", ""),
                _fmt(b.get("acreage"), decimals=0),
                b.get("cleanup_status", ""),
                b.get("land_use_current", ""),
            ] for b in brownfields]
            body += data_table(headers, rows, numeric_columns=[3])
        if not sites and not brownfields:
            body += callout("No real estate data available.", "warn")
        body += section_end()

        # Section 8: Tax & Incentives
        body += section_start(8, "Tax & Incentives", "incentives")
        programs = incentives.get("programs", [])
        if programs:
            headers = ["Program", "State", "Type", "Target Industries", "Max Benefit"]
            rows = [[
                p.get("program_name", ""),
                p.get("state", ""),
                p.get("program_type", ""),
                ", ".join(p["target_industries"]) if isinstance(p.get("target_industries"), list) else str(p.get("target_industries", "")),
                _fmt(p.get("max_benefit"), prefix="$"),
            ] for p in programs]
            body += data_table(headers, rows)
        deals = incentives.get("deals", [])
        if deals:
            body += "<h4 style='margin-top:20px'>Incentive Deals</h4>"
            headers = ["Company", "State", "County", "Subsidy", "Year", "Jobs"]
            rows = [[
                d.get("company_name") or "",
                d.get("state") or "",
                d.get("county") or "",
                _fmt(d.get("subsidy_value"), prefix="$"),
                d.get("year") or "",
                _fmt(d.get("jobs_announced")),
            ] for d in deals]
            body += data_table(headers, rows, numeric_columns=[3, 5])
        if not programs and not deals:
            body += callout("No incentive data available.", "warn")
        body += section_end()

        # Section 9: Environmental & Risk
        body += section_start(9, "Environmental & Risk", "environment")

        # Elevation summary
        elev_summary = environment.get("elev_summary", {})
        if elev_summary.get("counties"):
            body += callout(
                f"<strong>Elevation data across {_fmt(elev_summary.get('counties'))} counties:</strong> "
                f"Average elevation <strong>{_fmt(elev_summary.get('avg_elev'))} ft</strong>, "
                f"range from <strong>{_fmt(elev_summary.get('min_elev'))} ft</strong> "
                f"to <strong>{_fmt(elev_summary.get('max_elev'))} ft</strong>. "
                f"Low-elevation sites reduce grading costs; moderate elevation reduces flood risk.",
                "info",
            )

        # Elevation detail table
        elevation_data = environment.get("elevation", [])
        if elevation_data:
            body += "<h4 style='margin-top:20px'>County Elevation Profile (Lowest)</h4>"
            headers = ["County", "State", "FIPS", "Min (ft)", "Max (ft)", "Mean (ft)", "Range (ft)"]
            rows = [[
                e.get("county", ""),
                e.get("state", ""),
                e.get("fips_code", ""),
                _fmt(e.get("min_ft")),
                _fmt(e.get("max_ft")),
                _fmt(e.get("mean_ft")),
                _fmt(e.get("range_ft")),
            ] for e in elevation_data]
            body += data_table(headers, rows, numeric_columns=[3, 4, 5, 6])

        # FEMA NRI risk data
        risk_data = environment.get("risk_by_county", [])
        if risk_data:
            body += "<h4 style='margin-top:20px'>Natural Hazard Risk (FEMA NRI)</h4>"
            headers = ["County", "State", "Rating", "Risk Score", "Tornado", "Hurricane", "Wildfire"]
            rows = [[
                r.get("county_name", ""),
                r.get("state", ""),
                r.get("risk_rating", ""),
                _fmt(r.get("risk_score"), decimals=1),
                _fmt(r.get("tornado_score"), decimals=1),
                _fmt(r.get("hurricane_score"), decimals=1),
                _fmt(r.get("wildfire_score"), decimals=1),
            ] for r in risk_data]
            body += data_table(headers, rows, numeric_columns=[3, 4, 5, 6])

        # Flood zones
        flood_data = environment.get("flood_zones", [])
        if flood_data:
            body += "<h4 style='margin-top:20px'>FEMA Flood Zones</h4>"
            headers = ["State", "Zone Code", "Description", "High Risk", "Count"]
            rows = [[
                f.get("state", ""),
                f.get("zone_code", ""),
                f.get("zone_description", ""),
                "Yes" if f.get("is_high_risk") else "No",
                _fmt(f.get("zone_count")),
            ] for f in flood_data]
            body += data_table(headers, rows, numeric_columns=[4])

        # Wetlands
        wetland_data = environment.get("wetlands", [])
        if wetland_data:
            body += "<h4 style='margin-top:20px'>NWI Wetlands</h4>"
            headers = ["State", "Wetland Type", "Count", "Total Acres"]
            rows = [[
                w.get("state", ""),
                w.get("wetland_type", ""),
                _fmt(w.get("wetland_count")),
                _fmt(w.get("total_acres")),
            ] for w in wetland_data]
            body += data_table(headers, rows, numeric_columns=[2, 3])

        if not risk_data and not elevation_data and not flood_data and not wetland_data:
            body += callout("No environmental risk data available.", "warn")
        body += section_end()

        # Section 10: Workforce & Labor
        body += section_start(10, "Workforce & Labor", "workforce")
        if workforce:
            headers = ["Area", "Industry", "Employment", "Avg Weekly Wage", "Establishments"]
            rows = [[
                w.get("area_name", ""),
                w.get("industry_title", ""),
                _fmt(w.get("avg_monthly_employment")),
                _fmt(w.get("avg_weekly_wage"), prefix="$"),
                _fmt(w.get("establishments")),
            ] for w in workforce]
            body += data_table(headers, rows, numeric_columns=[2, 3, 4])
        else:
            body += callout("No workforce data available for tech sectors.", "warn")
        body += section_end()

        # Section 11: Existing DC Clusters
        body += section_start(11, "Existing DC Clusters", "dc-clusters")
        peeringdb = dc_clusters.get("peeringdb", [])
        if peeringdb:
            headers = ["Facility", "Operator", "City", "State"]
            rows = [[f.get("name", ""), f.get("operator", ""), f.get("city", ""), f.get("state", "")] for f in peeringdb]
            body += data_table(headers, rows)
        epoch = dc_clusters.get("epoch", [])
        if epoch:
            body += "<h4 style='margin-top:20px'>Epoch AI Datacenters</h4>"
            headers = ["Company", "Facility", "City", "State", "Power (MW)", "Year", "Status"]
            rows = [[
                e.get("company", ""),
                e.get("facility_name", ""),
                e.get("city", ""),
                e.get("state", ""),
                _fmt(e.get("power_capacity_mw"), decimals=0),
                e.get("year_opened", ""),
                e.get("status", ""),
            ] for e in epoch]
            body += data_table(headers, rows, numeric_columns=[4])
        if not peeringdb and not epoch:
            body += callout("No datacenter cluster data available.", "warn")
        body += section_end()

        # Section 12: Capital Model
        body += section_start(12, "Capital Model", "capital")
        if capital_model:
            body += '<div class="deal-scenario-grid">'
            for tier_name, model in capital_model.items():
                body += f"""
                <div class="deal-card">
                    <h4>{tier_name}</h4>
                    <div class="metric-grid">
                        <div class="metric-item"><div class="value">{_fmt(model['cost_per_mw'], prefix='$')}</div><div class="label">Cost/MW</div></div>
                        <div class="metric-item"><div class="value">{_fmt(model['total_build_cost'], prefix='$')}</div><div class="label">Total Build</div></div>
                        <div class="metric-item"><div class="value">{model['pue']}</div><div class="label">PUE</div></div>
                        <div class="metric-item"><div class="value">{model['uptime_sla']}%</div><div class="label">Uptime SLA</div></div>
                    </div>
                    <p class="muted-text" style="margin-top:12px">
                        Annual power @ 5c/kWh: {_fmt(model['annual_power_cost_5c'], prefix='$')} |
                        @ 8c/kWh: {_fmt(model['annual_power_cost_8c'], prefix='$')}
                    </p>
                </div>
                """
            body += "</div>"
        body += section_end()

        # Section 13: Deal Scenarios
        body += section_start(13, "Deal Scenarios", "deal-scenarios")
        if deal_scenarios:
            body += '<div class="deal-scenario-grid">'
            for i, scenario in enumerate(deal_scenarios):
                rec_cls = "recommended" if i == 0 else ""
                body += f"""
                <div class="deal-card {rec_cls}">
                    <h4>{scenario['county_name']}, {scenario['state']}</h4>
                    <p class="muted-text">Score: {scenario['overall_score']} | Grade: {scenario['grade']}</p>
                    <div class="metric-grid">
                        <div class="metric-item"><div class="value">{_fmt(scenario['build_cost'], prefix='$')}</div><div class="label">Build Cost</div></div>
                        <div class="metric-item"><div class="value">{_fmt(scenario['annual_revenue'], prefix='$')}</div><div class="label">Annual Revenue</div></div>
                        <div class="metric-item"><div class="value">{scenario['estimated_irr']}%</div><div class="label">Est. IRR</div></div>
                    </div>
                    <p class="muted-text" style="margin-top:8px">
                        Electricity: {scenario['electricity_rate']} c/kWh |
                        Annual Power: {_fmt(scenario['annual_power_cost'], prefix='$')} |
                        NOI: {_fmt(scenario['annual_noi'], prefix='$')}
                    </p>
                </div>
                """
            body += "</div>"
        else:
            body += callout("No deal scenarios available — score counties first.", "warn")
        body += section_end()

        # Section 14: Data Sources
        body += section_start(14, "Data Sources", "data-sources")
        if data_sources:
            headers = ["Source", "Table", "Weight", "Rows", "Last Collected"]
            rows = [[
                s["name"], s["table"], s["weight"],
                _fmt(s.get("row_count")), s.get("last_collected", "N/A"),
            ] for s in data_sources]
            body += data_table(headers, rows, numeric_columns=[3])

        # Weights table
        from app.ml.datacenter_site_metadata import DOMAIN_DOCUMENTATION
        body += "<h4 style='margin-top:20px'>Scoring Weights</h4>"
        weight_headers = ["Domain", "Weight", "Description"]
        weight_rows = [[
            k.replace("_", " ").title(),
            f"{int(v['weight'] * 100)}%",
            v["description"][:100],
        ] for k, v in DOMAIN_DOCUMENTATION.items()]
        body += data_table(weight_headers, weight_rows, numeric_columns=[1])

        body += callout(
            "All data sourced from free public APIs. Scores are relative rankings "
            "and should be validated with site-specific due diligence.",
            "info",
        )
        body += section_end()

        # Footer
        body += page_footer(
            ["Datacenter Site Selection Analysis — Nexdata Intelligence Platform"],
            f"Report generated {generated_at}",
        )

        return html_document(title, body, charts_js, DATACENTER_EXTRA_CSS)

    # ------------------------------------------------------------------
    # Excel Rendering
    # ------------------------------------------------------------------

    def render_excel(self, data: Dict) -> bytes:
        """Render Excel workbook with 7 worksheets."""
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")

        def write_sheet(ws, title, headers, rows):
            ws.title = title
            ws.append(headers)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            for row in rows:
                ws.append(row)
            for col in ws.columns:
                max_len = max(len(str(c.value or "")) for c in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

        # Sheet 1: Summary
        ws1 = wb.active
        summary = data.get("summary", {})
        write_sheet(ws1, "Summary", ["Metric", "Value"], [
            ["Counties Analyzed", summary.get("total_counties", 0)],
            ["A-Grade Counties", summary.get("a_grade", 0)],
            ["States Covered", summary.get("states_covered", 0)],
            ["Average Score", summary.get("avg_score", 0)],
            ["Top Score", summary.get("max_score", 0)],
            ["Verdict", data.get("ceo_overview", {}).get("verdict", "N/A")],
            ["Conviction", data.get("ceo_overview", {}).get("conviction", 0)],
        ])

        # Sheet 2: Rankings
        ws2 = wb.create_sheet()
        top = data.get("top_counties", [])
        headers = ["Rank", "County", "State", "Score", "Grade", "Power", "Connectivity", "Regulatory", "Labor", "Risk", "Cost"]
        rows = [[
            c.get("national_rank"), c.get("county_name"), c.get("state"),
            c.get("overall_score"), c.get("grade"),
            c.get("power_score"), c.get("connectivity_score"),
            c.get("regulatory_score"), c.get("labor_score"),
            c.get("risk_score"), c.get("cost_incentive_score"),
        ] for c in top]
        write_sheet(ws2, "Rankings", headers, rows)

        # Sheet 3: Power
        ws3 = wb.create_sheet()
        prices = data.get("power_analysis", {}).get("prices", [])
        write_sheet(ws3, "Power", ["State", "Avg Rate (c/kWh)"], [
            [p.get("state"), p.get("avg_rate")] for p in prices
        ])

        # Sheet 4: Connectivity
        ws4 = wb.create_sheet()
        ix_data = data.get("connectivity", {}).get("ix_by_state", [])
        write_sheet(ws4, "Connectivity", ["State", "IX Count"], [
            [d.get("state"), d.get("ix_count")] for d in ix_data
        ])

        # Sheet 5: Regulatory
        ws5 = wb.create_sheet()
        reg = data.get("regulatory", [])
        write_sheet(ws5, "Regulatory", ["County", "State", "Score", "Grade", "Permits", "Simplicity"], [
            [r.get("county_name"), r.get("state"), r.get("overall_score"),
             r.get("grade"), r.get("permit_velocity_score"),
             r.get("jurisdictional_simplicity_score")] for r in reg
        ])

        # Sheet 6: Cost Model
        ws6 = wb.create_sheet()
        scenarios = data.get("deal_scenarios", [])
        write_sheet(ws6, "Cost Model", ["County", "State", "Score", "Build Cost", "Annual Revenue", "IRR %"], [
            [s.get("county_name"), s.get("state"), s.get("overall_score"),
             s.get("build_cost"), s.get("annual_revenue"),
             s.get("estimated_irr")] for s in scenarios
        ])

        # Sheet 7: Sources
        ws7 = wb.create_sheet()
        sources = data.get("data_sources", [])
        write_sheet(ws7, "Sources", ["Source", "Table", "Weight", "Rows", "Last Collected"], [
            [s["name"], s["table"], s["weight"], s.get("row_count"), s.get("last_collected")] for s in sources
        ])

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
