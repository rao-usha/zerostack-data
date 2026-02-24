"""
Investor Profile Report Template.

Generates a one-pager with investor overview, portfolio summary,
top holdings, sector allocation, and team/leadership.
"""

import json
import logging
from datetime import datetime
from typing import Dict, Any, Optional
from io import BytesIO

from sqlalchemy import text
from sqlalchemy.orm import Session

from app.reports.design_system import (
    html_document, page_header, kpi_strip, kpi_card,
    toc, section_start, section_end,
    data_table, pill_badge, profile_card, callout,
    chart_container, chart_init_js, page_footer,
    build_doughnut_config, build_horizontal_bar_config,
    build_bar_fallback, build_chart_legend, CHART_COLORS,
)

logger = logging.getLogger(__name__)


def _format_aum(aum_millions: Optional[float]) -> str:
    """Format AUM in human-readable form (e.g. $1T, $553B, $40B)."""
    if aum_millions is None:
        return "N/A"
    if aum_millions >= 1_000_000:
        val = aum_millions / 1_000_000
        return f"${val:,.1f}T" if val != int(val) else f"${int(val)}T"
    if aum_millions >= 1_000:
        val = aum_millions / 1_000
        return f"${val:,.0f}B"
    return f"${aum_millions:,.0f}M"


def _format_market_cap(mcap: Optional[float]) -> str:
    """Format market cap in human-readable form (e.g. $2.1T, $45.3B, $800M)."""
    if mcap is None:
        return "-"
    if mcap >= 1e12:
        return f"${mcap / 1e12:,.1f}T"
    if mcap >= 1e9:
        return f"${mcap / 1e9:,.1f}B"
    if mcap >= 1e6:
        return f"${mcap / 1e6:,.0f}M"
    if mcap >= 1e3:
        return f"${mcap / 1e3:,.0f}K"
    return f"${mcap:,.0f}"


class InvestorProfileTemplate:
    """Investor profile report template."""

    name = "investor_profile"
    description = "One-pager investor profile with portfolio summary"

    def gather_data(self, db: Session, params: Dict[str, Any]) -> Dict[str, Any]:
        """Gather all data needed for the report."""
        investor_id = params.get("investor_id")
        investor_type = params.get("investor_type", "lp")

        investor = self._get_investor(db, investor_id, investor_type)

        data = {
            "generated_at": datetime.utcnow().isoformat(),
            "investor": investor,
            "portfolio_summary": self._get_portfolio_summary(
                db, investor_id, investor_type, investor.get("name")
            ),
            "top_holdings": self._get_top_holdings(
                db, investor_id, investor_type, investor.get("name")
            ),
            "sector_allocation": self._get_sector_allocation(
                db, investor_id, investor_type, investor.get("name")
            ),
            "recent_activity": self._get_recent_activity(
                db, investor_id, investor_type
            ),
            "team": self._get_team(db, investor_id, investor_type),
            "segments": self._get_segments(db, investor_id, investor_type),
        }

        return data

    def _get_investor(self, db: Session, investor_id: int, investor_type: str) -> Dict:
        """Get investor details."""
        if investor_type == "lp":
            result = db.execute(
                text("""
                SELECT id, name, lp_type, jurisdiction, website_url
                FROM lp_fund WHERE id = :id
            """),
                {"id": investor_id},
            )
            row = result.fetchone()
            if row:
                return {
                    "id": row[0],
                    "name": row[1],
                    "type": row[2],
                    "jurisdiction": row[3],
                    "website": row[4],
                    "aum_millions": None,
                }
        elif investor_type == "pe_firm":
            result = db.execute(
                text("""
                SELECT id, name, firm_type, primary_strategy,
                       headquarters_city, headquarters_state, headquarters_country,
                       website, aum_usd_millions, employee_count, founded_year
                FROM pe_firms WHERE id = :id
            """),
                {"id": investor_id},
            )
            row = result.fetchone()
            if row:
                location_parts = [p for p in [row[4], row[5], row[6]] if p]
                return {
                    "id": row[0],
                    "name": row[1],
                    "type": f"{row[2]} — {row[3]}" if row[3] else row[2],
                    "jurisdiction": ", ".join(location_parts) if location_parts else None,
                    "website": row[7],
                    "aum_millions": float(row[8]) if row[8] else None,
                    "employee_count": row[9],
                    "founded_year": row[10],
                }
        else:
            result = db.execute(
                text("""
                SELECT id, name, 'family_office' as lp_type,
                       CONCAT_WS(', ', city, state_province, country) as jurisdiction,
                       website, estimated_aum_millions
                FROM family_offices WHERE id = :id
            """),
                {"id": investor_id},
            )
            row = result.fetchone()
            if row:
                return {
                    "id": row[0],
                    "name": row[1],
                    "type": row[2],
                    "jurisdiction": row[3],
                    "website": row[4],
                    "aum_millions": row[5],
                }

        return {"id": investor_id, "name": "Unknown", "type": investor_type}

    def _get_pe_portfolio_query(self, investor_id: int, firm_name: Optional[str]):
        """Build a UNION query that pulls portfolio companies from both linkage paths.

        Path 1: pe_funds -> pe_fund_investments -> pe_portfolio_companies (13F, SEC)
        Path 2: pe_portfolio_companies.current_pe_owner = firm name (website-scraped)
        Deduplicates by company name.
        """
        # We use a CTE to union both sources, then deduplicate preferring
        # website-scraped (source='website') over 13F
        query = """
            WITH all_companies AS (
                -- Path 1: via fund investments
                SELECT DISTINCT ON (pc.id)
                    pc.id, pc.name, pc.industry,
                    CONCAT_WS(', ', pc.headquarters_city, pc.headquarters_state) as location,
                    fi.investment_type as stage,
                    pc.website as company_website,
                    pc.ownership_status,
                    fi.investment_date,
                    CASE WHEN fi.investment_type = '13F Holding' THEN 1 ELSE 0 END as is_13f,
                    pc.market_cap_usd,
                    pc.ticker
                FROM pe_fund_investments fi
                JOIN pe_funds f ON f.id = fi.fund_id
                JOIN pe_portfolio_companies pc ON pc.id = fi.company_id
                WHERE f.firm_id = :investor_id AND fi.status = 'Active'

                UNION ALL

                -- Path 2: via current_pe_owner (website-scraped)
                SELECT
                    pc.id, pc.name, pc.industry,
                    CONCAT_WS(', ', pc.headquarters_city, pc.headquarters_state) as location,
                    'Portfolio Company' as stage,
                    pc.website as company_website,
                    pc.ownership_status,
                    pc.created_at::date as investment_date,
                    0 as is_13f,
                    pc.market_cap_usd,
                    pc.ticker
                FROM pe_portfolio_companies pc
                WHERE pc.current_pe_owner = :firm_name
                    AND pc.id NOT IN (
                        SELECT fi2.company_id FROM pe_fund_investments fi2
                        JOIN pe_funds f2 ON f2.id = fi2.fund_id
                        WHERE f2.firm_id = :investor_id
                    )
            ),
            deduped AS (
                SELECT DISTINCT ON (UPPER(name))
                    id, name, industry, location, stage, company_website,
                    ownership_status, investment_date, is_13f,
                    market_cap_usd, ticker
                FROM all_companies
                ORDER BY UPPER(name), is_13f ASC, investment_date DESC NULLS LAST
            )
        """
        return query

    def _get_portfolio_summary(
        self, db: Session, investor_id: int, investor_type: str,
        firm_name: Optional[str] = None,
    ) -> Dict:
        """Get portfolio summary stats."""
        if investor_type == "pe_firm":
            base = self._get_pe_portfolio_query(investor_id, firm_name)
            result = db.execute(
                text(base + """
                SELECT
                    COUNT(*) as total_holdings,
                    COUNT(DISTINCT industry) FILTER (WHERE industry IS NOT NULL) as sectors,
                    COUNT(*) as current_holdings,
                    SUM(market_cap_usd) as total_market_cap,
                    COUNT(market_cap_usd) as companies_with_mcap
                FROM deduped
                """),
                {"investor_id": investor_id, "firm_name": firm_name or ""},
            )
        else:
            result = db.execute(
                text("""
                SELECT
                    COUNT(*) as total_holdings,
                    COUNT(DISTINCT company_industry) as sectors,
                    COUNT(CASE WHEN current_holding = 1 THEN 1 END) as current_holdings,
                    NULL as total_market_cap,
                    0 as companies_with_mcap
                FROM portfolio_companies
                WHERE investor_id = :investor_id AND investor_type = :investor_type
            """),
                {"investor_id": investor_id, "investor_type": investor_type},
            )

        row = result.fetchone()
        return {
            "total_holdings": row[0] if row else 0,
            "sectors": row[1] if row else 0,
            "current_holdings": row[2] if row else 0,
            "total_market_cap": float(row[3]) if row and row[3] else None,
            "companies_with_mcap": row[4] if row else 0,
        }

    def _get_top_holdings(
        self, db: Session, investor_id: int, investor_type: str,
        firm_name: Optional[str] = None,
    ) -> list:
        """Get all portfolio holdings with market cap and ticker."""
        if investor_type == "pe_firm":
            base = self._get_pe_portfolio_query(investor_id, firm_name)
            result = db.execute(
                text(base + """
                SELECT name, industry, location, stage, market_cap_usd, ticker, ownership_status
                FROM deduped
                WHERE is_13f = 0
                ORDER BY market_cap_usd DESC NULLS LAST, name
                """),
                {"investor_id": investor_id, "firm_name": firm_name or ""},
            )
            rows = result.fetchall()
            # If no non-13F holdings, fall back to showing 13F holdings
            if not rows:
                result = db.execute(
                    text(base + """
                    SELECT name, industry, location, stage, market_cap_usd, ticker, ownership_status
                    FROM deduped
                    ORDER BY market_cap_usd DESC NULLS LAST, name
                    """),
                    {"investor_id": investor_id, "firm_name": firm_name or ""},
                )
                rows = result.fetchall()
        else:
            result = db.execute(
                text("""
                SELECT company_name, company_industry, company_location, company_stage,
                       NULL as market_cap_usd, NULL as ticker, NULL as ownership_status
                FROM portfolio_companies
                WHERE investor_id = :investor_id
                    AND investor_type = :investor_type
                    AND current_holding = 1
                ORDER BY company_name
            """),
                {
                    "investor_id": investor_id,
                    "investor_type": investor_type,
                },
            )
            rows = result.fetchall()

        return [
            {
                "name": row[0],
                "industry": row[1],
                "location": row[2],
                "stage": row[3],
                "market_cap": float(row[4]) if row[4] else None,
                "ticker": row[5],
                "ownership_status": row[6],
            }
            for row in rows
        ]

    def _get_sector_allocation(
        self, db: Session, investor_id: int, investor_type: str,
        firm_name: Optional[str] = None,
    ) -> list:
        """Get sector allocation breakdown with market cap totals."""
        if investor_type == "pe_firm":
            base = self._get_pe_portfolio_query(investor_id, firm_name)
            result = db.execute(
                text(base + """
                SELECT
                    COALESCE(industry, 'Unknown') as sector,
                    COUNT(*) as count,
                    SUM(market_cap_usd) as sector_market_cap
                FROM deduped
                GROUP BY COALESCE(industry, 'Unknown')
                ORDER BY count DESC
                """),
                {"investor_id": investor_id, "firm_name": firm_name or ""},
            )
        else:
            result = db.execute(
                text("""
                SELECT
                    COALESCE(company_industry, 'Unknown') as sector,
                    COUNT(*) as count,
                    NULL as sector_market_cap
                FROM portfolio_companies
                WHERE investor_id = :investor_id
                    AND investor_type = :investor_type
                    AND current_holding = 1
                GROUP BY company_industry
                ORDER BY count DESC
            """),
                {"investor_id": investor_id, "investor_type": investor_type},
            )

        rows = result.fetchall()
        total = sum(r[1] for r in rows)

        return [
            {
                "sector": row[0],
                "count": row[1],
                "pct": round(row[1] / total * 100, 1) if total > 0 else 0,
                "market_cap": float(row[2]) if row[2] else None,
            }
            for row in rows
        ]

    def _get_recent_activity(
        self, db: Session, investor_id: int, investor_type: str, limit: int = 5
    ) -> list:
        """Get recent portfolio changes."""
        if investor_type == "pe_firm":
            result = db.execute(
                text("""
                SELECT pc.name, pc.industry, fi.investment_date
                FROM pe_fund_investments fi
                JOIN pe_funds f ON f.id = fi.fund_id
                JOIN pe_portfolio_companies pc ON pc.id = fi.company_id
                WHERE f.firm_id = :investor_id
                ORDER BY fi.investment_date DESC NULLS LAST
                LIMIT :limit
            """),
                {"investor_id": investor_id, "limit": limit},
            )
        else:
            result = db.execute(
                text("""
                SELECT company_name, company_industry, collected_date
                FROM portfolio_companies
                WHERE investor_id = :investor_id
                    AND investor_type = :investor_type
                ORDER BY collected_date DESC
                LIMIT :limit
            """),
                {
                    "investor_id": investor_id,
                    "investor_type": investor_type,
                    "limit": limit,
                },
            )

        return [
            {
                "company": row[0],
                "industry": row[1],
                "date": row[2].isoformat() if row[2] else None,
            }
            for row in result.fetchall()
        ]

    def _get_segments(
        self, db: Session, investor_id: int, investor_type: str,
    ) -> list:
        """Get business segment / fund breakdown for PE firms."""
        if investor_type != "pe_firm":
            return []

        result = db.execute(
            text("""
                SELECT name, strategy, final_close_usd_millions, status
                FROM pe_funds
                WHERE firm_id = :investor_id
                  AND name NOT LIKE '%13F%'
                ORDER BY final_close_usd_millions DESC NULLS LAST, name
            """),
            {"investor_id": investor_id},
        )

        segments = []
        for row in result.fetchall():
            aum = float(row[2]) if row[2] else None
            segments.append({
                "name": row[0],
                "strategy": row[1],
                "aum_millions": aum,
                "status": row[3],
            })
        return segments

    def _get_team(
        self, db: Session, investor_id: int, investor_type: str, limit: int = 50
    ) -> list:
        """Get key team members for PE firms with education and experience."""
        if investor_type != "pe_firm":
            return []

        result = db.execute(
            text("""
                SELECT p.id, p.full_name, fp.title, fp.seniority, fp.department,
                       p.linkedin_url, p.bio, fp.start_date
                FROM pe_firm_people fp
                JOIN pe_people p ON fp.person_id = p.id
                WHERE fp.firm_id = :investor_id AND fp.is_current = true
                ORDER BY
                    CASE fp.seniority
                        WHEN 'Partner' THEN 1
                        WHEN 'Managing Director' THEN 2
                        WHEN 'Principal' THEN 3
                        WHEN 'VP' THEN 4
                        WHEN 'Associate' THEN 5
                        ELSE 6
                    END,
                    p.full_name
                LIMIT :limit
            """),
            {"investor_id": investor_id, "limit": limit},
        )

        rows = result.fetchall()
        if not rows:
            return []

        # Collect person IDs for batch queries
        person_ids = [row[0] for row in rows]

        # Batch-fetch education
        edu_result = db.execute(
            text("""
                SELECT person_id, institution, degree, field_of_study, graduation_year
                FROM pe_person_education
                WHERE person_id = ANY(:pids)
                ORDER BY graduation_year DESC NULLS LAST
            """),
            {"pids": person_ids},
        )
        edu_map: dict = {}
        for erow in edu_result.fetchall():
            edu_map.setdefault(erow[0], []).append({
                "institution": erow[1],
                "degree": erow[2],
                "field": erow[3],
                "year": erow[4],
            })

        # Batch-fetch experience (prior roles, not current)
        exp_result = db.execute(
            text("""
                SELECT person_id, company, title, start_date, end_date
                FROM pe_person_experience
                WHERE person_id = ANY(:pids)
                ORDER BY start_date DESC NULLS LAST
            """),
            {"pids": person_ids},
        )
        exp_map: dict = {}
        for xrow in exp_result.fetchall():
            exp_map.setdefault(xrow[0], []).append({
                "company": xrow[1],
                "title": xrow[2],
                "start_year": xrow[3].year if xrow[3] else None,
                "end_year": xrow[4].year if xrow[4] else None,
            })

        team = []
        for row in rows:
            pid = row[0]
            start_date = row[7]
            start_year = start_date.year if start_date else None
            team.append({
                "name": row[1],
                "title": row[2],
                "seniority": row[3],
                "department": row[4],
                "linkedin": row[5],
                "bio": row[6],
                "start_year": start_year,
                "education": edu_map.get(pid, []),
                "experience": exp_map.get(pid, []),
            })
        return team

    def render_html(self, data: Dict[str, Any]) -> str:
        """Render report as HTML using the shared design system."""
        investor = data.get("investor", {})
        summary = data.get("portfolio_summary", {})
        holdings = data.get("top_holdings", [])
        sectors = data.get("sector_allocation", [])
        team = data.get("team", [])
        segments = data.get("segments", [])

        aum_display = _format_aum(investor.get("aum_millions"))
        charts_js = ""
        body = ""

        # ── Page Header ──────────────────────────────────────────
        sub_parts = [p for p in [investor.get("type"), investor.get("jurisdiction")] if p]
        subtitle = " · ".join(sub_parts) if sub_parts else None

        badge_parts = []
        if investor.get("aum_millions"):
            badge_parts.append(f"AUM {aum_display}")
        if investor.get("founded_year"):
            badge_parts.append(f"Founded {investor['founded_year']}")
        if investor.get("employee_count"):
            badge_parts.append(f"{investor['employee_count']:,} Employees")
        badge = " · ".join(badge_parts) if badge_parts else None

        body += page_header(
            title=f"{investor.get('name', 'Unknown')} \u2014 Investor Profile",
            subtitle=subtitle,
            badge=badge,
        )

        # ── KPI Strip ────────────────────────────────────────────
        total_mcap = summary.get("total_market_cap")
        exposure_display = _format_market_cap(total_mcap) if total_mcap else None

        cards = ""
        cards += kpi_card("Current Holdings", str(summary.get("current_holdings", 0)))
        cards += kpi_card("Sectors", str(summary.get("sectors", 0)))
        cards += kpi_card("Total Holdings", str(summary.get("total_holdings", 0)))
        if investor.get("aum_millions"):
            cards += kpi_card("AUM", aum_display)
        else:
            cards += kpi_card("Investor Type", str(investor.get("type", "N/A")))
        if exposure_display:
            cards += kpi_card("Public Equity*", exposure_display)
        elif investor.get("website"):
            cards += kpi_card("Website", investor["website"].replace("https://", "").replace("http://", ""))
        else:
            cards += kpi_card("Holdings (Active)", str(summary.get("current_holdings", 0)))

        body += '\n<div class="container">'
        body += "\n" + kpi_strip(cards)

        # ── Table of Contents ────────────────────────────────────
        toc_items = []
        sec_counter = 0
        if segments:
            sec_counter += 1
            toc_items.append({"number": sec_counter, "id": "aum-segments", "title": "AUM by Business Segment"})
        sec_counter += 1
        toc_items.append({"number": sec_counter, "id": "holdings", "title": "Portfolio Holdings"})
        if sectors:
            sec_counter += 1
            toc_items.append({"number": sec_counter, "id": "sectors", "title": "Sector Allocation"})
        if team:
            sec_counter += 1
            toc_items.append({"number": sec_counter, "id": "team", "title": "Key Team Members"})
        body += "\n" + toc(toc_items)

        # ── Section: AUM Segments ────────────────────────────────
        sec_num = 0
        if segments:
            sec_num += 1
            seg_total = sum(s.get("aum_millions") or 0 for s in segments)
            body += "\n" + section_start(sec_num, "AUM by Business Segment", "aum-segments")
            body += f'<p><strong>{len(segments)}</strong> business segments with combined AUM of <strong>{_format_aum(seg_total)}</strong>.</p>'

            seg_labels = [s.get("name", "N/A") for s in segments]
            seg_values = [s.get("aum_millions") or 0 for s in segments]
            seg_config = build_horizontal_bar_config(seg_labels, seg_values, dataset_label="AUM ($M)")
            seg_config_json = json.dumps(seg_config)
            bar_height = f"{max(len(seg_labels) * 48 + 40, 140)}px"
            body += chart_container("aumSegmentChart", seg_config_json, build_bar_fallback(seg_labels, seg_values), title="AUM by Segment ($M)", height=bar_height)
            charts_js += chart_init_js("aumSegmentChart", seg_config_json)

            # Segment detail cards
            seg_cards = ""
            for s in segments:
                aum = s.get("aum_millions")
                aum_seg = _format_aum(aum) if aum else "-"
                pct = (aum / seg_total * 100) if aum and seg_total else 0
                strategy = s.get("strategy") or "-"
                seg_cards += f"""<div class="segment-card">
    <div class="segment-header">
        <span class="segment-name">{s.get('name', 'N/A')}</span>
        <span class="segment-aum">{aum_seg}</span>
    </div>
    <div class="segment-strategy">{strategy}</div>
    <div class="segment-bar-track">
        <div class="segment-bar-fill" style="width: {pct:.0f}%"></div>
    </div>
    <div class="segment-pct">{pct:.0f}% of total AUM</div>
</div>"""

            body += f'<div class="segments-grid">{seg_cards}</div>'
            body += f'<div class="segment-total">Total: {_format_aum(seg_total)}</div>'
            body += f'<p class="footnote">Source: {investor.get("name", "Firm")} 10-K filing (Dec 31, 2024)</p>'
            body += "\n" + section_end()

        # ── Section: Holdings ────────────────────────────────────
        sec_num += 1
        body += "\n" + section_start(sec_num, "Portfolio Holdings", "holdings")
        body += f'<p><strong>{len(holdings)}</strong> portfolio companies across <strong>{summary.get("sectors", 0)}</strong> sectors.</p>'

        table_rows = []
        for h in holdings:
            ticker = h.get("ticker")
            ownership = h.get("ownership_status") or ""
            if ticker:
                badge_html = pill_badge("Public", "public")
                ticker_html = f'<span class="ticker">{ticker}</span>'
            elif "pe-backed" in ownership.lower():
                badge_html = pill_badge("PE-Backed", "pe")
                ticker_html = "-"
            elif "subsidiary" in ownership.lower():
                badge_html = pill_badge("Subsidiary", "sub")
                ticker_html = "-"
            else:
                badge_html = pill_badge("Private", "private")
                ticker_html = "-"

            table_rows.append([
                f'<span class="company-name">{h.get("name", "N/A")}</span>',
                ticker_html,
                h.get("industry") or "-",
                h.get("location") or "-",
                badge_html,
                _format_market_cap(h.get("market_cap")),
            ])

        body += data_table(
            headers=["Company", "Ticker", "Industry", "Location", "Status", "Est. Market Cap"],
            rows=table_rows,
            numeric_columns={5},
        )

        # Concentration callout
        if holdings:
            public_count = sum(1 for h in holdings if h.get("ticker"))
            private_count = len(holdings) - public_count
            body += callout(
                f"<strong>Insight:</strong> Portfolio is {public_count} public and {private_count} private companies. "
                f"Market cap figures represent total company value (Yahoo Finance), not position sizes.",
            )

        body += "\n" + section_end()

        # ── Section: Sector Allocation ───────────────────────────
        if sectors:
            sec_num += 1
            body += "\n" + section_start(sec_num, "Sector Allocation", "sectors")
            body += f'<p>Portfolio spans <strong>{len(sectors)}</strong> sectors.</p>'

            # Group into top 7 + "Other" for clean doughnut
            max_slices = 7
            if len(sectors) > max_slices:
                top = sectors[:max_slices]
                rest = sectors[max_slices:]
                other_count = sum(s.get("count", 0) for s in rest)
                other_pct = sum(s.get("pct", 0) for s in rest)
                chart_sectors = top + [{"sector": f"Other ({len(rest)})", "count": other_count, "pct": other_pct}]
            else:
                chart_sectors = sectors

            chart_labels = [s.get("sector", "N/A") for s in chart_sectors]
            chart_values = [float(s.get("count", 0)) for s in chart_sectors]
            chart_colors = [CHART_COLORS[i % len(CHART_COLORS)] for i in range(len(chart_labels))]

            doughnut_config = build_doughnut_config(chart_labels, chart_values, chart_colors)
            doughnut_json = json.dumps(doughnut_config)

            body += '<div class="chart-row">'
            body += '<div>'
            body += chart_container("sectorDoughnut", doughnut_json, build_bar_fallback(chart_labels, chart_values), size="medium", title="Sector Distribution")
            charts_js += chart_init_js("sectorDoughnut", doughnut_json)
            body += '</div>'
            body += '<div>'
            body += build_chart_legend(chart_labels, chart_values, chart_colors, show_pct=True)
            body += '</div>'
            body += '</div>'

            # Full sector detail table
            sector_rows = []
            for s in sectors:
                mcap_display = _format_market_cap(s.get("market_cap")) if s.get("market_cap") else "-"
                sector_rows.append([
                    s.get("sector", "N/A"),
                    str(s.get("count", 0)),
                    f'{s.get("pct", 0)}%',
                    mcap_display,
                ])
            body += data_table(
                headers=["Sector", "Companies", "% of Portfolio", "Est. Market Cap"],
                rows=sector_rows,
                numeric_columns={1, 2, 3},
            )
            body += "\n" + section_end()

        # ── Section: Team ────────────────────────────────────────
        if team:
            sec_num += 1
            body += "\n" + section_start(sec_num, "Key Team Members", "team")
            body += f'<p><strong>{len(team)}</strong> key team members.</p>'

            body += '<div class="team-grid">'
            for t in team:
                name = t.get("name", "N/A")
                name_parts = name.split()
                if len(name_parts) >= 2:
                    initials = (name_parts[0][0] + name_parts[-1][0]).upper()
                elif name_parts:
                    initials = name_parts[0][0].upper()
                else:
                    initials = "?"

                badges = []
                if t.get("seniority"):
                    badges.append(f'<span class="card-badge badge-seniority">{t["seniority"]}</span>')
                if t.get("department"):
                    badges.append(f'<span class="card-badge badge-dept">{t["department"]}</span>')
                if t.get("start_year"):
                    badges.append(f'<span class="card-badge badge-tenure">Since {t["start_year"]}</span>')

                body += "\n" + profile_card(
                    name=name,
                    title=t.get("title", "-"),
                    initials=initials,
                    badges=badges if badges else None,
                    bio=t.get("bio"),
                    experience=t.get("experience"),
                    education=t.get("education"),
                    linkedin=t.get("linkedin"),
                )
            body += '</div>'
            body += "\n" + section_end()

        body += '\n</div>'  # close container

        # ── Footer ───────────────────────────────────────────────
        notes = [
            "Data sourced from SEC EDGAR filings, company websites, and public records.",
            "Market cap figures from Yahoo Finance; represent total company value, not position sizes.",
            "Ownership status determined by cross-referencing SEC filings and web collection.",
        ]
        body += "\n" + page_footer(
            notes=notes,
            generated_line=f"Report generated {data.get('generated_at', 'N/A')} | Nexdata Investment Intelligence",
        )

        return html_document(
            title=f"{investor.get('name', 'Investor')} \u2014 Profile Report",
            body_content=body,
            charts_js=charts_js,
        )

    def render_excel(self, data: Dict[str, Any]) -> bytes:
        """Render report as Excel workbook."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()

        # Summary sheet
        ws = wb.active
        ws.title = "Summary"

        investor = data.get("investor", {})
        summary = data.get("portfolio_summary", {})

        # Header styling
        header_font = Font(bold=True, size=14)
        header_fill = PatternFill(
            start_color="3498DB", end_color="3498DB", fill_type="solid"
        )
        header_font_white = Font(bold=True, color="FFFFFF")

        # Title
        ws["A1"] = investor.get("name", "Investor Profile")
        ws["A1"].font = Font(bold=True, size=18)
        ws.merge_cells("A1:D1")

        # Investor details
        ws["A3"] = "Type"
        ws["B3"] = investor.get("type", "N/A")
        ws["A4"] = "Headquarters"
        ws["B4"] = investor.get("jurisdiction", "N/A")
        ws["A5"] = "AUM"
        ws["B5"] = _format_aum(investor.get("aum_millions"))
        ws["A6"] = "Website"
        ws["B6"] = investor.get("website", "N/A")
        row_offset = 6
        if investor.get("founded_year"):
            row_offset += 1
            ws[f"A{row_offset}"] = "Founded"
            ws[f"B{row_offset}"] = investor["founded_year"]
        if investor.get("employee_count"):
            row_offset += 1
            ws[f"A{row_offset}"] = "Employees"
            ws[f"B{row_offset}"] = investor["employee_count"]

        # Portfolio stats
        row_offset += 2
        ws[f"A{row_offset}"] = "Portfolio Summary"
        ws[f"A{row_offset}"].font = header_font
        row_offset += 1
        ws[f"A{row_offset}"] = "Current Holdings"
        ws[f"B{row_offset}"] = summary.get("current_holdings", 0)
        row_offset += 1
        ws[f"A{row_offset}"] = "Sectors"
        ws[f"B{row_offset}"] = summary.get("sectors", 0)
        row_offset += 1
        ws[f"A{row_offset}"] = "Total Holdings"
        ws[f"B{row_offset}"] = summary.get("total_holdings", 0)
        if summary.get("total_market_cap"):
            row_offset += 1
            ws[f"A{row_offset}"] = "Est. Public Equity Exposure"
            ws[f"B{row_offset}"] = _format_market_cap(summary["total_market_cap"])

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 40

        # Holdings sheet
        ws_holdings = wb.create_sheet("Holdings")
        holdings = data.get("top_holdings", [])

        headers = ["Company", "Ticker", "Industry", "Location", "Type", "Est. Market Cap"]
        for col, header in enumerate(headers, 1):
            cell = ws_holdings.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill

        for row, h in enumerate(holdings, 2):
            ws_holdings.cell(row=row, column=1, value=h.get("name"))
            ws_holdings.cell(row=row, column=2, value=h.get("ticker") or "-")
            ws_holdings.cell(row=row, column=3, value=h.get("industry") or "-")
            ws_holdings.cell(row=row, column=4, value=h.get("location") or "-")
            ws_holdings.cell(row=row, column=5, value=h.get("stage") or "-")
            ws_holdings.cell(row=row, column=6, value=h.get("market_cap"))

        ws_holdings.column_dimensions["A"].width = 30
        ws_holdings.column_dimensions["B"].width = 10
        ws_holdings.column_dimensions["C"].width = 20
        ws_holdings.column_dimensions["D"].width = 20
        ws_holdings.column_dimensions["E"].width = 18
        ws_holdings.column_dimensions["F"].width = 18

        # Sectors sheet
        ws_sectors = wb.create_sheet("Sectors")
        sectors = data.get("sector_allocation", [])

        headers = ["Sector", "Count", "Percentage", "Est. Value"]
        for col, header in enumerate(headers, 1):
            cell = ws_sectors.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill

        for row, s in enumerate(sectors, 2):
            ws_sectors.cell(row=row, column=1, value=s.get("sector"))
            ws_sectors.cell(row=row, column=2, value=s.get("count"))
            ws_sectors.cell(row=row, column=3, value=f"{s.get('pct', 0)}%")
            ws_sectors.cell(row=row, column=4, value=s.get("market_cap"))

        ws_sectors.column_dimensions["A"].width = 25
        ws_sectors.column_dimensions["B"].width = 10
        ws_sectors.column_dimensions["C"].width = 12
        ws_sectors.column_dimensions["D"].width = 18

        # Team sheet (PE firms only)
        team = data.get("team", [])
        if team:
            ws_team = wb.create_sheet("Team")
            headers = [
                "Name", "Title", "Seniority", "Department", "Start Year",
                "Education", "Prior Experience", "LinkedIn", "Bio",
            ]
            for col, header in enumerate(headers, 1):
                cell = ws_team.cell(row=1, column=col, value=header)
                cell.font = header_font_white
                cell.fill = header_fill

            for row, t in enumerate(team, 2):
                ws_team.cell(row=row, column=1, value=t.get("name"))
                ws_team.cell(row=row, column=2, value=t.get("title") or "-")
                ws_team.cell(row=row, column=3, value=t.get("seniority") or "-")
                ws_team.cell(row=row, column=4, value=t.get("department") or "-")
                ws_team.cell(row=row, column=5, value=t.get("start_year") or "-")
                # Format education
                edu_parts = []
                for edu in t.get("education", []):
                    deg = edu.get("degree", "")
                    inst = edu.get("institution", "")
                    yr = f" ({edu['year']})" if edu.get("year") else ""
                    edu_parts.append(f"{deg} — {inst}{yr}")
                ws_team.cell(row=row, column=6, value="; ".join(edu_parts) if edu_parts else "-")
                # Format experience
                exp_parts = []
                for exp in t.get("experience", [])[:3]:
                    yrs = ""
                    if exp.get("start_year") and exp.get("end_year"):
                        yrs = f" ({exp['start_year']}–{exp['end_year']})"
                    exp_parts.append(f"{exp.get('title', '')} at {exp.get('company', '')}{yrs}")
                ws_team.cell(row=row, column=7, value="; ".join(exp_parts) if exp_parts else "-")
                ws_team.cell(row=row, column=8, value=t.get("linkedin") or "")
                ws_team.cell(row=row, column=9, value=t.get("bio") or "")

            ws_team.column_dimensions["A"].width = 25
            ws_team.column_dimensions["B"].width = 30
            ws_team.column_dimensions["C"].width = 15
            ws_team.column_dimensions["D"].width = 18
            ws_team.column_dimensions["E"].width = 12
            ws_team.column_dimensions["F"].width = 45
            ws_team.column_dimensions["G"].width = 45
            ws_team.column_dimensions["H"].width = 35
            ws_team.column_dimensions["I"].width = 60

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        return output.getvalue()
