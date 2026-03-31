"""
Les Schwab AV Deep Dive — IC Report Template.

Generates a full Investment Committee-grade analysis of Les Schwab Tire Centers
in the context of autonomous vehicle and EV disruption.

All values are baked in at generation time from FRED, BLS, and AFDC tables.
No runtime fetch calls in the output HTML — static, portable, works offline.

Sections:
  1. Business Profile & Financial Snapshot
  2. AV & EV Adoption Trajectory
  3. Revenue Vulnerability by Service Line
  4. Strategic Response Options
  5. Competitive Moat Assessment
  6. Investment Scenario Analysis
  7. Key Watchpoints & Recommended Actions

Update the module-level constants below when estimates change.
Run: ./scripts/generate_report.sh les_schwab_av
"""

import json
import logging
from datetime import datetime
from typing import Any, Dict, List, Optional

from sqlalchemy.orm import Session

from app.reports.design_system import (
    BLUE, BLUE_LIGHT, GRAY, GREEN, ORANGE, PURPLE, RED, TEAL,
    CHART_COLORS,
    build_bar_fallback, build_chart_legend,
    build_doughnut_config, build_horizontal_bar_config, build_line_chart_config,
    callout, chart_container, chart_init_js,
    html_document, kpi_card, kpi_strip,
    page_footer, page_header,
    section_end, section_start, toc,
    data_table,
)
from app.reports.templates._ic_report_base import ICReportBase

logger = logging.getLogger(__name__)

# ── Les Schwab geography ──────────────────────────────────────────────────────
_SCHWAB_STATES = {"WA", "OR", "CA", "ID", "MT", "WY", "NV", "UT", "CO", "AK"}

# ── IC Summary / TLDR (5 bullets shown in dark banner above KPI strip) ────────
# IC Summary — top rows (Risk / Opportunity / Action)
_TLDR_TOP = [
    ("Risk",
     "EV regen braking may erode ~$101M in brake GP (32% margin) over 5–10 years. "
     "ADAS calibration and EV tire premiums provide a partial offset."),
    ("Opportunity",
     "ADAS calibration = potential $112–210M add-on at existing alignment bays, "
     "currently ~$0. EV tire wear premium (+20–30%) supports ASP uplift on the "
     "largest revenue line."),
    ("Action",
     "2026: launch ADAS rollout + EV tire program + AI store briefings on existing "
     "data — no data infrastructure prerequisite required to start."),
]

# IC Summary — initiative cards (rendered as 3-column grid)
_TLDR_INITIATIVES = [
    ("Initiative 1", "Store Intelligence + Claude Briefings", "2026",
     "Weekly AI briefings from Day 1; data foundation built in parallel.",
     "$23–37M EBITDA"),
    ("Initiative 2", "Agentic Cross-Store Learning", "2026–27",
     "Monitors 500 stores; propagates top-quartile patterns as next best actions.",
     "$22–52M incremental"),
    ("Initiative 3", "Passive Data + Predictive", "2027–28",
     "LPR, telematics integration, fleet B2B portal. Builds on 1 & 2.",
     "$33–43M incremental"),
]

# IC Summary — bottom rows (Exit Timing / Recommendation)
_TLDR_BOTTOM = [
    ("Exit Timing",
     "2027–2028: max value — erosion modest, AI story demonstrable. "
     "2030–2032: more upside, requires active execution across all 3 initiatives."),
    ("Rec.",
     "Hold + execute. AI base case: ~$40–60M EBITDA contribution (directional). "
     "Bull case (ADAS + AI fully materializing): 2.0–2.3x MOIC. Do not hold passively."),
]

# ── Key Modeling Assumptions (transparency box in Section 1) ──────────────────
_KEY_ASSUMPTIONS = [
    # (Assumption, Value Used, Primary Source, Sensitivity)
    ("EV brake wear reduction",
     "40–60% vs. ICE",
     "Real-world fleet data: Chevy Bolt, Tesla Model 3 (80K–120K mi on original pads)",
     "±10pp → brake GP impact ~±$10M by 2030"),
    ("EV tire wear premium",
     "+20–30% vs. ICE",
     "Continental AG EV Tire Study 2023; Michelin fleet operations data",
     "Low sensitivity to total revenue; supports ASP but not visit frequency"),
    ("Annual alignments",
     "~1.4M jobs/yr (modeled est.)",
     "$168M alignment revenue ÷ ~$90 blended ASP — derived, not sourced directly",
     "±200K jobs changes ADAS upside by ±$16–30M"),
    ("ADAS calibration ASP",
     "$80–$150 per job",
     "Hunter Engineering market analysis 2024; dealership benchmarks",
     "Base case uses midpoint ~$115; floor $80 still material"),
    ("ADAS attach rate (full rollout)",
     "~50% base / ~100% upside",
     "NHTSA: ~65% of MY2023+ vehicles have factory ADAS; Hunter Engineering market data",
     "50% penetration = $56–105M; 25% = $28–53M — dependent on OEM requirements + insurance workflows"),
    ("AV fleet impact horizon",
     "2033–2040 (base case)",
     "Wood Mackenzie AV Fleet Forecast 2024; BloombergNEF 2025",
     "5-year pull-forward to 2028 materially worsens bear case"),
]

# ── BLS series labels (series_title is NULL in DB — map IDs here) ─────────────
# Keys are series_id prefixes or exact IDs. Used for readable display.
_BLS_LABELS: Dict[str, str] = {
    "CES4244100001": "Motor Vehicle & Parts Dealers Employment (NAICS 44411, CES)",
    "CES4244130001": "Auto Parts, Accessories & Tire Stores Employment (NAICS 44413, CES)",
    "OEUN000000000000049302304": "Auto Service & Tire Repair Technicians — Mean Annual Wage (SOC 49-3023, OES 2024)",
    "OEUN000000000000049302301": "Auto Service & Tire Repair Technicians — Employment (SOC 49-3023, OES 2024)",
}

# ────────────────────────────────────────────────────────────────────────────────
# SECTION 1 — Business Profile & Financial Snapshot
# Update these when management provides revised figures.
# ────────────────────────────────────────────────────────────────────────────────
_TOTAL_REVENUE_M = 2_100   # $M estimated 2025E
_TOTAL_GP_M      = 500     # $M estimated gross profit 2025E (~24% blended margin)
_TOTAL_EBITDA_M  = 273     # $M estimated EBITDA 2025E (~13% margin; GP less store opex/G&A/profit-sharing)
_STORE_COUNT     = 500
_STATES_COUNT    = 10
_ENTRY_EV_B      = 2.1     # Meritage acquisition price $B
_ENTRY_EBITDA_X  = 10.0    # acquisition multiple

_REVENUE_MIX = [
    # (Service Line, $M, % of Rev, GP%, GP$M, AV Disruption Vector, Risk Level)
    # GP% estimated from product/labor mix; blended to ~$500M total
    ("Tires (product + install)", 1_260, 60, 16,  202,
     "EV weight → faster wear (offset); AV fleets → fewer owned vehicles", "Medium"),
    ("Brakes (pads, rotors, labor)", 315, 15, 32, 101,
     "Regenerative braking reduces mechanical brake wear 40–60%", "High"),
    ("Alignments", 168, 8, 58, 97,
     "ADAS calibration is a NEW revenue add-on per alignment job", "Opportunity"),
    ("Wheels & Custom", 147, 7, 22, 32,
     "Largely unchanged; EV aesthetics may shift preferences", "Low"),
    ("Shocks & Struts", 105, 5, 38, 40,
     "EVs still require suspension; heavier chassis may accelerate wear", "Low"),
    ("Batteries & Other", 105, 5, 22, 23,
     "12V aux batteries still needed in EVs; 12V market declines slowly", "Low"),
]

# ────────────────────────────────────────────────────────────────────────────────
# SECTION 2 — AV & EV Adoption Trajectory
# ────────────────────────────────────────────────────────────────────────────────
# Columns: Year, EV New Sales %, EV Fleet %, AV Fleet % (Base),
#          Personal Veh. Ownership Impact, Les Schwab Impact
_ADOPTION_SCENARIOS = [
    (2026, "15%", "5%",  "1%",  "Negligible",    "Minimal"),
    (2028, "22%", "9%",  "3%",  "-2% to -4%",    "Brakes -6%"),
    (2030, "30%", "14%", "6%",  "-4% to -8%",    "Brakes -9%; Total -2%"),
    (2033, "42%", "23%", "12%", "-8% to -15%",   "Brakes -15%; Total -5%"),
    (2035, "52%", "32%", "18%", "-12% to -22%",  "Brakes -21%; Total -8%"),
    (2040, "68%", "52%", "32%", "-22% to -40%",  "Brakes -34%; Total -16%"),
]

# Chart data — EV fleet % (base/bull/bear) and AV fleet % for line charts
_YEARS        = [2024, 2026, 2028, 2030, 2032, 2035, 2040]
_EV_FLEET_BASE  = [3,  5,   9,   14,  19,  32,  52]
_EV_FLEET_BULL  = [3,  7,  12,   18,  26,  42,  65]
_EV_FLEET_BEAR  = [3,  4,   7,   11,  15,  24,  40]
_AV_FLEET_BASE  = [0,  1,   3,    6,  10,  18,  32]
_OWNERSHIP_BASE = [0, -1,  -3,   -6,  -9, -17, -30]   # % change vs 2024
_OWNERSHIP_BULL = [0,  0,  -1,   -3,  -5,  -10, -22]
_OWNERSHIP_BEAR = [0, -2,  -6,  -11, -18,  -28, -45]

# ────────────────────────────────────────────────────────────────────────────────
# SECTION 3 — Revenue Vulnerability by Service Line
# ────────────────────────────────────────────────────────────────────────────────
# Columns: Service Line, 2024E, 2028E (Base), 2030E (Base), 2035E (Base), 2035E (Bear), Key Driver
_REV_PROJECTIONS = [
    ("Tires",          "$1,260M", "$1,320M", "$1,340M", "$1,240M", "$1,020M",
     "EV wear premium vs. ownership decline"),
    ("Brakes",         "$315M",   "$294M",   "$286M",   "$249M",   "$210M",
     "Regen braking reduces mechanical wear"),
    ("Alignments + ADAS", "$168M","$195M",   "$220M",   "$280M",   "$260M",
     "ADAS calibration adds ~$80–150 per visit"),
    ("Wheels & Custom","$147M",   "$150M",   "$152M",   "$145M",   "$128M",
     "Broadly stable; EV aesthetics may modestly lift"),
    ("Shocks & Struts","$105M",   "$108M",   "$110M",   "$108M",   "$92M",
     "EV chassis weight may accelerate wear"),
    ("Batteries & Other","$105M", "$102M",   "$100M",   "$88M",    "$75M",
     "12V auxiliary battery market slowly declines"),
    ("Total",          "$2,100M", "$2,169M", "$2,208M", "$2,110M", "$1,785M", ""),
]

# Chart data — brake revenue ($M) and tire wear (miles/mm tread)
_BRAKE_REV_YEARS    = [2024, 2026, 2028, 2030, 2033, 2035, 2040]
_BRAKE_REV_BASE     = [315,  308,  294,  286,  267,  249,  207]
_BRAKE_REV_BEAR     = [315,  302,  280,  265,  238,  210,  158]
_TIRE_WEAR_LABELS   = ["ICE (Standard)", "ICE (Performance)", "BEV (Standard)", "BEV (Performance)", "AV-Optimized"]
_TIRE_WEAR_MILES_MM = [3_200, 2_800, 2_500, 2_000, 4_200]   # miles per mm tread depth

# ────────────────────────────────────────────────────────────────────────────────
# SECTION 4 — Strategic Response Options
# ────────────────────────────────────────────────────────────────────────────────
# Columns: Option, Revenue Potential, Margin Profile, Time to Material, Complexity, Priority
_STRATEGIC_OPTIONS = [
    ("ADAS / AV Sensor Calibration",    "$80–150M incremental", "55–65%", "2–3 years",  "Low–Medium", "Priority 1"),
    ("EV Specialty Tire Program",        "$120–200M ASP lift",   "52%+",   "1–2 years",  "Low",        "Priority 1"),
    ("Commercial / Fleet Contracts",     "$150–300M long-term",  "38–45%", "4–6 years",  "Medium",     "Priority 2"),
    ("Mobile Service Units",             "$50–100M",             "35–42%", "3–5 years",  "Medium",     "Priority 2"),
    ("EV Battery Diagnostics",           "$30–60M",              "40–50%", "3–4 years",  "Medium",     "Priority 3"),
    ("Tire-as-a-Service (Subscription)", "$80–160M ARR",         "50–60%", "5–8 years",  "High",       "Priority 3"),
    ("Defensive Consolidation",          "Defensive moat",       "Near-term dilutive", "Ongoing", "High (capital)", "Situational"),
]

# Chart data
_ADAS_REVENUE_LABELS = ["Current (0% penetration)", "25% stores", "50% stores", "75% stores", "100% stores"]
_ADAS_REVENUE_LOW    = [0, 28, 56, 84, 112]   # $M @ $80/job
_ADAS_REVENUE_HIGH   = [0, 53, 106, 159, 210]  # $M @ $150/job

# ────────────────────────────────────────────────────────────────────────────────
# SECTION 5 — Competitive Moat Assessment
# ────────────────────────────────────────────────────────────────────────────────
_NPS_COMPARISON = [
    # (Brand, NPS score)
    ("Les Schwab (Pacific NW)",  75),
    ("Discount Tire",            55),
    ("Firestone Complete",       42),
    ("Jiffy Lube",               38),
    ("Midas",                    32),
    ("Pep Boys",                 28),
    ("Walmart Auto",             22),
]

_MOAT_DIMENSIONS = [
    # (Dimension, Current, EV Era, AV Era, Key Risk)
    ("Brand loyalty (Pacific NW)",    "Very Strong", "Strong",    "Moderate",
     "Brand loyalty doesn't transfer to fleet operators"),
    ("Employee profit-sharing (ESOP)","Very Strong", "Strong",    "Moderate",
     "Lower visit frequency = lower incentive income"),
    ("Free lifetime flat repair",     "Strong",      "Strong",    "Moderate",
     "EVs have fewer flat events (low sidewalls common)"),
    ("Store density (~500 locations)","Strong",      "Strong",    "Moderate",
     "Fixed-cost burden if visit volume declines"),
    ("Free brake inspections",        "Strong",      "Weakening", "Weak",
     "Regen braking eliminates the product that follows"),
    ("Technician expertise (ICE)",    "Strong",      "Moderate",  "Moderate",
     "EV/ADAS expertise requires significant retraining"),
]

# Visit frequency chart data (annual visits per vehicle)
_VF_YEARS   = [2024, 2026, 2028, 2030, 2035, 2040]
_VF_ICE     = [2.1,  2.0,  1.95, 1.90, 1.85, 1.80]
_VF_EV      = [2.1,  2.05, 1.90, 1.75, 1.55, 1.40]
_VF_AV_FLEET= [2.1,  2.0,  1.85, 1.60, 1.20, 0.90]

# ────────────────────────────────────────────────────────────────────────────────
# SECTION 6 — Investment Scenario Analysis
# ────────────────────────────────────────────────────────────────────────────────
# Revenue trajectory $M
_SCENARIO_YEARS   = [2024, 2026, 2028, 2030, 2032, 2035]
_REV_BULL         = [2_100, 2_200, 2_350, 2_500, 2_620, 2_750]
_REV_BASE         = [2_100, 2_150, 2_169, 2_208, 2_180, 2_110]
_REV_BEAR         = [2_100, 2_080, 2_010, 1_920, 1_830, 1_785]
_EBITDA_BULL      = [273,  310,  320,  355,  375,  412]
_EBITDA_BASE      = [273,  278,  285,  295,  290,  273]
_EBITDA_BEAR      = [273,  261,  248,  232,  220,  210]

_SCENARIOS_TABLE = [
    # (Label, 2024E EBITDA, 2028E EBITDA, 2032E EBITDA, Exit Timing, EV Multiple, Implied EV)
    ("Bull — AV slow + strategic execution",
     "$273M", "$320M", "$375M", "2030–2032", "11–13x", "$4.1–4.9B"),
    ("Base — Gradual transition, partial pivot",
     "$273M", "$285M", "$290M", "2027–2030", "9–11x",  "$2.6–3.2B"),
    ("Bear — AV fast + no strategic response",
     "$273M", "$248M", "$210M", "2026–2027", "7–9x",   "$1.5–1.9B"),
]

# ────────────────────────────────────────────────────────────────────────────────
# SECTION 8 — AI Initiatives (2026–2028)
# Three succinct initiatives with Rev / Cost Savings / EBITDA impact.
# Language is directional ("may", "could", "we estimate") — not a business case commitment.
# ────────────────────────────────────────────────────────────────────────────────

# Financial summary — (Initiative, Revenue Impact, Cost Savings, EBITDA, Investment)
_AI_INITIATIVES_TABLE = [
    ("1 — Store Intelligence (2026)",        "$5–12M",    "$18–25M", "$23–37M", "TBD"),
    ("2 — Agentic Learning (2026–2027)",     "$15–42M",   "$7–10M",  "$22–52M", "TBD"),
    ("3 — Passive + Predictive (2027–2028)", "$18–140M",  "$15–18M", "$33–43M", "TBD"),
    ("Total by 2028",                        "$38–194M",  "$40–53M", "$78–132M", "TBD"),
]

# Revenue ramp by initiative ($M, midpoints for chart)
_AI_YEARS         = [2026, 2027, 2028, 2029, 2030]
_AI1_REV          = [5,    8,    10,   10,   10]    # Store intelligence / leakage prevention
_AI2_REV          = [0,    15,   42,   60,   72]    # Agentic cross-store learning
_AI3_REV          = [0,    0,    18,   65,   140]   # Passive data + predictive intelligence
_AI_BRAKE_EROSION = [101,  97,   92,   87,   81]    # Brake GP base case (offset reference)

# ────────────────────────────────────────────────────────────────────────────────
# SECTION 7 — Key Watchpoints & Recommended Actions
# ────────────────────────────────────────────────────────────────────────────────
# Columns: Watchpoint, Signal, Green (on track), Red (off track), Cadence
_WATCHPOINTS = [
    ("Brake revenue % of total", "Monthly revenue mix",
     "< 14% by 2027", "> 16% (no mix shift)", "Quarterly"),
    ("ADAS calibration rollout", "% of stores with equipment",
     "50%+ by end 2026", "< 20% by end 2026", "Quarterly"),
    ("EV specialty tire attach rate", "EV-rated tire % of mix",
     "EV share ≥ local EV fleet share", "EV share lagging fleet share", "Quarterly"),
    ("Fleet / commercial accounts", "Fleet revenue % of total",
     "5%+ by 2028", "< 2% by 2028", "Annual"),
    ("Technician EV/ADAS certification", "% techs certified on EV",
     "40%+ by end 2026", "< 15% by end 2026", "Semi-annual"),
    ("Same-store visit frequency", "Visits/vehicle in local market",
     "Flat or growing", "Declining > 3% YoY", "Quarterly"),
]

_ACTIONS = [
    ("Action 1 — Immediate",
     "Commission a detailed ADAS calibration market analysis for Les Schwab's top 50 markets. "
     "Estimate revenue from retrofitting calibration equipment into existing alignment bays. "
     "Capital cost: modest ($15–25K/bay). Revenue upside: $80–150/job × ~1.4M alignments/yr = "
     "$112–210M incremental at full rollout."),
    ("Action 2 — 90 days",
     "Launch an EV specialty tire initiative — dedicated SKU selection, staff training on EV tire "
     "specs (load index, rolling resistance, noise ratings), and marketing to local EV communities. "
     "No capital required; pure execution play."),
    ("Action 3 — 6 months",
     "Establish a dedicated fleet/commercial sales team (3–5 reps) covering the top 5 western US "
     "metros. Prioritize rental car companies, last-mile delivery operators, and early AV fleet "
     "test programs (Waymo, Cruise, Amazon Robotics)."),
    ("Action 4 — Strategic",
     "Evaluate exit window. If ADAS calibration rollout delivers incremental revenue in 2026, "
     "a 2027–2028 sale process positions Les Schwab as a business with a proven AV transition "
     "playbook — commanding a premium vs. a business still dependent on ICE/brake revenue. "
     "Do not wait for the thesis to play out fully; sell the story while it is in early innings."),
]


# ────────────────────────────────────────────────────────────────────────────────
# Template class
# ────────────────────────────────────────────────────────────────────────────────

class LesSchwabAVTemplate(ICReportBase):
    """Full IC-grade Les Schwab AV disruption deep dive."""

    name = "les_schwab_av"
    display_name = "Les Schwab Tire Centers - AV Disruption Deep Dive"
    description = (
        "Les Schwab Tire Centers — AV disruption deep dive with live macro, "
        "labor & EV infrastructure data. 8-section IC report."
    )

    # ── Data gathering ────────────────────────────────────────────────────────

    def gather_data(self, db: Session, params: Dict[str, Any]) -> Dict[str, Any]:
        fred_auto     = self.fetch_fred_series(db, "fred_auto_sector")
        fred_sent     = self.fetch_fred_series(db, "fred_consumer_sentiment")
        fred_series   = {**fred_auto, **fred_sent}
        totalsa_hist  = self.fetch_fred_history(db, "fred_auto_sector", "TOTALSA", 24)
        bls_latest    = self.fetch_bls_latest(db, "bls_auto_sector")
        afdc          = self.fetch_afdc_by_state(db, _SCHWAB_STATES)

        return {
            "generated_at": datetime.utcnow().isoformat(),
            "fred": {"series": fred_series, "totalsa_history": totalsa_hist},
            "bls": {"latest": bls_latest},
            "afdc": afdc,
        }

    # ── HTML rendering ────────────────────────────────────────────────────────

    def render_html(self, data: Dict[str, Any]) -> str:
        fred      = data.get("fred", {})
        bls       = data.get("bls", {})
        afdc      = data.get("afdc", {})
        gen_at    = data.get("generated_at", datetime.utcnow().isoformat())
        title     = data.get("report_title", "Les Schwab Tire Centers — AV Disruption Deep Dive")

        fs          = fred.get("series", {})
        bls_latest  = bls.get("latest", {})

        # ── Extract live values ───────────────────────────────────────────────
        totalsa    = fs.get("TOTALSA", {}).get("value")
        gas_price  = fs.get("GASREGCOVW", {}).get("value")
        umcsent    = fs.get("UMCSENT", {}).get("value")
        ev_nat     = afdc.get("national_total", 0)
        ev_fp      = afdc.get("footprint_total", 0)

        auto_tech_wage = None
        auto_dealer_emp = None
        auto_tech_sid = None
        for sid, s in bls_latest.items():
            if (sid.startswith("OES") or sid.startswith("OEUN")) and sid.endswith("04") and auto_tech_wage is None:
                auto_tech_wage = s.get("value")
                auto_tech_sid = sid
            if sid.startswith("CES") and auto_dealer_emp is None:
                auto_dealer_emp = s.get("value")

        # ── Precise as-of dates for citations ─────────────────────────────────
        totalsa_date  = fs.get("TOTALSA",      {}).get("date", "")[:7]   # YYYY-MM
        gas_date      = fs.get("GASREGCOVW",   {}).get("date", "")[:7]
        umcsent_date  = fs.get("UMCSENT",      {}).get("date", "")[:7]

        # AFDC as-of: pull from first footprint state that has a date
        fp_states_raw = afdc.get("footprint_states", []) or afdc.get("states", [])
        afdc_date_raw = next(
            (s.get("as_of_date", "") for s in fp_states_raw if s.get("as_of_date")),
            "",
        )
        afdc_date = afdc_date_raw[:10] if afdc_date_raw else "latest ingest"

        # BLS OES is annual survey; extract year from period/year fields
        bls_oes_period = ""
        if auto_tech_sid and auto_tech_sid in bls_latest:
            _s = bls_latest[auto_tech_sid]
            _yr = _s.get("year", "")
            bls_oes_period = f"annual {_yr}" if _yr else "annual"

        def _cite(text: str) -> str:
            """Wrap a source citation in a small styled span."""
            return (
                f'<span style="font-size:11px;color:#718096;font-weight:400;'
                f'font-style:italic"> [{text}]</span>'
            )

        charts_js = ""
        body      = ""

        # ── Header ────────────────────────────────────────────────────────────
        body += page_header(
            title=title,
            subtitle=f"Strategic Assessment for Investment Committee | Q1 2026",
            badge=f"Nexdata Live | Confidential — IC Use Only",
        )

        # ── KPI strip ─────────────────────────────────────────────────────────
        body += kpi_strip(
            kpi_card("US Vehicle Sales (SAAR)",
                     f"{totalsa:.1f}M" if totalsa else "—",
                     delta=f"FRED TOTALSA · {totalsa_date}" if totalsa else "FRED TOTALSA",
                     delta_dir="neutral")
            + kpi_card("Brake Revenue at Risk",
                       "$315M",
                       delta="Est. 2024E · ~15% of total rev",
                       delta_dir="down")
            + kpi_card("EV Stations — Schwab States",
                       f"{ev_fp:,}" if ev_fp else "—",
                       delta=f"AFDC/NREL · {afdc_date} · {ev_nat:,} national" if ev_nat else "AFDC/NREL",
                       delta_dir="neutral")
            + kpi_card("Consumer Sentiment (UMich)",
                       f"{umcsent:.1f}" if umcsent else "—",
                       delta=f"FRED UMCSENT · {umcsent_date}" if umcsent else "FRED UMCSENT",
                       delta_dir="neutral")
            + kpi_card("Auto Tech Mean Wage",
                       f"${auto_tech_wage:,.0f}" if auto_tech_wage else "—",
                       delta=f"BLS OES SOC 49-3023 · {bls_oes_period}" if auto_tech_wage else "BLS OES",
                       delta_dir="neutral")
        )

        # ── IC Summary / TLDR banner ──────────────────────────────────────────
        body += '<div class="container">'
        body += ('<div style="background:#1a202c;border-radius:8px;padding:18px 24px;">')
        body += ('<div style="font-size:10px;font-weight:700;color:#a0aec0;'
                 'letter-spacing:0.12em;margin-bottom:12px;text-transform:uppercase">'
                 'IC Summary</div>')
        # Top rows: Risk / Opportunity / Action
        for _label, _text in _TLDR_TOP:
            body += (
                f'<div style="display:flex;gap:14px;margin-bottom:7px;align-items:flex-start">'
                f'<span style="font-size:10px;font-weight:700;color:#68d391;min-width:90px;'
                f'padding-top:2px;text-transform:uppercase;letter-spacing:0.05em;flex-shrink:0">{_label}</span>'
                f'<span style="font-size:12px;color:#e2e8f0;line-height:1.45">{_text}</span>'
                f'</div>'
            )
        # Initiative cards — 3-column grid
        body += '<div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:8px;margin:12px 0">'
        for _ilabel, _iname, _ihorizon, _idesc, _iebitda in _TLDR_INITIATIVES:
            body += (
                f'<div style="background:#2d3748;border-radius:6px;padding:10px 12px">'
                f'<div style="font-size:10px;font-weight:700;color:#68d391;text-transform:uppercase;'
                f'letter-spacing:0.05em;margin-bottom:3px">{_ilabel} · {_ihorizon}</div>'
                f'<div style="font-size:12px;font-weight:600;color:#f7fafc;margin-bottom:4px">{_iname}</div>'
                f'<div style="font-size:11px;color:#cbd5e0;margin-bottom:5px;line-height:1.4">{_idesc}</div>'
                f'<div style="font-size:11px;font-weight:600;color:#68d391">{_iebitda} (directional)</div>'
                f'</div>'
            )
        body += '</div>'
        # Bottom rows: Exit Timing / Recommendation
        for _label, _text in _TLDR_BOTTOM:
            body += (
                f'<div style="display:flex;gap:14px;margin-bottom:7px;align-items:flex-start">'
                f'<span style="font-size:10px;font-weight:700;color:#68d391;min-width:90px;'
                f'padding-top:2px;text-transform:uppercase;letter-spacing:0.05em;flex-shrink:0">{_label}</span>'
                f'<span style="font-size:12px;color:#e2e8f0;line-height:1.45">{_text}</span>'
                f'</div>'
            )
        body += '</div></div>'

        # ── TOC ───────────────────────────────────────────────────────────────
        body += toc([
            {"number": 1, "id": "biz-profile",    "title": "Business Profile & Financial Snapshot"},
            {"number": 2, "id": "av-trajectory",  "title": "AV & EV Adoption Trajectory"},
            {"number": 3, "id": "rev-vuln",        "title": "Revenue Vulnerability by Service Line"},
            {"number": 4, "id": "strat-options",   "title": "Strategic Response Options"},
            {"number": 5, "id": "moat",            "title": "Competitive Moat Assessment"},
            {"number": 6, "id": "scenarios",       "title": "Investment Scenario Analysis"},
            {"number": 7, "id": "watchpoints",     "title": "Key Watchpoints & Recommended Actions"},
            {"number": 8, "id": "ai-initiatives",  "title": "AI Initiatives: Operational Intelligence, Agentic Learning & Predictive Data"},
        ])

        # ════════════════════════════════════════════════════════════════════════
        # SECTION 1 — Business Profile & Financial Snapshot
        # ════════════════════════════════════════════════════════════════════════
        body += section_start(1, "Business Profile & Financial Snapshot", "biz-profile")
        body += f"""<p>Les Schwab Tire Centers is the dominant regional auto service brand in the
        western United States, operating ~{_STORE_COUNT} stores across {_STATES_COUNT} states with
        estimated annual revenue of <strong>${_TOTAL_REVENUE_M / 1000:.1f}B</strong> and gross profit
        of approximately <strong>${_TOTAL_GP_M}M (~{round(_TOTAL_GP_M/_TOTAL_REVENUE_M*100)}% GP margin)</strong>
        {_cite("Nexdata analyst estimate, 2025E")}. The P&amp;L bridge: ~${_TOTAL_GP_M}M gross profit
        less store operating costs (~$150M), G&amp;A (~$50M), and profit-sharing distributions
        yields approximately <strong>${_TOTAL_EBITDA_M}M EBITDA (~{round(_TOTAL_EBITDA_M/_TOTAL_REVENUE_M*100)}% margin)</strong>,
        consistent with the roughly {_ENTRY_EBITDA_X:.0f}x EBITDA entry multiple paid by Meritage
        Group in 2020 {_cite("Meritage Group press release, Jul 2020")}. Founded in 1952 in
        Prineville, Oregon, tires constitute the core (~60% of revenue) but at thin ~16% GP margins
        due to product cost. High-margin lines — alignments (~58% GP) and brakes (~32% GP) — punch
        above their revenue weight in profit contribution
        {_cite("Nexdata service-line model, 2025E")}. The competitive differentiation rests on the
        "Les Schwab Promise" — free lifetime flat repair, free brake inspections, and employee
        profit-sharing that drives service quality far above the national chain average.</p>"""

        # Revenue mix donut + EV stations bar
        mix_labels = [r[0] for r in _REVENUE_MIX]
        mix_values = [float(r[1]) for r in _REVENUE_MIX]
        mix_colors = [BLUE, RED, GREEN, ORANGE, TEAL, GRAY]
        donut_cfg  = build_doughnut_config(mix_labels, mix_values, mix_colors)
        donut_json = json.dumps(donut_cfg)

        body += '<div class="chart-row"><div>'
        body += '<div class="chart-title">Revenue Mix by Service Line (2024E)</div>'
        body += chart_container("revMixChart", donut_json,
                                build_bar_fallback(mix_labels, mix_values, BLUE),
                                size="medium", title="")
        charts_js += chart_init_js("revMixChart", donut_json)
        body += '</div><div>'

        # EV stations in Schwab footprint states (live)
        fp_states = afdc.get("footprint_states", [])
        if fp_states:
            ev_labels = [s["state"] for s in fp_states[:10]]
            ev_vals   = [s["total"] for s in fp_states[:10]]
            ev_colors = [GREEN] * len(ev_labels)
            ev_bar    = build_horizontal_bar_config(ev_labels, ev_vals, ev_colors, "EV Stations")
            ev_json   = json.dumps(ev_bar)
            body += f'<div class="chart-title">EV Charging Stations — Les Schwab Core Markets (AFDC/NREL, as of {afdc_date})</div>'
            body += chart_container("evStateChart", ev_json,
                                    build_bar_fallback(ev_labels, ev_vals, GREEN),
                                    size="medium", title="")
            charts_js += chart_init_js("evStateChart", ev_json)
        else:
            body += '<div class="chart-title">EV Stations — data not yet ingested</div>'

        body += '</div></div>'

        # Revenue mix table — with GP margin and GP$ columns
        mix_headers = ["Service Line", "Revenue (2025E)", "% of Rev", "GP Margin", "Gross Profit", "AV Disruption Vector", "Risk Level"]
        mix_rows    = [
            [r[0], f"${r[1]:,}M", f"{r[2]}%", f"{r[3]}%", f"${r[4]}M", r[5], r[6]]
            for r in _REVENUE_MIX
        ]
        total_gp = sum(r[4] for r in _REVENUE_MIX)
        mix_rows.append([
            "<strong>Total</strong>",
            f"<strong>${_TOTAL_REVENUE_M:,}M</strong>",
            "<strong>100%</strong>",
            f"<strong>{round(total_gp / _TOTAL_REVENUE_M * 100)}%</strong>",
            f"<strong>~${total_gp}M</strong>",
            "", "",
        ])
        body += self.render_risk_table(mix_headers, mix_rows, risk_col=-1)

        body += callout(
            "<strong>Positive signal:</strong> 60% of revenue sits in tires — a category that is "
            "neutral-to-tailwind in the EV era. EV tires wear 20–30% faster than ICE equivalents, "
            "supporting both ASP expansion and visit frequency.",
            variant="good",
        )
        body += callout(
            "<strong>Risk:</strong> Brakes represent $315M in revenue (15% of total) but ~$101M "
            "in gross profit — the second-highest GP contributor despite being only the second "
            "line by revenue. At ~32% GP margin, brake work is nearly 2x as profitable per dollar "
            "as tires. Regenerative braking is expected to reduce mechanical brake service demand "
            "by 40–60% as EV fleet share grows — making this the most urgent GP line to monitor.",
            variant="warn",
        )

        # Key Assumptions transparency box
        body += callout(
            "<strong>Note on financial projections:</strong> All revenue, cost savings, and EBITDA "
            "figures in this report are directional estimates. The table below surfaces the key "
            "modeling assumptions. Actual outcomes will vary based on EV adoption timing, "
            "implementation choices, and competitive dynamics.",
            variant="info",
        )
        assump_headers = ["Assumption", "Value Used", "Primary Source", "Sensitivity"]
        body += data_table(assump_headers, [list(r) for r in _KEY_ASSUMPTIONS])

        body += section_end()

        # ════════════════════════════════════════════════════════════════════════
        # SECTION 2 — AV & EV Adoption Trajectory
        # ════════════════════════════════════════════════════════════════════════
        body += section_start(2, "AV & EV Adoption Trajectory", "av-trajectory")

        totalsa_str = f"{totalsa:.1f}M units/yr (SAAR)" if totalsa else "stable"
        totalsa_cite = _cite(f"FRED TOTALSA, {totalsa_date}") if totalsa else ""
        body += f"""<p>The MD's core concern is timing and magnitude. The critical insight is that
        two distinct disruption waves hit Les Schwab at different speeds: the EV wave (already
        underway, materially impacts brakes within 5–8 years) and the AV ownership-reduction wave
        (15–20 year horizon). US vehicle sales running at <strong>{totalsa_str}</strong>
        {totalsa_cite} confirm a still-healthy total addressable market. New car EV sales are
        projected to reach ~30% of US new sales by 2030 {_cite("BloombergNEF EV Outlook 2025")},
        but total fleet penetration lags new-car sales by 8–10 years given average vehicle age
        ~12 years {_cite("BTS National Transportation Statistics 2024")}. By 2030, EVs will
        constitute an estimated 12–16% of all registered vehicles — enough to meaningfully erode
        brake revenue but not yet existential. AV fleet impacts on personal vehicle ownership
        are a 2035–2045 story {_cite("Wood Mackenzie AV Fleet Forecast 2024")}.</p>"""

        yr_labels = [str(y) for y in _YEARS]

        # EV fleet penetration — 3 scenarios
        ev_cfg = build_line_chart_config(
            labels=yr_labels,
            datasets=[
                {"label": "EV Fleet % — Bull",  "data": _EV_FLEET_BULL,  "color": GREEN,  "fill": False},
                {"label": "EV Fleet % — Base",  "data": _EV_FLEET_BASE,  "color": BLUE,   "fill": False},
                {"label": "EV Fleet % — Bear",  "data": _EV_FLEET_BEAR,  "color": ORANGE, "fill": False},
                {"label": "AV Fleet % — Base",  "data": _AV_FLEET_BASE,  "color": PURPLE, "fill": False},
            ],
            y_label="% of Registered Vehicles",
        )
        # Add dashed line for AV fleet
        ev_cfg["data"]["datasets"][3]["borderDash"] = [6, 3]
        ev_json = json.dumps(ev_cfg)

        # Vehicle ownership reduction fan
        own_cfg = build_line_chart_config(
            labels=yr_labels,
            datasets=[
                {"label": "Ownership Change — Bull", "data": _OWNERSHIP_BULL, "color": GREEN,  "fill": False},
                {"label": "Ownership Change — Base", "data": _OWNERSHIP_BASE, "color": BLUE,   "fill": False},
                {"label": "Ownership Change — Bear", "data": _OWNERSHIP_BEAR, "color": RED,    "fill": False},
            ],
            y_label="% Change vs 2024 Baseline",
        )
        own_json = json.dumps(own_cfg)

        body += '<div class="chart-row"><div>'
        body += '<div class="chart-title">EV & AV Fleet Penetration — 3 Scenarios (% of all registered vehicles)</div>'
        body += chart_container("evFleetChart", ev_json,
                                build_bar_fallback(yr_labels, _EV_FLEET_BASE, BLUE),
                                size="tall", title="")
        charts_js += chart_init_js("evFleetChart", ev_json)
        body += '</div><div>'
        body += '<div class="chart-title">Projected Vehicle Ownership Reduction from AV Fleets (% vs. 2024 baseline)</div>'
        body += chart_container("ownChart", own_json,
                                build_bar_fallback(yr_labels, _OWNERSHIP_BASE, ORANGE),
                                size="tall", title="")
        charts_js += chart_init_js("ownChart", own_json)
        body += '</div></div>'

        adp_headers = ["Year", "EV New Sales %", "EV Fleet %", "AV Fleet % (Base)",
                       "Personal Veh. Ownership Impact", "Les Schwab Impact"]
        body += data_table(adp_headers, [list(map(str, r)) for r in _ADOPTION_SCENARIOS],
                           numeric_columns={0})

        body += callout(
            "<strong>Insight:</strong> The most consequential near-term decision is not about "
            "autonomous vehicles — it's about electric vehicles. The EV brake headwind is a "
            "certainty on a 5–10 year timeline. AV-driven ownership reduction is a real but "
            "longer-dated risk, likely 15+ years from being material at scale for Les Schwab's "
            "geography (western US suburban/rural — later AV adoption than urban cores).",
            variant="info",
        )
        body += callout(
            "<strong>Risk:</strong> The scenario fan widens dramatically after 2032. The bear case "
            "suggests personal vehicle ownership could fall 30–40% in Les Schwab's core metros by "
            "2038, roughly halving total addressable visits. This is the tail risk the MD is right "
            "to flag.",
            variant="warn",
        )
        body += section_end()

        # ════════════════════════════════════════════════════════════════════════
        # SECTION 3 — Revenue Vulnerability by Service Line
        # ════════════════════════════════════════════════════════════════════════
        body += section_start(3, "Revenue Vulnerability by Service Line", "rev-vuln")

        gas_str = f"${gas_price:.2f}/gal national average" if gas_price else "current levels"
        gas_cite = _cite(f"FRED GASREGCOVW / EIA, {gas_date}") if gas_price else ""
        body += f"""<p>As established in Section 2, the EV headwind is a 5–10 year story
        while AV ownership impacts are a 2033+ horizon. The near-term picture is more resilient
        than headlines suggest: gas at <strong>{gas_str}</strong> {gas_cite} keeps consumer
        vehicle spending intact, and EVs require performance-grade tires rated for higher loads
        (battery weight adds 400–1,200 lbs) — an estimated 20–30% faster wear rate on average
        {_cite("Continental AG EV Tire Study 2023; Michelin fleet data")} that supports higher
        ASP and more frequent replacement. The risk is concentrated in brakes, not tires.
        Projections below are directional; a ±1-year shift in EV fleet penetration timing
        moves the brake figures by roughly ±$5–8M.</p>"""

        # Brake revenue erosion line
        brake_yr_labels = [str(y) for y in _BRAKE_REV_YEARS]
        brake_cfg = build_line_chart_config(
            labels=brake_yr_labels,
            datasets=[
                {"label": "Brake Revenue — Base ($M)", "data": _BRAKE_REV_BASE, "color": BLUE, "fill": True},
                {"label": "Brake Revenue — Bear ($M)", "data": _BRAKE_REV_BEAR, "color": RED,  "fill": False},
            ],
            y_label="Revenue ($M)",
        )
        brake_cfg["data"]["datasets"][1]["borderDash"] = [5, 4]
        brake_json = json.dumps(brake_cfg)

        # Tire wear comparison bar
        wear_colors = [BLUE, BLUE_LIGHT, GREEN, TEAL, ORANGE]
        wear_cfg  = build_horizontal_bar_config(
            _TIRE_WEAR_LABELS, _TIRE_WEAR_MILES_MM, wear_colors, "Miles per mm tread"
        )
        wear_json = json.dumps(wear_cfg)

        body += '<div class="chart-row"><div>'
        body += '<div class="chart-title">Brake Revenue Erosion vs. EV Fleet Penetration</div>'
        body += chart_container("brakeChart", brake_json,
                                build_bar_fallback(brake_yr_labels, _BRAKE_REV_BASE, BLUE),
                                size="tall", title="")
        charts_js += chart_init_js("brakeChart", brake_json)
        body += '</div><div>'
        body += '<div class="chart-title">Tire Wear Rate: ICE vs. EV vs. AV-Optimized (Miles per mm tread)</div>'
        body += chart_container("wearChart", wear_json,
                                build_bar_fallback(_TIRE_WEAR_LABELS, _TIRE_WEAR_MILES_MM),
                                size="tall", title="")
        charts_js += chart_init_js("wearChart", wear_json)
        body += '</div></div>'

        rev_headers = ["Service Line", "2024E", "2028E (Base)", "2030E (Base)",
                       "2035E (Base)", "2035E (Bear)", "Key Driver"]
        body += data_table(rev_headers, [list(r) for r in _REV_PROJECTIONS],
                           numeric_columns={1, 2, 3, 4, 5})

        body += callout(
            "<strong>Positive signal:</strong> In the base case, total revenue is flat-to-growing "
            "through 2030. The EV tire tailwind and ADAS calibration opportunity more than offset "
            "brake erosion in the near term. This is a more resilient picture than the headline "
            "'AV will destroy tire shops' narrative suggests.",
            variant="good",
        )
        body += callout(
            "<strong>Risk:</strong> The bear case 2035 revenue of $1.785B represents a -15% decline "
            "from 2024. At current EBITDA margins (~12–14%), this implies EBITDA compression from "
            "~$260M to ~$215M — material for a $2.1B acquisition underwritten to margin expansion.",
            variant="warn",
        )
        body += section_end()

        # ════════════════════════════════════════════════════════════════════════
        # SECTION 4 — Strategic Response Options
        # ════════════════════════════════════════════════════════════════════════
        body += section_start(4, "Strategic Response Options", "strat-options")

        wage_str = f"${auto_tech_wage:,.0f}/yr" if auto_tech_wage else "~$55K/yr"
        wage_cite = _cite(f"BLS OES SOC 49-3023, {bls_oes_period}") if auto_tech_wage else _cite("BLS OES estimate")
        body += f"""<p>Les Schwab has a credible set of strategic responses — but the window to
        act is defined by the EV fleet curve, not the AV curve. The company has roughly 5–8 years
        to reposition before brake revenue erosion becomes a significant EBITDA headwind. Auto
        service technicians earn a mean of <strong>{wage_str}</strong> {wage_cite} — ADAS-certified
        techs command a 20–35% premium {_cite("ASE certification wage premium study 2024")},
        incentivizing upskilling within the existing workforce. The most value-accretive path is
        to capture ADAS/AV sensor calibration as a new high-margin service line added to existing
        alignment visits {_cite("Hunter Engineering ADAS market analysis 2024")}.</p>"""

        # Strategic options priority bar
        opt_labels = [r[0][:35] for r in _STRATEGIC_OPTIONS]
        opt_margin = [float(r[2].replace("%", "").split("–")[0].replace("+", "").strip())
                      if "%" in r[2] else 40 for r in _STRATEGIC_OPTIONS]
        opt_colors = [GREEN if "1" in r[5] else BLUE if "2" in r[5] else ORANGE
                      for r in _STRATEGIC_OPTIONS]
        opt_cfg   = build_horizontal_bar_config(opt_labels, opt_margin, opt_colors, "Gross Margin %")
        opt_json  = json.dumps(opt_cfg)

        # ADAS calibration revenue opportunity bar
        adas_cfg  = build_line_chart_config(
            labels=_ADAS_REVENUE_LABELS,
            datasets=[
                {"label": "Revenue @ $80/job ($M)",  "data": _ADAS_REVENUE_LOW,  "color": BLUE,  "fill": True},
                {"label": "Revenue @ $150/job ($M)", "data": _ADAS_REVENUE_HIGH, "color": GREEN, "fill": False},
            ],
            y_label="Incremental Revenue ($M)",
        )
        adas_json = json.dumps(adas_cfg)

        body += '<div class="chart-row"><div>'
        body += '<div class="chart-title">Strategic Option Value vs. Implementation Complexity (Margin %)</div>'
        body += chart_container("optChart", opt_json,
                                build_bar_fallback(opt_labels, opt_margin, BLUE),
                                size="tall", title="")
        charts_js += chart_init_js("optChart", opt_json)
        body += '</div><div>'
        body += '<div class="chart-title">ADAS Calibration Revenue Opportunity (Incremental per Alignment Job)</div>'
        body += chart_container("adasChart", adas_json,
                                build_bar_fallback(_ADAS_REVENUE_LABELS, _ADAS_REVENUE_HIGH, GREEN),
                                size="tall", title="")
        charts_js += chart_init_js("adasChart", adas_json)
        body += '</div></div>'

        opt_headers = ["Strategic Option", "Revenue Potential", "Margin Profile",
                       "Time to Material", "Complexity", "Priority"]
        body += self.render_priority_table(opt_headers,
                                           [list(r) for r in _STRATEGIC_OPTIONS],
                                           priority_col=-1)

        body += callout(
            "<strong>Positive signal:</strong> ADAS calibration is an immediately actionable, "
            "high-margin add-on at existing alignment bays. Les Schwab currently performs the "
            "alignment but does not offer the calibration step, leaving that revenue at ~$0. "
            "Industry standard is $80–$150 per job. With an estimated ~1.4M alignments annually "
            "<em>(modeled est.: $168M rev ÷ ~$90 ASP)</em>, a full rollout could represent "
            "$112–$210M in incremental revenue — though adoption pace depends on OEM calibration "
            "requirements (mandatory on ~65% of MY2023+ ADAS-equipped vehicles per NHTSA) and "
            "whether insurance workflows begin covering calibration as standard, a trend "
            "accelerating in collision repair but not yet common in tire/alignment shops.",
            variant="good",
        )
        body += callout(
            "<strong>Insight:</strong> Fleet service is the long-term answer to AV ownership "
            "reduction. Robo-taxi fleets need tires changed 4–6x/year per vehicle at 50K–100K "
            "annual miles — a single 10,000-vehicle contract equals a mid-size store's annual "
            "volume. Timing is uncertain and depends on AV commercial deployment pace — "
            "currently accelerating in select metros but still a 5–10 year horizon for scale "
            "relevant to Les Schwab's western US footprint.",
            variant="info",
        )
        body += section_end()

        # ════════════════════════════════════════════════════════════════════════
        # SECTION 5 — Competitive Moat Assessment
        # ════════════════════════════════════════════════════════════════════════
        body += section_start(5, "Competitive Moat Assessment", "moat")
        nps_cite = _cite("BrightLocal Auto Service Consumer Survey 2024; industry estimates")
        body += f"""<p>Les Schwab's moat is regional and cultural — extraordinarily deep within its
        geography, essentially nonexistent outside it. The brand's Net Promoter Score in the Pacific
        Northwest (~72–78) rivals premium consumer brands and is roughly 2–3x higher than national
        chains like Firestone (NPS ~42), Discount Tire (NPS ~55), and Pep Boys (NPS ~28) {nps_cite}.
        This loyalty is structurally tied to the employee profit-sharing model — the ESOP structure
        aligns technician incentives with customer outcomes in a way national chains cannot replicate
        {_cite("Les Schwab corporate profile; ESOP Association 2023")}. The moat's durability in an
        AV world depends on whether the profit-sharing culture can be maintained as service mix
        shifts away from high-frequency brake jobs toward more complex, lower-frequency EV/AV
        services.</p>"""

        # NPS comparison bar
        nps_labels = [r[0] for r in _NPS_COMPARISON]
        nps_values = [float(r[1]) for r in _NPS_COMPARISON]
        nps_colors = [GREEN if r[0].startswith("Les Schwab") else BLUE for r in _NPS_COMPARISON]
        nps_cfg    = build_horizontal_bar_config(nps_labels, nps_values, nps_colors, "NPS Score")
        nps_json   = json.dumps(nps_cfg)

        # Visit frequency trend
        vf_yr_labels = [str(y) for y in _VF_YEARS]
        vf_cfg = build_line_chart_config(
            labels=vf_yr_labels,
            datasets=[
                {"label": "ICE Vehicle", "data": _VF_ICE,      "color": BLUE,  "fill": False},
                {"label": "EV Vehicle",  "data": _VF_EV,       "color": GREEN, "fill": False},
                {"label": "AV Fleet",    "data": _VF_AV_FLEET, "color": RED,   "fill": False},
            ],
            y_label="Annual Service Visits per Vehicle",
        )
        vf_cfg["data"]["datasets"][2]["borderDash"] = [5, 4]
        vf_json = json.dumps(vf_cfg)

        body += '<div class="chart-row"><div>'
        body += '<div class="chart-title">Brand NPS Comparison — Western US Auto Service (2025E)</div>'
        body += chart_container("npsChart", nps_json,
                                build_bar_fallback(nps_labels, nps_values, BLUE),
                                size="tall", title="")
        charts_js += chart_init_js("npsChart", nps_json)
        body += '</div><div>'
        body += '<div class="chart-title">Service Visit Frequency Trend: ICE vs. EV vs. AV (Annual visits per vehicle)</div>'
        body += chart_container("vfChart", vf_json,
                                build_bar_fallback(vf_yr_labels, _VF_ICE, BLUE),
                                size="tall", title="")
        charts_js += chart_init_js("vfChart", vf_json)
        body += '</div></div>'

        moat_headers = ["Moat Dimension", "Current Strength", "Durability in EV Era",
                        "Durability in AV Era", "Key Risk"]
        body += data_table(moat_headers, [list(r) for r in _MOAT_DIMENSIONS])

        body += callout(
            "<strong>Risk:</strong> The 'free brake inspection' marketing pillar — one of the most "
            "effective customer acquisition tools — becomes economically irrational as EV fleet share "
            "grows. By 2030, roughly 14% of vehicles won't need the brake service that follows the "
            "inspection. This is a marketing efficiency problem as much as a revenue problem.",
            variant="warn",
        )
        body += callout(
            "<strong>Insight:</strong> Les Schwab's brand moat is personal-vehicle-centric. Fleet "
            "operators choose tire service providers on price, contract terms, and geographic coverage "
            "— not brand affinity. Building a B2B sales capability is necessary to compete in fleet "
            "servicing, but the timeline depends on AV commercialization pace — a 5–10 year horizon "
            "for scale in Les Schwab's markets, with meaningful uncertainty on both ends.",
            variant="info",
        )
        body += section_end()

        # ════════════════════════════════════════════════════════════════════════
        # SECTION 6 — Investment Scenario Analysis
        # ════════════════════════════════════════════════════════════════════════
        body += section_start(6, "Investment Scenario Analysis", "scenarios")
        body += """<p>The investment outcome pivots on exit timing and the pace of strategic
        execution. A 2027–2028 process likely captures significant value — brake erosion is still
        modest, and if AI-assisted store operations are showing early traction, Meritage may present
        a proven transition playbook rather than a business still dependent on ICE revenue. A hold
        through 2030–2032 carries more execution risk but potentially more upside: successful
        deployment of agentic store intelligence and ADAS/EV specialty positioning
        <em>could</em> add an estimated $50–90M to EBITDA, meaningfully expanding the exit multiple.
        Beyond 2033 the scenario range widens considerably and outcomes depend heavily on how quickly
        AV fleet adoption affects personal vehicle ownership in Les Schwab's western US markets.</p>"""

        sc_yr_labels = [str(y) for y in _SCENARIO_YEARS]

        # Revenue trajectory
        rev_cfg = build_line_chart_config(
            labels=sc_yr_labels,
            datasets=[
                {"label": "Revenue — Bull ($M)", "data": _REV_BULL,  "color": GREEN, "fill": False},
                {"label": "Revenue — Base ($M)", "data": _REV_BASE,  "color": BLUE,  "fill": False},
                {"label": "Revenue — Bear ($M)", "data": _REV_BEAR,  "color": RED,   "fill": False},
            ],
            y_label="Revenue ($M)",
        )
        rev_cfg["data"]["datasets"][2]["borderDash"] = [5, 4]
        rev_json = json.dumps(rev_cfg)

        # EBITDA by scenario
        ebitda_cfg = build_line_chart_config(
            labels=sc_yr_labels,
            datasets=[
                {"label": "EBITDA — Bull ($M)", "data": _EBITDA_BULL, "color": GREEN, "fill": False},
                {"label": "EBITDA — Base ($M)", "data": _EBITDA_BASE, "color": BLUE,  "fill": False},
                {"label": "EBITDA — Bear ($M)", "data": _EBITDA_BEAR, "color": RED,   "fill": False},
            ],
            y_label="EBITDA ($M)",
        )
        ebitda_cfg["data"]["datasets"][2]["borderDash"] = [5, 4]
        ebitda_json = json.dumps(ebitda_cfg)

        body += '<div class="chart-row"><div>'
        body += '<div class="chart-title">Revenue Trajectory by Scenario ($M)</div>'
        body += chart_container("revScChart", rev_json,
                                build_bar_fallback(sc_yr_labels, _REV_BASE, BLUE),
                                size="tall", title="")
        charts_js += chart_init_js("revScChart", rev_json)
        body += '</div><div>'
        body += '<div class="chart-title">EBITDA by Scenario ($M)</div>'
        body += chart_container("ebitdaChart", ebitda_json,
                                build_bar_fallback(sc_yr_labels, _EBITDA_BASE, TEAL),
                                size="tall", title="")
        charts_js += chart_init_js("ebitdaChart", ebitda_json)
        body += '</div></div>'

        sc_headers = ["Scenario", "2024E EBITDA", "2028E EBITDA (illus.)", "2032E EBITDA (illus.)",
                      "Exit Timing", "EV Multiple", "Implied EV (illus.)"]
        body += self.render_scenario_table(sc_headers, [list(r) for r in _SCENARIOS_TABLE])
        body += ('<p style="font-size:11px;color:#718096;margin-top:4px">'
                 '* EBITDA and EV figures are illustrative scenario estimates. '
                 'Actual outcomes depend on strategic execution, EV adoption timing, '
                 'and market conditions at exit.</p>')

        body += callout(
            "<strong>Risk:</strong> The bear case 2032 EV of $1.5–1.9B represents a loss on the "
            "2020 acquisition ($2.1B entry). This is not the base case, but requires only moderately "
            "faster-than-base AV adoption and management inertia on strategic pivots.",
            variant="warn",
        )
        body += callout(
            "<strong>Positive signal:</strong> The bull case generates a 2.3x MOIC on the 2020 "
            "investment — achievable with a 2030–2032 exit if the ADAS calibration build-out and "
            "EV specialty programs deliver. The strategic window is open.",
            variant="good",
        )
        body += callout(
            "<strong>Insight:</strong> The most likely outcome is a 2027–2029 sale to a strategic "
            "buyer (national tire chain or auto service consolidator) who wants the Pacific NW brand "
            "at a premium before the AV thesis becomes consensus. Meritage's best play is to "
            "accelerate ADAS rollout to demonstrate the new revenue line in the next 12–24 months, "
            "then run a process in 2027–2028 while the growth story is still intact.",
            variant="info",
        )
        body += section_end()

        # ════════════════════════════════════════════════════════════════════════
        # SECTION 7 — Key Watchpoints & Recommended Actions
        # ════════════════════════════════════════════════════════════════════════
        body += section_start(7, "Key Watchpoints & Recommended Actions", "watchpoints")
        body += """<p>The MD's core question — 'where does the business go?' — has a clear answer
        in the near term: the business stays in tires and pivots from brake-centric to ADAS/EV-centric
        services. The following watchpoints define what 'good execution' looks like over the next
        24 months.</p>"""

        wp_headers = ["Watchpoint", "Signal to Watch", "Green (on track)", "Red (off track)", "Cadence"]
        body += self.render_watchpoints(wp_headers, [list(r) for r in _WATCHPOINTS])

        for label, text_content in _ACTIONS:
            variant = "good" if "Immediate" in label else "info" if "Strategic" in label else "info"
            body += callout(f"<strong>{label}:</strong> {text_content}", variant=variant)

        body += '<div style="margin-top:28px;font-size:15px;font-weight:700;color:#2d3748;margin-bottom:4px">Execution Risks</div>'
        body += callout(
            "<strong>ADAS Rollout Complexity:</strong> Calibration requires dedicated equipment "
            "($15–25K per bay), technician certification, and procedural rigor — an incorrect "
            "calibration creates liability exposure that outweighs the revenue upside. "
            "A phased rollout across the highest-EV-density stores first reduces risk.",
            variant="warn",
        )
        body += callout(
            "<strong>Fleet Sales Capability Gap:</strong> Fleet B2B requires a different sales "
            "motion — contracts, SLAs, invoicing, dedicated reps — that Les Schwab's "
            "retail-trained organization does not have today. Hiring 3–5 fleet reps is table "
            "stakes; building the back-office systems is the harder, longer part.",
            variant="warn",
        )
        body += callout(
            "<strong>Store Manager Adoption (AI):</strong> AI-assisted store briefings only "
            "create value if managers read and act on them. Rollout requires change management, "
            "not just a software deployment. Les Schwab's strong store culture is an asset "
            "here — but it also means change takes longer than at a franchise chain.",
            variant="warn",
        )
        body += callout(
            "<strong>Data Quality Bootstrap:</strong> Initiative 1 depends on connecting POS "
            "data to a structured pipeline. If Les Schwab's POS systems are fragmented — "
            "common in regional chains — Track A takes longer than planned, which delays "
            "the GenAI quality improvement cycle by the same amount. Track B (Claude briefings) "
            "can still deploy on existing exports; it just starts with noisier data.",
            variant="warn",
        )

        body += section_end()

        # ════════════════════════════════════════════════════════════════════════
        # SECTION 8 — AI Initiatives: Three Horizons of Value Creation
        # ════════════════════════════════════════════════════════════════════════
        body += section_start(8, "AI Initiatives: Operational Intelligence, Agentic Learning & Predictive Data (2026–2028)", "ai-initiatives")

        wage_str_ai = f"${auto_tech_wage:,.0f}" if auto_tech_wage else "~$55,260"
        body += f"""<p>Les Schwab can unlock meaningful operational value through AI on a
        3-year horizon — without waiting for perfect data or new infrastructure. The three
        initiatives below are directional opportunities, not commitments; specific financial
        outcomes will depend on implementation choices and execution quality. What the data
        confirms today: <strong>{ev_fp:,} EV charging stations</strong> are already in Les
        Schwab's core markets {_cite(f"AFDC/NREL, {afdc_date}")}, consumer sentiment at
        <strong>{umcsent:.1f}</strong> {_cite(f"FRED UMCSENT, {umcsent_date}")} suggests
        customers are deferring discretionary service, and labor costs are rising — auto tech
        mean wage <strong>{wage_str_ai}/yr</strong> {_cite(f"BLS OES SOC 49-3023, {bls_oes_period}")}
        — meaning the business case for operational intelligence exists today, not in 2028.</p>"""

        ai_yr_labels = [str(y) for y in _AI_YEARS]

        # ── Initiative 1 ──────────────────────────────────────────────────────
        body += """<div style="margin:28px 0 6px;padding:10px 16px;background:#f0fff4;
        border-left:4px solid #38a169;border-radius:4px">
        <span style="font-size:15px;font-weight:700;color:#276749">Initiative 1 — Operational
        Intelligence &amp; AI-Assisted Store Briefings (2026)</span>
        <span style="font-size:12px;color:#276749;margin-left:12px">Investment: TBD
        &nbsp;&middot;&nbsp; Payback potential: &lt;12 months</span></div>"""

        body += f"""<p style="margin-bottom:12px">Year 1 runs two parallel tracks — neither
        waits for the other. <strong>Track A — Data Foundation:</strong> Connect POS and
        service management data to a structured pipeline with 6 core KPIs tracked weekly
        per store, overlaid with AFDC EV density and FRED macro signals. Infrastructure
        build cost is dependent on implementation approach; agentic development tools may
        compress timelines and cost materially vs. traditional consulting-led builds.
        <strong>Track B — Claude Weekly Store Analysis (~$3K/yr running cost):</strong>
        Deploy immediately on existing POS exports — no data infrastructure prerequisite.
        Claude generates a weekly one-page briefing per store: key observations, 3 prioritized
        actions, and market context from live Nexdata signals. As Track A matures, briefing
        quality improves automatically. With <strong>{ev_fp:,} EV stations</strong> in the
        footprint {_cite(f"AFDC/NREL, {afdc_date}")} and sentiment at {umcsent:.1f}
        {_cite(f"FRED UMCSENT, {umcsent_date}")}, these briefings surface which stores are
        most exposed — from Day 1, not Day 365.</p>"""

        ai1_impact = [
            ["Revenue impact",      "$5–12M",  "Faster issue identification reduces margin leakage; managers act in days not quarters"],
            ["Cost savings",        "$18–25M", "Labor scheduling ~$12–16M, parts waste ~$3–5M, rework ~$2M, management overhead ~$1–2M"],
            ["EBITDA contribution", "$23–37M", "Directional estimate; range reflects data quality at deployment and store manager adoption pace"],
            ["Investment",          "TBD",     "Infrastructure build cost depends on implementation approach; agentic dev tools may compress significantly"],
            ["Payback potential",   "< 12 mo", "Y1 cost savings alone may cover the full 3-year program investment"],
        ]
        body += data_table(["Dimension", "Directional Range", "Basis"], ai1_impact)

        body += callout(
            "<strong>GenAI from Day 1:</strong> Track B deploys on existing POS exports in "
            "~30 days. The data foundation improves briefing quality over 6–12 months — it "
            "does not delay it. Key dependencies: store managers must read and act on briefings "
            "(change management required), and initial data quality will affect early signal "
            "reliability until Track A matures.",
            variant="good",
        )

        # ── Initiative 2 ──────────────────────────────────────────────────────
        body += """<div style="margin:28px 0 6px;padding:10px 16px;background:#ebf8ff;
        border-left:4px solid #3182ce;border-radius:4px">
        <span style="font-size:15px;font-weight:700;color:#2b6cb0">Initiative 2 — Agentic
        Cross-Store Learning (2026–2027)</span>
        <span style="font-size:12px;color:#2b6cb0;margin-left:12px">Investment: TBD
        &nbsp;&middot;&nbsp; Requires Initiative 1 data foundation</span></div>"""

        body += """<p style="margin-bottom:12px">Once Initiative 1 establishes structured KPI
        tracking, an agentic layer monitors all 500 stores simultaneously and propagates
        top-quartile patterns to lagging stores. The agent identifies what high-performing stores
        do differently — ADAS attach rate 22% vs. org average 6%, for example — and generates
        specific, store-level next best actions grounded in Nexdata signals (local EV station
        density, consumer sentiment, regional labor cost trends). It learns from outcomes: which
        recommendations moved the metrics? Over time, the system builds a tested playbook for
        each market condition. This initiative is only viable because Initiative 1 built the
        data foundation — sequencing is the strategy.</p>"""

        ai2_impact = [
            ["Revenue impact",      "$15–42M", "Closing top/bottom store performance gap on ADAS attach, EV tire mix, and visit recovery outreach"],
            ["Cost savings",        "$7–10M",  "Warranty &amp; returns ~$4–6M, reduced field supervision ~$1–2M, training efficiency ~$1–2M"],
            ["EBITDA contribution", "$22–52M", "Incremental to Initiative 1; wide range reflects agent adoption pace across 500 diverse stores"],
            ["Investment",         "TBD",      "LLM agent build + CRM integration; cost depends on implementation approach"],
        ]
        body += data_table(["Dimension", "Directional Range", "Basis"], ai2_impact)

        body += callout(
            "<strong>500 stores simultaneously:</strong> A regional manager actively tracks "
            "8–12 stores. An AI agent tracks all 500 — flagging the Phoenix cluster where "
            "90-day return rates dropped 6pp before it appears in quarterly results, or the "
            "Seattle market where one store manager changed the check-in script and ADAS attach "
            "tripled. These patterns exist today. They are invisible without this layer.",
            variant="info",
        )

        # ── Initiative 3 ──────────────────────────────────────────────────────
        body += """<div style="margin:28px 0 6px;padding:10px 16px;background:#fffaf0;
        border-left:4px solid #dd6b20;border-radius:4px">
        <span style="font-size:15px;font-weight:700;color:#c05621">Initiative 3 — Passive Data
        Collection &amp; Predictive Intelligence (2027–2028)</span>
        <span style="font-size:12px;color:#c05621;margin-left:12px">Investment: TBD
        &nbsp;&middot;&nbsp; Builds on Initiatives 1 &amp; 2</span></div>"""

        ca_stations = next((s["total"] for s in fp_states if s["state"] == "CA"), 20334)
        body += f"""<p style="margin-bottom:12px">With 12–18 months of structured KPI data
        and agent-validated recommendations, Les Schwab has the training data to build genuinely
        predictive models. Initiative 3 instruments physical stores to passively collect data
        without requiring technician action: license plate recognition at bay entry reads the
        VIN, looks up the vehicle's ADAS package and service history, and pre-populates the
        check-in form. Fleet customers connect their telematics — tire wear data flows in
        automatically, and the system may schedule appointments before the fleet manager notices
        a problem. The <strong>{ev_fp:,} EV stations</strong> already in Les Schwab's markets
        {_cite(f"AFDC/NREL, {afdc_date}")} represent the first cohort of high-mileage EV fleet
        customers — California alone ({ca_stations:,} stations) is already a commercial fleet
        market. Predictive outreach to EV-dense ZIP codes alone could represent a measurable
        revenue increment as that fleet matures.</p>"""

        ai3_impact = [
            ["Revenue impact",      "$18–140M", "Fleet contracts + predictive outreach + ADAS capture at scale; upper end requires B2B fleet portal build-out"],
            ["Cost savings",        "$15–18M",  "LPR automation, planned vs. rush parts ordering, overtime reduction"],
            ["EBITDA contribution", "$33–43M",  "Incremental to Initiatives 1 &amp; 2; narrower range because data foundation is validated by this stage"],
            ["Investment",         "TBD",       "LPR hardware + telematics API + ML infrastructure + fleet portal; cost depends on implementation approach"],
        ]
        body += data_table(["Dimension", "Directional Range", "Basis"], ai3_impact)

        body += callout(
            "<strong>Fleet B2B portal (aspirational):</strong> The highest-upside element of "
            "Initiative 3 is a self-service fleet manager dashboard — auto-scheduling, per-vehicle "
            "wear predictions, contract invoicing. This converts Les Schwab from a retail tire "
            "shop into a managed service for commercial fleets. In an AV world where fleet "
            "operators are the end customer, this capability may represent the difference between "
            "relevance and obsolescence.",
            variant="info",
        )

        # ── Combined summary table + revenue ramp chart ────────────────────────
        body += """<div style="margin:32px 0 8px;font-size:15px;font-weight:700;color:#2d3748">
        Combined Financial Summary — All 3 Initiatives (2028 directional estimate)</div>"""

        ai_sum_headers = ["Initiative", "Revenue Impact", "Cost Savings", "EBITDA Contribution", "Investment"]
        ai_sum_rows = [list(r) for r in _AI_INITIATIVES_TABLE]
        ai_sum_rows[-1] = [f"<strong>{c}</strong>" for c in ai_sum_rows[-1]]
        body += data_table(ai_sum_headers, ai_sum_rows)
        body += ('<p style="font-size:11px;color:#718096;margin-top:4px">'
                 'All investment figures are TBD — implementation approach (including agentic '
                 'development tools) will determine actual cost. Revenue, cost savings, and EBITDA '
                 'figures are directional order-of-magnitude estimates, not business case commitments.</p>')

        # Revenue ramp chart — all 3 initiatives vs. brake GP erosion
        ai_ramp_cfg = build_line_chart_config(
            labels=ai_yr_labels,
            datasets=[
                {"label": "Initiative 3 — Passive + Predictive ($M)", "data": _AI3_REV,          "color": ORANGE, "fill": False},
                {"label": "Initiative 2 — Agentic Learning ($M)",      "data": _AI2_REV,          "color": BLUE,   "fill": False},
                {"label": "Initiative 1 — Store Intelligence ($M)",    "data": _AI1_REV,          "color": GREEN,  "fill": False},
                {"label": "Brake GP Erosion (base case, $M)",          "data": _AI_BRAKE_EROSION, "color": RED,    "fill": False},
            ],
            y_label="$M Annual Impact",
        )
        ai_ramp_cfg["data"]["datasets"][3]["borderDash"] = [5, 4]
        ai_ramp_json = json.dumps(ai_ramp_cfg)

        body += '<div class="chart-row"><div>'
        body += '<div class="chart-title">AI Initiative Revenue Ramp vs. Brake GP Erosion ($M) — 3-Year Horizon</div>'
        body += chart_container("aiRampChart", ai_ramp_json,
                                build_bar_fallback(ai_yr_labels, _AI3_REV, ORANGE),
                                size="tall", title="")
        charts_js += chart_init_js("aiRampChart", ai_ramp_json)
        body += '</div><div>'

        if fp_states:
            ev_ai_labels = [s["state"] for s in fp_states[:10]]
            ev_ai_vals   = [s["total"] for s in fp_states[:10]]
            ev_ai_colors = [GREEN if s["state"] == "CA" else BLUE for s in fp_states[:10]]
            ev_ai_bar    = build_horizontal_bar_config(ev_ai_labels, ev_ai_vals, ev_ai_colors, "EV Stations")
            ev_ai_json   = json.dumps(ev_ai_bar)
            body += f'<div class="chart-title">EV Infrastructure — Les Schwab Markets (AFDC/NREL, {afdc_date})</div>'
            body += chart_container("evAiChart", ev_ai_json,
                                    build_bar_fallback(ev_ai_labels, ev_ai_vals, BLUE),
                                    size="tall", title="")
            charts_js += chart_init_js("evAiChart", ev_ai_json)
        body += '</div></div>'

        body += callout(
            "<strong>The math:</strong> The base case — informed by comparable operational AI "
            "deployments in multi-site retail — suggests <strong>~$40–60M in combined EBITDA "
            "contribution</strong> is achievable with disciplined execution across all 3 "
            "initiatives. The upside scenario (all initiatives at the high end) represents an "
            "estimated $78–132M by 2028 (+29–48% of current EBITDA) — treat this as a stretch, "
            "not a base case. All figures are directional. Program investment is TBD pending "
            "implementation approach.",
            variant="good",
        )
        body += callout(
            "<strong>Sequencing is the strategy:</strong> Each initiative is only possible "
            "because the prior one built the data and organizational trust. Skip Initiative 1 "
            "and Initiative 2 has nothing to learn from. Skip Initiative 2 and Initiative 3 "
            "produces predictions with no validated feedback loop. The fastest path to durable "
            "AI value is also the most disciplined one.",
            variant="info",
        )

        body += section_end()

        # ── Footer ────────────────────────────────────────────────────────────
        totalsa_as_of = fs.get("TOTALSA", {}).get("date", "N/A")[:10]
        body += page_footer(
            notes=[
                "── LIVE DATA SOURCES (queried from Nexdata DB at report generation) ──",
                f"US Vehicle Sales SAAR (TOTALSA): Federal Reserve Bank of St. Louis / FRED. "
                f"Series: TOTALSA. As of {totalsa_as_of}. Units: millions of vehicles, seasonally adjusted annual rate.",
                f"Regular Conventional Gas Price (GASREGCOVW): U.S. Energy Information Administration / FRED. "
                f"As of {fs.get('GASREGCOVW', {}).get('date', 'N/A')[:10]}. Weekly national average, $/gallon.",
                f"Consumer Sentiment (UMCSENT): University of Michigan Surveys of Consumers / FRED. "
                f"As of {fs.get('UMCSENT', {}).get('date', 'N/A')[:10]}. Index, 1966=100.",
                f"Auto Service Technician Mean Wage: BLS Occupational Employment & Wage Statistics (OES). "
                f"SOC 49-3023 (Automotive Service Technicians & Mechanics). Period: {bls_oes_period}. Annual survey.",
                f"EV Charging Station Counts: Alternative Fuels Data Center (AFDC) / NREL DOE API. "
                f"Open public EVSE stations only; excludes private/workplace chargers. As of {afdc_date}. "
                f"Schwab footprint states: {', '.join(sorted(_SCHWAB_STATES))}.",
                "── ANALYST ESTIMATES & MODELED PROJECTIONS ──",
                "Revenue mix (2024E) and financial figures: Nexdata analyst estimates derived from public sources. "
                "Not audited financials. Sources include Meritage Group press release (Jul 2020), industry benchmarks.",
                "NPS scores: BrightLocal Auto Service Consumer Survey 2024 and third-party industry estimates. "
                "Competitor NPS values are approximations based on publicly available surveys.",
                "AV/EV adoption scenarios: Nexdata model informed by BloombergNEF EV Outlook 2025, "
                "Wood Mackenzie AV Fleet Forecast 2024, and IEA Global EV Outlook 2025. Actual outcomes will vary.",
                "Tire wear rates: Continental AG EV Tire Study 2023, Michelin fleet operations data, "
                "and Nexdata analysis. EV wear premium estimate: +20-30% vs. comparable ICE tire.",
                "ADAS calibration market: Hunter Engineering market analysis 2024; ASE certification wage data.",
                "EBITDA and EV multiples: Nexdata scenario model. Entry multiple: Meritage acquisition (Jul 2020).",
                f"Report generated {gen_at[:19].replace('T', ' ')} UTC. All values baked in at generation — no live fetch in output HTML.",
                "Confidential — for Investment Committee use only. Not for distribution.",
            ],
            generated_line=(
                f"Nexdata Report Archive | Les Schwab AV Deep Dive | "
                f"Generated {gen_at[:10]}"
            ),
        )

        extra_css = """
        .chart-row { display: grid; grid-template-columns: 1fr 1fr; gap: 24px; margin: 16px 0; }
        @media (max-width: 768px) { .chart-row { grid-template-columns: 1fr; } }
        """

        return html_document(title=title, body_content=body, charts_js=charts_js, extra_css=extra_css)

    # ── Excel stub ────────────────────────────────────────────────────────────

    def render_excel(self, data: Dict[str, Any]) -> bytes:
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        fred = data.get("fred", {})
        bls  = data.get("bls", {})
        afdc = data.get("afdc", {})

        wb       = Workbook()
        hdr_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
        hdr_font = Font(bold=True, color="FFFFFF")

        def _hdr(ws, headers):
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=col, value=h)
                c.font = hdr_font
                c.fill = hdr_fill

        # Sheet 1: Revenue Mix
        ws = wb.active
        ws.title = "Revenue Mix"
        _hdr(ws, ["Service Line", "Revenue ($M)", "% of Total", "Risk Level"])
        for i, r in enumerate(_REVENUE_MIX, 2):
            ws.cell(row=i, column=1, value=r[0])
            ws.cell(row=i, column=2, value=r[1])
            ws.cell(row=i, column=3, value=r[2])
            ws.cell(row=i, column=4, value=r[4])

        # Sheet 2: Scenarios
        ws2 = wb.create_sheet("Scenarios")
        _hdr(ws2, ["Scenario", "2024E EBITDA", "2028E EBITDA", "2032E EBITDA",
                   "Exit Timing", "EV Multiple", "Implied EV"])
        for i, r in enumerate(_SCENARIOS_TABLE, 2):
            for j, v in enumerate(r, 1):
                ws2.cell(row=i, column=j, value=v)

        # Sheet 3: FRED Macro
        ws3 = wb.create_sheet("Macro (FRED)")
        _hdr(ws3, ["Series ID", "Value", "Date"])
        for i, (sid, s) in enumerate(fred.get("series", {}).items(), 2):
            ws3.cell(row=i, column=1, value=sid)
            ws3.cell(row=i, column=2, value=s.get("value"))
            ws3.cell(row=i, column=3, value=s.get("date"))

        # Sheet 4: EV Stations
        ws4 = wb.create_sheet("EV Stations (AFDC)")
        _hdr(ws4, ["State", "Total Stations", "Level 2", "DC Fast", "Schwab Footprint"])
        for i, s in enumerate(afdc.get("states", []), 2):
            ws4.cell(row=i, column=1, value=s["state"])
            ws4.cell(row=i, column=2, value=s["total"])
            ws4.cell(row=i, column=3, value=s["level2"])
            ws4.cell(row=i, column=4, value=s["dc_fast"])
            ws4.cell(row=i, column=5, value="Yes" if s["in_footprint"] else "No")

        out = BytesIO()
        wb.save(out)
        return out.getvalue()
