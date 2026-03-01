"""
Med-Spa Market Analysis Report Template.

Generates an investor-ready aesthetics roll-up thesis report synthesizing
medspa_prospects (Yelp-sourced acquisition targets), zip_medspa_scores
(ZIP affluence model), pe_portfolio_companies (PE platform comps), and
Census ACS income data into a comprehensive market analysis.

Sections:
  1. Executive Summary
  2. Market Map (Chart.js visualizations)
  3. Top Acquisition Targets
  4. Market Concentration Analysis
  5. ZIP Affluence Profile
  6. Competitive Landscape
  7. Data Sources & Methodology
  8. Whitespace Analysis
  9. Workforce Economics
  10. Service Category Breakdown
  11. PE Platform Benchmarking
  12. Review Velocity & Growth Signals
  13. Deal Model — Unit Economics
  14. Deal Model — Capital Requirements
  15. Deal Model — Returns Analysis
  16. Stealth Wealth Signal (IRS SOI non-wage income analysis)
  17. Migration Alpha (IRS county-to-county wealth flow leading indicator)
  18. Medical Provider Density Signal (CMS Medicare provider-to-medspa imbalance)
  19. Real Estate Appreciation Alpha (Redfin/FHFA home price timing signal)
  20. Deposit Wealth Concentration (FDIC branch deposits per capita)
  21. Business Formation Velocity (IRS SOI business income density)
  22. Opportunity Zone Overlay (tax-advantaged roll-up targets)
  23. Demographic Demand Model (education-driven demand proxy)
  24. PE Competitive Heat Map (aesthetics deal activity)
  25. Construction Momentum Signal (HUD permit growth leading indicator)
  26. Medical CPI Pricing Power (medical vs general inflation spread)
  27. Talent Pipeline Pressure (healthcare JOLTS labor scarcity)
"""

import json
import logging
from datetime import datetime
from typing import Dict, Any, List, Optional
from io import BytesIO

from sqlalchemy import text
from sqlalchemy.orm import Session

from app.reports.design_system import (
    html_document, page_header, kpi_strip, kpi_card,
    toc, section_start, section_end,
    data_table, pill_badge, callout,
    chart_container, chart_init_js, page_footer,
    build_doughnut_config, build_horizontal_bar_config,
    build_bar_fallback, build_chart_legend, CHART_COLORS,
    BLUE, BLUE_LIGHT, BLUE_DARK, ORANGE, GREEN, RED, GRAY,
    PURPLE, TEAL, PINK,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Grade color mapping for badges
# ---------------------------------------------------------------------------

GRADE_BADGE_COLORS = {
    "A": "public",    # blue
    "B": "private",   # green
    "C": "pe",        # amber
    "D": "sub",       # gray
    "F": "default",   # gray
}

# Prospect scoring weights (mirrored from metadata for methodology section)
PROSPECT_WEIGHTS = {
    "zip_affluence": 0.30,
    "yelp_rating": 0.25,
    "review_volume": 0.20,
    "low_competition": 0.15,
    "price_tier": 0.10,
}

ZIP_SCORE_WEIGHTS = {
    "affluence_density": 0.30,
    "discretionary_wealth": 0.25,
    "market_size": 0.20,
    "professional_density": 0.15,
    "wealth_concentration": 0.10,
}

# ---------------------------------------------------------------------------
# Deal Model Benchmarks (Industry averages by Yelp price tier)
# Sources: AmSpa State of the Industry, IBISWorld, PE deal comps
# ---------------------------------------------------------------------------

MEDSPA_BENCHMARKS = {
    "$":    {"revenue": 400_000, "ebitda_margin": 0.12, "entry_multiple": 3.0},
    "$$":   {"revenue": 700_000, "ebitda_margin": 0.18, "entry_multiple": 3.5},
    "$$$":  {"revenue": 1_200_000, "ebitda_margin": 0.22, "entry_multiple": 4.0},
    "$$$$": {"revenue": 2_000_000, "ebitda_margin": 0.25, "entry_multiple": 4.5},
    None:   {"revenue": 600_000, "ebitda_margin": 0.15, "entry_multiple": 3.5},
}

DEAL_ASSUMPTIONS = {
    "debt_pct": 0.60,
    "equity_pct": 0.40,
    "transaction_cost_pct": 0.05,
    "working_capital_months": 3,
    "sga_pct": 0.32,
    "cogs_pct": 0.40,
    "scenarios": {
        "conservative": {"entry_multiple": 3.0, "exit_multiple": 7, "margin_improvement": 0.03, "hold_years": 5},
        "base":         {"entry_multiple": 4.0, "exit_multiple": 10, "margin_improvement": 0.05, "hold_years": 5},
        "aggressive":   {"entry_multiple": 3.5, "exit_multiple": 12, "margin_improvement": 0.07, "hold_years": 4},
    },
}

# Extra CSS for med-spa report specifics
MEDSPA_EXTRA_CSS = """
/* Med-Spa Report Extras */
.thesis-box {
    background: linear-gradient(135deg, #ebf8ff 0%, #f0fff4 100%);
    border: 1px solid #bee3f8;
    border-radius: 10px;
    padding: 24px;
    margin: 16px 0;
}
[data-theme="dark"] .thesis-box {
    background: linear-gradient(135deg, #1a365d 0%, #22543d 100%);
    border-color: #2a4365;
}
.thesis-box h3 {
    font-size: 15px;
    font-weight: 700;
    color: var(--primary);
    margin-bottom: 10px;
}
.thesis-box ul {
    padding-left: 20px;
    margin: 0;
}
.thesis-box li {
    font-size: 13px;
    color: var(--gray-700);
    margin-bottom: 6px;
    line-height: 1.5;
}
.thesis-box li strong { color: var(--gray-900); }

.score-bar {
    display: inline-block;
    height: 8px;
    border-radius: 4px;
    background: var(--primary-light);
    vertical-align: middle;
    margin-left: 6px;
}

.metric-grid {
    display: grid;
    grid-template-columns: repeat(auto-fill, minmax(260px, 1fr));
    gap: 16px;
    margin: 16px 0;
}
.metric-card {
    background: var(--white);
    border-radius: 10px;
    padding: 20px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.05);
}
.metric-card .metric-label {
    font-size: 12px;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    color: var(--gray-500);
    font-weight: 600;
}
.metric-card .metric-value {
    font-size: 24px;
    font-weight: 700;
    color: var(--primary);
    margin: 4px 0;
}
.metric-card .metric-detail {
    font-size: 12px;
    color: var(--gray-500);
}

.weight-row {
    display: flex;
    align-items: center;
    gap: 12px;
    padding: 8px 0;
    border-bottom: 1px solid var(--gray-100);
    font-size: 13px;
}
.weight-row:last-child { border-bottom: none; }
.weight-label { flex: 1; color: var(--gray-700); font-weight: 500; }
.weight-bar-track {
    flex: 2;
    height: 8px;
    background: var(--gray-100);
    border-radius: 4px;
    overflow: hidden;
}
.weight-bar-fill {
    height: 100%;
    border-radius: 4px;
    background: var(--primary-light);
}
.weight-pct {
    min-width: 40px;
    text-align: right;
    font-weight: 600;
    color: var(--gray-900);
}

/* Deal Model Sections */
.deal-scenario-grid {
    display: grid;
    grid-template-columns: repeat(auto-fill, minmax(280px, 1fr));
    gap: 16px;
    margin: 16px 0;
}
.scenario-card {
    background: var(--white);
    border-radius: 10px;
    padding: 20px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.05);
    border-top: 4px solid var(--primary);
}
.scenario-card.conservative { border-top-color: #48bb78; }
.scenario-card.base { border-top-color: #4299e1; }
.scenario-card.aggressive { border-top-color: #ed8936; }
.scenario-card .scenario-label {
    font-size: 11px;
    text-transform: uppercase;
    letter-spacing: 0.8px;
    font-weight: 700;
    margin-bottom: 12px;
}
.scenario-card.conservative .scenario-label { color: #48bb78; }
.scenario-card.base .scenario-label { color: #4299e1; }
.scenario-card.aggressive .scenario-label { color: #ed8936; }
.scenario-card .scenario-metric {
    display: flex;
    justify-content: space-between;
    padding: 6px 0;
    border-bottom: 1px solid var(--gray-100);
    font-size: 13px;
}
.scenario-card .scenario-metric:last-child { border-bottom: none; }
.scenario-card .scenario-metric .label { color: var(--gray-500); }
.scenario-card .scenario-metric .value { font-weight: 600; color: var(--gray-900); }

.capital-stack {
    display: flex;
    height: 40px;
    border-radius: 8px;
    overflow: hidden;
    margin: 16px 0;
}
.capital-stack .stack-segment {
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 11px;
    font-weight: 600;
    color: #fff;
    transition: width 0.3s ease;
}

.pnl-waterfall {
    display: flex;
    align-items: flex-end;
    gap: 8px;
    height: 180px;
    padding: 16px 0;
}
.pnl-waterfall .waterfall-bar {
    flex: 1;
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: flex-end;
}
.pnl-waterfall .waterfall-bar .bar {
    width: 100%;
    max-width: 80px;
    border-radius: 4px 4px 0 0;
    transition: height 0.3s ease;
}
.pnl-waterfall .waterfall-bar .bar-label {
    font-size: 11px;
    color: var(--gray-500);
    margin-top: 6px;
    text-align: center;
}
.pnl-waterfall .waterfall-bar .bar-value {
    font-size: 12px;
    font-weight: 600;
    color: var(--gray-900);
    margin-bottom: 4px;
}

/* Stealth Wealth & Migration Alpha Sections */
.wealth-composition-grid {
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 24px;
    margin: 16px 0;
    align-items: start;
}
@media (max-width: 768px) {
    .wealth-composition-grid { grid-template-columns: 1fr; }
}
.highlight-table tr.emerging-highlight {
    background: rgba(72, 187, 120, 0.08);
}
.highlight-table tr.emerging-highlight td:first-child {
    border-left: 3px solid #48bb78;
}
[data-theme="dark"] .highlight-table tr.emerging-highlight {
    background: rgba(72, 187, 120, 0.12);
}
.migration-bar-pos { background: #48bb78; }
.migration-bar-neg { background: #fc8181; }

/* Fix: kpi-strip inside sections should not use page-header's negative margin */
.section-body .kpi-strip {
    margin: 0 0 20px 0;
}

/* Fix: chart-container.large was missing a height rule */
.chart-container.large { height: 400px; }

/* Fix: prevent chart/legend overflow into adjacent content */
.chart-container { overflow: hidden; }
.chart-row { overflow: hidden; align-items: start; }

/* Sections 18-21: Provider / RE / Deposit / Business Formation */
.provider-split-grid {
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 24px;
    margin: 16px 0;
    align-items: start;
}
@media (max-width: 768px) {
    .provider-split-grid { grid-template-columns: 1fr; }
}
.highlight-table tr.opportunity-highlight {
    background: rgba(66, 153, 225, 0.08);
}
.highlight-table tr.opportunity-highlight td:first-child {
    border-left: 3px solid #4299e1;
}
[data-theme="dark"] .highlight-table tr.opportunity-highlight {
    background: rgba(66, 153, 225, 0.12);
}
"""


def _fmt(n, decimals=0) -> str:
    """Format a number with commas."""
    if n is None:
        return "-"
    if decimals == 0:
        return f"{int(n):,}"
    return f"{n:,.{decimals}f}"


def _fmt_currency(n) -> str:
    """Format as dollar amount."""
    if n is None:
        return "-"
    if n >= 1_000_000:
        return f"${n / 1_000_000:,.1f}M"
    if n >= 1_000:
        return f"${n / 1_000:,.0f}K"
    return f"${n:,.0f}"


def _grade_badge(grade: str) -> str:
    """Return a pill badge colored by grade."""
    variant = GRADE_BADGE_COLORS.get(grade, "default")
    return pill_badge(f"Grade {grade}", variant)


class MedSpaMarketTemplate:
    """Med-spa market analysis report template."""

    name = "medspa_market"
    description = "Investor-ready med-spa market analysis with acquisition targets and roll-up thesis"

    # ------------------------------------------------------------------
    # Data Gathering
    # ------------------------------------------------------------------

    def gather_data(self, db: Session, params: Dict[str, Any]) -> Dict[str, Any]:
        """Gather all data needed for the report."""
        # Optional filters
        state_filter = params.get("state")  # e.g. "TX" to scope to one state
        min_grade = params.get("min_grade", "F")  # minimum grade to include
        top_n = params.get("top_n", 25)  # number of top targets to display

        data = {
            "generated_at": datetime.utcnow().isoformat(),
            "params": params,
            "summary": self._get_summary_stats(db, state_filter),
            "prospects_by_state": self._get_prospects_by_state(db, state_filter),
            "grade_distribution": self._get_grade_distribution(db, state_filter),
            "score_histogram": self._get_score_histogram(db, state_filter),
            "top_targets": self._get_top_targets(db, top_n, state_filter),
            "zip_concentration": self._get_zip_concentration(db, state_filter),
            "state_avg_scores": self._get_state_avg_scores(db, state_filter),
            "a_grade_by_state": self._get_a_grade_by_state(db, state_filter),
            "zip_affluence_by_state": self._get_zip_affluence_by_state(db, state_filter),
            "census_income": self._get_census_income_correlation(db, state_filter),
            "high_income_zip_penetration": self._get_high_income_zip_penetration(db, state_filter),
            "pe_comps": self._get_pe_aesthetics_comps(db),
            "recent_deals": self._get_recent_deals(db),
            "data_freshness": self._get_data_freshness(db),
            # Section 8: Whitespace Analysis
            **self._get_whitespace_data(db),
            # Section 9: Workforce Economics
            **self._get_bls_wage_data(db),
            # Section 10: Service Categories
            "category_breakdown": self._get_category_breakdown(db, state_filter),
            # Section 11: PE Benchmarking
            **self._get_pe_financial_benchmarks(db),
            # Section 12: Growth Signals
            **self._get_growth_signals(db, state_filter),
            # Sections 13-15: Deal Model
            "deal_model": self._get_deal_model_data(db, state_filter),
            # Section 16: Stealth Wealth Signal
            **self._get_stealth_wealth_data(db, state_filter),
            # Section 17: Migration Alpha
            **self._get_migration_alpha_data(db, state_filter),
            # Section 18: Medical Provider Density Signal
            **self._get_provider_density_data(db, state_filter),
            # Section 19: Real Estate Appreciation Alpha
            **self._get_real_estate_alpha_data(db, state_filter),
            # Section 20: Deposit Wealth Concentration
            **self._get_deposit_wealth_data(db, state_filter),
            # Section 21: Business Formation Velocity
            **self._get_business_formation_data(db, state_filter),
            # Section 22: Opportunity Zone Overlay
            **self._get_opportunity_zone_data(db, state_filter),
            # Section 23: Demographic Demand Model
            **self._get_demographic_demand_data(db, state_filter),
            # Section 24: PE Competitive Heat Map
            **self._get_pe_competitive_data(db, state_filter),
            # Section 25: Construction Momentum Signal
            **self._get_construction_momentum_data(db, state_filter),
            # Section 26: Medical CPI Pricing Power (national)
            **self._get_medical_cpi_data(db),
            # Section 27: Talent Pipeline Pressure (national)
            **self._get_talent_pipeline_data(db),
        }

        return data

    def _get_summary_stats(self, db: Session, state: Optional[str]) -> Dict:
        """Get total addressable market overview stats."""
        state_clause = "AND state = :state" if state else ""
        params: Dict[str, Any] = {}
        if state:
            params["state"] = state.upper()

        result = db.execute(
            text(f"""
                SELECT
                    COUNT(*) as total_prospects,
                    COUNT(*) FILTER (WHERE acquisition_grade = 'A') as a_grade,
                    COUNT(*) FILTER (WHERE acquisition_grade = 'B') as b_grade,
                    COUNT(*) FILTER (WHERE acquisition_grade IN ('A', 'B')) as ab_grade,
                    COUNT(DISTINCT state) as states_covered,
                    COUNT(DISTINCT zip_code) as zips_covered,
                    ROUND(AVG(acquisition_score), 1) as avg_score,
                    ROUND(MAX(acquisition_score), 1) as max_score,
                    ROUND(AVG(rating), 2) as avg_rating,
                    ROUND(AVG(review_count), 0) as avg_reviews,
                    ROUND(AVG(zip_overall_score), 1) as avg_zip_score,
                    SUM(review_count) as total_reviews
                FROM medspa_prospects
                WHERE 1=1 {state_clause}
            """),
            params,
        )
        row = result.fetchone()
        if not row or row[0] == 0:
            return {
                "total_prospects": 0, "a_grade": 0, "b_grade": 0,
                "ab_grade": 0, "states_covered": 0, "zips_covered": 0,
                "avg_score": 0, "max_score": 0, "avg_rating": 0,
                "avg_reviews": 0, "avg_zip_score": 0, "total_reviews": 0,
            }

        return {
            "total_prospects": row[0],
            "a_grade": row[1],
            "b_grade": row[2],
            "ab_grade": row[3],
            "states_covered": row[4],
            "zips_covered": row[5],
            "avg_score": float(row[6]) if row[6] else 0,
            "max_score": float(row[7]) if row[7] else 0,
            "avg_rating": float(row[8]) if row[8] else 0,
            "avg_reviews": int(row[9]) if row[9] else 0,
            "avg_zip_score": float(row[10]) if row[10] else 0,
            "total_reviews": int(row[11]) if row[11] else 0,
        }

    def _get_prospects_by_state(self, db: Session, state: Optional[str]) -> List[Dict]:
        """Get prospect count by state (top 20)."""
        state_clause = "AND state = :state" if state else ""
        params: Dict[str, Any] = {}
        if state:
            params["state"] = state.upper()

        result = db.execute(
            text(f"""
                SELECT state, COUNT(*) as cnt
                FROM medspa_prospects
                WHERE state IS NOT NULL {state_clause}
                GROUP BY state
                ORDER BY cnt DESC
                LIMIT 20
            """),
            params,
        )
        return [{"state": row[0], "count": row[1]} for row in result.fetchall()]

    def _get_grade_distribution(self, db: Session, state: Optional[str]) -> List[Dict]:
        """Get grade distribution."""
        state_clause = "AND state = :state" if state else ""
        params: Dict[str, Any] = {}
        if state:
            params["state"] = state.upper()

        result = db.execute(
            text(f"""
                SELECT acquisition_grade, COUNT(*) as cnt
                FROM medspa_prospects
                WHERE 1=1 {state_clause}
                GROUP BY acquisition_grade
                ORDER BY
                    CASE acquisition_grade
                        WHEN 'A' THEN 1
                        WHEN 'B' THEN 2
                        WHEN 'C' THEN 3
                        WHEN 'D' THEN 4
                        WHEN 'F' THEN 5
                    END
            """),
            params,
        )
        rows = result.fetchall()
        total = sum(r[1] for r in rows)
        return [
            {
                "grade": row[0],
                "count": row[1],
                "pct": round(row[1] / total * 100, 1) if total > 0 else 0,
            }
            for row in rows
        ]

    def _get_score_histogram(self, db: Session, state: Optional[str]) -> List[Dict]:
        """Get score distribution in 10-point buckets."""
        state_clause = "AND state = :state" if state else ""
        params: Dict[str, Any] = {}
        if state:
            params["state"] = state.upper()

        result = db.execute(
            text(f"""
                SELECT
                    FLOOR(acquisition_score / 10) * 10 as bucket,
                    COUNT(*) as cnt
                FROM medspa_prospects
                WHERE 1=1 {state_clause}
                GROUP BY bucket
                ORDER BY bucket
            """),
            params,
        )
        return [
            {"bucket": f"{int(row[0])}-{int(row[0]) + 9}", "count": row[1]}
            for row in result.fetchall()
        ]

    def _get_top_targets(
        self, db: Session, top_n: int, state: Optional[str]
    ) -> List[Dict]:
        """Get top acquisition targets."""
        state_clause = "AND state = :state" if state else ""
        params: Dict[str, Any] = {"limit": top_n}
        if state:
            params["state"] = state.upper()

        result = db.execute(
            text(f"""
                SELECT
                    name, city, state, zip_code,
                    acquisition_score, acquisition_grade,
                    rating, review_count, price,
                    zip_overall_score, zip_grade,
                    zip_affluence_sub, yelp_rating_sub,
                    review_volume_sub, low_competition_sub,
                    price_tier_sub, competitor_count_in_zip,
                    phone, url
                FROM medspa_prospects
                WHERE acquisition_grade IN ('A', 'B') {state_clause}
                ORDER BY acquisition_score DESC
                LIMIT :limit
            """),
            params,
        )
        return [
            {
                "name": row[0],
                "city": row[1],
                "state": row[2],
                "zip_code": row[3],
                "score": float(row[4]) if row[4] else 0,
                "grade": row[5],
                "rating": float(row[6]) if row[6] else 0,
                "reviews": int(row[7]) if row[7] else 0,
                "price": row[8],
                "zip_score": float(row[9]) if row[9] else 0,
                "zip_grade": row[10],
                "zip_affluence_sub": float(row[11]) if row[11] else 0,
                "yelp_rating_sub": float(row[12]) if row[12] else 0,
                "review_volume_sub": float(row[13]) if row[13] else 0,
                "low_competition_sub": float(row[14]) if row[14] else 0,
                "price_tier_sub": float(row[15]) if row[15] else 0,
                "competitors": int(row[16]) if row[16] else 0,
                "phone": row[17],
                "url": row[18],
            }
            for row in result.fetchall()
        ]

    def _get_zip_concentration(self, db: Session, state: Optional[str]) -> List[Dict]:
        """Get top ZIPs by prospect count (competition density)."""
        state_clause = "AND state = :state" if state else ""
        params: Dict[str, Any] = {}
        if state:
            params["state"] = state.upper()

        result = db.execute(
            text(f"""
                SELECT
                    zip_code, state, city,
                    COUNT(*) as prospect_count,
                    ROUND(AVG(acquisition_score), 1) as avg_score,
                    ROUND(AVG(zip_overall_score), 1) as avg_zip_score
                FROM medspa_prospects
                WHERE zip_code IS NOT NULL {state_clause}
                GROUP BY zip_code, state, city
                ORDER BY prospect_count DESC
                LIMIT 20
            """),
            params,
        )
        return [
            {
                "zip_code": row[0],
                "state": row[1],
                "city": row[2],
                "prospect_count": row[3],
                "avg_score": float(row[4]) if row[4] else 0,
                "avg_zip_score": float(row[5]) if row[5] else 0,
            }
            for row in result.fetchall()
        ]

    def _get_state_avg_scores(self, db: Session, state: Optional[str]) -> List[Dict]:
        """Get states ranked by avg acquisition score."""
        state_clause = "AND state = :state" if state else ""
        params: Dict[str, Any] = {}
        if state:
            params["state"] = state.upper()

        result = db.execute(
            text(f"""
                SELECT
                    state,
                    COUNT(*) as prospect_count,
                    ROUND(AVG(acquisition_score), 1) as avg_score,
                    ROUND(MAX(acquisition_score), 1) as max_score,
                    COUNT(*) FILTER (WHERE acquisition_grade = 'A') as a_count
                FROM medspa_prospects
                WHERE state IS NOT NULL {state_clause}
                GROUP BY state
                ORDER BY avg_score DESC
                LIMIT 15
            """),
            params,
        )
        return [
            {
                "state": row[0],
                "prospect_count": row[1],
                "avg_score": float(row[2]) if row[2] else 0,
                "max_score": float(row[3]) if row[3] else 0,
                "a_count": row[4],
            }
            for row in result.fetchall()
        ]

    def _get_a_grade_by_state(self, db: Session, state: Optional[str]) -> List[Dict]:
        """Get A-grade concentration by state."""
        state_clause = "AND state = :state" if state else ""
        params: Dict[str, Any] = {}
        if state:
            params["state"] = state.upper()

        result = db.execute(
            text(f"""
                SELECT
                    state,
                    COUNT(*) FILTER (WHERE acquisition_grade = 'A') as a_count,
                    COUNT(*) as total,
                    ROUND(
                        COUNT(*) FILTER (WHERE acquisition_grade = 'A')::numeric
                        / NULLIF(COUNT(*), 0) * 100, 1
                    ) as a_pct
                FROM medspa_prospects
                WHERE state IS NOT NULL {state_clause}
                GROUP BY state
                HAVING COUNT(*) FILTER (WHERE acquisition_grade = 'A') > 0
                ORDER BY a_count DESC
                LIMIT 15
            """),
            params,
        )
        return [
            {
                "state": row[0],
                "a_count": row[1],
                "total": row[2],
                "a_pct": float(row[3]) if row[3] else 0,
            }
            for row in result.fetchall()
        ]

    def _get_zip_affluence_by_state(self, db: Session, state: Optional[str]) -> List[Dict]:
        """Get average ZIP affluence score by state from zip_medspa_scores."""
        state_clause = "AND state_abbr = :state" if state else ""
        params: Dict[str, Any] = {}
        if state:
            params["state"] = state.upper()

        try:
            result = db.execute(
                text(f"""
                    SELECT
                        state_abbr,
                        COUNT(*) as zip_count,
                        ROUND(AVG(overall_score), 1) as avg_score,
                        ROUND(AVG(affluence_density_score), 1) as avg_affluence,
                        ROUND(AVG(avg_agi), 0) as avg_agi,
                        COUNT(*) FILTER (WHERE grade = 'A') as a_zips
                    FROM zip_medspa_scores
                    WHERE state_abbr IS NOT NULL {state_clause}
                    GROUP BY state_abbr
                    ORDER BY avg_score DESC
                    LIMIT 15
                """),
                params,
            )
            return [
                {
                    "state": row[0],
                    "zip_count": row[1],
                    "avg_score": float(row[2]) if row[2] else 0,
                    "avg_affluence": float(row[3]) if row[3] else 0,
                    "avg_agi": float(row[4]) if row[4] else 0,
                    "a_zips": row[5],
                }
                for row in result.fetchall()
            ]
        except Exception as e:
            logger.warning(f"Could not fetch ZIP affluence data: {e}")
            db.rollback()
            return []

    def _get_census_income_correlation(self, db: Session, state: Optional[str]) -> List[Dict]:
        """Get median household income from Census ACS by state, correlated with ZIP scores."""
        try:
            # Try to pull ACS median income data
            result = db.execute(
                text("""
                    SELECT
                        z.state_abbr,
                        ROUND(AVG(z.avg_agi), 0) as avg_zip_agi,
                        ROUND(AVG(z.overall_score), 1) as avg_zip_score,
                        COUNT(DISTINCT z.zip_code) as zip_count
                    FROM zip_medspa_scores z
                    WHERE z.state_abbr IS NOT NULL
                    GROUP BY z.state_abbr
                    HAVING COUNT(*) >= 10
                    ORDER BY avg_zip_score DESC
                    LIMIT 15
                """),
            )
            return [
                {
                    "state": row[0],
                    "avg_agi": float(row[1]) if row[1] else 0,
                    "avg_zip_score": float(row[2]) if row[2] else 0,
                    "zip_count": row[3],
                }
                for row in result.fetchall()
            ]
        except Exception as e:
            logger.warning(f"Could not fetch census income data: {e}")
            db.rollback()
            return []

    def _get_high_income_zip_penetration(self, db: Session, state: Optional[str]) -> Dict:
        """Get penetration of prospects into high-income ZIPs."""
        try:
            result = db.execute(
                text("""
                    SELECT
                        COUNT(DISTINCT z.zip_code) as total_a_zips,
                        COUNT(DISTINCT mp.zip_code) as zips_with_prospects,
                        ROUND(
                            COUNT(DISTINCT mp.zip_code)::numeric
                            / NULLIF(COUNT(DISTINCT z.zip_code), 0) * 100, 1
                        ) as penetration_pct
                    FROM zip_medspa_scores z
                    LEFT JOIN medspa_prospects mp ON z.zip_code = mp.zip_code
                    WHERE z.grade = 'A'
                """),
            )
            row = result.fetchone()
            if row:
                return {
                    "total_a_zips": int(row[0]) if row[0] else 0,
                    "zips_with_prospects": int(row[1]) if row[1] else 0,
                    "penetration_pct": float(row[2]) if row[2] else 0,
                }
        except Exception as e:
            logger.warning(f"Could not fetch high-income penetration: {e}")
            db.rollback()
        return {"total_a_zips": 0, "zips_with_prospects": 0, "penetration_pct": 0}

    def _get_pe_aesthetics_comps(self, db: Session) -> List[Dict]:
        """Get PE-backed aesthetics/med-spa companies from pe_portfolio_companies."""
        try:
            result = db.execute(
                text("""
                    SELECT
                        pc.name, pc.industry, pc.sub_industry,
                        pc.headquarters_city, pc.headquarters_state,
                        pc.current_pe_owner, pc.ownership_status,
                        pc.employee_count, pc.founded_year,
                        pc.website, pc.status
                    FROM pe_portfolio_companies pc
                    WHERE (
                        LOWER(pc.industry) LIKE '%aesthet%'
                        OR LOWER(pc.industry) LIKE '%med%spa%'
                        OR LOWER(pc.industry) LIKE '%dermatol%'
                        OR LOWER(pc.industry) LIKE '%cosmetic%'
                        OR LOWER(pc.sub_industry) LIKE '%aesthet%'
                        OR LOWER(pc.sub_industry) LIKE '%med%spa%'
                        OR LOWER(pc.name) LIKE '%medspa%'
                        OR LOWER(pc.name) LIKE '%med spa%'
                        OR LOWER(pc.name) LIKE '%aesthet%'
                        OR LOWER(pc.sector) LIKE '%aesthet%'
                        OR LOWER(pc.sector) LIKE '%healthcare%services%'
                    )
                    ORDER BY pc.name
                """),
            )
            return [
                {
                    "name": row[0],
                    "industry": row[1],
                    "sub_industry": row[2],
                    "city": row[3],
                    "state": row[4],
                    "pe_owner": row[5],
                    "ownership_status": row[6],
                    "employees": row[7],
                    "founded": row[8],
                    "website": row[9],
                    "status": row[10],
                }
                for row in result.fetchall()
            ]
        except Exception as e:
            logger.warning(f"Could not fetch PE aesthetics comps: {e}")
            db.rollback()
            return []

    def _get_recent_deals(self, db: Session) -> List[Dict]:
        """Get recent aesthetics/med-spa deals."""
        try:
            result = db.execute(
                text("""
                    SELECT
                        d.deal_name, d.deal_type, d.deal_sub_type,
                        d.announced_date, d.closed_date,
                        d.enterprise_value_usd, d.ev_ebitda_multiple,
                        d.ev_revenue_multiple,
                        pc.name as company_name, pc.industry
                    FROM pe_deals d
                    JOIN pe_portfolio_companies pc ON d.company_id = pc.id
                    WHERE (
                        LOWER(pc.industry) LIKE '%aesthet%'
                        OR LOWER(pc.industry) LIKE '%med%spa%'
                        OR LOWER(pc.industry) LIKE '%dermatol%'
                        OR LOWER(pc.industry) LIKE '%cosmetic%'
                        OR LOWER(pc.sub_industry) LIKE '%aesthet%'
                        OR LOWER(pc.name) LIKE '%medspa%'
                        OR LOWER(pc.name) LIKE '%med spa%'
                    )
                    ORDER BY d.announced_date DESC NULLS LAST
                    LIMIT 10
                """),
            )
            return [
                {
                    "deal_name": row[0],
                    "deal_type": row[1],
                    "deal_sub_type": row[2],
                    "announced_date": row[3].isoformat() if row[3] else None,
                    "closed_date": row[4].isoformat() if row[4] else None,
                    "ev": float(row[5]) if row[5] else None,
                    "ev_ebitda": float(row[6]) if row[6] else None,
                    "ev_revenue": float(row[7]) if row[7] else None,
                    "company": row[8],
                    "industry": row[9],
                }
                for row in result.fetchall()
            ]
        except Exception as e:
            logger.warning(f"Could not fetch recent deals: {e}")
            db.rollback()
            return []

    def _get_data_freshness(self, db: Session) -> Dict:
        """Get data freshness timestamps."""
        freshness = {}

        # Prospects freshness
        try:
            result = db.execute(
                text("""
                    SELECT
                        MIN(discovered_at) as earliest,
                        MAX(discovered_at) as latest,
                        COUNT(*) as total
                    FROM medspa_prospects
                """),
            )
            row = result.fetchone()
            if row and row[2]:
                freshness["prospects"] = {
                    "earliest": row[0].isoformat() if row[0] else None,
                    "latest": row[1].isoformat() if row[1] else None,
                    "total": row[2],
                }
        except Exception:
            db.rollback()

        # ZIP scores freshness
        try:
            result = db.execute(
                text("""
                    SELECT
                        MIN(score_date) as earliest,
                        MAX(score_date) as latest,
                        COUNT(*) as total
                    FROM zip_medspa_scores
                """),
            )
            row = result.fetchone()
            if row and row[2]:
                freshness["zip_scores"] = {
                    "earliest": row[0].isoformat() if row[0] else None,
                    "latest": row[1].isoformat() if row[1] else None,
                    "total": row[2],
                }
        except Exception:
            db.rollback()

        # IRS SOI ZIP Income freshness
        try:
            result = db.execute(
                text("""
                    SELECT
                        MIN(tax_year) as earliest_year,
                        MAX(tax_year) as latest_year,
                        COUNT(*) as total
                    FROM irs_soi_zip_income
                """),
            )
            row = result.fetchone()
            if row and row[2]:
                freshness["irs_soi_zip_income"] = {
                    "earliest": str(row[0]) if row[0] else None,
                    "latest": str(row[1]) if row[1] else None,
                    "total": row[2],
                }
        except Exception:
            db.rollback()

        # IRS SOI Migration freshness
        try:
            result = db.execute(
                text("""
                    SELECT
                        MIN(tax_year) as earliest_year,
                        MAX(tax_year) as latest_year,
                        COUNT(*) as total
                    FROM irs_soi_migration
                """),
            )
            row = result.fetchone()
            if row and row[2]:
                freshness["irs_soi_migration"] = {
                    "earliest": str(row[0]) if row[0] else None,
                    "latest": str(row[1]) if row[1] else None,
                    "total": row[2],
                }
        except Exception:
            db.rollback()

        # CMS Medicare Utilization freshness
        try:
            result = db.execute(
                text("SELECT COUNT(*) as total FROM cms_medicare_utilization"),
            )
            row = result.fetchone()
            if row and row[0]:
                freshness["cms_medicare"] = {
                    "earliest": None, "latest": None, "total": row[0],
                }
        except Exception:
            db.rollback()

        # Redfin freshness
        try:
            result = db.execute(
                text("""
                    SELECT MIN(period_end) as earliest, MAX(period_end) as latest,
                           COUNT(*) as total
                    FROM realestate_redfin
                """),
            )
            row = result.fetchone()
            if row and row[2]:
                freshness["redfin"] = {
                    "earliest": row[0].isoformat() if row[0] else None,
                    "latest": row[1].isoformat() if row[1] else None,
                    "total": row[2],
                }
        except Exception:
            db.rollback()

        # FHFA HPI freshness
        try:
            result = db.execute(
                text("""
                    SELECT MIN(year) as earliest, MAX(year) as latest,
                           COUNT(*) as total
                    FROM realestate_fhfa_hpi
                """),
            )
            row = result.fetchone()
            if row and row[2]:
                freshness["fhfa_hpi"] = {
                    "earliest": str(row[0]) if row[0] else None,
                    "latest": str(row[1]) if row[1] else None,
                    "total": row[2],
                }
        except Exception:
            db.rollback()

        # FDIC Summary Deposits freshness
        try:
            result = db.execute(
                text("""
                    SELECT MIN(year) as earliest, MAX(year) as latest,
                           COUNT(*) as total
                    FROM fdic_summary_deposits
                """),
            )
            row = result.fetchone()
            if row and row[2]:
                freshness["fdic_deposits"] = {
                    "earliest": str(row[0]) if row[0] else None,
                    "latest": str(row[1]) if row[1] else None,
                    "total": row[2],
                }
        except Exception:
            db.rollback()

        # IRS SOI Business Income freshness
        try:
            result = db.execute(
                text("""
                    SELECT MIN(tax_year) as earliest, MAX(tax_year) as latest,
                           COUNT(*) as total
                    FROM irs_soi_business_income
                """),
            )
            row = result.fetchone()
            if row and row[2]:
                freshness["irs_soi_business_income"] = {
                    "earliest": str(row[0]) if row[0] else None,
                    "latest": str(row[1]) if row[1] else None,
                    "total": row[2],
                }
        except Exception:
            db.rollback()

        # Opportunity Zone freshness
        try:
            result = db.execute(
                text("SELECT COUNT(*) as total FROM opportunity_zone"),
            )
            row = result.fetchone()
            if row and row[0]:
                freshness["opportunity_zone"] = {
                    "earliest": None, "latest": None, "total": row[0],
                }
        except Exception:
            db.rollback()

        # Educational Attainment freshness
        try:
            result = db.execute(
                text("""
                    SELECT MIN(period_year) as earliest, MAX(period_year) as latest,
                           COUNT(*) as total
                    FROM educational_attainment
                """),
            )
            row = result.fetchone()
            if row and row[2]:
                freshness["educational_attainment"] = {
                    "earliest": str(row[0]) if row[0] else None,
                    "latest": str(row[1]) if row[1] else None,
                    "total": row[2],
                }
        except Exception:
            db.rollback()

        # HUD Building Permits freshness
        try:
            result = db.execute(
                text("""
                    SELECT MIN(date) as earliest, MAX(date) as latest,
                           COUNT(*) as total
                    FROM realestate_hud_permits
                """),
            )
            row = result.fetchone()
            if row and row[2]:
                freshness["hud_permits"] = {
                    "earliest": row[0].isoformat() if row[0] else None,
                    "latest": row[1].isoformat() if row[1] else None,
                    "total": row[2],
                }
        except Exception:
            db.rollback()

        # BLS CPI freshness
        try:
            result = db.execute(
                text("""
                    SELECT MIN(year) as earliest, MAX(year) as latest,
                           COUNT(*) as total
                    FROM bls_cpi
                """),
            )
            row = result.fetchone()
            if row and row[2]:
                freshness["bls_cpi"] = {
                    "earliest": str(row[0]) if row[0] else None,
                    "latest": str(row[1]) if row[1] else None,
                    "total": row[2],
                }
        except Exception:
            db.rollback()

        return freshness

    # ------------------------------------------------------------------
    # Section 8: Whitespace Analysis
    # ------------------------------------------------------------------

    def _get_whitespace_data(self, db: Session) -> Dict:
        """Get A-grade ZIPs with no discovered prospects (greenfield opportunities)."""
        try:
            # Whitespace ZIPs: A-grade with no prospects
            result = db.execute(
                text("""
                    SELECT z.zip_code, z.state_abbr, z.overall_score, z.avg_agi,
                           z.total_returns, z.affluence_density_score
                    FROM zip_medspa_scores z
                    WHERE z.grade = 'A'
                      AND z.zip_code NOT IN (
                          SELECT DISTINCT zip_code FROM medspa_prospects
                          WHERE zip_code IS NOT NULL
                      )
                    ORDER BY z.overall_score DESC
                """),
            )
            whitespace_zips = [
                {
                    "zip_code": row[0],
                    "state": row[1],
                    "score": float(row[2]) if row[2] else 0,
                    "avg_agi": float(row[3]) if row[3] else 0,
                    "total_returns": int(row[4]) if row[4] else 0,
                    "affluence_score": float(row[5]) if row[5] else 0,
                }
                for row in result.fetchall()
            ]

            # By-state rollup
            state_counts: Dict[str, int] = {}
            for z in whitespace_zips:
                st = z["state"] or "??"
                state_counts[st] = state_counts.get(st, 0) + 1
            whitespace_by_state = sorted(
                [{"state": k, "count": v} for k, v in state_counts.items()],
                key=lambda x: x["count"],
                reverse=True,
            )

            # Summary stats
            total_a_result = db.execute(
                text("SELECT COUNT(*) FROM zip_medspa_scores WHERE grade = 'A'"),
            )
            total_a = total_a_result.scalar() or 0

            return {
                "whitespace_zips": whitespace_zips,
                "whitespace_by_state": whitespace_by_state,
                "whitespace_summary": {
                    "total_a_zips": total_a,
                    "a_with_prospects": total_a - len(whitespace_zips),
                    "whitespace_count": len(whitespace_zips),
                },
            }
        except Exception as e:
            logger.warning(f"Could not fetch whitespace data: {e}")
            db.rollback()
            return {
                "whitespace_zips": [],
                "whitespace_by_state": [],
                "whitespace_summary": {"total_a_zips": 0, "a_with_prospects": 0, "whitespace_count": 0},
            }

    # ------------------------------------------------------------------
    # Section 9: Workforce Economics (BLS OES)
    # ------------------------------------------------------------------

    def _get_bls_wage_data(self, db: Session) -> Dict:
        """Get BLS OES wage trends for aesthetics-adjacent occupations."""
        # Target series IDs: annual mean wages (data type 04) for key occupations
        WAGE_SERIES = {
            "OEUM000000000000029122904": "Dermatologists",
            "OEUM000000000000029114104": "Registered Nurses",
            "OEUM000000000000029117104": "Nurse Practitioners",
            "OEUM000000000000031901104": "Massage Therapists",
            "OEUM000000000000039501204": "Cosmetologists",
            "OEUM000000000000031909904": "Healthcare Support",
            "OEUM000000000000029107104": "Physician Assistants",
            "OEUM000000000000029209904": "Health Technicians",
        }
        EMPLOYMENT_SERIES = {
            "OEUM000000000000029122901": "Dermatologists",
            "OEUM000000000000029114101": "Registered Nurses",
            "OEUM000000000000029117101": "Nurse Practitioners",
            "OEUM000000000000031901101": "Massage Therapists",
            "OEUM000000000000039501201": "Cosmetologists",
            "OEUM000000000000031909901": "Healthcare Support",
            "OEUM000000000000029107101": "Physician Assistants",
            "OEUM000000000000029209901": "Health Technicians",
        }
        try:
            # Wages
            wage_ids = list(WAGE_SERIES.keys())
            result = db.execute(
                text("""
                    SELECT series_id, year, value
                    FROM bls_oes
                    WHERE series_id = ANY(:ids)
                      AND period = 'M13'
                    ORDER BY year
                """),
                {"ids": wage_ids},
            )
            bls_wages: Dict[str, Dict[int, float]] = {}
            for row in result.fetchall():
                occ = WAGE_SERIES.get(row[0], row[0])
                yr = int(row[1])
                val = float(row[2]) if row[2] else 0
                bls_wages.setdefault(occ, {})[yr] = val

            # Employment counts
            emp_ids = list(EMPLOYMENT_SERIES.keys())
            result = db.execute(
                text("""
                    SELECT series_id, year, value
                    FROM bls_oes
                    WHERE series_id = ANY(:ids)
                      AND period = 'M13'
                    ORDER BY year
                """),
                {"ids": emp_ids},
            )
            bls_employment: Dict[str, Dict[int, float]] = {}
            for row in result.fetchall():
                occ = EMPLOYMENT_SERIES.get(row[0], row[0])
                yr = int(row[1])
                val = float(row[2]) if row[2] else 0
                bls_employment.setdefault(occ, {})[yr] = val

            return {
                "bls_wages": bls_wages,
                "bls_employment": bls_employment,
            }
        except Exception as e:
            logger.warning(f"Could not fetch BLS wage data: {e}")
            db.rollback()
            return {"bls_wages": {}, "bls_employment": {}}

    # ------------------------------------------------------------------
    # Section 10: Service Category Breakdown
    # ------------------------------------------------------------------

    def _get_category_breakdown(self, db: Session, state: Optional[str]) -> List[Dict]:
        """Get prospect counts by Yelp category."""
        state_clause = "WHERE state = :state" if state else ""
        params: Dict[str, Any] = {}
        if state:
            params["state"] = state.upper()

        try:
            result = db.execute(
                text(f"""
                    SELECT unnest(categories) AS category, COUNT(*) AS cnt,
                           ROUND(AVG(acquisition_score), 1) AS avg_score,
                           ROUND(AVG(rating), 2) AS avg_rating,
                           ROUND(AVG(review_count), 0) AS avg_reviews
                    FROM medspa_prospects
                    {state_clause}
                    GROUP BY category
                    ORDER BY cnt DESC
                """),
                params,
            )
            return [
                {
                    "category": row[0],
                    "count": row[1],
                    "avg_score": float(row[2]) if row[2] else 0,
                    "avg_rating": float(row[3]) if row[3] else 0,
                    "avg_reviews": int(row[4]) if row[4] else 0,
                }
                for row in result.fetchall()
            ]
        except Exception as e:
            logger.warning(f"Could not fetch category breakdown: {e}")
            db.rollback()
            return []

    # ------------------------------------------------------------------
    # Section 11: PE Platform Benchmarking
    # ------------------------------------------------------------------

    def _get_pe_financial_benchmarks(self, db: Session) -> Dict:
        """Get financial benchmarks for PE-backed aesthetics platforms."""
        try:
            result = db.execute(
                text("""
                    SELECT c.name, c.headquarters_state, c.employee_count,
                           c.current_pe_owner,
                           f.revenue_usd, f.revenue_growth_pct, f.ebitda_margin_pct,
                           f.gross_margin_pct, f.debt_to_ebitda, f.is_estimated,
                           f.confidence
                    FROM pe_portfolio_companies c
                    JOIN pe_company_financials f ON f.company_id = c.id
                    WHERE (c.industry ILIKE '%aesthetics%'
                           OR c.sub_industry ILIKE '%med%spa%'
                           OR c.sub_industry ILIKE '%dermatolog%')
                      AND f.fiscal_year = (
                          SELECT MAX(fiscal_year) FROM pe_company_financials
                          WHERE company_id = c.id
                      )
                    ORDER BY f.revenue_usd DESC NULLS LAST
                """),
            )
            pe_financials = [
                {
                    "name": row[0],
                    "state": row[1],
                    "employees": int(row[2]) if row[2] else None,
                    "pe_owner": row[3],
                    "revenue": float(row[4]) if row[4] else None,
                    "rev_growth": float(row[5]) if row[5] else None,
                    "ebitda_margin": float(row[6]) if row[6] else None,
                    "gross_margin": float(row[7]) if row[7] else None,
                    "debt_ebitda": float(row[8]) if row[8] else None,
                    "is_estimated": bool(row[9]) if row[9] is not None else True,
                    "confidence": row[10] or "low",
                }
                for row in result.fetchall()
            ]

            # Compute summary stats
            ebitda_margins = [f["ebitda_margin"] for f in pe_financials if f["ebitda_margin"] is not None]
            rev_growths = [f["rev_growth"] for f in pe_financials if f["rev_growth"] is not None]
            debt_ratios = [f["debt_ebitda"] for f in pe_financials if f["debt_ebitda"] is not None]

            pe_summary = {
                "avg_ebitda_margin": round(sum(ebitda_margins) / len(ebitda_margins), 1) if ebitda_margins else None,
                "avg_rev_growth": round(sum(rev_growths) / len(rev_growths), 1) if rev_growths else None,
                "median_debt_ebitda": round(
                    sorted(debt_ratios)[len(debt_ratios) // 2], 1
                ) if debt_ratios else None,
            }

            return {
                "pe_financials": pe_financials,
                "pe_financial_summary": pe_summary,
            }
        except Exception as e:
            logger.warning(f"Could not fetch PE financial benchmarks: {e}")
            db.rollback()
            return {
                "pe_financials": [],
                "pe_financial_summary": {"avg_ebitda_margin": None, "avg_rev_growth": None, "median_debt_ebitda": None},
            }

    # ------------------------------------------------------------------
    # Section 12: Review Velocity & Growth Signals
    # ------------------------------------------------------------------

    def _get_growth_signals(self, db: Session, state: Optional[str]) -> Dict:
        """Get review velocity and low-competition opportunities."""
        state_clause = "AND state = :state" if state else ""
        params: Dict[str, Any] = {}
        if state:
            params["state"] = state.upper()

        try:
            # Top by review count
            result = db.execute(
                text(f"""
                    SELECT name, city, state, review_count, rating,
                           acquisition_score, acquisition_grade,
                           competitor_count_in_zip, zip_overall_score
                    FROM medspa_prospects
                    WHERE acquisition_grade IN ('A', 'B') {state_clause}
                    ORDER BY review_count DESC
                    LIMIT 25
                """),
                params,
            )
            top_by_reviews = [
                {
                    "name": row[0], "city": row[1], "state": row[2],
                    "reviews": int(row[3]) if row[3] else 0,
                    "rating": float(row[4]) if row[4] else 0,
                    "score": float(row[5]) if row[5] else 0,
                    "grade": row[6],
                    "competitors": int(row[7]) if row[7] else 0,
                    "zip_score": float(row[8]) if row[8] else 0,
                }
                for row in result.fetchall()
            ]

            # Review volume distribution
            result = db.execute(
                text(f"""
                    SELECT
                        CASE
                            WHEN review_count >= 500 THEN '500+'
                            WHEN review_count >= 200 THEN '200-499'
                            WHEN review_count >= 100 THEN '100-199'
                            WHEN review_count >= 50  THEN '50-99'
                            ELSE '<50'
                        END AS bucket,
                        COUNT(*) AS cnt,
                        ROUND(AVG(acquisition_score), 1) AS avg_score,
                        ROUND(AVG(rating), 2) AS avg_rating
                    FROM medspa_prospects
                    WHERE 1=1 {state_clause}
                    GROUP BY bucket
                    ORDER BY MIN(review_count) DESC
                """),
                params,
            )
            review_buckets = [
                {
                    "bucket": row[0], "count": row[1],
                    "avg_score": float(row[2]) if row[2] else 0,
                    "avg_rating": float(row[3]) if row[3] else 0,
                }
                for row in result.fetchall()
            ]

            # Low competition gems
            result = db.execute(
                text(f"""
                    SELECT name, city, state, acquisition_score, acquisition_grade,
                           rating, review_count, competitor_count_in_zip, zip_overall_score
                    FROM medspa_prospects
                    WHERE acquisition_grade IN ('A', 'B')
                      AND competitor_count_in_zip <= 3
                      {state_clause}
                    ORDER BY acquisition_score DESC
                    LIMIT 15
                """),
                params,
            )
            low_competition_gems = [
                {
                    "name": row[0], "city": row[1], "state": row[2],
                    "score": float(row[3]) if row[3] else 0,
                    "grade": row[4],
                    "rating": float(row[5]) if row[5] else 0,
                    "reviews": int(row[6]) if row[6] else 0,
                    "competitors": int(row[7]) if row[7] else 0,
                    "zip_score": float(row[8]) if row[8] else 0,
                }
                for row in result.fetchall()
            ]

            return {
                "top_by_reviews": top_by_reviews,
                "review_buckets": review_buckets,
                "low_competition_gems": low_competition_gems,
            }
        except Exception as e:
            logger.warning(f"Could not fetch growth signals: {e}")
            db.rollback()
            return {
                "top_by_reviews": [],
                "review_buckets": [],
                "low_competition_gems": [],
            }

    # ------------------------------------------------------------------
    # Sections 13-15: Deal Model
    # ------------------------------------------------------------------

    def _get_deal_model_data(self, db: Session, state: Optional[str]) -> Dict:
        """Build full PE roll-up deal model from A-grade prospects and industry benchmarks.

        Delegates to DealEngine for core calculations while preserving the same
        return shape for backward compatibility with existing report sections.
        """
        try:
            from app.services.deal_engine import DealEngine

            engine = DealEngine(db)
            portfolio = engine.get_target_portfolio(state=state)
            economics = engine.compute_tier_economics(portfolio["tier_counts"])
            capital = engine.compute_capital_stack(
                economics["total_acquisition_cost"],
                economics["total_ebitda"],
                economics["total_revenue"],
            )
            # Use the report's own DEAL_ASSUMPTIONS for backward compatibility
            scenarios = engine.run_scenarios(economics, DEAL_ASSUMPTIONS["scenarios"])

            # P&L waterfall per average location
            da = DEAL_ASSUMPTIONS
            total_locs = economics["total_locations"]
            avg_rev = economics["total_revenue"] / total_locs if total_locs > 0 else 0

            return {
                "tier_economics": economics["tier_economics"],
                "total_locations": economics["total_locations"],
                "total_revenue": economics["total_revenue"],
                "total_ebitda": economics["total_ebitda"],
                "weighted_margin": economics["weighted_margin"],
                "total_acquisition_cost": economics["total_acquisition_cost"],
                "capital_stack": {
                    "debt": capital["debt"],
                    "equity": capital["equity"],
                    "transaction_costs": capital["transaction_costs"],
                    "working_capital": capital["working_capital"],
                    "total_capital_required": capital["total_capital_required"],
                },
                "leverage_ratio": capital["leverage_ratio"],
                "pnl_waterfall": {
                    "revenue": avg_rev,
                    "cogs": avg_rev * da["cogs_pct"],
                    "gross_profit": avg_rev * (1 - da["cogs_pct"]),
                    "sga": avg_rev * da["sga_pct"],
                    "ebitda": avg_rev * (1 - da["cogs_pct"]) - avg_rev * da["sga_pct"],
                },
                "scenarios": scenarios,
                "a_grade_states": portfolio["a_grade_states"],
            }
        except Exception as e:
            logger.warning(f"Could not compute deal model data: {e}")
            db.rollback()
            return {
                "tier_economics": [],
                "total_locations": 0,
                "total_revenue": 0,
                "total_ebitda": 0,
                "weighted_margin": 0,
                "total_acquisition_cost": 0,
                "capital_stack": {
                    "debt": 0, "equity": 0, "transaction_costs": 0,
                    "working_capital": 0, "total_capital_required": 0,
                },
                "leverage_ratio": 0,
                "pnl_waterfall": {
                    "revenue": 0, "cogs": 0, "gross_profit": 0,
                    "sga": 0, "ebitda": 0,
                },
                "scenarios": {},
                "a_grade_states": [],
            }

    # ------------------------------------------------------------------
    # Section 16: Stealth Wealth Signal
    # ------------------------------------------------------------------

    def _get_stealth_wealth_data(self, db: Session, state: Optional[str]) -> Dict:
        """Cross-reference IRS SOI non-wage income with medspa scores to find hidden demand."""
        try:
            # Check if table exists
            check = db.execute(text(
                "SELECT 1 FROM information_schema.tables "
                "WHERE table_name = 'irs_soi_zip_income'"
            ))
            if not check.fetchone():
                return {"stealth_wealth": {}}

            # Get latest tax year
            yr_row = db.execute(text(
                "SELECT MAX(tax_year) FROM irs_soi_zip_income"
            )).fetchone()
            if not yr_row or not yr_row[0]:
                return {"stealth_wealth": {}}
            latest_year = yr_row[0]

            state_clause = "AND zi.state_abbr = :state" if state else ""
            params: Dict[str, Any] = {"year": latest_year}
            if state:
                params["state"] = state.upper()

            # --- Wealth composition summary (national/state) ---
            comp_row = db.execute(text(f"""
                SELECT
                    SUM(total_wages) AS wages,
                    SUM(total_capital_gains) AS cap_gains,
                    SUM(total_dividends) AS dividends,
                    SUM(total_business_income) AS biz_income,
                    SUM(total_agi) AS total_agi,
                    SUM(num_returns) AS total_returns
                FROM irs_soi_zip_income zi
                WHERE tax_year = :year AND agi_class = '0'
                    {state_clause}
            """), params).fetchone()

            wealth_composition = {}
            if comp_row and comp_row[4] and float(comp_row[4]) > 0:
                total_agi = float(comp_row[4])
                wages = float(comp_row[0] or 0)
                cap_gains = float(comp_row[1] or 0)
                dividends = float(comp_row[2] or 0)
                biz_income = float(comp_row[3] or 0)
                other = total_agi - wages - cap_gains - dividends - biz_income
                wealth_composition = {
                    "wages_pct": round(wages / total_agi * 100, 1),
                    "cap_gains_pct": round(cap_gains / total_agi * 100, 1),
                    "dividends_pct": round(dividends / total_agi * 100, 1),
                    "biz_income_pct": round(biz_income / total_agi * 100, 1),
                    "other_pct": round(max(other, 0) / total_agi * 100, 1),
                    "total_returns": int(comp_row[5] or 0),
                    "total_agi_billions": round(total_agi * 1000 / 1e9, 1),
                    "tax_year": latest_year,
                }

            # --- Stealth ZIPs: high non-wage income but low medspa score ---
            stealth_rows = db.execute(text(f"""
                WITH zip_wealth AS (
                    SELECT
                        zi.zip_code,
                        zi.state_abbr,
                        zi.num_returns,
                        zi.total_agi,
                        zi.total_wages,
                        (COALESCE(zi.total_capital_gains, 0)
                         + COALESCE(zi.total_dividends, 0)
                         + COALESCE(zi.total_business_income, 0)) AS non_wage_total,
                        CASE WHEN zi.num_returns > 0
                            THEN (COALESCE(zi.total_capital_gains, 0)
                                  + COALESCE(zi.total_dividends, 0)
                                  + COALESCE(zi.total_business_income, 0))
                                 * 1000.0 / zi.num_returns
                            ELSE 0 END AS non_wage_per_return,
                        CASE WHEN zi.total_agi > 0
                            THEN (COALESCE(zi.total_capital_gains, 0)
                                  + COALESCE(zi.total_dividends, 0)
                                  + COALESCE(zi.total_business_income, 0))
                                 * 100.0 / zi.total_agi
                            ELSE 0 END AS non_wage_pct
                    FROM irs_soi_zip_income zi
                    WHERE zi.tax_year = :year
                        AND zi.agi_class = '0'
                        AND zi.num_returns > 100
                        {state_clause}
                )
                SELECT
                    zw.zip_code,
                    zw.state_abbr,
                    ROUND(zw.non_wage_per_return::numeric, 0) AS non_wage_per_return,
                    ROUND(zw.non_wage_pct::numeric, 1) AS non_wage_pct,
                    zw.num_returns,
                    ROUND(zw.total_agi * 1000.0 / NULLIF(zw.num_returns, 0), 0) AS avg_agi,
                    COALESCE(zms.overall_score, 0) AS medspa_score,
                    COALESCE(zms.grade, '-') AS medspa_grade,
                    COALESCE(mp_cnt.a_count, 0) AS a_grade_medspas
                FROM zip_wealth zw
                LEFT JOIN zip_medspa_scores zms ON zms.zip_code = zw.zip_code
                LEFT JOIN (
                    SELECT zip_code, COUNT(*) AS a_count
                    FROM medspa_prospects
                    WHERE acquisition_grade = 'A'
                    GROUP BY zip_code
                ) mp_cnt ON mp_cnt.zip_code = zw.zip_code
                WHERE zw.non_wage_pct > 25
                    AND COALESCE(zms.overall_score, 0) < 70
                ORDER BY zw.non_wage_per_return DESC
                LIMIT 50
            """), params).fetchall()

            stealth_zips = [
                {
                    "zip_code": r[0],
                    "state": r[1],
                    "non_wage_per_return": float(r[2] or 0),
                    "non_wage_pct": float(r[3] or 0),
                    "num_returns": int(r[4] or 0),
                    "avg_agi": float(r[5] or 0),
                    "medspa_score": float(r[6] or 0),
                    "medspa_grade": r[7] or "-",
                    "a_grade_medspas": int(r[8] or 0),
                }
                for r in stealth_rows
            ]

            # Count validated ZIPs (high non-wage AND high medspa score)
            validated_row = db.execute(text(f"""
                SELECT COUNT(*)
                FROM irs_soi_zip_income zi
                JOIN zip_medspa_scores zms ON zms.zip_code = zi.zip_code
                WHERE zi.tax_year = :year AND zi.agi_class = '0'
                    AND zi.num_returns > 100
                    AND zms.overall_score >= 70
                    AND (COALESCE(zi.total_capital_gains, 0)
                         + COALESCE(zi.total_dividends, 0)
                         + COALESCE(zi.total_business_income, 0))
                         * 100.0 / NULLIF(zi.total_agi, 0) > 25
                    {state_clause}
            """), params).fetchone()
            validated_count = int(validated_row[0]) if validated_row else 0

            # Top states by stealth ZIP count
            top_states_rows = db.execute(text(f"""
                WITH zip_wealth AS (
                    SELECT zi.zip_code, zi.state_abbr,
                        CASE WHEN zi.total_agi > 0
                            THEN (COALESCE(zi.total_capital_gains, 0)
                                  + COALESCE(zi.total_dividends, 0)
                                  + COALESCE(zi.total_business_income, 0))
                                 * 100.0 / zi.total_agi
                            ELSE 0 END AS non_wage_pct
                    FROM irs_soi_zip_income zi
                    WHERE zi.tax_year = :year AND zi.agi_class = '0'
                        AND zi.num_returns > 100
                        {state_clause}
                )
                SELECT zw.state_abbr, COUNT(*) AS stealth_count
                FROM zip_wealth zw
                LEFT JOIN zip_medspa_scores zms ON zms.zip_code = zw.zip_code
                WHERE zw.non_wage_pct > 25
                    AND COALESCE(zms.overall_score, 0) < 70
                GROUP BY zw.state_abbr
                ORDER BY stealth_count DESC
                LIMIT 15
            """), params).fetchall()

            top_states = [
                {"state": r[0], "count": int(r[1])}
                for r in top_states_rows
            ]

            avg_non_wage = (
                round(sum(z["non_wage_per_return"] for z in stealth_zips) / len(stealth_zips), 0)
                if stealth_zips else 0
            )

            return {
                "stealth_wealth": {
                    "stealth_zips": stealth_zips,
                    "validated_count": validated_count,
                    "wealth_composition": wealth_composition,
                    "top_states": top_states,
                    "summary": {
                        "total_stealth": len(stealth_zips),
                        "validated": validated_count,
                        "avg_non_wage_income": avg_non_wage,
                        "tax_year": latest_year,
                    },
                }
            }
        except Exception as e:
            logger.warning(f"Could not compute stealth wealth data: {e}")
            db.rollback()
            return {"stealth_wealth": {}}

    # ------------------------------------------------------------------
    # Section 17: Migration Alpha
    # ------------------------------------------------------------------

    def _get_migration_alpha_data(self, db: Session, state: Optional[str]) -> Dict:
        """Use IRS county-to-county migration to identify wealth inflow vs medspa density."""
        try:
            # Check if table exists
            check = db.execute(text(
                "SELECT 1 FROM information_schema.tables "
                "WHERE table_name = 'irs_soi_migration'"
            ))
            if not check.fetchone():
                return {"migration_alpha": {}}

            # Get latest tax year
            yr_row = db.execute(text(
                "SELECT MAX(tax_year) FROM irs_soi_migration"
            )).fetchone()
            if not yr_row or not yr_row[0]:
                return {"migration_alpha": {}}
            latest_year = yr_row[0]

            state_clause_dest = "AND m.dest_state_abbr = :state" if state else ""
            params: Dict[str, Any] = {"year": latest_year}
            if state:
                params["state"] = state.upper()

            # --- State-level net AGI flows ---
            state_flows_rows = db.execute(text(f"""
                WITH inflows AS (
                    SELECT dest_state_abbr AS state_abbr,
                           SUM(total_agi) AS inflow_agi,
                           SUM(num_returns) AS inflow_returns
                    FROM irs_soi_migration m
                    WHERE tax_year = :year AND flow_type = 'inflow'
                        {state_clause_dest}
                    GROUP BY dest_state_abbr
                ),
                outflows AS (
                    SELECT dest_state_abbr AS state_abbr,
                           SUM(total_agi) AS outflow_agi,
                           SUM(num_returns) AS outflow_returns
                    FROM irs_soi_migration m
                    WHERE tax_year = :year AND flow_type = 'outflow'
                        {state_clause_dest}
                    GROUP BY dest_state_abbr
                ),
                medspa_density AS (
                    SELECT state, COUNT(*) AS total_medspas,
                           COUNT(*) FILTER (WHERE acquisition_grade = 'A') AS a_grade_count
                    FROM medspa_prospects
                    GROUP BY state
                )
                SELECT
                    COALESCE(i.state_abbr, o.state_abbr) AS state_abbr,
                    COALESCE(i.inflow_agi, 0) AS inflow_agi,
                    COALESCE(o.outflow_agi, 0) AS outflow_agi,
                    (COALESCE(i.inflow_agi, 0) - COALESCE(o.outflow_agi, 0)) AS net_agi,
                    COALESCE(i.inflow_returns, 0) AS inflow_returns,
                    COALESCE(o.outflow_returns, 0) AS outflow_returns,
                    COALESCE(md.total_medspas, 0) AS total_medspas,
                    COALESCE(md.a_grade_count, 0) AS a_grade_count,
                    CASE WHEN COALESCE(md.a_grade_count, 0) > 0
                        THEN (COALESCE(i.inflow_agi, 0) - COALESCE(o.outflow_agi, 0))
                             * 1.0 / md.a_grade_count
                        ELSE (COALESCE(i.inflow_agi, 0) - COALESCE(o.outflow_agi, 0)) * 1.0
                    END AS migration_alpha
                FROM inflows i
                FULL OUTER JOIN outflows o ON i.state_abbr = o.state_abbr
                LEFT JOIN medspa_density md
                    ON md.state = COALESCE(i.state_abbr, o.state_abbr)
                WHERE COALESCE(i.state_abbr, o.state_abbr) IS NOT NULL
                ORDER BY net_agi DESC
            """), params).fetchall()

            state_flows = [
                {
                    "state": r[0],
                    "inflow_agi_m": round(float(r[1] or 0) * 1000 / 1e6, 1),
                    "outflow_agi_m": round(float(r[2] or 0) * 1000 / 1e6, 1),
                    "net_agi_m": round(float(r[3] or 0) * 1000 / 1e6, 1),
                    "inflow_returns": int(r[4] or 0),
                    "outflow_returns": int(r[5] or 0),
                    "total_medspas": int(r[6] or 0),
                    "a_grade_count": int(r[7] or 0),
                    "migration_alpha": round(float(r[8] or 0), 1),
                }
                for r in state_flows_rows
            ]

            # --- Top county-level inflows ---
            county_rows = db.execute(text(f"""
                SELECT
                    m.dest_state_abbr,
                    m.dest_county_name,
                    m.orig_state_abbr,
                    SUM(m.total_agi) AS inflow_agi,
                    SUM(m.num_returns) AS inflow_returns
                FROM irs_soi_migration m
                WHERE m.tax_year = :year AND m.flow_type = 'inflow'
                    {state_clause_dest}
                GROUP BY m.dest_state_abbr, m.dest_county_name, m.orig_state_abbr
                ORDER BY inflow_agi DESC
                LIMIT 20
            """), params).fetchall()

            top_county_inflows = [
                {
                    "dest_state": r[0],
                    "county": r[1],
                    "orig_state": r[2],
                    "agi_m": round(float(r[3] or 0) * 1000 / 1e6, 1),
                    "returns": int(r[4] or 0),
                }
                for r in county_rows
            ]

            # --- Classify emerging markets ---
            emerging = [
                s for s in state_flows
                if s["net_agi_m"] > 0 and s["a_grade_count"] < 20
            ]
            emerging.sort(key=lambda x: x["migration_alpha"], reverse=True)

            # --- Summary stats ---
            total_net = sum(s["net_agi_m"] for s in state_flows)
            top_gainer = state_flows[0]["state"] if state_flows and state_flows[0]["net_agi_m"] > 0 else "-"
            top_loser = state_flows[-1]["state"] if state_flows and state_flows[-1]["net_agi_m"] < 0 else "-"

            # Wealth exodus: states losing the most
            wealth_exodus = [s for s in state_flows if s["net_agi_m"] < 0]
            wealth_exodus.sort(key=lambda x: x["net_agi_m"])

            return {
                "migration_alpha": {
                    "state_flows": state_flows,
                    "top_county_inflows": top_county_inflows,
                    "wealth_exodus": wealth_exodus[:10],
                    "emerging_markets": emerging[:15],
                    "summary": {
                        "total_net_agi_m": round(total_net, 1),
                        "top_gainer": top_gainer,
                        "top_loser": top_loser,
                        "emerging_count": len(emerging),
                        "tax_year": latest_year,
                    },
                }
            }
        except Exception as e:
            logger.warning(f"Could not compute migration alpha data: {e}")
            db.rollback()
            return {"migration_alpha": {}}

    # ------------------------------------------------------------------
    # Section 18: Medical Provider Density Signal (CMS Medicare)
    # ------------------------------------------------------------------

    def _get_provider_density_data(self, db: Session, state: Optional[str]) -> Dict:
        """Cross-reference CMS Medicare provider counts with medspa density by ZIP."""
        try:
            check = db.execute(text(
                "SELECT 1 FROM information_schema.tables "
                "WHERE table_name = 'cms_medicare_utilization'"
            ))
            if not check.fetchone():
                return {"provider_density": {}}

            aesthetics_types = [
                'Dermatology',
                'Plastic and Reconstructive Surgery',
                'Nurse Practitioner',
                'Physician Assistant',
            ]

            state_clause = "AND cu.rndrng_prvdr_state_abrvtn = :state" if state else ""
            params: Dict[str, Any] = {"provider_types": aesthetics_types}
            if state:
                params["state"] = state.upper()

            # Provider counts by type (for bar chart)
            type_rows = db.execute(text(f"""
                SELECT
                    cu.rndrng_prvdr_type,
                    COUNT(DISTINCT cu.rndrng_npi) AS provider_count,
                    SUM(cu.tot_benes) AS total_benes
                FROM cms_medicare_utilization cu
                WHERE cu.rndrng_prvdr_type = ANY(:provider_types)
                    {state_clause}
                GROUP BY cu.rndrng_prvdr_type
                ORDER BY provider_count DESC
            """), params).fetchall()

            provider_by_type = [
                {
                    "type": r[0],
                    "count": int(r[1] or 0),
                    "beneficiaries": int(r[2] or 0),
                }
                for r in type_rows
            ]

            # ZIP-level aggregation with medspa cross-reference
            zip_rows = db.execute(text(f"""
                WITH zip_providers AS (
                    SELECT
                        cu.rndrng_prvdr_zip5 AS zip_code,
                        COUNT(DISTINCT cu.rndrng_npi) AS provider_count,
                        SUM(cu.tot_benes) AS total_benes,
                        STRING_AGG(DISTINCT cu.rndrng_prvdr_type, ', ') AS provider_types
                    FROM cms_medicare_utilization cu
                    WHERE cu.rndrng_prvdr_type = ANY(:provider_types)
                        {state_clause}
                    GROUP BY cu.rndrng_prvdr_zip5
                    HAVING COUNT(DISTINCT cu.rndrng_npi) >= 3
                )
                SELECT
                    zp.zip_code,
                    zp.provider_count,
                    zp.total_benes,
                    zp.provider_types,
                    COALESCE(zms.overall_score, 0) AS medspa_score,
                    COALESCE(zms.grade, '-') AS medspa_grade,
                    COALESCE(mp_cnt.medspa_count, 0) AS medspa_count,
                    CASE WHEN COALESCE(mp_cnt.medspa_count, 0) > 0
                        THEN ROUND(zp.provider_count * 1.0 / mp_cnt.medspa_count, 1)
                        ELSE zp.provider_count * 1.0
                    END AS imbalance_score,
                    COALESCE(irs.num_returns, 0) AS num_returns
                FROM zip_providers zp
                LEFT JOIN zip_medspa_scores zms ON zms.zip_code = zp.zip_code
                LEFT JOIN (
                    SELECT zip_code, COUNT(*) AS medspa_count
                    FROM medspa_prospects
                    GROUP BY zip_code
                ) mp_cnt ON mp_cnt.zip_code = zp.zip_code
                LEFT JOIN (
                    SELECT zip_code, num_returns
                    FROM irs_soi_zip_income
                    WHERE agi_class = '0'
                        AND tax_year = (SELECT MAX(tax_year) FROM irs_soi_zip_income)
                ) irs ON irs.zip_code = zp.zip_code
                ORDER BY imbalance_score DESC
                LIMIT 50
            """), params).fetchall()

            opportunity_zips = [
                {
                    "zip_code": r[0],
                    "provider_count": int(r[1] or 0),
                    "beneficiaries": int(r[2] or 0),
                    "provider_types": r[3] or "",
                    "medspa_score": float(r[4] or 0),
                    "medspa_grade": r[5] or "-",
                    "medspa_count": int(r[6] or 0),
                    "imbalance_score": float(r[7] or 0),
                    "num_returns": int(r[8] or 0),
                }
                for r in zip_rows
            ]

            # Filter true opportunity ZIPs (low medspa score)
            opp_zips = [z for z in opportunity_zips if z["medspa_score"] < 60]

            total_providers = sum(t["count"] for t in provider_by_type)
            total_benes = sum(t["beneficiaries"] for t in provider_by_type)
            avg_per_zip = (
                round(total_providers / len(opportunity_zips), 1)
                if opportunity_zips else 0
            )
            top_type = provider_by_type[0]["type"] if provider_by_type else "-"

            return {
                "provider_density": {
                    "opportunity_zips": opportunity_zips,
                    "provider_by_type": provider_by_type,
                    "summary": {
                        "total_provider_zip_pairs": len(opportunity_zips),
                        "opportunity_zips": len(opp_zips),
                        "avg_providers_per_zip": avg_per_zip,
                        "top_provider_type": top_type,
                        "total_providers": total_providers,
                        "total_beneficiaries": total_benes,
                    },
                }
            }
        except Exception as e:
            logger.warning(f"Could not compute provider density data: {e}")
            db.rollback()
            return {"provider_density": {}}

    # ------------------------------------------------------------------
    # Section 19: Real Estate Appreciation Alpha (Redfin / FHFA)
    # ------------------------------------------------------------------

    def _get_real_estate_alpha_data(self, db: Session, state: Optional[str]) -> Dict:
        """Cross-reference real estate appreciation with medspa scores for timing signals."""
        try:
            from collections import defaultdict

            # Try Redfin first (ZIP-level), fall back to FHFA (ZIP3-level)
            source_used = None

            # Check Redfin
            check_redfin = db.execute(text(
                "SELECT 1 FROM information_schema.tables "
                "WHERE table_name = 'realestate_redfin'"
            ))
            has_redfin = check_redfin.fetchone() is not None

            # Check FHFA
            check_fhfa = db.execute(text(
                "SELECT 1 FROM information_schema.tables "
                "WHERE table_name = 'realestate_fhfa_hpi'"
            ))
            has_fhfa = check_fhfa.fetchone() is not None

            if not has_redfin and not has_fhfa:
                return {"real_estate_alpha": {}}

            state_clause = ""
            params: Dict[str, Any] = {}
            if state:
                params["state"] = state.upper()

            zip_re_data = []

            if has_redfin:
                source_used = "Redfin"
                state_clause = "AND r.state_code = :state" if state else ""
                rows = db.execute(text(f"""
                    WITH latest AS (
                        SELECT
                            r.zip_code,
                            r.state_code,
                            r.median_sale_price,
                            r.period_end,
                            ROW_NUMBER() OVER (
                                PARTITION BY r.zip_code
                                ORDER BY r.period_end DESC
                            ) AS rn
                        FROM realestate_redfin r
                        WHERE r.median_sale_price IS NOT NULL
                            AND r.zip_code IS NOT NULL
                            {state_clause}
                    ),
                    prior_year AS (
                        SELECT
                            r.zip_code,
                            r.median_sale_price AS prior_price,
                            ROW_NUMBER() OVER (
                                PARTITION BY r.zip_code
                                ORDER BY r.period_end DESC
                            ) AS rn
                        FROM realestate_redfin r
                        WHERE r.median_sale_price IS NOT NULL
                            AND r.zip_code IS NOT NULL
                            AND r.period_end <= (
                                SELECT MAX(period_end) - INTERVAL '11 months'
                                FROM realestate_redfin
                            )
                            {state_clause}
                    )
                    SELECT
                        l.zip_code,
                        l.state_code,
                        l.median_sale_price,
                        p.prior_price,
                        CASE WHEN p.prior_price > 0
                            THEN ROUND(((l.median_sale_price - p.prior_price)
                                 * 100.0 / p.prior_price)::numeric, 1)
                            ELSE NULL
                        END AS yoy_change
                    FROM latest l
                    LEFT JOIN prior_year p ON p.zip_code = l.zip_code AND p.rn = 1
                    WHERE l.rn = 1
                    ORDER BY l.median_sale_price DESC
                """), params).fetchall()

                zip_re_data = [
                    {
                        "zip_code": r[0],
                        "state": r[1],
                        "median_price": float(r[2] or 0),
                        "yoy_change": float(r[4]) if r[4] is not None else None,
                    }
                    for r in rows
                ]
            elif has_fhfa:
                source_used = "FHFA"
                state_clause = "AND h.state_abbr = :state" if state else ""
                rows = db.execute(text(f"""
                    WITH latest AS (
                        SELECT
                            h.geography_id AS zip3,
                            h.state_abbr,
                            h.index_value,
                            h.year,
                            h.quarter,
                            ROW_NUMBER() OVER (
                                PARTITION BY h.geography_id
                                ORDER BY h.year DESC, h.quarter DESC
                            ) AS rn
                        FROM realestate_fhfa_hpi h
                        WHERE h.geography_type = 'ZIP3'
                            AND h.index_value IS NOT NULL
                            {state_clause}
                    ),
                    prior AS (
                        SELECT
                            h.geography_id AS zip3,
                            h.index_value AS prior_value,
                            ROW_NUMBER() OVER (
                                PARTITION BY h.geography_id
                                ORDER BY h.year DESC, h.quarter DESC
                            ) AS rn
                        FROM realestate_fhfa_hpi h
                        WHERE h.geography_type = 'ZIP3'
                            AND h.index_value IS NOT NULL
                            AND (h.year * 4 + h.quarter) <= (
                                (SELECT MAX(year) FROM realestate_fhfa_hpi
                                 WHERE geography_type = 'ZIP3') * 4
                                + (SELECT MAX(quarter) FROM realestate_fhfa_hpi
                                   WHERE geography_type = 'ZIP3'
                                   AND year = (SELECT MAX(year) FROM realestate_fhfa_hpi
                                               WHERE geography_type = 'ZIP3'))
                            ) - 4
                            {state_clause}
                    )
                    SELECT
                        l.zip3,
                        l.state_abbr,
                        l.index_value,
                        p.prior_value,
                        CASE WHEN p.prior_value > 0
                            THEN ROUND(((l.index_value - p.prior_value)
                                 * 100.0 / p.prior_value)::numeric, 1)
                            ELSE NULL
                        END AS yoy_change
                    FROM latest l
                    LEFT JOIN prior p ON p.zip3 = l.zip3 AND p.rn = 1
                    WHERE l.rn = 1
                    ORDER BY l.index_value DESC
                """), params).fetchall()

                zip_re_data = [
                    {
                        "zip_code": r[0],
                        "state": r[1],
                        "median_price": float(r[2] or 0),
                        "yoy_change": float(r[4]) if r[4] is not None else None,
                    }
                    for r in rows
                ]

            if not zip_re_data:
                return {"real_estate_alpha": {}}

            # Load medspa scores for cross-reference (client-side join)
            score_rows = db.execute(text(
                "SELECT zip_code, overall_score, grade FROM zip_medspa_scores"
            )).fetchall()
            score_map = {r[0]: {"score": float(r[1] or 0), "grade": r[2] or "-"} for r in score_rows}

            # Enrich with medspa scores
            for z in zip_re_data:
                zk = z["zip_code"]
                ms = score_map.get(zk, {"score": 0, "grade": "-"})
                z["medspa_score"] = ms["score"]
                z["medspa_grade"] = ms["grade"]

            # Timing opportunity ZIPs: appreciating + medspa score < 60
            timing_zips = [
                z for z in zip_re_data
                if z.get("yoy_change") is not None
                and z["yoy_change"] > 0
                and z["medspa_score"] < 60
            ]
            timing_zips.sort(key=lambda x: x["yoy_change"], reverse=True)

            # State-level summary
            state_agg = defaultdict(lambda: {"prices": [], "yoys": []})
            for z in zip_re_data:
                st = z.get("state")
                if st:
                    state_agg[st]["prices"].append(z["median_price"])
                    if z.get("yoy_change") is not None:
                        state_agg[st]["yoys"].append(z["yoy_change"])

            state_summary = []
            for st, vals in state_agg.items():
                avg_price = round(sum(vals["prices"]) / len(vals["prices"]), 0) if vals["prices"] else 0
                avg_yoy = round(sum(vals["yoys"]) / len(vals["yoys"]), 1) if vals["yoys"] else 0
                state_summary.append({
                    "state": st,
                    "avg_median_price": avg_price,
                    "avg_yoy_change": avg_yoy,
                    "zip_count": len(vals["prices"]),
                })
            state_summary.sort(key=lambda x: x["avg_yoy_change"], reverse=True)

            # Summary KPIs
            all_prices = [z["median_price"] for z in zip_re_data if z["median_price"] > 0]
            all_yoys = [z["yoy_change"] for z in zip_re_data if z.get("yoy_change") is not None]
            avg_price = round(sum(all_prices) / len(all_prices), 0) if all_prices else 0
            avg_yoy = round(sum(all_yoys) / len(all_yoys), 1) if all_yoys else 0
            hottest = state_summary[0]["state"] if state_summary else "-"

            return {
                "real_estate_alpha": {
                    "timing_zips": timing_zips[:50],
                    "state_summary": state_summary,
                    "source": source_used,
                    "summary": {
                        "avg_median_price": avg_price,
                        "avg_yoy_change": avg_yoy,
                        "timing_opportunity_zips": len(timing_zips),
                        "hottest_state": hottest,
                        "zips_analyzed": len(zip_re_data),
                    },
                }
            }
        except Exception as e:
            logger.warning(f"Could not compute real estate alpha data: {e}")
            db.rollback()
            return {"real_estate_alpha": {}}

    # ------------------------------------------------------------------
    # Section 20: Deposit Wealth Concentration (FDIC)
    # ------------------------------------------------------------------

    def _get_deposit_wealth_data(self, db: Session, state: Optional[str]) -> Dict:
        """Cross-reference FDIC branch deposits with medspa scores for wealth signals."""
        try:
            check = db.execute(text(
                "SELECT 1 FROM information_schema.tables "
                "WHERE table_name = 'fdic_summary_deposits'"
            ))
            if not check.fetchone():
                return {"deposit_wealth": {}}

            # Get latest year
            yr_row = db.execute(text(
                "SELECT MAX(year) FROM fdic_summary_deposits"
            )).fetchone()
            if not yr_row or not yr_row[0]:
                return {"deposit_wealth": {}}
            latest_year = yr_row[0]

            state_clause = "AND fd.stalpbr = :state" if state else ""
            params: Dict[str, Any] = {"year": latest_year}
            if state:
                params["state"] = state.upper()

            # ZIP-level deposit aggregation
            zip_rows = db.execute(text(f"""
                WITH zip_deposits AS (
                    SELECT
                        fd.zipbr AS zip_code,
                        fd.stalpbr AS state_abbr,
                        SUM(fd.depsum) AS total_deposits,
                        COUNT(*) AS branch_count
                    FROM fdic_summary_deposits fd
                    WHERE fd.year = :year
                        AND LENGTH(fd.zipbr) = 5
                        {state_clause}
                    GROUP BY fd.zipbr, fd.stalpbr
                )
                SELECT
                    zd.zip_code,
                    zd.state_abbr,
                    zd.total_deposits,
                    zd.branch_count,
                    COALESCE(irs.num_returns, 0) AS num_returns,
                    CASE WHEN COALESCE(irs.num_returns, 0) > 0
                        THEN ROUND((zd.total_deposits / irs.num_returns)::numeric, 0)
                        ELSE 0
                    END AS deposits_per_return,
                    COALESCE(zms.overall_score, 0) AS medspa_score,
                    COALESCE(zms.grade, '-') AS medspa_grade
                FROM zip_deposits zd
                LEFT JOIN (
                    SELECT zip_code, num_returns
                    FROM irs_soi_zip_income
                    WHERE agi_class = '0'
                        AND tax_year = (SELECT MAX(tax_year) FROM irs_soi_zip_income)
                ) irs ON irs.zip_code = zd.zip_code
                LEFT JOIN zip_medspa_scores zms ON zms.zip_code = zd.zip_code
                WHERE zd.total_deposits > 0
                ORDER BY deposits_per_return DESC
                LIMIT 100
            """), params).fetchall()

            all_zips = [
                {
                    "zip_code": r[0],
                    "state": r[1],
                    "total_deposits": float(r[2] or 0),
                    "branch_count": int(r[3] or 0),
                    "num_returns": int(r[4] or 0),
                    "deposits_per_return": float(r[5] or 0),
                    "medspa_score": float(r[6] or 0),
                    "medspa_grade": r[7] or "-",
                }
                for r in zip_rows
            ]

            # Underserved = high deposits + medspa score < 60
            underserved = [z for z in all_zips if z["medspa_score"] < 60]

            # State deposit concentration
            state_dep_rows = db.execute(text(f"""
                SELECT
                    fd.stalpbr AS state_abbr,
                    SUM(fd.depsum) AS total_deposits
                FROM fdic_summary_deposits fd
                WHERE fd.year = :year
                    AND LENGTH(fd.zipbr) = 5
                    {state_clause}
                GROUP BY fd.stalpbr
                ORDER BY total_deposits DESC
            """), params).fetchall()

            state_deposits = [
                {"state": r[0], "deposits": float(r[1] or 0)}
                for r in state_dep_rows
            ]

            # Summary KPIs
            total_deposits = sum(s["deposits"] for s in state_deposits)
            avg_dpr = (
                round(sum(z["deposits_per_return"] for z in all_zips if z["deposits_per_return"] > 0)
                      / max(len([z for z in all_zips if z["deposits_per_return"] > 0]), 1), 0)
            )

            return {
                "deposit_wealth": {
                    "underserved_zips": underserved[:50],
                    "all_zips": all_zips[:50],
                    "state_deposits": state_deposits,
                    "summary": {
                        "zips_analyzed": len(all_zips),
                        "total_deposits_t": round(total_deposits / 1e12, 2),
                        "avg_deposits_per_return": avg_dpr,
                        "underserved_zips": len(underserved),
                        "year": latest_year,
                    },
                }
            }
        except Exception as e:
            logger.warning(f"Could not compute deposit wealth data: {e}")
            db.rollback()
            return {"deposit_wealth": {}}

    # ------------------------------------------------------------------
    # Section 21: Business Formation Velocity (IRS SOI Business Income)
    # ------------------------------------------------------------------

    def _get_business_formation_data(self, db: Session, state: Optional[str]) -> Dict:
        """Compute business density from IRS SOI business income data vs medspa scores."""
        try:
            check = db.execute(text(
                "SELECT 1 FROM information_schema.tables "
                "WHERE table_name = 'irs_soi_business_income'"
            ))
            if not check.fetchone():
                return {"business_formation": {}}

            # Get latest and prior tax years
            yr_rows = db.execute(text(
                "SELECT DISTINCT tax_year FROM irs_soi_business_income "
                "ORDER BY tax_year DESC LIMIT 2"
            )).fetchall()
            if not yr_rows:
                return {"business_formation": {}}

            latest_year = yr_rows[0][0]
            prior_year = yr_rows[1][0] if len(yr_rows) > 1 else None

            state_clause = "AND bi.state_abbr = :state" if state else ""
            params: Dict[str, Any] = {"year": latest_year}
            if state:
                params["state"] = state.upper()

            has_prior = prior_year is not None
            if has_prior:
                params["prior_year"] = prior_year
                prior_join = """
                    LEFT JOIN irs_soi_business_income bp
                        ON bp.zip_code = bi.zip_code
                        AND bp.tax_year = :prior_year
                        AND bp.num_returns > 100
                """
                prior_select = """,
                    CASE WHEN bp.num_returns > 0
                        THEN ROUND((bp.num_with_business_income * 100.0
                              / bp.num_returns)::numeric, 1)
                        ELSE NULL
                    END AS prior_biz_density
                """
            else:
                prior_join = ""
                prior_select = ", NULL AS prior_biz_density"

            zip_rows = db.execute(text(f"""
                SELECT
                    bi.zip_code,
                    bi.state_abbr,
                    bi.num_returns,
                    bi.num_with_business_income,
                    ROUND((bi.num_with_business_income * 100.0
                          / NULLIF(bi.num_returns, 0))::numeric, 1) AS biz_density,
                    COALESCE(bi.total_schedule_c_income, 0) * 1000 AS schedule_c,
                    COALESCE(bi.total_partnership_income, 0) * 1000 AS partnership,
                    COALESCE(bi.total_scorp_income, 0) * 1000 AS scorp,
                    COALESCE(zms.overall_score, 0) AS medspa_score,
                    COALESCE(zms.grade, '-') AS medspa_grade
                    {prior_select}
                FROM irs_soi_business_income bi
                {prior_join}
                LEFT JOIN zip_medspa_scores zms ON zms.zip_code = bi.zip_code
                WHERE bi.tax_year = :year
                    AND bi.num_returns > 100
                    {state_clause}
                ORDER BY biz_density DESC
                LIMIT 100
            """), params).fetchall()

            all_zips = []
            for r in zip_rows:
                biz_density = float(r[4] or 0)
                prior_density = float(r[10]) if r[10] is not None else None
                yoy_growth = None
                if prior_density is not None and prior_density > 0:
                    yoy_growth = round(
                        (biz_density - prior_density) / prior_density * 100, 1
                    )
                all_zips.append({
                    "zip_code": r[0],
                    "state": r[1],
                    "num_returns": int(r[2] or 0),
                    "num_with_biz": int(r[3] or 0),
                    "biz_density": biz_density,
                    "schedule_c": float(r[5] or 0),
                    "partnership": float(r[6] or 0),
                    "scorp": float(r[7] or 0),
                    "medspa_score": float(r[8] or 0),
                    "medspa_grade": r[9] or "-",
                    "yoy_growth": yoy_growth,
                })

            # Entrepreneurial ZIPs: biz_density > 20% and medspa < 60
            entrepreneurial = [
                z for z in all_zips
                if z["biz_density"] > 20 and z["medspa_score"] < 60
            ]

            # State-level avg biz density
            state_density_rows = db.execute(text(f"""
                SELECT
                    bi.state_abbr,
                    ROUND(AVG(CASE WHEN bi.num_returns > 0
                        THEN bi.num_with_business_income * 100.0 / bi.num_returns
                        ELSE 0 END)::numeric, 1) AS avg_biz_density,
                    COUNT(*) AS zip_count
                FROM irs_soi_business_income bi
                WHERE bi.tax_year = :year
                    AND bi.num_returns > 100
                    {state_clause}
                GROUP BY bi.state_abbr
                ORDER BY avg_biz_density DESC
            """), params).fetchall()

            state_density = [
                {"state": r[0], "avg_density": float(r[1] or 0), "zip_count": int(r[2] or 0)}
                for r in state_density_rows
            ]

            # Income type composition (national)
            comp_row = db.execute(text(f"""
                SELECT
                    SUM(COALESCE(total_schedule_c_income, 0)) AS schedule_c,
                    SUM(COALESCE(total_partnership_income, 0)) AS partnership,
                    SUM(COALESCE(total_scorp_income, 0)) AS scorp
                FROM irs_soi_business_income bi
                WHERE bi.tax_year = :year
                    AND bi.num_returns > 100
                    {state_clause}
            """), params).fetchone()

            income_composition = {}
            if comp_row:
                sc = float(comp_row[0] or 0)
                pt = float(comp_row[1] or 0)
                sp = float(comp_row[2] or 0)
                total_biz = sc + pt + sp
                if total_biz > 0:
                    income_composition = {
                        "schedule_c_pct": round(sc / total_biz * 100, 1),
                        "partnership_pct": round(pt / total_biz * 100, 1),
                        "scorp_pct": round(sp / total_biz * 100, 1),
                    }

            # Summary KPIs
            avg_density = (
                round(sum(z["biz_density"] for z in all_zips) / len(all_zips), 1)
                if all_zips else 0
            )
            high_growth = [z for z in all_zips if z.get("yoy_growth") is not None and z["yoy_growth"] > 10]
            avg_sc = (
                round(sum(z["schedule_c"] for z in all_zips) / len(all_zips), 0)
                if all_zips else 0
            )

            return {
                "business_formation": {
                    "entrepreneurial_zips": entrepreneurial[:50],
                    "all_zips": all_zips[:50],
                    "state_density": state_density,
                    "income_composition": income_composition,
                    "summary": {
                        "zips_analyzed": len(all_zips),
                        "avg_biz_density": avg_density,
                        "high_growth_zips": len(high_growth),
                        "avg_schedule_c": avg_sc,
                        "tax_year": latest_year,
                        "has_prior_year": has_prior,
                    },
                }
            }
        except Exception as e:
            logger.warning(f"Could not compute business formation data: {e}")
            db.rollback()
            return {"business_formation": {}}

    # ------------------------------------------------------------------
    # Section 22: Opportunity Zone Overlay
    # ------------------------------------------------------------------

    def _get_opportunity_zone_data(self, db: Session, state: Optional[str]) -> Dict:
        """Cross-reference Opportunity Zone tracts with medspa prospect states."""
        try:
            check = db.execute(text(
                "SELECT 1 FROM information_schema.tables "
                "WHERE table_name = 'opportunity_zone'"
            ))
            if not check.fetchone():
                return {"opportunity_zones": {}}

            state_clause = "AND oz.state = :state" if state else ""
            params: Dict[str, Any] = {}
            if state:
                params["state"] = state.upper()

            # Aggregate OZ tracts by state + cross with medspa prospect counts
            rows = db.execute(text(f"""
                SELECT
                    oz.state,
                    COUNT(*) AS oz_tracts,
                    SUM(CASE WHEN oz.is_low_income THEN 1 ELSE 0 END) AS low_income_tracts,
                    SUM(CASE WHEN NOT oz.is_low_income THEN 1 ELSE 0 END) AS contiguous_tracts,
                    COALESCE(mp.prospect_count, 0) AS prospect_count,
                    COALESCE(mp.a_grade_count, 0) AS a_grade_count
                FROM opportunity_zone oz
                LEFT JOIN (
                    SELECT state,
                           COUNT(*) AS prospect_count,
                           COUNT(*) FILTER (WHERE acquisition_grade = 'A') AS a_grade_count
                    FROM medspa_prospects
                    GROUP BY state
                ) mp ON mp.state = oz.state
                WHERE 1=1 {state_clause}
                GROUP BY oz.state, mp.prospect_count, mp.a_grade_count
                ORDER BY oz_tracts DESC
            """), params).fetchall()

            if not rows:
                return {"opportunity_zones": {}}

            state_data = []
            total_tracts = 0
            total_low_income = 0
            total_contiguous = 0
            states_with_oz = 0
            medspa_oz_states = 0
            tax_advantaged_states = 0

            for r in rows:
                tracts = int(r[1] or 0)
                low_income = int(r[2] or 0)
                contiguous = int(r[3] or 0)
                prospects = int(r[4] or 0)
                a_grade = int(r[5] or 0)
                oz_per_prospect = round(tracts / prospects, 2) if prospects > 0 else 0

                state_data.append({
                    "state": r[0],
                    "oz_tracts": tracts,
                    "low_income_tracts": low_income,
                    "contiguous_tracts": contiguous,
                    "prospect_count": prospects,
                    "a_grade_count": a_grade,
                    "oz_per_prospect": oz_per_prospect,
                })

                total_tracts += tracts
                total_low_income += low_income
                total_contiguous += contiguous
                states_with_oz += 1
                if prospects > 0:
                    medspa_oz_states += 1
                if a_grade > 5 and tracts > 50:
                    tax_advantaged_states += 1

            return {
                "opportunity_zones": {
                    "state_data": state_data,
                    "summary": {
                        "total_tracts": total_tracts,
                        "states_with_oz": states_with_oz,
                        "medspa_oz_states": medspa_oz_states,
                        "tax_advantaged_states": tax_advantaged_states,
                        "total_low_income": total_low_income,
                        "total_contiguous": total_contiguous,
                    },
                }
            }
        except Exception as e:
            logger.warning(f"Could not compute opportunity zone data: {e}")
            db.rollback()
            return {"opportunity_zones": {}}

    # ------------------------------------------------------------------
    # Section 23: Demographic Demand Model
    # ------------------------------------------------------------------

    def _get_demographic_demand_data(self, db: Session, state: Optional[str]) -> Dict:
        """Cross-reference educational attainment with medspa prospect density."""
        try:
            check = db.execute(text(
                "SELECT 1 FROM information_schema.tables "
                "WHERE table_name = 'educational_attainment'"
            ))
            if not check.fetchone():
                return {"demographic_demand": {}}

            # Get latest period year
            yr_row = db.execute(text(
                "SELECT MAX(period_year) FROM educational_attainment"
            )).fetchone()
            if not yr_row or not yr_row[0]:
                return {"demographic_demand": {}}

            latest_year = yr_row[0]
            state_clause = "AND LEFT(ea.area_fips, 2) = :state_fips" if state else ""
            params: Dict[str, Any] = {"year": latest_year}

            # State FIPS lookup for filtering (only needed if state filter active)
            if state:
                fips_row = db.execute(text(
                    "SELECT DISTINCT LEFT(area_fips, 2) FROM educational_attainment "
                    "WHERE area_name ILIKE :state_name LIMIT 1"
                ), {"state_name": f"%{state}%"}).fetchone()
                if fips_row:
                    params["state_fips"] = fips_row[0]
                else:
                    return {"demographic_demand": {}}

            # Aggregate by state-level: use area_type that gives state-level data
            rows = db.execute(text(f"""
                WITH state_edu AS (
                    SELECT
                        CASE WHEN ea.area_type IN ('State', 'state')
                            THEN ea.area_name
                            ELSE ea.area_name
                        END AS state_name,
                        LEFT(ea.area_fips, 2) AS state_fips,
                        AVG(ea.pct_bachelors) AS avg_bachelors,
                        AVG(ea.pct_graduate) AS avg_graduate,
                        SUM(ea.population_25_plus) AS total_pop_25plus,
                        COUNT(*) AS area_count
                    FROM educational_attainment ea
                    WHERE ea.period_year = :year
                        {state_clause}
                    GROUP BY LEFT(ea.area_fips, 2), ea.area_name
                    HAVING AVG(ea.pct_bachelors) IS NOT NULL
                )
                SELECT
                    se.state_name,
                    se.state_fips,
                    ROUND(se.avg_bachelors::numeric, 1) AS avg_bachelors,
                    ROUND(se.avg_graduate::numeric, 1) AS avg_graduate,
                    se.total_pop_25plus,
                    se.area_count,
                    COALESCE(mp.prospect_count, 0) AS prospect_count,
                    COALESCE(mp.a_grade_count, 0) AS a_grade_count
                FROM state_edu se
                LEFT JOIN (
                    SELECT state,
                           COUNT(*) AS prospect_count,
                           COUNT(*) FILTER (WHERE acquisition_grade = 'A') AS a_grade_count
                    FROM medspa_prospects
                    GROUP BY state
                ) mp ON mp.state = se.state_fips
                ORDER BY avg_bachelors DESC
            """), params).fetchall()

            if not rows:
                return {"demographic_demand": {}}

            state_data = []
            all_bachelors = []
            all_graduate = []
            underserved_educated = 0

            for r in rows:
                bachelors = float(r[2] or 0)
                graduate = float(r[3] or 0)
                prospects = int(r[6] or 0)
                a_grade = int(r[7] or 0)
                # Gap score: high education rank - low medspa density = opportunity
                edu_score = bachelors + graduate
                gap_score = round(edu_score - (prospects * 0.1), 1) if prospects < 200 else 0

                state_data.append({
                    "state_name": r[0],
                    "state_fips": r[1],
                    "avg_bachelors": bachelors,
                    "avg_graduate": graduate,
                    "total_pop_25plus": int(r[4] or 0),
                    "area_count": int(r[5] or 0),
                    "prospect_count": prospects,
                    "a_grade_count": a_grade,
                    "gap_score": gap_score,
                })
                all_bachelors.append(bachelors)
                all_graduate.append(graduate)
                if bachelors > 30 and prospects < 50:
                    underserved_educated += 1

            avg_b = round(sum(all_bachelors) / len(all_bachelors), 1) if all_bachelors else 0
            avg_g = round(sum(all_graduate) / len(all_graduate), 1) if all_graduate else 0

            return {
                "demographic_demand": {
                    "state_data": state_data,
                    "summary": {
                        "avg_bachelors_pct": avg_b,
                        "avg_graduate_pct": avg_g,
                        "states_analyzed": len(state_data),
                        "underserved_educated": underserved_educated,
                        "period_year": latest_year,
                    },
                }
            }
        except Exception as e:
            logger.warning(f"Could not compute demographic demand data: {e}")
            db.rollback()
            return {"demographic_demand": {}}

    # ------------------------------------------------------------------
    # Section 24: PE Competitive Heat Map
    # ------------------------------------------------------------------

    def _get_pe_competitive_data(self, db: Session, state: Optional[str]) -> Dict:
        """Filter PE deals for aesthetics-related keywords and compute competitive landscape."""
        try:
            check = db.execute(text(
                "SELECT 1 FROM information_schema.tables "
                "WHERE table_name = 'pe_portfolio_companies'"
            ))
            if not check.fetchone():
                return {"pe_competitive": {}}

            aesthetics_filter = """
                (pc.industry ILIKE '%aesthetic%'
                 OR pc.industry ILIKE '%medspa%'
                 OR pc.industry ILIKE '%med spa%'
                 OR pc.industry ILIKE '%dermatolog%'
                 OR pc.industry ILIKE '%cosmetic%'
                 OR pc.industry ILIKE '%skin%'
                 OR pc.industry ILIKE '%beauty%'
                 OR pc.industry ILIKE '%wellness%'
                 OR pc.sub_industry ILIKE '%aesthetic%'
                 OR pc.sub_industry ILIKE '%medspa%'
                 OR pc.sub_industry ILIKE '%med spa%'
                 OR pc.sub_industry ILIKE '%dermatolog%'
                 OR pc.sub_industry ILIKE '%cosmetic%'
                 OR pc.sub_industry ILIKE '%skin%'
                 OR pc.sub_industry ILIKE '%beauty%'
                 OR pc.sub_industry ILIKE '%wellness%')
            """
            state_clause = "AND pc.headquarters_state = :state" if state else ""
            params: Dict[str, Any] = {}
            if state:
                params["state"] = state.upper()

            # Get aesthetics PE portfolio companies with deal info
            rows = db.execute(text(f"""
                SELECT
                    pc.name,
                    pc.industry,
                    pc.sub_industry,
                    pc.headquarters_state,
                    pc.headquarters_city,
                    pc.current_pe_owner,
                    pc.employee_count,
                    pc.ownership_status,
                    pc.is_platform_company,
                    d.deal_type,
                    d.enterprise_value_usd,
                    d.ev_ebitda_multiple,
                    d.announced_date,
                    d.buyer_name,
                    d.status AS deal_status
                FROM pe_portfolio_companies pc
                LEFT JOIN pe_deals d ON d.company_id = pc.id
                WHERE {aesthetics_filter}
                    {state_clause}
                ORDER BY d.announced_date DESC NULLS LAST
                LIMIT 100
            """), params).fetchall()

            if not rows:
                return {"pe_competitive": {}}

            deals = []
            multiples = []
            total_ev = 0
            deal_types: Dict[str, int] = {}
            state_counts: Dict[str, int] = {}
            buyers: Dict[str, int] = {}

            for r in rows:
                deal_type = r[9] or "Unknown"
                ev = float(r[10] or 0)
                multiple = float(r[11] or 0)
                buyer = r[13] or r[5] or "Unknown"

                deals.append({
                    "company": r[0],
                    "industry": r[1] or "-",
                    "state": r[3] or "-",
                    "city": r[4] or "-",
                    "pe_owner": r[5] or "-",
                    "deal_type": deal_type,
                    "ev_usd": ev,
                    "ev_ebitda": multiple,
                    "date": r[12].isoformat() if r[12] else "-",
                    "buyer": buyer,
                    "is_platform": bool(r[8]),
                })

                if multiple > 0:
                    multiples.append(multiple)
                if ev > 0:
                    total_ev += ev
                deal_types[deal_type] = deal_types.get(deal_type, 0) + 1
                if r[3]:
                    state_counts[r[3]] = state_counts.get(r[3], 0) + 1
                buyers[buyer] = buyers.get(buyer, 0) + 1

            # Deal timeline by year
            year_counts: Dict[str, int] = {}
            for d in deals:
                if d["date"] != "-":
                    yr = d["date"][:4]
                    year_counts[yr] = year_counts.get(yr, 0) + 1

            avg_multiple = round(sum(multiples) / len(multiples), 1) if multiples else 0
            most_active = max(buyers, key=buyers.get) if buyers else "-"
            platforms = sum(1 for d in deals if d["is_platform"])

            return {
                "pe_competitive": {
                    "deals": deals[:50],
                    "deal_type_breakdown": deal_types,
                    "state_counts": dict(sorted(state_counts.items(), key=lambda x: x[1], reverse=True)),
                    "year_counts": dict(sorted(year_counts.items())),
                    "summary": {
                        "pe_platforms": platforms,
                        "avg_ev_ebitda": avg_multiple,
                        "total_deal_value": total_ev,
                        "most_active_buyer": most_active,
                        "total_deals": len(deals),
                    },
                }
            }
        except Exception as e:
            logger.warning(f"Could not compute PE competitive data: {e}")
            db.rollback()
            return {"pe_competitive": {}}

    # ------------------------------------------------------------------
    # Section 25: Construction Momentum Signal
    # ------------------------------------------------------------------

    def _get_construction_momentum_data(self, db: Session, state: Optional[str]) -> Dict:
        """Compute YoY building permit growth and cross with medspa density."""
        try:
            check = db.execute(text(
                "SELECT 1 FROM information_schema.tables "
                "WHERE table_name = 'realestate_hud_permits'"
            ))
            if not check.fetchone():
                return {"construction_momentum": {}}

            # Get latest two years of state-level data
            yr_rows = db.execute(text("""
                SELECT DISTINCT EXTRACT(YEAR FROM date)::int AS yr
                FROM realestate_hud_permits
                WHERE geography_type = 'State'
                ORDER BY yr DESC LIMIT 2
            """)).fetchall()
            if not yr_rows:
                return {"construction_momentum": {}}

            latest_year = yr_rows[0][0]
            prior_year = yr_rows[1][0] if len(yr_rows) > 1 else None

            state_clause = "AND hp.geography_name ILIKE :state_name" if state else ""
            params: Dict[str, Any] = {"latest_year": latest_year}
            if state:
                params["state_name"] = f"%{state}%"

            has_prior = prior_year is not None
            if has_prior:
                params["prior_year"] = prior_year

            prior_join = ""
            prior_select = ", NULL AS prior_permits"
            if has_prior:
                prior_join = """
                    LEFT JOIN (
                        SELECT geography_name,
                               SUM(permits_total) AS prior_total
                        FROM realestate_hud_permits
                        WHERE EXTRACT(YEAR FROM date) = :prior_year
                            AND geography_type = 'State'
                        GROUP BY geography_name
                    ) pp ON pp.geography_name = hp.geography_name
                """
                prior_select = ", pp.prior_total AS prior_permits"

            rows = db.execute(text(f"""
                SELECT
                    hp.geography_name,
                    SUM(hp.permits_total) AS total_permits,
                    SUM(hp.permits_1unit) AS permits_1unit,
                    SUM(COALESCE(hp.permits_2to4units, 0)) AS permits_2to4,
                    SUM(hp.permits_5plus) AS permits_5plus
                    {prior_select}
                FROM realestate_hud_permits hp
                {prior_join}
                WHERE EXTRACT(YEAR FROM date) = :latest_year
                    AND hp.geography_type = 'State'
                    {state_clause}
                GROUP BY hp.geography_name
                    {', pp.prior_total' if has_prior else ''}
                ORDER BY total_permits DESC
            """), params).fetchall()

            if not rows:
                return {"construction_momentum": {}}

            # Cross with medspa prospect counts
            medspa_counts = {}
            try:
                ms_rows = db.execute(text(
                    "SELECT state, COUNT(*) AS cnt, "
                    "COUNT(*) FILTER (WHERE acquisition_grade = 'A') AS a_cnt "
                    "FROM medspa_prospects GROUP BY state"
                )).fetchall()
                for mr in ms_rows:
                    medspa_counts[mr[0]] = {"count": int(mr[1]), "a_grade": int(mr[2])}
            except Exception:
                pass

            state_data = []
            yoy_values = []
            total_permits_all = 0
            high_growth_count = 0

            for r in rows:
                geo_name = r[0] or ""
                total_permits = int(r[1] or 0)
                p_1unit = int(r[2] or 0)
                p_2to4 = int(r[3] or 0)
                p_5plus = int(r[4] or 0)
                prior_permits = int(r[5]) if r[5] is not None else None

                yoy_growth = None
                if prior_permits is not None and prior_permits > 0:
                    yoy_growth = round(
                        (total_permits - prior_permits) / prior_permits * 100, 1
                    )
                    yoy_values.append(yoy_growth)
                    if yoy_growth > 10:
                        high_growth_count += 1

                # Try to match state abbreviation from geography_name
                # geography_name is usually full state name
                state_abbr = geo_name[:2].upper() if len(geo_name) == 2 else ""
                ms = medspa_counts.get(state_abbr, {})

                state_data.append({
                    "state_name": geo_name,
                    "state_abbr": state_abbr,
                    "total_permits": total_permits,
                    "permits_1unit": p_1unit,
                    "permits_2to4": p_2to4,
                    "permits_5plus": p_5plus,
                    "yoy_growth": yoy_growth,
                    "medspa_count": ms.get("count", 0),
                    "a_grade_count": ms.get("a_grade", 0),
                })
                total_permits_all += total_permits

            # Sort by YoY growth for display
            state_data.sort(key=lambda x: x.get("yoy_growth") or -999, reverse=True)

            avg_yoy = round(sum(yoy_values) / len(yoy_values), 1) if yoy_values else 0
            top_state = state_data[0]["state_name"] if state_data else "-"

            # Permit type composition
            total_1unit = sum(s["permits_1unit"] for s in state_data)
            total_2to4 = sum(s["permits_2to4"] for s in state_data)
            total_5plus = sum(s["permits_5plus"] for s in state_data)
            total_all = total_1unit + total_2to4 + total_5plus
            permit_composition = {}
            if total_all > 0:
                permit_composition = {
                    "single_family_pct": round(total_1unit / total_all * 100, 1),
                    "two_to_four_pct": round(total_2to4 / total_all * 100, 1),
                    "five_plus_pct": round(total_5plus / total_all * 100, 1),
                }

            return {
                "construction_momentum": {
                    "state_data": state_data,
                    "permit_composition": permit_composition,
                    "summary": {
                        "states_analyzed": len(state_data),
                        "avg_yoy_growth": avg_yoy,
                        "high_growth_states": high_growth_count,
                        "top_state": top_state,
                        "latest_year": latest_year,
                        "has_prior_year": has_prior,
                    },
                }
            }
        except Exception as e:
            logger.warning(f"Could not compute construction momentum data: {e}")
            db.rollback()
            return {"construction_momentum": {}}

    # ------------------------------------------------------------------
    # Section 26: Medical CPI Pricing Power
    # ------------------------------------------------------------------

    def _get_medical_cpi_data(self, db: Session) -> Dict:
        """Compare medical care CPI vs general CPI for pricing power analysis."""
        try:
            check = db.execute(text(
                "SELECT 1 FROM information_schema.tables "
                "WHERE table_name = 'bls_cpi'"
            ))
            if not check.fetchone():
                return {"medical_cpi": {}}

            # Query medical care (CUSR0000SAM) and all items (CUSR0000SA0)
            # Use annual average (period='M13') or December (period='M12') as fallback
            rows = db.execute(text("""
                SELECT
                    series_id,
                    year,
                    period,
                    value
                FROM bls_cpi
                WHERE series_id IN ('CUSR0000SAM', 'CUSR0000SA0')
                    AND (period = 'M13' OR period = 'M12')
                ORDER BY series_id, year DESC, period DESC
            """)).fetchall()

            if not rows:
                return {"medical_cpi": {}}

            # Organize by series and year (prefer M13 over M12)
            medical: Dict[int, float] = {}
            general: Dict[int, float] = {}

            for r in rows:
                series = r[0]
                year = int(r[1])
                value = float(r[3])
                target = medical if series == 'CUSR0000SAM' else general
                if year not in target:  # First hit wins (M13 preferred due to ORDER BY)
                    target[year] = value

            if not medical or not general:
                return {"medical_cpi": {}}

            # Compute annual comparisons for overlapping years
            common_years = sorted(set(medical.keys()) & set(general.keys()), reverse=True)
            if not common_years:
                return {"medical_cpi": {}}

            annual_data = []
            cumulative_divergence = 0
            for i, year in enumerate(common_years[:6]):  # Last 6 years
                med_val = medical[year]
                gen_val = general[year]
                spread = round(med_val - gen_val, 2)

                # YoY change
                med_yoy = None
                gen_yoy = None
                if i + 1 < len(common_years):
                    prior_year = common_years[i + 1]
                    if prior_year in medical and medical[prior_year] > 0:
                        med_yoy = round(
                            (med_val - medical[prior_year]) / medical[prior_year] * 100, 2
                        )
                    if prior_year in general and general[prior_year] > 0:
                        gen_yoy = round(
                            (gen_val - general[prior_year]) / general[prior_year] * 100, 2
                        )

                yoy_spread = round(med_yoy - gen_yoy, 2) if med_yoy is not None and gen_yoy is not None else None
                if yoy_spread is not None:
                    cumulative_divergence += yoy_spread

                annual_data.append({
                    "year": year,
                    "medical_cpi": med_val,
                    "general_cpi": gen_val,
                    "spread": spread,
                    "medical_yoy": med_yoy,
                    "general_yoy": gen_yoy,
                    "yoy_spread": yoy_spread,
                    "cumulative_divergence": round(cumulative_divergence, 2),
                })

            # Current values
            latest_year = common_years[0]
            current_medical = medical[latest_year]
            latest_med_yoy = annual_data[0].get("medical_yoy", 0) if annual_data else 0
            latest_spread = annual_data[0].get("yoy_spread", 0) if annual_data else 0

            # 5-year CAGR if available
            cagr_5yr = None
            if len(common_years) >= 6:
                start_val = medical[common_years[5]]
                end_val = medical[common_years[0]]
                if start_val > 0:
                    cagr_5yr = round(
                        ((end_val / start_val) ** (1 / 5) - 1) * 100, 2
                    )

            return {
                "medical_cpi": {
                    "annual_data": annual_data,
                    "summary": {
                        "current_medical_cpi": current_medical,
                        "yoy_medical_change": latest_med_yoy,
                        "medical_vs_general_spread": latest_spread,
                        "cagr_5yr": cagr_5yr,
                        "latest_year": latest_year,
                    },
                }
            }
        except Exception as e:
            logger.warning(f"Could not compute medical CPI data: {e}")
            db.rollback()
            return {"medical_cpi": {}}

    # ------------------------------------------------------------------
    # Section 27: Talent Pipeline Pressure
    # ------------------------------------------------------------------

    def _get_talent_pipeline_data(self, db: Session) -> Dict:
        """Analyze healthcare JOLTS data for talent scarcity signals."""
        try:
            check = db.execute(text(
                "SELECT 1 FROM information_schema.tables "
                "WHERE table_name = 'bls_jolts'"
            ))
            if not check.fetchone():
                return {"talent_pipeline": {}}

            # Healthcare openings: JTU6200000000000JOL, hires: JTU6200000000000HIR
            rows = db.execute(text("""
                SELECT
                    series_id,
                    year,
                    period,
                    value
                FROM bls_jolts
                WHERE series_id IN ('JTU6200000000000JOL', 'JTU6200000000000HIR')
                    AND period LIKE 'M%%'
                ORDER BY year DESC, period DESC
            """)).fetchall()

            if not rows:
                return {"talent_pipeline": {}}

            # Organize by series, year, and month
            openings_monthly: Dict[str, float] = {}  # "YYYY-MM" -> value
            hires_monthly: Dict[str, float] = {}

            for r in rows:
                series = r[0]
                year = int(r[1])
                period = r[2]  # M01, M02, etc.
                value = float(r[3])
                key = f"{year}-{period}"
                if series == 'JTU6200000000000JOL':
                    openings_monthly[key] = value
                else:
                    hires_monthly[key] = value

            if not openings_monthly:
                return {"talent_pipeline": {}}

            # Build quarterly data from monthly (Q = avg of 3 months)
            quarterly_data = []
            all_keys = sorted(openings_monthly.keys(), reverse=True)

            # Group by quarter
            quarter_map: Dict[str, List[float]] = {}
            quarter_hires_map: Dict[str, List[float]] = {}
            for key in all_keys:
                parts = key.split("-")
                year = parts[0]
                month = int(parts[1].replace("M", ""))
                q = (month - 1) // 3 + 1
                q_key = f"{year}-Q{q}"
                quarter_map.setdefault(q_key, []).append(openings_monthly[key])
                if key in hires_monthly:
                    quarter_hires_map.setdefault(q_key, []).append(hires_monthly[key])

            sorted_quarters = sorted(quarter_map.keys(), reverse=True)
            for q_key in sorted_quarters[:12]:  # Last 12 quarters
                avg_openings = round(sum(quarter_map[q_key]) / len(quarter_map[q_key]), 1)
                avg_hires = 0
                if q_key in quarter_hires_map and quarter_hires_map[q_key]:
                    avg_hires = round(
                        sum(quarter_hires_map[q_key]) / len(quarter_hires_map[q_key]), 1
                    )
                ratio = round(avg_openings / avg_hires, 2) if avg_hires > 0 else 0

                quarterly_data.append({
                    "period": q_key,
                    "openings": avg_openings,
                    "hires": avg_hires,
                    "ratio": ratio,
                })

            # YoY change for latest quarter
            yoy_change = None
            if len(quarterly_data) >= 5:
                current = quarterly_data[0]["openings"]
                year_ago = quarterly_data[4]["openings"]
                if year_ago > 0:
                    yoy_change = round(
                        (current - year_ago) / year_ago * 100, 1
                    )

            # Trend: compare first half ratios vs second half
            trend = "stable"
            if len(quarterly_data) >= 4:
                recent_ratios = [q["ratio"] for q in quarterly_data[:4] if q["ratio"] > 0]
                older_ratios = [q["ratio"] for q in quarterly_data[4:8] if q["ratio"] > 0]
                if recent_ratios and older_ratios:
                    recent_avg = sum(recent_ratios) / len(recent_ratios)
                    older_avg = sum(older_ratios) / len(older_ratios)
                    if recent_avg > older_avg * 1.05:
                        trend = "tightening"
                    elif recent_avg < older_avg * 0.95:
                        trend = "easing"

            latest = quarterly_data[0] if quarterly_data else {}

            return {
                "talent_pipeline": {
                    "quarterly_data": quarterly_data,
                    "summary": {
                        "latest_openings": latest.get("openings", 0),
                        "openings_to_hires_ratio": latest.get("ratio", 0),
                        "yoy_change": yoy_change,
                        "trend": trend,
                    },
                }
            }
        except Exception as e:
            logger.warning(f"Could not compute talent pipeline data: {e}")
            db.rollback()
            return {"talent_pipeline": {}}

    # ------------------------------------------------------------------
    # HTML Rendering
    # ------------------------------------------------------------------

    def render_html(self, data: Dict[str, Any]) -> str:
        """Render report as HTML using the shared design system."""
        summary = data.get("summary", {})
        prospects_by_state = data.get("prospects_by_state", [])
        grade_dist = data.get("grade_distribution", [])
        score_hist = data.get("score_histogram", [])
        top_targets = data.get("top_targets", [])
        zip_conc = data.get("zip_concentration", [])
        state_avgs = data.get("state_avg_scores", [])
        a_grade_state = data.get("a_grade_by_state", [])
        zip_affluence = data.get("zip_affluence_by_state", [])
        census_income = data.get("census_income", [])
        hi_zip_pen = data.get("high_income_zip_penetration", {})
        pe_comps = data.get("pe_comps", [])
        recent_deals = data.get("recent_deals", [])
        data_fresh = data.get("data_freshness", {})
        params = data.get("params", {})

        # New sections (8-12) data
        whitespace_zips = data.get("whitespace_zips", [])
        whitespace_by_state = data.get("whitespace_by_state", [])
        whitespace_summary = data.get("whitespace_summary", {})
        bls_wages = data.get("bls_wages", {})
        bls_employment = data.get("bls_employment", {})
        category_breakdown = data.get("category_breakdown", [])
        pe_financials = data.get("pe_financials", [])
        pe_fin_summary = data.get("pe_financial_summary", {})
        top_by_reviews = data.get("top_by_reviews", [])
        review_buckets = data.get("review_buckets", [])
        low_competition_gems = data.get("low_competition_gems", [])

        # Deal model data (sections 13-15)
        deal_model = data.get("deal_model", {})

        # Sections 16-17 data
        stealth_wealth = data.get("stealth_wealth", {})
        migration_alpha = data.get("migration_alpha", {})

        # Sections 18-21 data
        provider_density = data.get("provider_density", {})
        real_estate_alpha = data.get("real_estate_alpha", {})
        deposit_wealth = data.get("deposit_wealth", {})
        business_formation = data.get("business_formation", {})

        # Sections 22-27 data
        opportunity_zones = data.get("opportunity_zones", {})
        demographic_demand = data.get("demographic_demand", {})
        pe_competitive = data.get("pe_competitive", {})
        construction_momentum = data.get("construction_momentum", {})
        medical_cpi = data.get("medical_cpi", {})
        talent_pipeline = data.get("talent_pipeline", {})

        charts_js = ""
        body = ""

        # ---- Determine scope for header ----
        state_filter = params.get("state")
        scope_label = f" ({state_filter})" if state_filter else " (National)"

        # ---- Page Header ----
        body += page_header(
            title=f"Aesthetics Roll-Up Thesis{scope_label}",
            subtitle="Med-Spa Market Analysis \u00b7 Acquisition Target Intelligence",
            badge=f"{_fmt(summary.get('total_prospects'))} Prospects Discovered",
        )

        # ---- KPI Strip ----
        cards = ""
        cards += kpi_card("Total Prospects", _fmt(summary.get("total_prospects")))
        cards += kpi_card("A-Grade Targets", _fmt(summary.get("a_grade")))
        cards += kpi_card("States Covered", _fmt(summary.get("states_covered")))
        cards += kpi_card("Avg Acq. Score", str(summary.get("avg_score", 0)))
        cards += kpi_card("Avg Yelp Rating", f"{summary.get('avg_rating', 0):.1f}")

        body += '\n<div class="container">'
        body += "\n" + kpi_strip(cards)

        # ---- Table of Contents ----
        toc_items = [
            {"number": 1, "id": "exec-summary", "title": "Executive Summary"},
            {"number": 2, "id": "market-map", "title": "Market Map"},
            {"number": 3, "id": "top-targets", "title": "Top Acquisition Targets"},
            {"number": 4, "id": "concentration", "title": "Market Concentration Analysis"},
            {"number": 5, "id": "zip-affluence", "title": "ZIP Affluence Profile"},
            {"number": 6, "id": "competitive", "title": "Competitive Landscape"},
            {"number": 7, "id": "methodology", "title": "Data Sources & Methodology"},
            {"number": 8, "id": "whitespace", "title": "Whitespace Analysis"},
            {"number": 9, "id": "workforce", "title": "Workforce Economics"},
            {"number": 10, "id": "categories", "title": "Service Category Breakdown"},
            {"number": 11, "id": "pe-benchmarks", "title": "PE Platform Benchmarking"},
            {"number": 12, "id": "growth-signals", "title": "Review Velocity & Growth Signals"},
            {"number": 13, "id": "deal-unit-econ", "title": "Deal Model \u2014 Unit Economics"},
            {"number": 14, "id": "deal-capital", "title": "Deal Model \u2014 Capital Requirements"},
            {"number": 15, "id": "deal-returns", "title": "Deal Model \u2014 Returns Analysis"},
            {"number": 16, "id": "stealth-wealth", "title": "Stealth Wealth Signal"},
            {"number": 17, "id": "migration-alpha", "title": "Migration Alpha"},
            {"number": 18, "id": "provider-density", "title": "Medical Provider Density Signal"},
            {"number": 19, "id": "re-alpha", "title": "Real Estate Appreciation Alpha"},
            {"number": 20, "id": "deposit-wealth", "title": "Deposit Wealth Concentration"},
            {"number": 21, "id": "biz-formation", "title": "Business Formation Velocity"},
            {"number": 22, "id": "oz-overlay", "title": "Opportunity Zone Overlay"},
            {"number": 23, "id": "demographic-demand", "title": "Demographic Demand Model"},
            {"number": 24, "id": "pe-heatmap", "title": "PE Competitive Heat Map"},
            {"number": 25, "id": "construction-momentum", "title": "Construction Momentum Signal"},
            {"number": 26, "id": "medical-cpi", "title": "Medical CPI Pricing Power"},
            {"number": 27, "id": "talent-pipeline", "title": "Talent Pipeline Pressure"},
        ]
        body += "\n" + toc(toc_items)

        # ==================================================================
        # Section 1: Executive Summary
        # ==================================================================
        body += "\n" + section_start(1, "Executive Summary", "exec-summary")

        total_p = summary.get("total_prospects", 0)
        a_count = summary.get("a_grade", 0)
        ab_count = summary.get("ab_grade", 0)
        states_n = summary.get("states_covered", 0)

        body += f"""<p>Nexdata has identified <strong>{_fmt(total_p)}</strong> med-spa acquisition prospects
across <strong>{states_n}</strong> states, scored using a 5-factor weighted composite model combining
ZIP-level affluence data (IRS SOI), Yelp consumer signals, and competitive density analysis.</p>"""

        # Investment thesis box
        body += '<div class="thesis-box">'
        body += "<h3>Investment Thesis: Aesthetics Roll-Up</h3>"
        body += "<ul>"
        body += f"<li><strong>{_fmt(a_count)} A-grade</strong> and <strong>{_fmt(ab_count)} A/B-grade</strong> prospects identified as high-conviction acquisition targets</li>"

        # Top 5 states
        top5_states = prospects_by_state[:5]
        if top5_states:
            states_str = ", ".join(
                f"{s['state']} ({s['count']})" for s in top5_states
            )
            body += f"<li><strong>Top target markets:</strong> {states_str}</li>"

        body += f"<li><strong>Average acquisition score:</strong> {summary.get('avg_score', 0)}/100 across all prospects</li>"
        body += f"<li><strong>Average Yelp rating:</strong> {summary.get('avg_rating', 0):.1f} stars with {_fmt(summary.get('avg_reviews'))} avg reviews per location</li>"

        if hi_zip_pen.get("penetration_pct"):
            body += f"<li><strong>Market coverage:</strong> {hi_zip_pen['penetration_pct']:.0f}% of A-grade affluent ZIPs already have discoverable prospects</li>"

        body += "</ul></div>"

        # Grade distribution summary as metric cards
        body += '<div class="metric-grid">'
        for gd in grade_dist:
            body += f"""<div class="metric-card">
    <div class="metric-label">Grade {gd['grade']} Prospects</div>
    <div class="metric-value">{_fmt(gd['count'])}</div>
    <div class="metric-detail">{gd['pct']}% of total pipeline</div>
</div>"""
        body += "</div>"

        body += "\n" + section_end()

        # ==================================================================
        # Section 2: Market Map (Charts)
        # ==================================================================
        body += "\n" + section_start(2, "Market Map", "market-map")
        body += '<p>Visual distribution of prospects by geography, grade, and score.</p>'

        # Chart 2a: Prospects by state (horizontal bar)
        if prospects_by_state:
            state_labels = [s["state"] for s in prospects_by_state[:15]]
            state_values = [float(s["count"]) for s in prospects_by_state[:15]]
            state_bar_config = build_horizontal_bar_config(
                state_labels, state_values, dataset_label="Prospects"
            )
            state_bar_json = json.dumps(state_bar_config)
            bar_height = f"{max(len(state_labels) * 48 + 40, 200)}px"

            body += chart_container(
                "stateBarChart", state_bar_json,
                build_bar_fallback(state_labels, state_values),
                title="Prospects by State (Top 15)",
                height=bar_height,
            )
            charts_js += chart_init_js("stateBarChart", state_bar_json)

        # Chart row: Grade donut + Score histogram
        body += '<div class="chart-row">'

        # Chart 2b: Grade distribution (donut)
        if grade_dist:
            grade_labels = [f"Grade {g['grade']}" for g in grade_dist]
            grade_values = [float(g["count"]) for g in grade_dist]
            grade_colors_map = {
                "A": BLUE, "B": GREEN, "C": ORANGE, "D": GRAY, "F": RED,
            }
            grade_colors = [
                grade_colors_map.get(g["grade"], GRAY) for g in grade_dist
            ]
            donut_config = build_doughnut_config(grade_labels, grade_values, grade_colors)
            donut_json = json.dumps(donut_config)

            body += "<div>"
            body += chart_container(
                "gradeDonut", donut_json,
                build_bar_fallback(grade_labels, grade_values),
                size="medium",
                title="Grade Distribution",
            )
            charts_js += chart_init_js("gradeDonut", donut_json)
            body += build_chart_legend(
                grade_labels, grade_values, grade_colors, show_pct=True
            )
            body += "</div>"

        # Chart 2c: Score histogram (vertical bar)
        if score_hist:
            hist_labels = [h["bucket"] for h in score_hist]
            hist_values = [float(h["count"]) for h in score_hist]
            hist_config = {
                "type": "bar",
                "data": {
                    "labels": hist_labels,
                    "datasets": [{
                        "label": "Prospects",
                        "data": hist_values,
                        "backgroundColor": BLUE_LIGHT,
                        "borderWidth": 0,
                        "borderRadius": 4,
                    }],
                },
                "options": {
                    "responsive": True,
                    "maintainAspectRatio": False,
                    "plugins": {"legend": {"display": False}},
                    "scales": {
                        "x": {
                            "grid": {"display": False},
                            "ticks": {"color": "#4a5568"},
                        },
                        "y": {
                            "grid": {"color": "#edf2f7"},
                            "ticks": {"color": "#4a5568"},
                            "beginAtZero": True,
                        },
                    },
                },
            }
            hist_json = json.dumps(hist_config)

            body += "<div>"
            body += chart_container(
                "scoreHistogram", hist_json,
                build_bar_fallback(hist_labels, hist_values),
                size="medium",
                title="Score Distribution (10-pt Buckets)",
            )
            charts_js += chart_init_js("scoreHistogram", hist_json)
            body += "</div>"

        body += "</div>"  # close chart-row
        body += "\n" + section_end()

        # ==================================================================
        # Section 3: Top Acquisition Targets
        # ==================================================================
        body += "\n" + section_start(3, "Top Acquisition Targets", "top-targets")
        body += f'<p><strong>{len(top_targets)}</strong> highest-scoring A/B-grade prospects ranked by composite acquisition score.</p>'

        if top_targets:
            table_rows = []
            for i, t in enumerate(top_targets, 1):
                grade_html = _grade_badge(t["grade"])
                zip_grade_html = _grade_badge(t["zip_grade"]) if t.get("zip_grade") else "-"
                score_bar_width = min(t["score"], 100)
                score_html = (
                    f'{t["score"]:.0f}'
                    f'<span class="score-bar" style="width:{score_bar_width}px"></span>'
                )
                price_display = t.get("price") or "-"

                table_rows.append([
                    str(i),
                    f'<span class="company-name">{t["name"]}</span>',
                    t.get("city") or "-",
                    t.get("state") or "-",
                    score_html,
                    grade_html,
                    f'{t["rating"]:.1f}',
                    _fmt(t["reviews"]),
                    zip_grade_html,
                    price_display,
                ])

            body += data_table(
                headers=[
                    "#", "Name", "City", "State", "Score",
                    "Grade", "Rating", "Reviews", "ZIP Grade", "Price",
                ],
                rows=table_rows,
                numeric_columns={0, 4, 6, 7},
            )

            # Insight callout
            if top_targets:
                top_state_counts: Dict[str, int] = {}
                for t in top_targets:
                    st = t.get("state", "??")
                    top_state_counts[st] = top_state_counts.get(st, 0) + 1
                top_state = max(top_state_counts, key=top_state_counts.get)
                body += callout(
                    f"<strong>Insight:</strong> {top_state_counts[top_state]} of the top "
                    f"{len(top_targets)} targets are in <strong>{top_state}</strong>. "
                    f"Average score among top targets: <strong>"
                    f"{sum(t['score'] for t in top_targets) / len(top_targets):.1f}</strong>.",
                )
        else:
            body += callout(
                "<strong>No A/B-grade targets found.</strong> Run the med-spa discovery "
                "pipeline to populate prospects.",
                variant="warn",
            )

        body += "\n" + section_end()

        # ==================================================================
        # Section 4: Market Concentration Analysis
        # ==================================================================
        body += "\n" + section_start(4, "Market Concentration Analysis", "concentration")
        body += '<p>Understanding competition density and geographic distribution of targets.</p>'

        # 4a: Top ZIPs by prospect count
        if zip_conc:
            body += '<h3 style="font-size:15px;font-weight:600;color:var(--primary);margin:16px 0 8px">Top ZIPs by Prospect Density</h3>'

            zip_rows = []
            for z in zip_conc[:15]:
                zip_rows.append([
                    z["zip_code"],
                    z.get("city") or "-",
                    z.get("state") or "-",
                    str(z["prospect_count"]),
                    str(z["avg_score"]),
                    str(z["avg_zip_score"]),
                ])
            body += data_table(
                headers=["ZIP", "City", "State", "Prospects", "Avg Score", "ZIP Score"],
                rows=zip_rows,
                numeric_columns={3, 4, 5},
            )

        # 4b: States ranked by avg score
        if state_avgs:
            body += '<h3 style="font-size:15px;font-weight:600;color:var(--primary);margin:16px 0 8px">States Ranked by Average Acquisition Score</h3>'

            state_avg_labels = [s["state"] for s in state_avgs]
            state_avg_values = [s["avg_score"] for s in state_avgs]
            state_avg_config = build_horizontal_bar_config(
                state_avg_labels, state_avg_values, dataset_label="Avg Score"
            )
            state_avg_json = json.dumps(state_avg_config)
            sa_height = f"{max(len(state_avg_labels) * 48 + 40, 200)}px"

            body += chart_container(
                "stateAvgChart", state_avg_json,
                build_bar_fallback(state_avg_labels, state_avg_values),
                title="Average Acquisition Score by State",
                height=sa_height,
            )
            charts_js += chart_init_js("stateAvgChart", state_avg_json)

        # 4c: A-grade concentration
        if a_grade_state:
            body += '<h3 style="font-size:15px;font-weight:600;color:var(--primary);margin:16px 0 8px">A-Grade Concentration by State</h3>'

            ag_rows = []
            for a in a_grade_state:
                ag_rows.append([
                    a["state"],
                    str(a["a_count"]),
                    str(a["total"]),
                    f'{a["a_pct"]}%',
                ])
            body += data_table(
                headers=["State", "A-Grade", "Total", "A-Grade %"],
                rows=ag_rows,
                numeric_columns={1, 2, 3},
            )

        body += "\n" + section_end()

        # ==================================================================
        # Section 5: ZIP Affluence Profile
        # ==================================================================
        body += "\n" + section_start(5, "ZIP Affluence Profile", "zip-affluence")
        body += '<p>ZIP-level affluence analysis from IRS SOI income data, correlated with med-spa market potential.</p>'

        # 5a: ZIP affluence by state
        if zip_affluence:
            body += '<h3 style="font-size:15px;font-weight:600;color:var(--primary);margin:16px 0 8px">Average ZIP Affluence Score by State (Top 15)</h3>'

            za_rows = []
            for z in zip_affluence:
                za_rows.append([
                    z["state"],
                    str(z["zip_count"]),
                    str(z["avg_score"]),
                    str(z["avg_affluence"]),
                    _fmt_currency(z["avg_agi"]),
                    str(z["a_zips"]),
                ])
            body += data_table(
                headers=["State", "ZIPs Scored", "Avg Score", "Affluence Score", "Avg AGI", "A-Grade ZIPs"],
                rows=za_rows,
                numeric_columns={1, 2, 3, 4, 5},
            )

        # 5b: Income correlation
        if census_income:
            body += '<h3 style="font-size:15px;font-weight:600;color:var(--primary);margin:16px 0 8px">Income vs. Med-Spa Score Correlation</h3>'

            inc_labels = [c["state"] for c in census_income]
            inc_values = [c["avg_zip_score"] for c in census_income]
            inc_config = build_horizontal_bar_config(
                inc_labels, inc_values, dataset_label="Avg ZIP Score"
            )
            inc_json = json.dumps(inc_config)
            inc_height = f"{max(len(inc_labels) * 48 + 40, 200)}px"

            body += chart_container(
                "incomeChart", inc_json,
                build_bar_fallback(inc_labels, inc_values),
                title="Average ZIP Med-Spa Score by State",
                height=inc_height,
            )
            charts_js += chart_init_js("incomeChart", inc_json)

        # 5c: High-income ZIP penetration
        if hi_zip_pen.get("total_a_zips"):
            body += '<h3 style="font-size:15px;font-weight:600;color:var(--primary);margin:16px 0 8px">High-Income ZIP Penetration</h3>'
            body += '<div class="metric-grid">'
            body += f"""<div class="metric-card">
    <div class="metric-label">A-Grade Affluent ZIPs</div>
    <div class="metric-value">{_fmt(hi_zip_pen['total_a_zips'])}</div>
    <div class="metric-detail">ZIPs scoring 80+ on affluence model</div>
</div>"""
            body += f"""<div class="metric-card">
    <div class="metric-label">ZIPs with Prospects</div>
    <div class="metric-value">{_fmt(hi_zip_pen['zips_with_prospects'])}</div>
    <div class="metric-detail">A-grade ZIPs where we found med-spas</div>
</div>"""
            body += f"""<div class="metric-card">
    <div class="metric-label">Market Penetration</div>
    <div class="metric-value">{hi_zip_pen['penetration_pct']:.0f}%</div>
    <div class="metric-detail">Coverage of top-affluence ZIP codes</div>
</div>"""
            body += "</div>"

            if hi_zip_pen["penetration_pct"] < 50:
                body += callout(
                    f"<strong>Opportunity:</strong> Only {hi_zip_pen['penetration_pct']:.0f}% of "
                    f"A-grade affluent ZIPs have discoverable med-spa businesses. "
                    f"This represents <strong>{hi_zip_pen['total_a_zips'] - hi_zip_pen['zips_with_prospects']}</strong> "
                    f"underserved high-income markets.",
                    variant="good",
                )
            else:
                body += callout(
                    f"<strong>Market Maturity:</strong> {hi_zip_pen['penetration_pct']:.0f}% of "
                    f"A-grade affluent ZIPs already have med-spa businesses. "
                    f"Focus on acquisition quality over greenfield expansion.",
                )

        body += "\n" + section_end()

        # ==================================================================
        # Section 6: Competitive Landscape
        # ==================================================================
        body += "\n" + section_start(6, "Competitive Landscape", "competitive")
        body += '<p>PE-backed aesthetics platforms and recent deal activity in the sector.</p>'

        # 6a: PE platform comps
        if pe_comps:
            body += '<h3 style="font-size:15px;font-weight:600;color:var(--primary);margin:16px 0 8px">PE-Backed Aesthetics Platforms</h3>'

            comp_rows = []
            for c in pe_comps:
                owner = c.get("pe_owner") or "-"
                location = ", ".join(
                    p for p in [c.get("city"), c.get("state")] if p
                ) or "-"
                emp = _fmt(c.get("employees")) if c.get("employees") else "-"
                status_badge = pill_badge(
                    c.get("status") or "Active",
                    "public" if c.get("status") == "Active" else "default",
                )
                comp_rows.append([
                    f'<span class="company-name">{c["name"]}</span>',
                    c.get("industry") or "-",
                    location,
                    owner,
                    emp,
                    status_badge,
                ])

            body += data_table(
                headers=["Company", "Industry", "Location", "PE Owner", "Employees", "Status"],
                rows=comp_rows,
            )

            body += callout(
                f"<strong>Competitive context:</strong> {len(pe_comps)} PE-backed platforms "
                f"identified in aesthetics/med-spa sector. Roll-up activity signals "
                f"strong institutional interest in the space.",
            )
        else:
            body += callout(
                "<strong>No PE-backed aesthetics companies found in database.</strong> "
                "Run PE portfolio collection with aesthetics sector filter to populate comps.",
                variant="warn",
            )

        # 6b: Recent deals
        if recent_deals:
            body += '<h3 style="font-size:15px;font-weight:600;color:var(--primary);margin:16px 0 8px">Recent Deal Activity</h3>'

            deal_rows = []
            for d in recent_deals:
                ev_display = _fmt_currency(d.get("ev")) if d.get("ev") else "-"
                mult_display = f'{d["ev_ebitda"]:.1f}x' if d.get("ev_ebitda") else "-"
                date_display = d.get("announced_date") or d.get("closed_date") or "-"
                deal_rows.append([
                    f'<span class="company-name">{d.get("company") or d.get("deal_name") or "N/A"}</span>',
                    d.get("deal_type") or "-",
                    date_display,
                    ev_display,
                    mult_display,
                ])

            body += data_table(
                headers=["Company", "Deal Type", "Date", "EV", "EV/EBITDA"],
                rows=deal_rows,
                numeric_columns={3, 4},
            )

        # Market consolidation trends narrative
        body += '<h3 style="font-size:15px;font-weight:600;color:var(--primary);margin:16px 0 8px">Market Consolidation Trends</h3>'
        body += """<div class="thesis-box">
<h3>Why Aesthetics is Attractive for PE Roll-Up</h3>
<ul>
    <li><strong>Fragmented market:</strong> The US aesthetics market is estimated at $20B+ with thousands of independent operators, creating ample buy-and-build opportunity</li>
    <li><strong>Recurring revenue:</strong> Med-spas benefit from membership models, repeat visit patterns, and consumable product revenue (Botox, fillers, skincare)</li>
    <li><strong>Demographic tailwinds:</strong> Growing demand across age cohorts, male demographics, and geographic markets driven by social media normalization</li>
    <li><strong>Operating leverage:</strong> Multi-unit platforms achieve procurement savings (15-25%), shared marketing, and provider network effects</li>
    <li><strong>Multiple arbitrage:</strong> Single-unit med-spas trade at 3-5x EBITDA; scaled platforms command 8-12x, creating immediate value on consolidation</li>
</ul>
</div>"""

        body += "\n" + section_end()

        # ==================================================================
        # Section 7: Data Sources & Methodology
        # ==================================================================
        body += "\n" + section_start(7, "Data Sources & Methodology", "methodology")

        # 7a: Scoring model explanation
        body += '<h3 style="font-size:15px;font-weight:600;color:var(--primary);margin:16px 0 8px">Acquisition Prospect Scoring Model</h3>'
        body += '<p>Each med-spa prospect receives a composite score (0-100) based on five weighted factors:</p>'

        # Weight visualization
        for factor, weight in PROSPECT_WEIGHTS.items():
            label = factor.replace("_", " ").title()
            pct = int(weight * 100)
            body += f"""<div class="weight-row">
    <span class="weight-label">{label}</span>
    <div class="weight-bar-track"><div class="weight-bar-fill" style="width:{pct * 3}px"></div></div>
    <span class="weight-pct">{pct}%</span>
</div>"""

        body += '<p style="margin-top:16px">Scores are converted to letter grades using standard thresholds:</p>'

        grade_table_rows = [
            ["A", "80-100", "Top-tier acquisition target"],
            ["B", "65-79", "Strong prospect"],
            ["C", "50-64", "Moderate potential"],
            ["D", "35-49", "Below-average prospect"],
            ["F", "0-34", "Not recommended"],
        ]
        body += data_table(
            headers=["Grade", "Score Range", "Interpretation"],
            rows=grade_table_rows,
        )

        # ZIP affluence model
        body += '<h3 style="font-size:15px;font-weight:600;color:var(--primary);margin:16px 0 8px">ZIP Revenue Potential Score</h3>'
        body += '<p>Each US ZIP code is scored (0-100) for med-spa revenue potential using five IRS SOI-derived signals:</p>'

        for factor, weight in ZIP_SCORE_WEIGHTS.items():
            label = factor.replace("_", " ").title()
            pct = int(weight * 100)
            body += f"""<div class="weight-row">
    <span class="weight-label">{label}</span>
    <div class="weight-bar-track"><div class="weight-bar-fill" style="width:{pct * 3}px"></div></div>
    <span class="weight-pct">{pct}%</span>
</div>"""

        # 7b: Data freshness
        body += '<h3 style="font-size:15px;font-weight:600;color:var(--primary);margin:16px 0 8px">Data Freshness</h3>'

        freshness_rows = []
        if data_fresh.get("prospects"):
            p = data_fresh["prospects"]
            freshness_rows.append([
                "Med-Spa Prospects",
                "Yelp Business Search API",
                _fmt(p.get("total")),
                p.get("latest", "-"),
            ])
        if data_fresh.get("zip_scores"):
            z = data_fresh["zip_scores"]
            freshness_rows.append([
                "ZIP Affluence Scores",
                "IRS SOI ZIP Income Data (2021 tax year)",
                _fmt(z.get("total")),
                z.get("latest", "-"),
            ])

        # Always show these static rows
        freshness_rows.append([
            "PE Portfolio Companies",
            "SEC EDGAR, Company Websites",
            "-",
            "-",
        ])
        freshness_rows.append([
            "Census ACS Income",
            "US Census Bureau ACS 5-Year",
            "-",
            "-",
        ])
        if data_fresh.get("irs_soi_zip_income"):
            iz = data_fresh["irs_soi_zip_income"]
            freshness_rows.append([
                "IRS SOI ZIP Income (Stealth Wealth)",
                f"IRS SOI, tax years {iz.get('earliest', '?')}-{iz.get('latest', '?')}",
                _fmt(iz.get("total")),
                iz.get("latest", "-"),
            ])
        if data_fresh.get("irs_soi_migration"):
            im = data_fresh["irs_soi_migration"]
            freshness_rows.append([
                "IRS SOI Migration (Migration Alpha)",
                f"IRS SOI, tax years {im.get('earliest', '?')}-{im.get('latest', '?')}",
                _fmt(im.get("total")),
                im.get("latest", "-"),
            ])
        if data_fresh.get("cms_medicare"):
            cm = data_fresh["cms_medicare"]
            freshness_rows.append([
                "CMS Medicare Utilization (Provider Density)",
                "CMS Medicare Provider Utilization",
                _fmt(cm.get("total")),
                "-",
            ])
        if data_fresh.get("redfin"):
            rf = data_fresh["redfin"]
            freshness_rows.append([
                "Redfin Real Estate (RE Alpha)",
                "Redfin Housing Market Data",
                _fmt(rf.get("total")),
                rf.get("latest", "-"),
            ])
        if data_fresh.get("fhfa_hpi"):
            fh = data_fresh["fhfa_hpi"]
            freshness_rows.append([
                "FHFA HPI (RE Alpha fallback)",
                f"FHFA House Price Index, {fh.get('earliest', '?')}-{fh.get('latest', '?')}",
                _fmt(fh.get("total")),
                fh.get("latest", "-"),
            ])
        if data_fresh.get("fdic_deposits"):
            fd = data_fresh["fdic_deposits"]
            freshness_rows.append([
                "FDIC Summary Deposits (Deposit Wealth)",
                f"FDIC SOD, years {fd.get('earliest', '?')}-{fd.get('latest', '?')}",
                _fmt(fd.get("total")),
                fd.get("latest", "-"),
            ])
        if data_fresh.get("irs_soi_business_income"):
            ib = data_fresh["irs_soi_business_income"]
            freshness_rows.append([
                "IRS SOI Business Income (Biz Formation)",
                f"IRS SOI, tax years {ib.get('earliest', '?')}-{ib.get('latest', '?')}",
                _fmt(ib.get("total")),
                ib.get("latest", "-"),
            ])
        if data_fresh.get("opportunity_zone"):
            oz = data_fresh["opportunity_zone"]
            freshness_rows.append([
                "Opportunity Zones (OZ Overlay)",
                "CDFI Fund / Treasury Dept",
                _fmt(oz.get("total")),
                "-",
            ])
        if data_fresh.get("educational_attainment"):
            ea = data_fresh["educational_attainment"]
            freshness_rows.append([
                "Educational Attainment (Demographic Demand)",
                f"Census ACS, years {ea.get('earliest', '?')}-{ea.get('latest', '?')}",
                _fmt(ea.get("total")),
                ea.get("latest", "-"),
            ])
        if data_fresh.get("hud_permits"):
            hp = data_fresh["hud_permits"]
            freshness_rows.append([
                "HUD Building Permits (Construction Momentum)",
                "HUD State of the Cities Data Systems",
                _fmt(hp.get("total")),
                hp.get("latest", "-"),
            ])
        if data_fresh.get("bls_cpi"):
            bc = data_fresh["bls_cpi"]
            freshness_rows.append([
                "BLS CPI (Medical Pricing Power)",
                f"BLS Consumer Price Index, {bc.get('earliest', '?')}-{bc.get('latest', '?')}",
                _fmt(bc.get("total")),
                bc.get("latest", "-"),
            ])

        body += data_table(
            headers=["Dataset", "Source", "Records", "Last Updated"],
            rows=freshness_rows,
        )

        # 7c: Coverage statistics
        body += '<h3 style="font-size:15px;font-weight:600;color:var(--primary);margin:16px 0 8px">Coverage Statistics</h3>'
        body += '<div class="metric-grid">'
        body += f"""<div class="metric-card">
    <div class="metric-label">Prospects Discovered</div>
    <div class="metric-value">{_fmt(summary.get('total_prospects'))}</div>
    <div class="metric-detail">Unique med-spa locations from Yelp</div>
</div>"""
        body += f"""<div class="metric-card">
    <div class="metric-label">ZIPs Covered</div>
    <div class="metric-value">{_fmt(summary.get('zips_covered'))}</div>
    <div class="metric-detail">Unique ZIP codes with prospects</div>
</div>"""
        body += f"""<div class="metric-card">
    <div class="metric-label">Total Consumer Reviews</div>
    <div class="metric-value">{_fmt(summary.get('total_reviews'))}</div>
    <div class="metric-detail">Aggregate Yelp reviews analyzed</div>
</div>"""
        body += "</div>"

        body += "\n" + section_end()

        # ==================================================================
        # Section 8: Whitespace Analysis
        # ==================================================================
        body += "\n" + section_start(8, "Whitespace Analysis", "whitespace")
        body += '<p>Grade-A affluent ZIPs with zero discovered med-spa prospects — greenfield acquisition opportunities.</p>'

        ws_total = whitespace_summary.get("total_a_zips", 0)
        ws_with = whitespace_summary.get("a_with_prospects", 0)
        ws_count = whitespace_summary.get("whitespace_count", 0)

        if ws_total > 0:
            # KPI cards
            body += '<div class="metric-grid">'
            body += f"""<div class="metric-card">
    <div class="metric-label">Total A-Grade ZIPs</div>
    <div class="metric-value">{_fmt(ws_total)}</div>
    <div class="metric-detail">ZIPs scoring 80+ on affluence model</div>
</div>"""
            body += f"""<div class="metric-card">
    <div class="metric-label">A-Grade w/ Prospects</div>
    <div class="metric-value">{_fmt(ws_with)}</div>
    <div class="metric-detail">Already have discoverable med-spas</div>
</div>"""
            body += f"""<div class="metric-card">
    <div class="metric-label">Whitespace ZIPs</div>
    <div class="metric-value" style="color:{GREEN}">{_fmt(ws_count)}</div>
    <div class="metric-detail">Greenfield opportunity zones</div>
</div>"""
            body += "</div>"

            # Data table: top 25 whitespace ZIPs
            if whitespace_zips:
                ws_rows = []
                for z in whitespace_zips[:25]:
                    ws_rows.append([
                        z["zip_code"],
                        z.get("state") or "-",
                        f'{z["score"]:.1f}',
                        _fmt_currency(z["avg_agi"]),
                        _fmt(z["total_returns"]),
                        f'{z["affluence_score"]:.1f}',
                    ])
                body += data_table(
                    headers=["ZIP", "State", "Score", "Avg AGI", "Tax Returns", "Affluence Score"],
                    rows=ws_rows,
                    numeric_columns={2, 3, 4, 5},
                )

            # Horizontal bar: whitespace by state (top 15)
            if whitespace_by_state:
                ws_state_labels = [s["state"] for s in whitespace_by_state[:15]]
                ws_state_values = [float(s["count"]) for s in whitespace_by_state[:15]]
                ws_bar_config = build_horizontal_bar_config(
                    ws_state_labels, ws_state_values, dataset_label="Whitespace ZIPs"
                )
                ws_bar_json = json.dumps(ws_bar_config)
                ws_bar_height = f"{max(len(ws_state_labels) * 48 + 40, 200)}px"

                body += chart_container(
                    "whitespaceStateChart", ws_bar_json,
                    build_bar_fallback(ws_state_labels, ws_state_values),
                    title="Whitespace ZIPs by State (Top 15)",
                    height=ws_bar_height,
                )
                charts_js += chart_init_js("whitespaceStateChart", ws_bar_json)

            body += callout(
                f"<strong>Opportunity:</strong> {_fmt(ws_count)} high-affluence A-grade ZIPs have "
                f"no discoverable med-spa businesses — these represent greenfield targets for "
                f"de novo clinic launches or franchise expansion.",
                variant="good",
            )
        else:
            body += callout(
                "<strong>No whitespace data available.</strong> Run ZIP affluence scoring "
                "and med-spa discovery to populate this analysis.",
                variant="warn",
            )

        body += "\n" + section_end()

        # ==================================================================
        # Section 9: Workforce Economics
        # ==================================================================
        body += "\n" + section_start(9, "Workforce Economics", "workforce")
        body += '<p>BLS Occupational Employment & Wage Statistics for aesthetics-adjacent roles — labor cost trends that impact unit economics.</p>'

        if bls_wages:
            # Collect all years across all occupations
            all_years = sorted(set(yr for occ_data in bls_wages.values() for yr in occ_data.keys()))

            # Wage table: occupation × year matrix
            wage_rows = []
            for occ in sorted(bls_wages.keys()):
                yr_data = bls_wages[occ]
                row_vals = [occ]
                for yr in all_years:
                    val = yr_data.get(yr)
                    row_vals.append(_fmt_currency(val) if val else "-")
                # Growth column: first year → last year
                first_val = float(yr_data.get(all_years[0], 0) or 0) if all_years else 0
                last_val = float(yr_data.get(all_years[-1], 0) or 0) if all_years else 0
                if first_val > 0 and last_val > 0:
                    growth = (last_val - first_val) / first_val * 100
                    row_vals.append(f"{growth:+.1f}%")
                else:
                    row_vals.append("-")
                wage_rows.append(row_vals)

            wage_headers = ["Occupation"] + [str(y) for y in all_years] + ["Growth"]
            body += data_table(
                headers=wage_headers,
                rows=wage_rows,
                numeric_columns=set(range(1, len(wage_headers))),
            )

            # Horizontal bar: latest year wages per occupation
            if all_years:
                latest_yr = all_years[-1]
                bar_occs = []
                bar_wages = []
                for occ in sorted(bls_wages.keys()):
                    val = bls_wages[occ].get(latest_yr)
                    if val:
                        bar_occs.append(occ)
                        bar_wages.append(val)
                if bar_occs:
                    wage_bar_config = build_horizontal_bar_config(
                        bar_occs, bar_wages, dataset_label=f"Annual Mean Wage ({latest_yr})"
                    )
                    wage_bar_json = json.dumps(wage_bar_config)
                    wage_bar_height = f"{max(len(bar_occs) * 48 + 40, 200)}px"

                    body += chart_container(
                        "wageBarChart", wage_bar_json,
                        build_bar_fallback(bar_occs, [float(v) for v in bar_wages]),
                        title=f"Annual Mean Wage by Occupation ({latest_yr})",
                        height=wage_bar_height,
                    )
                    charts_js += chart_init_js("wageBarChart", wage_bar_json)

            # Compute insight for NPs and PAs
            np_growth = None
            pa_growth = None
            for occ in bls_wages:
                yr_data = bls_wages[occ]
                fv = float(yr_data.get(all_years[0], 0) or 0) if all_years else 0
                lv = float(yr_data.get(all_years[-1], 0) or 0) if all_years else 0
                if fv > 0 and lv > 0:
                    g = (lv - fv) / fv * 100
                    if "Nurse Practitioner" in occ:
                        np_growth = g
                    elif "Physician Assistant" in occ:
                        pa_growth = g

            insight_parts = []
            if np_growth is not None:
                insight_parts.append(f"Nurse Practitioners ({np_growth:+.1f}%)")
            if pa_growth is not None:
                insight_parts.append(f"Physician Assistants ({pa_growth:+.1f}%)")
            if insight_parts:
                body += callout(
                    f"<strong>Labor Cost Signal:</strong> Key mid-level providers — "
                    f"{' and '.join(insight_parts)} — show significant wage growth, "
                    f"signaling rising talent competition that impacts med-spa unit economics.",
                )
            else:
                body += callout(
                    "<strong>Labor Cost Signal:</strong> BLS wage data shows multi-year "
                    "trends in aesthetics-adjacent occupations. Monitor provider wages "
                    "as a key input to unit economics modeling.",
                )
        else:
            body += callout(
                "<strong>No BLS wage data available.</strong> Ingest BLS OES data to "
                "populate workforce economics analysis.",
                variant="warn",
            )

        body += "\n" + section_end()

        # ==================================================================
        # Section 10: Service Category Breakdown
        # ==================================================================
        body += "\n" + section_start(10, "Service Category Breakdown", "categories")
        body += '<p>Disaggregation of discovered prospects by Yelp business category — reveals service mix and niche positioning.</p>'

        if category_breakdown:
            total_cat_count = sum(c["count"] for c in category_breakdown)

            # Doughnut chart: top 8 categories
            top_cats = category_breakdown[:8]
            other_count = sum(c["count"] for c in category_breakdown[8:])
            cat_labels = [c["category"] for c in top_cats]
            cat_values = [float(c["count"]) for c in top_cats]
            if other_count > 0:
                cat_labels.append("Other")
                cat_values.append(float(other_count))

            cat_colors = list(CHART_COLORS[:len(cat_labels)])
            donut_config = build_doughnut_config(cat_labels, cat_values, cat_colors)
            donut_json = json.dumps(donut_config)

            body += '<div class="chart-row">'
            body += "<div>"
            body += chart_container(
                "categoryDonut", donut_json,
                build_bar_fallback(cat_labels, cat_values),
                size="medium",
                title="Category Distribution",
            )
            charts_js += chart_init_js("categoryDonut", donut_json)
            body += build_chart_legend(cat_labels, cat_values, cat_colors, show_pct=True)
            body += "</div>"
            body += "</div>"

            # Data table
            cat_rows = []
            for c in category_breakdown[:20]:
                pct = round(c["count"] / total_cat_count * 100, 1) if total_cat_count > 0 else 0
                cat_rows.append([
                    c["category"],
                    _fmt(c["count"]),
                    f"{pct}%",
                    f'{c["avg_score"]:.1f}',
                    f'{c["avg_rating"]:.1f}',
                    _fmt(c["avg_reviews"]),
                ])
            body += data_table(
                headers=["Category", "Count", "% Share", "Avg Score", "Avg Rating", "Avg Reviews"],
                rows=cat_rows,
                numeric_columns={1, 2, 3, 4, 5},
            )

            # Insight
            top_cat = category_breakdown[0] if category_breakdown else None
            if top_cat:
                body += callout(
                    f"<strong>Insight:</strong> <strong>{top_cat['category']}</strong> is the dominant "
                    f"category with {_fmt(top_cat['count'])} prospects "
                    f"({round(top_cat['count'] / total_cat_count * 100, 1)}% share). "
                    f"Niche categories may represent differentiated positioning or "
                    f"underserved sub-segments worth investigating.",
                )
        else:
            body += callout(
                "<strong>No category data available.</strong> Prospect categories are populated "
                "during med-spa discovery.",
                variant="warn",
            )

        body += "\n" + section_end()

        # ==================================================================
        # Section 11: PE Platform Benchmarking
        # ==================================================================
        body += "\n" + section_start(11, "PE Platform Benchmarking", "pe-benchmarks")
        body += '<p>Financial benchmarks for PE-backed aesthetics platforms — revenue, margins, and leverage metrics.</p>'

        if pe_financials:
            # Metric cards
            body += '<div class="metric-grid">'
            avg_em = pe_fin_summary.get("avg_ebitda_margin")
            avg_rg = pe_fin_summary.get("avg_rev_growth")
            med_de = pe_fin_summary.get("median_debt_ebitda")
            body += f"""<div class="metric-card">
    <div class="metric-label">Avg EBITDA Margin</div>
    <div class="metric-value">{f'{avg_em:.1f}%' if avg_em is not None else '-'}</div>
    <div class="metric-detail">Across PE-backed platforms</div>
</div>"""
            body += f"""<div class="metric-card">
    <div class="metric-label">Avg Revenue Growth</div>
    <div class="metric-value">{f'{avg_rg:+.1f}%' if avg_rg is not None else '-'}</div>
    <div class="metric-detail">Year-over-year</div>
</div>"""
            body += f"""<div class="metric-card">
    <div class="metric-label">Median Debt/EBITDA</div>
    <div class="metric-value">{f'{med_de:.1f}x' if med_de is not None else '-'}</div>
    <div class="metric-detail">Leverage benchmark</div>
</div>"""
            body += "</div>"

            # Data table
            pef_rows = []
            for f in pe_financials:
                rev_display = _fmt_currency(f["revenue"]) if f["revenue"] else "-"
                rg_display = f'{f["rev_growth"]:+.1f}%' if f["rev_growth"] is not None else "-"
                em_display = f'{f["ebitda_margin"]:.1f}%' if f["ebitda_margin"] is not None else "-"
                emp_display = _fmt(f["employees"]) if f["employees"] else "-"
                de_display = f'{f["debt_ebitda"]:.1f}x' if f["debt_ebitda"] is not None else "-"
                conf_badge = pill_badge(f["confidence"], "public" if f["confidence"] == "high" else "default")
                pef_rows.append([
                    f'<span class="company-name">{f["name"]}</span>',
                    f.get("pe_owner") or "-",
                    rev_display,
                    rg_display,
                    em_display,
                    emp_display,
                    de_display,
                    conf_badge,
                ])
            body += data_table(
                headers=["Company", "PE Owner", "Revenue", "Rev Growth", "EBITDA Margin",
                         "Employees", "Debt/EBITDA", "Confidence"],
                rows=pef_rows,
                numeric_columns={2, 3, 4, 5, 6},
            )

            # Revenue bar chart (top 8)
            rev_companies = [f for f in pe_financials if f["revenue"]][:8]
            if rev_companies:
                rev_labels = [f["name"] for f in rev_companies]
                rev_values = [float(f["revenue"]) for f in rev_companies]
                rev_bar_config = build_horizontal_bar_config(
                    rev_labels, rev_values, dataset_label="Revenue (USD)"
                )
                rev_bar_json = json.dumps(rev_bar_config)
                rev_bar_height = f"{max(len(rev_labels) * 48 + 40, 200)}px"

                body += chart_container(
                    "peRevenueChart", rev_bar_json,
                    build_bar_fallback(rev_labels, rev_values),
                    title="Revenue by Platform (Top 8)",
                    height=rev_bar_height,
                )
                charts_js += chart_init_js("peRevenueChart", rev_bar_json)

            body += callout(
                "<strong>Data Confidence Note:</strong> Financial data for private PE-backed "
                "platforms is estimated based on industry analysis, employee counts, and "
                "comparable transactions. Treat as directional benchmarks, not audited figures.",
                variant="warn",
            )
        else:
            body += callout(
                "<strong>No PE financial data available.</strong> Seed PE company financials "
                "to populate benchmarking analysis.",
                variant="warn",
            )

        body += "\n" + section_end()

        # ==================================================================
        # Section 12: Review Velocity & Growth Signals
        # ==================================================================
        body += "\n" + section_start(12, "Review Velocity & Growth Signals", "growth-signals")
        body += '<p>Demand momentum signals via review volume, competitive dynamics, and market positioning.</p>'

        if top_by_reviews:
            # Data table: top 25 by review volume
            rev_rows = []
            for t in top_by_reviews:
                grade_html = _grade_badge(t["grade"])
                rev_rows.append([
                    f'<span class="company-name">{t["name"]}</span>',
                    t.get("city") or "-",
                    t.get("state") or "-",
                    _fmt(t["reviews"]),
                    f'{t["rating"]:.1f}',
                    f'{t["score"]:.0f}',
                    grade_html,
                    str(t["competitors"]),
                ])
            body += '<h3 style="font-size:15px;font-weight:600;color:var(--primary);margin:16px 0 8px">Top 25 by Review Volume (A/B Grade)</h3>'
            body += data_table(
                headers=["Name", "City", "State", "Reviews", "Rating", "Score", "Grade", "Competitors"],
                rows=rev_rows,
                numeric_columns={3, 4, 5, 7},
            )

        # Review distribution chart
        if review_buckets:
            bucket_labels = [b["bucket"] for b in review_buckets]
            bucket_values = [float(b["count"]) for b in review_buckets]
            bucket_config = {
                "type": "bar",
                "data": {
                    "labels": bucket_labels,
                    "datasets": [{
                        "label": "Prospects",
                        "data": bucket_values,
                        "backgroundColor": BLUE_LIGHT,
                        "borderWidth": 0,
                        "borderRadius": 4,
                    }],
                },
                "options": {
                    "responsive": True,
                    "maintainAspectRatio": False,
                    "plugins": {"legend": {"display": False}},
                    "scales": {
                        "x": {"grid": {"display": False}, "ticks": {"color": "#4a5568"}},
                        "y": {"grid": {"color": "#edf2f7"}, "ticks": {"color": "#4a5568"}, "beginAtZero": True},
                    },
                },
            }
            bucket_json = json.dumps(bucket_config)

            body += chart_container(
                "reviewBucketChart", bucket_json,
                build_bar_fallback(bucket_labels, bucket_values),
                title="Review Volume Distribution",
            )
            charts_js += chart_init_js("reviewBucketChart", bucket_json)

        # Low competition gems
        if low_competition_gems:
            body += '<h3 style="font-size:15px;font-weight:600;color:var(--primary);margin:16px 0 8px">Low Competition Gems (A/B Grade, \u22643 Competitors)</h3>'
            gem_rows = []
            for g in low_competition_gems:
                grade_html = _grade_badge(g["grade"])
                gem_rows.append([
                    f'<span class="company-name">{g["name"]}</span>',
                    g.get("city") or "-",
                    g.get("state") or "-",
                    f'{g["score"]:.0f}',
                    grade_html,
                    f'{g["rating"]:.1f}',
                    _fmt(g["reviews"]),
                    str(g["competitors"]),
                ])
            body += data_table(
                headers=["Name", "City", "State", "Score", "Grade", "Rating", "Reviews", "Competitors"],
                rows=gem_rows,
                numeric_columns={3, 5, 6, 7},
            )

        # Insight callout
        high_review_count = sum(1 for t in top_by_reviews if t["reviews"] >= 200)
        gem_count = len(low_competition_gems)
        if top_by_reviews or low_competition_gems:
            body += callout(
                f"<strong>Demand Signals:</strong> {high_review_count} A/B-grade prospects have 200+ "
                f"reviews, indicating established consumer demand and brand recognition. "
                f"{gem_count} prospects are A/B-grade with 3 or fewer competitors in their ZIP — "
                f"prime bolt-on acquisition targets with limited competitive pressure.",
                variant="good",
            )
        else:
            body += callout(
                "<strong>No growth signal data available.</strong> Run med-spa discovery "
                "to populate review and competition analysis.",
                variant="warn",
            )

        body += "\n" + section_end()

        # ==================================================================
        # Section 13: Deal Model — Unit Economics
        # ==================================================================
        body += "\n" + section_start(13, "Deal Model \u2014 Unit Economics", "deal-unit-econ")
        body += '<p>Estimated portfolio economics for all A-grade acquisition targets using industry benchmark revenue and margins by Yelp price tier.</p>'

        dm_tiers = deal_model.get("tier_economics", [])
        dm_total_loc = deal_model.get("total_locations", 0)
        dm_total_rev = deal_model.get("total_revenue", 0)
        dm_total_ebitda = deal_model.get("total_ebitda", 0)
        dm_w_margin = deal_model.get("weighted_margin", 0)
        dm_pnl = deal_model.get("pnl_waterfall", {})

        if dm_tiers:
            # KPI cards
            body += '<div class="metric-grid">'
            body += f"""<div class="metric-card">
    <div class="metric-label">Total A-Grade Locations</div>
    <div class="metric-value">{_fmt(dm_total_loc)}</div>
    <div class="metric-detail">Across all price tiers</div>
</div>"""
            body += f"""<div class="metric-card">
    <div class="metric-label">Est. Portfolio Revenue</div>
    <div class="metric-value">{_fmt_currency(dm_total_rev)}</div>
    <div class="metric-detail">Based on industry benchmarks</div>
</div>"""
            body += f"""<div class="metric-card">
    <div class="metric-label">Est. Portfolio EBITDA</div>
    <div class="metric-value">{_fmt_currency(dm_total_ebitda)}</div>
    <div class="metric-detail">Weighted avg margin: {dm_w_margin:.1%}</div>
</div>"""
            body += "</div>"

            # Price tier distribution table
            body += '<h3 style="font-size:15px;font-weight:600;color:var(--primary);margin:16px 0 8px">Revenue Estimates by Price Tier</h3>'

            tier_rows = []
            for t in dm_tiers:
                tier_rows.append([
                    t["tier"],
                    _fmt(t["count"]),
                    _fmt_currency(t["avg_revenue"]),
                    f'{t["ebitda_margin"]:.0%}',
                    _fmt_currency(t["avg_ebitda"]),
                    f'{t["entry_multiple"]:.1f}x',
                    _fmt_currency(t["total_revenue"]),
                    _fmt_currency(t["total_ebitda"]),
                ])
            # Totals row
            tier_rows.append([
                "<strong>Total</strong>",
                f'<strong>{_fmt(dm_total_loc)}</strong>',
                "-",
                f'<strong>{dm_w_margin:.0%}</strong>',
                "-",
                "-",
                f'<strong>{_fmt_currency(dm_total_rev)}</strong>',
                f'<strong>{_fmt_currency(dm_total_ebitda)}</strong>',
            ])
            body += data_table(
                headers=["Price Tier", "Locations", "Avg Revenue", "EBITDA Margin",
                         "Avg EBITDA", "Entry Multiple", "Total Revenue", "Total EBITDA"],
                rows=tier_rows,
                numeric_columns={1, 2, 3, 4, 5, 6, 7},
            )

            # Price tier doughnut chart
            if len(dm_tiers) > 1:
                tier_labels = [t["tier"] for t in dm_tiers]
                tier_values = [float(t["count"]) for t in dm_tiers]
                tier_colors = [BLUE, GREEN, ORANGE, PURPLE, GRAY][:len(tier_labels)]
                tier_donut_config = build_doughnut_config(tier_labels, tier_values, tier_colors)
                tier_donut_json = json.dumps(tier_donut_config)

                body += '<div class="chart-row">'
                body += "<div>"
                body += chart_container(
                    "tierDonut", tier_donut_json,
                    build_bar_fallback(tier_labels, tier_values),
                    size="medium",
                    title="A-Grade Locations by Price Tier",
                )
                charts_js += chart_init_js("tierDonut", tier_donut_json)
                body += build_chart_legend(tier_labels, tier_values, tier_colors, show_pct=True)
                body += "</div>"
                body += "</div>"

            # P&L waterfall (per average location)
            if dm_pnl.get("revenue", 0) > 0:
                body += '<h3 style="font-size:15px;font-weight:600;color:var(--primary);margin:16px 0 8px">Per-Location P&L Waterfall (Portfolio Average)</h3>'

                pnl_rev = dm_pnl["revenue"]
                pnl_items = [
                    ("Revenue", pnl_rev, BLUE),
                    ("COGS (40%)", dm_pnl["cogs"], RED),
                    ("Gross Profit", dm_pnl["gross_profit"], GREEN),
                    ("SG&A (32%)", dm_pnl["sga"], ORANGE),
                    ("EBITDA", dm_pnl["ebitda"], TEAL),
                ]
                max_val = max(v for _, v, _ in pnl_items) if pnl_items else 1

                body += '<div class="pnl-waterfall">'
                for label, value, color in pnl_items:
                    bar_h = max(int(value / max_val * 150), 8) if max_val > 0 else 8
                    body += f"""<div class="waterfall-bar">
    <div class="bar-value">{_fmt_currency(value)}</div>
    <div class="bar" style="height:{bar_h}px;background:{color}"></div>
    <div class="bar-label">{label}</div>
</div>"""
                body += "</div>"

            body += callout(
                "<strong>Methodology:</strong> Revenue and margin estimates are based on AmSpa "
                "State of the Industry Report, IBISWorld Medspa Industry Analysis, and PE deal "
                "comps benchmarks. Actual financials will vary by location — these are planning estimates.",
                variant="warn",
            )
        else:
            body += callout(
                "<strong>No A-grade targets found.</strong> Run med-spa discovery and scoring "
                "to populate the deal model.",
                variant="warn",
            )

        body += "\n" + section_end()

        # ==================================================================
        # Section 14: Deal Model — Capital Requirements
        # ==================================================================
        body += "\n" + section_start(14, "Deal Model \u2014 Capital Requirements", "deal-capital")
        body += '<p>Total capital needed to acquire all A-grade targets, including financing structure, transaction costs, and working capital reserves.</p>'

        dm_acq_cost = deal_model.get("total_acquisition_cost", 0)
        dm_cap = deal_model.get("capital_stack", {})
        dm_leverage = deal_model.get("leverage_ratio", 0)

        if dm_acq_cost > 0:
            # Acquisition cost by tier
            body += '<h3 style="font-size:15px;font-weight:600;color:var(--primary);margin:16px 0 8px">Acquisition Cost by Price Tier</h3>'

            acq_rows = []
            for t in dm_tiers:
                acq_rows.append([
                    t["tier"],
                    _fmt(t["count"]),
                    _fmt_currency(t["avg_ebitda"]),
                    f'{t["entry_multiple"]:.1f}x',
                    _fmt_currency(t["total_acq_cost"]),
                ])
            acq_rows.append([
                "<strong>Total</strong>",
                f'<strong>{_fmt(dm_total_loc)}</strong>',
                "-",
                "-",
                f'<strong>{_fmt_currency(dm_acq_cost)}</strong>',
            ])
            body += data_table(
                headers=["Price Tier", "Locations", "Avg EBITDA", "Entry Multiple", "Total Acquisition Cost"],
                rows=acq_rows,
                numeric_columns={1, 2, 3, 4},
            )

            # Capital stack KPIs
            equity = dm_cap.get("equity", 0)
            debt = dm_cap.get("debt", 0)
            txn_costs = dm_cap.get("transaction_costs", 0)
            wc = dm_cap.get("working_capital", 0)
            total_cap = dm_cap.get("total_capital_required", 0)

            body += '<h3 style="font-size:15px;font-weight:600;color:var(--primary);margin:16px 0 8px">Capital Stack</h3>'
            body += '<div class="metric-grid">'
            body += f"""<div class="metric-card">
    <div class="metric-label">Total Acquisition Cost</div>
    <div class="metric-value">{_fmt_currency(dm_acq_cost)}</div>
    <div class="metric-detail">{_fmt(dm_total_loc)} locations × avg EBITDA × entry multiple</div>
</div>"""
            body += f"""<div class="metric-card">
    <div class="metric-label">Equity Check (40%)</div>
    <div class="metric-value">{_fmt_currency(equity)}</div>
    <div class="metric-detail">Sponsor equity contribution</div>
</div>"""
            body += f"""<div class="metric-card">
    <div class="metric-label">Senior Debt (60%)</div>
    <div class="metric-value">{_fmt_currency(debt)}</div>
    <div class="metric-detail">Leverage: {dm_leverage:.1f}x Debt/EBITDA</div>
</div>"""
            body += f"""<div class="metric-card">
    <div class="metric-label">Total Capital Required</div>
    <div class="metric-value" style="color:{GREEN}">{_fmt_currency(total_cap)}</div>
    <div class="metric-detail">Equity + transaction costs + working capital</div>
</div>"""
            body += "</div>"

            # Detailed capital table
            cap_rows = [
                ["Equity Contribution (40%)", _fmt_currency(equity)],
                ["Senior Debt (60%)", _fmt_currency(debt)],
                ["Transaction Costs (5%)", _fmt_currency(txn_costs)],
                [f"Working Capital ({DEAL_ASSUMPTIONS['working_capital_months']}mo SG&A)", _fmt_currency(wc)],
                ["<strong>Total Capital Required</strong>", f'<strong>{_fmt_currency(total_cap)}</strong>'],
            ]
            body += data_table(
                headers=["Component", "Amount"],
                rows=cap_rows,
                numeric_columns={1},
            )

            # Capital stack visualization (horizontal bar)
            cap_items = [
                ("Equity", equity, BLUE),
                ("Senior Debt", debt, TEAL),
                ("Transaction Costs", txn_costs, ORANGE),
                ("Working Capital", wc, GRAY),
            ]
            cap_total = sum(v for _, v, _ in cap_items)

            if cap_total > 0:
                body += '<h3 style="font-size:15px;font-weight:600;color:var(--primary);margin:16px 0 8px">Capital Stack Breakdown</h3>'
                body += '<div class="capital-stack">'
                for label, value, color in cap_items:
                    pct = value / cap_total * 100
                    if pct >= 3:  # Only show label if segment is wide enough
                        body += f'<div class="stack-segment" style="width:{pct:.1f}%;background:{color}" title="{label}: {_fmt_currency(value)}">{label} {_fmt_currency(value)}</div>'
                    elif pct > 0:
                        body += f'<div class="stack-segment" style="width:{pct:.1f}%;background:{color}" title="{label}: {_fmt_currency(value)}"></div>'
                body += '</div>'

                # Chart.js horizontal bar version
                cap_labels = [c[0] for c in cap_items]
                cap_values = [c[1] for c in cap_items]
                cap_bar_config = build_horizontal_bar_config(
                    cap_labels, cap_values, dataset_label="Capital ($)"
                )
                cap_bar_json = json.dumps(cap_bar_config)

                body += chart_container(
                    "capitalStackChart", cap_bar_json,
                    build_bar_fallback(cap_labels, cap_values),
                    title="Capital Stack Components",
                    height="280px",
                )
                charts_js += chart_init_js("capitalStackChart", cap_bar_json)

            # Leverage check callout
            if dm_leverage > 4:
                body += callout(
                    f"<strong>Leverage Warning:</strong> At {dm_leverage:.1f}x Debt/EBITDA, the portfolio "
                    f"exceeds the typical 4.0x leverage threshold. Consider reducing the debt percentage "
                    f"or phasing acquisitions to maintain healthy coverage ratios.",
                    variant="warn",
                )
            else:
                body += callout(
                    f"<strong>Leverage Check:</strong> At {dm_leverage:.1f}x Debt/EBITDA, the portfolio "
                    f"is within the typical 4.0x leverage comfort zone for PE-backed healthcare roll-ups.",
                    variant="good",
                )
        else:
            body += callout(
                "<strong>No acquisition cost data available.</strong> "
                "Deal model requires A-grade targets with price tier data.",
                variant="warn",
            )

        body += "\n" + section_end()

        # ==================================================================
        # Section 15: Deal Model — Returns Analysis
        # ==================================================================
        body += "\n" + section_start(15, "Deal Model \u2014 Returns Analysis", "deal-returns")
        body += '<p>Projected returns across three scenarios with varying exit multiples, margin improvement, and hold periods.</p>'

        dm_scenarios = deal_model.get("scenarios", {})
        dm_a_states = deal_model.get("a_grade_states", [])

        if dm_scenarios:
            # Scenario cards
            body += '<div class="deal-scenario-grid">'
            for scenario_name in ["conservative", "base", "aggressive"]:
                s = dm_scenarios.get(scenario_name, {})
                if not s:
                    continue
                label_map = {"conservative": "Conservative", "base": "Base Case", "aggressive": "Aggressive"}
                body += f'<div class="scenario-card {scenario_name}">'
                body += f'<div class="scenario-label">{label_map.get(scenario_name, scenario_name)}</div>'
                metrics = [
                    ("Entry EV", _fmt_currency(s.get("entry_ev", 0))),
                    ("Exit Multiple", f'{s.get("exit_multiple", 0)}x'),
                    ("EBITDA Improvement", f'{s.get("margin_improvement", 0):.0%}/yr'),
                    ("Hold Period", f'{s.get("hold_years", 0)} years'),
                    ("Exit EV", _fmt_currency(s.get("exit_ev", 0))),
                    ("Gross MOIC", f'{s.get("gross_moic", 0):.1f}x'),
                    ("Net IRR", f'{s.get("net_irr", 0):.1%}'),
                ]
                for label, value in metrics:
                    body += f"""<div class="scenario-metric">
    <span class="label">{label}</span>
    <span class="value">{value}</span>
</div>"""
                body += "</div>"
            body += "</div>"

            # Scenario comparison table
            body += '<h3 style="font-size:15px;font-weight:600;color:var(--primary);margin:16px 0 8px">Scenario Comparison</h3>'

            scenario_rows = []
            for scenario_name in ["conservative", "base", "aggressive"]:
                s = dm_scenarios.get(scenario_name, {})
                if not s:
                    continue
                label_map = {"conservative": "Conservative", "base": "Base Case", "aggressive": "Aggressive"}
                scenario_rows.append([
                    f'<strong>{label_map.get(scenario_name)}</strong>',
                    _fmt_currency(s.get("entry_ev", 0)),
                    f'{s.get("exit_multiple", 0)}x',
                    f'{s.get("margin_improvement", 0):.0%}/yr',
                    f'{s.get("hold_years", 0)} yrs',
                    _fmt_currency(s.get("exit_ev", 0)),
                    f'{s.get("gross_moic", 0):.1f}x',
                    f'{s.get("net_irr", 0):.1%}',
                ])

            body += data_table(
                headers=["Scenario", "Entry EV", "Exit Multiple", "EBITDA Improvement",
                         "Hold Period", "Exit EV", "Gross MOIC", "Net IRR"],
                rows=scenario_rows,
                numeric_columns={1, 2, 3, 4, 5, 6, 7},
            )

            # Synergy callout
            body += callout(
                "<strong>Synergy Upside (Not Modeled):</strong> Procurement savings (15-25% on "
                "injectables/consumables), shared marketing infrastructure, centralized scheduling "
                "& billing technology platform, and provider network effects could add 300-500bps "
                "of additional margin improvement beyond baseline projections.",
                variant="good",
            )

            # Phased rollout strategy
            body += '<h3 style="font-size:15px;font-weight:600;color:var(--primary);margin:16px 0 8px">Phased Rollout Strategy</h3>'
            body += """<div class="thesis-box">
<h3>Recommended Acquisition Cadence</h3>
<ul>
    <li><strong>Year 1 — Platform Build (Top 100):</strong> Acquire highest-scoring A-grade targets in top 5 states. Establish shared services infrastructure, negotiate vendor contracts, deploy tech platform.</li>
    <li><strong>Year 2 — Scale (+200):</strong> Expand to 300 total locations. Leverage platform economics for procurement savings. Begin cross-selling across locations.</li>
    <li><strong>Years 3-5 — Full Portfolio:</strong> Complete remaining ~{0} acquisitions. Drive margin improvement through operational integration. Prepare for exit.</li>
</ul>
</div>""".format(max(dm_total_loc - 300, 0))

            # Exit value by state doughnut
            if dm_a_states and dm_scenarios.get("base"):
                base_exit_ev = dm_scenarios["base"].get("exit_ev", 0)
                total_a = sum(s["count"] for s in dm_a_states)

                if total_a > 0 and base_exit_ev > 0:
                    top_exit_states = dm_a_states[:10]
                    other_count = sum(s["count"] for s in dm_a_states[10:])

                    exit_labels = [s["state"] for s in top_exit_states]
                    exit_values = [round(s["count"] / total_a * base_exit_ev / 1_000_000, 1) for s in top_exit_states]
                    if other_count > 0:
                        exit_labels.append("Other")
                        exit_values.append(round(other_count / total_a * base_exit_ev / 1_000_000, 1))

                    exit_colors = list(CHART_COLORS[:len(exit_labels)])
                    exit_donut_config = build_doughnut_config(exit_labels, exit_values, exit_colors)
                    exit_donut_json = json.dumps(exit_donut_config)

                    body += '<div class="chart-row">'
                    body += "<div>"
                    body += chart_container(
                        "exitValueDonut", exit_donut_json,
                        build_bar_fallback(exit_labels, exit_values),
                        size="medium",
                        title="Base Case Exit Value by State ($M)",
                    )
                    charts_js += chart_init_js("exitValueDonut", exit_donut_json)
                    body += build_chart_legend(exit_labels, exit_values, exit_colors, show_pct=True)
                    body += "</div>"
                    body += "</div>"
        else:
            body += callout(
                "<strong>No scenario data available.</strong> Deal model requires A-grade "
                "targets to generate returns analysis.",
                variant="warn",
            )

        body += "\n" + section_end()

        # ==================================================================
        # Section 16: Stealth Wealth Signal
        # ==================================================================
        body += "\n" + section_start(16, "Stealth Wealth Signal", "stealth-wealth")

        sw_summary = stealth_wealth.get("summary", {})
        sw_zips = stealth_wealth.get("stealth_zips", [])
        sw_comp = stealth_wealth.get("wealth_composition", {})
        sw_states = stealth_wealth.get("top_states", [])

        if not stealth_wealth:
            body += callout(
                "<strong>IRS SOI ZIP income data not yet ingested.</strong> "
                "Run <code>POST /api/v1/irs-soi/zip-income/ingest</code> to enable "
                "stealth wealth analysis. This cross-references non-wage income "
                "(capital gains, dividends, partnership income) with medspa scores "
                "to reveal hidden demand in affluent ZIPs.",
                variant="info",
            )
        else:
            # KPI cards
            sw_cards = ""
            sw_cards += kpi_card("ZIPs Analyzed", _fmt(sw_comp.get("total_returns")))
            sw_cards += kpi_card(
                "Stealth ZIPs Found",
                _fmt(sw_summary.get("total_stealth")),
            )
            sw_cards += kpi_card(
                "Validated (High Score)",
                _fmt(sw_summary.get("validated")),
            )
            sw_cards += kpi_card(
                "Avg Non-Wage Income",
                _fmt_currency(sw_summary.get("avg_non_wage_income", 0)),
            )
            body += kpi_strip(sw_cards)

            body += f"""<p style="margin:12px 0;font-size:14px;color:var(--gray-600)">
Stealth wealth ZIPs have <strong>&gt;25% non-wage income</strong> (capital gains + dividends +
partnership income) but a medspa score below 70 — indicating affluent populations
underserved by current med-spa supply. Tax year: <strong>{sw_summary.get('tax_year', '?')}</strong>.</p>"""

            # Wealth composition doughnut + explanation
            if sw_comp:
                body += '<div class="wealth-composition-grid">'

                # Doughnut chart
                comp_labels = ["W-2 Wages", "Capital Gains", "Dividends", "Business/Partnership", "Other"]
                comp_values = [
                    sw_comp.get("wages_pct", 0),
                    sw_comp.get("cap_gains_pct", 0),
                    sw_comp.get("dividends_pct", 0),
                    sw_comp.get("biz_income_pct", 0),
                    sw_comp.get("other_pct", 0),
                ]
                comp_colors = [BLUE, GREEN, ORANGE, PURPLE, GRAY]
                comp_config = build_doughnut_config(comp_labels, comp_values, comp_colors)
                comp_json = json.dumps(comp_config)

                body += "<div>"
                body += chart_container(
                    "wealthCompDonut", comp_json,
                    build_bar_fallback(comp_labels, comp_values),
                    size="medium",
                    title="Income Composition (% of AGI)",
                )
                charts_js += chart_init_js("wealthCompDonut", comp_json)
                body += build_chart_legend(comp_labels, comp_values, comp_colors, show_pct=True)
                body += "</div>"

                # Explanation card
                body += """<div>
<div class="thesis-box">
    <h3>Why Non-Wage Income Matters</h3>
    <ul>
        <li><strong>Capital gains &amp; dividends</strong> indicate investable wealth — these households have assets producing returns, not just paychecks.</li>
        <li><strong>Partnership/business income</strong> signals entrepreneurs and professionals with high discretionary spend.</li>
        <li>Traditional models use W-2 wages or median household income, <strong>missing 25-40% of actual purchasing power</strong> in wealthy ZIPs.</li>
        <li>Stealth wealth ZIPs are undervalued by competitors using standard data — a <strong>first-mover advantage</strong> for informed acquirers.</li>
    </ul>
</div>
</div>"""
                body += "</div>"  # close wealth-composition-grid

            # Stealth ZIPs data table
            if sw_zips:
                body += '<h3 style="font-size:15px;font-weight:600;color:var(--primary);margin:20px 0 8px">Top Stealth Wealth ZIPs</h3>'
                sw_rows = []
                for z in sw_zips[:25]:
                    sw_rows.append([
                        z["zip_code"],
                        z["state"],
                        _fmt_currency(z["non_wage_per_return"]),
                        f"{z['non_wage_pct']:.1f}%",
                        _fmt(z["num_returns"]),
                        _fmt_currency(z["avg_agi"]),
                        str(z["medspa_score"]),
                        z["medspa_grade"],
                        str(z["a_grade_medspas"]),
                    ])
                body += data_table(
                    headers=[
                        "ZIP", "State", "Non-Wage/Return", "Non-Wage %",
                        "Returns", "Avg AGI", "Medspa Score", "Grade", "A-Grade Medspas",
                    ],
                    rows=sw_rows,
                )

            # States bar chart
            if sw_states:
                st_labels = [s["state"] for s in sw_states[:12]]
                st_values = [s["count"] for s in sw_states[:12]]
                st_colors = [BLUE] * len(st_labels)
                st_config = build_horizontal_bar_config(
                    st_labels, st_values, st_colors,
                    dataset_label="Stealth ZIPs",
                )
                st_json = json.dumps(st_config)

                body += chart_container(
                    "stealthByState", st_json,
                    build_bar_fallback(st_labels, st_values),
                    size="large",
                    title="Stealth Wealth ZIPs by State",
                )
                charts_js += chart_init_js("stealthByState", st_json)

            # Methodology callout
            body += callout(
                "<strong>Methodology:</strong> ZIPs are flagged as \"stealth\" when &gt;25% of total AGI "
                "comes from non-wage sources (capital gains + dividends + business income) AND the "
                "current medspa acquisition score is below 70. This identifies affluent areas where "
                "traditional income-based models underestimate demand. Minimum 100 tax returns per ZIP "
                "to ensure statistical significance.",
                variant="info",
            )

            # Opportunity callout
            if sw_zips:
                body += callout(
                    f"<strong>Opportunity:</strong> {sw_summary.get('total_stealth', 0)} stealth wealth ZIPs "
                    f"identified with an average non-wage income of "
                    f"{_fmt_currency(sw_summary.get('avg_non_wage_income', 0))} per return. "
                    f"These ZIPs represent a first-mover acquisition opportunity invisible to "
                    f"competitors using standard income data.",
                    variant="tip",
                )

        body += "\n" + section_end()

        # ==================================================================
        # Section 17: Migration Alpha
        # ==================================================================
        body += "\n" + section_start(17, "Migration Alpha", "migration-alpha")

        ma_summary = migration_alpha.get("summary", {})
        ma_flows = migration_alpha.get("state_flows", [])
        ma_emerging = migration_alpha.get("emerging_markets", [])
        ma_counties = migration_alpha.get("top_county_inflows", [])
        ma_exodus = migration_alpha.get("wealth_exodus", [])

        if not migration_alpha:
            body += callout(
                "<strong>IRS SOI migration data not yet ingested.</strong> "
                "Run <code>POST /api/v1/irs-soi/migration/ingest</code> to enable "
                "migration alpha analysis. This uses county-to-county wealth flows "
                "as a 2-3 year leading indicator of medspa demand.",
                variant="info",
            )
        else:
            # KPI cards
            ma_cards = ""
            ma_cards += kpi_card(
                "Net Wealth Movement",
                f"${abs(ma_summary.get('total_net_agi_m', 0)):,.0f}M",
            )
            ma_cards += kpi_card("Top Gainer", ma_summary.get("top_gainer", "-"))
            ma_cards += kpi_card("Top Loser", ma_summary.get("top_loser", "-"))
            ma_cards += kpi_card(
                "Emerging Markets",
                _fmt(ma_summary.get("emerging_count")),
            )
            body += kpi_strip(ma_cards)

            body += f"""<p style="margin:12px 0;font-size:14px;color:var(--gray-600)">
Migration Alpha measures <strong>net wealth inflows</strong> relative to existing medspa density.
States receiving large AGI inflows but with few A-grade medspas represent a
<strong>2-3 year leading indicator</strong> of unmet demand. Tax year: <strong>{ma_summary.get('tax_year', '?')}</strong>.</p>"""

            # Dual-color horizontal bar: top 15 net flows
            if ma_flows:
                top_flow_states = ma_flows[:10]
                bottom_flow_states = sorted(ma_flows, key=lambda x: x["net_agi_m"])[:5]
                bar_states = top_flow_states + [s for s in bottom_flow_states if s not in top_flow_states]
                bar_states.sort(key=lambda x: x["net_agi_m"], reverse=True)

                bar_labels = [s["state"] for s in bar_states]
                bar_values = [s["net_agi_m"] for s in bar_states]
                bar_colors = ["#48bb78" if v >= 0 else "#fc8181" for v in bar_values]
                bar_config = build_horizontal_bar_config(
                    bar_labels, bar_values, bar_colors,
                    dataset_label="Net AGI ($M)",
                )
                bar_json = json.dumps(bar_config)

                body += chart_container(
                    "migrationNetFlow", bar_json,
                    build_bar_fallback(bar_labels, bar_values),
                    size="large",
                    title="Net Wealth Migration by State ($M AGI)",
                )
                charts_js += chart_init_js("migrationNetFlow", bar_json)

            # State flows table
            if ma_flows:
                body += '<h3 style="font-size:15px;font-weight:600;color:var(--primary);margin:20px 0 8px">State Wealth Flows</h3>'
                flow_rows = []
                for s in ma_flows[:20]:
                    net_color = "color:#48bb78" if s["net_agi_m"] >= 0 else "color:#fc8181"
                    flow_rows.append([
                        s["state"],
                        f"${s['inflow_agi_m']:,.1f}M",
                        f"${s['outflow_agi_m']:,.1f}M",
                        f'<span style="{net_color};font-weight:600">${s["net_agi_m"]:+,.1f}M</span>',
                        _fmt(s["inflow_returns"]),
                        _fmt(s["total_medspas"]),
                        _fmt(s["a_grade_count"]),
                        f"{s['migration_alpha']:,.0f}",
                    ])
                body += data_table(
                    headers=[
                        "State", "Inflow AGI", "Outflow AGI", "Net AGI",
                        "Inflow Returns", "Medspas", "A-Grade", "Migration Alpha",
                    ],
                    rows=flow_rows,
                )

            # Emerging markets highlight table
            if ma_emerging:
                body += '<h3 style="font-size:15px;font-weight:600;color:var(--primary);margin:20px 0 8px">Emerging Markets (Positive Inflow, &lt;20 A-Grade Medspas)</h3>'
                em_rows = []
                for s in ma_emerging[:10]:
                    em_rows.append([
                        f'<strong>{s["state"]}</strong>',
                        f"${s['net_agi_m']:+,.1f}M",
                        _fmt(s["inflow_returns"]),
                        _fmt(s["total_medspas"]),
                        _fmt(s["a_grade_count"]),
                        f"{s['migration_alpha']:,.0f}",
                    ])
                body += '<div class="highlight-table">'
                body += data_table(
                    headers=[
                        "State", "Net AGI", "Inflow Returns",
                        "Total Medspas", "A-Grade", "Migration Alpha",
                    ],
                    rows=em_rows,
                )
                body += "</div>"

                body += callout(
                    f"<strong>{len(ma_emerging)} emerging markets identified.</strong> "
                    "These states are receiving net wealth inflows but have fewer than 20 "
                    "A-grade medspa targets — signaling greenfield opportunity before "
                    "competitors recognize the demand shift.",
                    variant="tip",
                )

            # County drill-down
            if ma_counties:
                body += '<h3 style="font-size:15px;font-weight:600;color:var(--primary);margin:20px 0 8px">Top County-Level Inflows</h3>'
                county_rows = []
                for c in ma_counties[:15]:
                    county_rows.append([
                        c["dest_state"],
                        c["county"],
                        c["orig_state"],
                        f"${c['agi_m']:,.1f}M",
                        _fmt(c["returns"]),
                    ])
                body += data_table(
                    headers=["Dest State", "County", "Origin State", "AGI Inflow", "Returns"],
                    rows=county_rows,
                )

            # Leading indicator callout
            body += callout(
                "<strong>Leading Indicator:</strong> Wealth migration is a 2-3 year "
                "leading indicator of local service demand. High-AGI movers "
                "typically establish primary residence before seeking premium "
                "services like med-spas. States with high migration alpha today "
                "will see increased medspa demand in 2-3 years.",
                variant="info",
            )

        body += "\n" + section_end()

        # ==================================================================
        # Section 18: Medical Provider Density Signal
        # ==================================================================
        body += "\n" + section_start(18, "Medical Provider Density Signal", "provider-density")

        pd_summary = provider_density.get("summary", {})
        pd_opp_zips = provider_density.get("opportunity_zips", [])
        pd_by_type = provider_density.get("provider_by_type", [])

        if not provider_density:
            body += callout(
                "<strong>CMS Medicare utilization data not yet ingested.</strong> "
                "Run <code>POST /api/v1/cms/ingest/medicare-utilization</code> to enable "
                "provider density analysis. This cross-references dermatologists, plastic "
                "surgeons, NPs, and PAs billing Medicare with medspa density to find "
                "referral opportunity ZIPs.",
                variant="info",
            )
        else:
            # KPI cards
            pd_cards = ""
            pd_cards += kpi_card("Provider-ZIP Pairs", _fmt(pd_summary.get("total_provider_zip_pairs")))
            pd_cards += kpi_card(
                "Opportunity ZIPs",
                _fmt(pd_summary.get("opportunity_zips")),
                delta="Medspa score < 60",
            )
            pd_cards += kpi_card("Avg Providers/ZIP", str(pd_summary.get("avg_providers_per_zip", 0)))
            pd_cards += kpi_card("Top Provider Type", pd_summary.get("top_provider_type", "-"))
            body += "\n" + kpi_strip(pd_cards)

            body += """<p style="font-size:14px;color:var(--gray-600);margin:12px 0">
            <strong>Provider-rich, medspa-poor ZIPs</strong> represent referral corridors where
            medical providers (dermatologists, plastic surgeons, NPs, PAs) are billing Medicare
            but few medspas exist to capture the aesthetics crossover demand. Higher imbalance
            scores indicate greater unmet demand.</p>"""

            # Charts: provider bar + beneficiary doughnut
            if pd_by_type:
                body += '<div class="provider-split-grid">'
                # Horizontal bar: provider counts by type
                bar_labels = [t["type"] for t in pd_by_type]
                bar_values = [t["count"] for t in pd_by_type]
                bar_colors = [BLUE, TEAL, PURPLE, ORANGE][:len(pd_by_type)]
                bar_config = build_horizontal_bar_config(
                    bar_labels, bar_values, bar_colors,
                    dataset_label="Provider Count",
                )
                bar_json = json.dumps(bar_config)
                body += "<div>"
                body += chart_container(
                    "providerTypeBar", bar_json,
                    build_bar_fallback(bar_labels, bar_values),
                    size="medium",
                    title="Provider Counts by Specialty",
                )
                charts_js += chart_init_js("providerTypeBar", bar_json)
                body += "</div>"

                # Doughnut: beneficiary mix by type
                bene_labels = [t["type"] for t in pd_by_type]
                bene_values = [t["beneficiaries"] for t in pd_by_type]
                bene_colors = [BLUE, TEAL, PURPLE, ORANGE][:len(pd_by_type)]
                donut_config = build_doughnut_config(bene_labels, bene_values, bene_colors)
                donut_json = json.dumps(donut_config)
                body += "<div>"
                body += chart_container(
                    "providerBeneDonut", donut_json,
                    build_bar_fallback(bene_labels, bene_values),
                    size="medium",
                    title="Beneficiary Mix by Provider Type",
                )
                charts_js += chart_init_js("providerBeneDonut", donut_json)
                body += build_chart_legend(bene_labels, bene_values, bene_colors, show_pct=True)
                body += "</div>"
                body += "</div>"

            # Data table: top 25 opportunity ZIPs
            if pd_opp_zips:
                body += '<h3 style="font-size:15px;font-weight:600;color:var(--primary);margin:20px 0 8px">Top Opportunity ZIPs (Provider-Rich, Medspa-Poor)</h3>'
                pd_rows = []
                for z in pd_opp_zips[:25]:
                    pd_rows.append([
                        z["zip_code"],
                        str(z["provider_count"]),
                        z["provider_types"][:40],
                        _fmt(z["beneficiaries"]),
                        str(z["medspa_count"]),
                        f"{z['imbalance_score']:.1f}",
                        f"{z['medspa_score']:.0f}",
                        z["medspa_grade"],
                    ])
                body += '<div class="highlight-table">'
                body += data_table(
                    headers=["ZIP", "Providers", "Types", "Benes", "Medspas",
                             "Imbalance", "Score", "Grade"],
                    rows=pd_rows,
                )
                body += "</div>"

            body += callout(
                f"<strong>{pd_summary.get('opportunity_zips', 0)} opportunity ZIPs identified</strong> "
                "where medical providers outnumber medspas. These represent referral corridors — "
                "an acquiring platform can build provider relationships for patient referrals "
                "in aesthetics-adjacent treatments.",
                variant="tip",
            )

            body += callout(
                "<strong>Methodology:</strong> CMS Medicare Provider Utilization data filtered "
                "for Dermatology, Plastic Surgery, Nurse Practitioners, and Physician Assistants. "
                "Imbalance score = providers ÷ medspas per ZIP (higher = more opportunity).",
                variant="info",
            )

        body += "\n" + section_end()

        # ==================================================================
        # Section 19: Real Estate Appreciation Alpha
        # ==================================================================
        body += "\n" + section_start(19, "Real Estate Appreciation Alpha", "re-alpha")

        re_summary = real_estate_alpha.get("summary", {})
        re_timing = real_estate_alpha.get("timing_zips", [])
        re_states = real_estate_alpha.get("state_summary", [])
        re_source = real_estate_alpha.get("source", "")

        if not real_estate_alpha:
            body += callout(
                "<strong>Real estate data not yet ingested.</strong> "
                "Run <code>POST /api/v1/realestate/redfin/ingest</code> (preferred) or "
                "<code>POST /api/v1/realestate/fhfa/ingest</code> (fallback) to enable "
                "real estate appreciation alpha. This identifies appreciating ZIPs with "
                "low medspa penetration — timing signals for market entry.",
                variant="info",
            )
        else:
            # KPI cards
            re_cards = ""
            re_cards += kpi_card("Avg Median Price", _fmt_currency(re_summary.get("avg_median_price", 0)))
            re_cards += kpi_card(
                "Avg YoY Appreciation",
                f"{re_summary.get('avg_yoy_change', 0):+.1f}%",
            )
            re_cards += kpi_card(
                "Timing Opportunity ZIPs",
                _fmt(re_summary.get("timing_opportunity_zips")),
                delta="Appreciating + score < 60",
            )
            re_cards += kpi_card("Hottest State", re_summary.get("hottest_state", "-"))
            body += "\n" + kpi_strip(re_cards)

            body += f"""<p style="font-size:14px;color:var(--gray-600);margin:12px 0">
            <strong>Timing opportunity ZIPs</strong> are areas where home values are appreciating
            (indicating growing wealth) but medspa penetration remains low (score &lt; 60).
            Rising property values are a 1-2 year leading indicator of premium service demand.
            Source: <strong>{re_source}</strong>{"" if re_source == "Redfin" else " (ZIP3-level, less granular)"}.</p>"""

            # Chart: top 15 states by avg YoY appreciation
            if re_states:
                top_15 = re_states[:15]
                bar_labels = [s["state"] for s in top_15]
                bar_values = [s["avg_yoy_change"] for s in top_15]
                bar_colors = [GREEN if v >= 0 else RED for v in bar_values]
                bar_config = build_horizontal_bar_config(
                    bar_labels, bar_values, bar_colors,
                    dataset_label="Avg YoY Appreciation %",
                )
                bar_json = json.dumps(bar_config)
                body += chart_container(
                    "reAlphaBar", bar_json,
                    build_bar_fallback(bar_labels, [f"{v:+.1f}%" for v in bar_values]),
                    size="large",
                    title="Top States by Avg Home Price Appreciation (YoY %)",
                )
                charts_js += chart_init_js("reAlphaBar", bar_json)

            # Data table: top 25 timing opportunity ZIPs
            if re_timing:
                body += '<h3 style="font-size:15px;font-weight:600;color:var(--primary);margin:20px 0 8px">Top Timing Opportunity ZIPs</h3>'
                re_rows = []
                for z in re_timing[:25]:
                    re_rows.append([
                        z["zip_code"],
                        z.get("state", "-"),
                        _fmt_currency(z["median_price"]),
                        f"{z['yoy_change']:+.1f}%" if z.get("yoy_change") is not None else "-",
                        f"{z['medspa_score']:.0f}",
                        z.get("medspa_grade", "-"),
                    ])
                body += '<div class="highlight-table">'
                body += data_table(
                    headers=["ZIP", "State", "Median Price", "YoY Change",
                             "Medspa Score", "Grade"],
                    rows=re_rows,
                )
                body += "</div>"

            body += callout(
                f"<strong>{re_summary.get('timing_opportunity_zips', 0)} timing opportunities</strong> "
                "where home prices are rising but medspa competition remains thin. "
                "Early movers in appreciating markets capture affluent demographics before "
                "competitors recognize the demand signal.",
                variant="tip",
            )

        body += "\n" + section_end()

        # ==================================================================
        # Section 20: Deposit Wealth Concentration
        # ==================================================================
        body += "\n" + section_start(20, "Deposit Wealth Concentration", "deposit-wealth")

        dw_summary = deposit_wealth.get("summary", {})
        dw_underserved = deposit_wealth.get("underserved_zips", [])
        dw_state_deps = deposit_wealth.get("state_deposits", [])

        if not deposit_wealth:
            body += callout(
                "<strong>FDIC deposit data not yet ingested.</strong> "
                "Run <code>POST /api/v1/fdic/deposits/ingest</code> to enable "
                "deposit wealth analysis. This identifies ZIPs with high bank deposits "
                "per capita but low medspa penetration — liquid wealth signals.",
                variant="info",
            )
        else:
            # KPI cards
            dw_cards = ""
            dw_cards += kpi_card("ZIPs Analyzed", _fmt(dw_summary.get("zips_analyzed")))
            dw_cards += kpi_card(
                "Total Deposits",
                f"${dw_summary.get('total_deposits_t', 0):.2f}T",
            )
            dw_cards += kpi_card(
                "Avg Deposits/Return",
                _fmt_currency(dw_summary.get("avg_deposits_per_return", 0)),
            )
            dw_cards += kpi_card(
                "Underserved ZIPs",
                _fmt(dw_summary.get("underserved_zips")),
                delta="High deposits + score < 60",
            )
            body += "\n" + kpi_strip(dw_cards)

            body += """<p style="font-size:14px;color:var(--gray-600);margin:12px 0">
            <strong>Deposit wealth concentration</strong> measures liquid wealth (bank deposits)
            per tax return by ZIP. ZIPs with high deposits but low medspa scores represent
            underserved affluent populations with disposable income for aesthetic services.</p>"""

            # Charts: doughnut (state concentration) + bar (top ZIPs by deposits/return)
            if dw_state_deps:
                body += '<div class="provider-split-grid">'

                # Doughnut: top 10 states + other
                top_10 = dw_state_deps[:10]
                other_total = sum(s["deposits"] for s in dw_state_deps[10:])
                donut_labels = [s["state"] for s in top_10]
                donut_values = [round(s["deposits"] / 1e9, 1) for s in top_10]
                if other_total > 0:
                    donut_labels.append("Other")
                    donut_values.append(round(other_total / 1e9, 1))
                donut_colors = list(CHART_COLORS[:len(donut_labels)])
                donut_config = build_doughnut_config(donut_labels, donut_values, donut_colors)
                donut_json = json.dumps(donut_config)
                body += "<div>"
                body += chart_container(
                    "depositStateDonut", donut_json,
                    build_bar_fallback(donut_labels, [f"${v}B" for v in donut_values]),
                    size="medium",
                    title="State Deposit Concentration ($B)",
                )
                charts_js += chart_init_js("depositStateDonut", donut_json)
                body += build_chart_legend(donut_labels, donut_values, donut_colors, show_pct=True)
                body += "</div>"

                # Bar: top ZIPs by deposits/return
                if dw_underserved:
                    top_bar = dw_underserved[:15]
                    bar_labels = [z["zip_code"] for z in top_bar]
                    bar_values = [z["deposits_per_return"] for z in top_bar]
                    bar_colors = [BLUE] * len(top_bar)
                    bar_config = build_horizontal_bar_config(
                        bar_labels, bar_values, bar_colors,
                        dataset_label="Deposits per Return ($)",
                    )
                    bar_json = json.dumps(bar_config)
                    body += "<div>"
                    body += chart_container(
                        "depositZipBar", bar_json,
                        build_bar_fallback(bar_labels, [_fmt_currency(v) for v in bar_values]),
                        size="medium",
                        title="Top Underserved ZIPs by Deposits/Return",
                    )
                    charts_js += chart_init_js("depositZipBar", bar_json)
                    body += "</div>"

                body += "</div>"

            # Data table: top 25 underserved ZIPs
            if dw_underserved:
                body += '<h3 style="font-size:15px;font-weight:600;color:var(--primary);margin:20px 0 8px">Top Underserved ZIPs (High Deposits, Low Medspa Score)</h3>'
                dw_rows = []
                for z in dw_underserved[:25]:
                    dw_rows.append([
                        z["zip_code"],
                        z.get("state", "-"),
                        _fmt_currency(z["total_deposits"]),
                        _fmt_currency(z["deposits_per_return"]),
                        str(z["branch_count"]),
                        f"{z['medspa_score']:.0f}",
                        z.get("medspa_grade", "-"),
                    ])
                body += '<div class="highlight-table">'
                body += data_table(
                    headers=["ZIP", "State", "Total Deposits", "Deposits/Return",
                             "Branches", "Medspa Score", "Grade"],
                    rows=dw_rows,
                )
                body += "</div>"

            body += callout(
                f"<strong>{dw_summary.get('underserved_zips', 0)} underserved ZIPs</strong> "
                "with high deposit concentrations but low medspa penetration. "
                "Bank deposits represent liquid, discretionary wealth — a direct proxy "
                "for medspa spending capacity.",
                variant="tip",
            )

            body += callout(
                f"<strong>Methodology:</strong> FDIC Summary of Deposits ({dw_summary.get('year', '-')}), "
                "aggregated by ZIP. Deposits per return normalized using IRS SOI return counts "
                "where available. Underserved = deposits/return in top quartile + medspa score < 60.",
                variant="info",
            )

        body += "\n" + section_end()

        # ==================================================================
        # Section 21: Business Formation Velocity
        # ==================================================================
        body += "\n" + section_start(21, "Business Formation Velocity", "biz-formation")

        bf_summary = business_formation.get("summary", {})
        bf_entrep = business_formation.get("entrepreneurial_zips", [])
        bf_states = business_formation.get("state_density", [])
        bf_comp = business_formation.get("income_composition", {})

        if not business_formation:
            body += callout(
                "<strong>IRS SOI business income data not yet ingested.</strong> "
                "Run <code>POST /api/v1/irs-soi/business-income/ingest</code> to enable "
                "business formation analysis. This identifies ZIPs with high entrepreneurial "
                "density (Schedule C, partnerships, S-corps) where affluent business owners "
                "would be natural medspa clientele.",
                variant="info",
            )
        else:
            # KPI cards
            bf_cards = ""
            bf_cards += kpi_card("ZIPs Analyzed", _fmt(bf_summary.get("zips_analyzed")))
            bf_cards += kpi_card(
                "Avg Business Density",
                f"{bf_summary.get('avg_biz_density', 0):.1f}%",
            )
            bf_cards += kpi_card(
                "High-Growth ZIPs",
                _fmt(bf_summary.get("high_growth_zips")),
                delta=">10% YoY growth",
            )
            bf_cards += kpi_card(
                "Avg Schedule C Income",
                _fmt_currency(bf_summary.get("avg_schedule_c", 0)),
            )
            body += "\n" + kpi_strip(bf_cards)

            body += """<p style="font-size:14px;color:var(--gray-600);margin:12px 0">
            <strong>Business formation velocity</strong> measures the percentage of tax returns
            reporting business income (Schedule C, partnerships, S-corps) per ZIP.
            ZIPs with high entrepreneurial density and low medspa penetration signal
            affluent, self-employed populations with flexible schedules and disposable income —
            ideal medspa demographics.</p>"""

            # Charts: bar (states by density) + doughnut (income type mix)
            body += '<div class="provider-split-grid">'

            if bf_states:
                top_15 = bf_states[:15]
                bar_labels = [s["state"] for s in top_15]
                bar_values = [s["avg_density"] for s in top_15]
                bar_colors = [TEAL] * len(top_15)
                bar_config = build_horizontal_bar_config(
                    bar_labels, bar_values, bar_colors,
                    dataset_label="Avg Business Density %",
                )
                bar_json = json.dumps(bar_config)
                body += "<div>"
                body += chart_container(
                    "bizDensityBar", bar_json,
                    build_bar_fallback(bar_labels, [f"{v}%" for v in bar_values]),
                    size="medium",
                    title="Top States by Avg Business Density",
                )
                charts_js += chart_init_js("bizDensityBar", bar_json)
                body += "</div>"

            if bf_comp:
                comp_labels = ["Schedule C (Sole Props)", "Partnerships", "S-Corps"]
                comp_values = [
                    bf_comp.get("schedule_c_pct", 0),
                    bf_comp.get("partnership_pct", 0),
                    bf_comp.get("scorp_pct", 0),
                ]
                comp_colors = [ORANGE, PURPLE, TEAL]
                donut_config = build_doughnut_config(comp_labels, comp_values, comp_colors)
                donut_json = json.dumps(donut_config)
                body += "<div>"
                body += chart_container(
                    "bizCompDonut", donut_json,
                    build_bar_fallback(comp_labels, [f"{v}%" for v in comp_values]),
                    size="medium",
                    title="Business Income Type Composition",
                )
                charts_js += chart_init_js("bizCompDonut", donut_json)
                body += build_chart_legend(comp_labels, comp_values, comp_colors, show_pct=True)
                body += "</div>"

            body += "</div>"

            # Data table: top 25 entrepreneurial ZIPs
            if bf_entrep:
                body += '<h3 style="font-size:15px;font-weight:600;color:var(--primary);margin:20px 0 8px">Top Entrepreneurial ZIPs (High Biz Density, Low Medspa Score)</h3>'
                bf_rows = []
                for z in bf_entrep[:25]:
                    yoy_str = f"{z['yoy_growth']:+.1f}%" if z.get("yoy_growth") is not None else "-"
                    bf_rows.append([
                        z["zip_code"],
                        z.get("state", "-"),
                        f"{z['biz_density']:.1f}%",
                        _fmt_currency(z["schedule_c"]),
                        _fmt_currency(z["partnership"]),
                        yoy_str,
                        f"{z['medspa_score']:.0f}",
                        z.get("medspa_grade", "-"),
                    ])
                body += '<div class="highlight-table">'
                body += data_table(
                    headers=["ZIP", "State", "Biz Density", "Sched C", "Partnership",
                             "YoY Growth", "Medspa Score", "Grade"],
                    rows=bf_rows,
                )
                body += "</div>"

            body += callout(
                f"<strong>{len(bf_entrep)} entrepreneurial ZIPs identified</strong> "
                "with business density >20% and medspa score <60. "
                "Entrepreneurial populations correlate with premium service demand — "
                "these ZIPs have affluent, time-flexible residents underserved by medspas.",
                variant="tip",
            )

            body += callout(
                f"<strong>Methodology:</strong> IRS SOI Business Income "
                f"(tax year {bf_summary.get('tax_year', '-')}). "
                "Business density = returns with business income ÷ total returns. "
                "Dollar amounts reported in actual dollars (source data × 1,000). "
                "Minimum 100 returns per ZIP for statistical significance.",
                variant="info",
            )

        body += "\n" + section_end()

        # ==================================================================
        # Section 22: Opportunity Zone Overlay
        # ==================================================================
        body += "\n" + section_start(22, "Opportunity Zone Overlay", "oz-overlay")

        oz_summary = opportunity_zones.get("summary", {})
        oz_states = opportunity_zones.get("state_data", [])

        if not opportunity_zones:
            body += callout(
                "<strong>Opportunity Zone data not yet ingested.</strong> "
                "Run <code>POST /api/v1/site-intel/opportunity-zones/ingest</code> to enable "
                "OZ overlay analysis. This identifies states where Qualified Opportunity Zones "
                "overlap with medspa acquisition targets — unlocking 10-year capital gains "
                "deferral for PE roll-up investments.",
                variant="info",
            )
        else:
            oz_cards = ""
            oz_cards += kpi_card("Total OZ Tracts", _fmt(oz_summary.get("total_tracts")))
            oz_cards += kpi_card("States with OZs", _fmt(oz_summary.get("states_with_oz")))
            oz_cards += kpi_card(
                "OZ Tracts in Medspa States",
                _fmt(oz_summary.get("medspa_oz_states")),
            )
            oz_cards += kpi_card(
                "Tax-Advantaged States",
                _fmt(oz_summary.get("tax_advantaged_states")),
                delta=">5 A-grade + >50 OZ tracts",
            )
            body += "\n" + kpi_strip(oz_cards)

            body += """<p style="font-size:14px;color:var(--gray-600);margin:12px 0">
            <strong>Opportunity Zones</strong> offer 10-year capital gains tax deferral for
            investments in designated Census tracts. States with high OZ density AND high
            A-grade medspa counts represent <strong>tax-advantaged roll-up targets</strong> —
            PE investors can defer gains while building platform value.</p>"""

            # Charts: bar (top 15 states by OZ tracts) + doughnut (low-income vs contiguous)
            body += '<div class="provider-split-grid">'

            if oz_states:
                top_15 = oz_states[:15]
                bar_labels = [s["state"] for s in top_15]
                bar_values = [float(s["oz_tracts"]) for s in top_15]
                bar_config = build_horizontal_bar_config(
                    bar_labels, bar_values,
                    dataset_label="OZ Tracts",
                )
                bar_json = json.dumps(bar_config)
                body += "<div>"
                body += chart_container(
                    "ozStateBar", bar_json,
                    build_bar_fallback(bar_labels, bar_values),
                    size="medium",
                    title="Top 15 States by OZ Tract Count",
                )
                charts_js += chart_init_js("ozStateBar", bar_json)
                body += "</div>"

            total_li = oz_summary.get("total_low_income", 0)
            total_ct = oz_summary.get("total_contiguous", 0)
            if total_li or total_ct:
                donut_labels = ["Low-Income Tracts", "Contiguous Tracts"]
                donut_values = [float(total_li), float(total_ct)]
                donut_colors = [BLUE, ORANGE]
                donut_config = build_doughnut_config(donut_labels, donut_values, donut_colors)
                donut_json = json.dumps(donut_config)
                body += "<div>"
                body += chart_container(
                    "ozDesignDonut", donut_json,
                    build_bar_fallback(donut_labels, donut_values),
                    size="medium",
                    title="OZ Designation Type",
                )
                charts_js += chart_init_js("ozDesignDonut", donut_json)
                body += build_chart_legend(donut_labels, donut_values, donut_colors, show_pct=True)
                body += "</div>"

            body += "</div>"

            # Data table: top 25 states
            if oz_states:
                body += '<h3 style="font-size:15px;font-weight:600;color:var(--primary);margin:20px 0 8px">Top States: OZ Tracts vs Medspa Prospects</h3>'
                oz_rows = []
                for s in oz_states[:25]:
                    oz_rows.append([
                        s["state"],
                        _fmt(s["oz_tracts"]),
                        _fmt(s["low_income_tracts"]),
                        _fmt(s["prospect_count"]),
                        _fmt(s["a_grade_count"]),
                        f"{s['oz_per_prospect']:.1f}" if s["oz_per_prospect"] else "-",
                    ])
                body += '<div class="highlight-table">'
                body += data_table(
                    headers=["State", "OZ Tracts", "Low-Income", "Medspa Prospects",
                             "A-Grade", "OZ/Prospect Ratio"],
                    rows=oz_rows,
                )
                body += "</div>"

            body += callout(
                f"<strong>{oz_summary.get('tax_advantaged_states', 0)} tax-advantaged states</strong> "
                "identified with both significant OZ density and A-grade medspa targets. "
                "Qualified Opportunity Fund investments in these zones can defer and reduce "
                "capital gains taxes — a powerful PE incentive for location selection.",
                variant="tip",
            )

        body += "\n" + section_end()

        # ==================================================================
        # Section 23: Demographic Demand Model
        # ==================================================================
        body += "\n" + section_start(23, "Demographic Demand Model", "demographic-demand")

        dd_summary = demographic_demand.get("summary", {})
        dd_states = demographic_demand.get("state_data", [])

        if not demographic_demand:
            body += callout(
                "<strong>Educational attainment data not yet ingested.</strong> "
                "Run <code>POST /api/v1/site-intel/educational-attainment/ingest</code> to enable "
                "demographic demand modeling. Bachelor's+ education density is the strongest "
                "proxy for medspa demand — correlated with income, health awareness, and "
                "aesthetics spending.",
                variant="info",
            )
        else:
            dd_cards = ""
            dd_cards += kpi_card(
                "Avg Bachelor's+ %",
                f"{dd_summary.get('avg_bachelors_pct', 0):.1f}%",
            )
            dd_cards += kpi_card(
                "Avg Graduate %",
                f"{dd_summary.get('avg_graduate_pct', 0):.1f}%",
            )
            dd_cards += kpi_card("States Analyzed", _fmt(dd_summary.get("states_analyzed")))
            dd_cards += kpi_card(
                "Underserved Educated States",
                _fmt(dd_summary.get("underserved_educated")),
                delta="Bachelor's >30%, <50 prospects",
            )
            body += "\n" + kpi_strip(dd_cards)

            body += f"""<p style="font-size:14px;color:var(--gray-600);margin:12px 0">
            <strong>Demographic demand modeling</strong> uses educational attainment as the
            strongest medspa demand proxy. Bachelor's+ populations correlate with higher income,
            health consciousness, and aesthetics spending. States with high education but low
            medspa penetration signal <strong>underserved demand</strong>.
            Period: <strong>{dd_summary.get('period_year', '?')}</strong>.</p>"""

            # Charts: bar (top 15 by bachelor's %) + doughnut (education distribution)
            body += '<div class="provider-split-grid">'

            if dd_states:
                top_15 = dd_states[:15]
                bar_labels = [s["state_name"][:20] for s in top_15]
                bar_values = [s["avg_bachelors"] for s in top_15]
                bar_config = build_horizontal_bar_config(
                    bar_labels, bar_values,
                    dataset_label="Bachelor's+ %",
                )
                bar_json = json.dumps(bar_config)
                body += "<div>"
                body += chart_container(
                    "demoEduBar", bar_json,
                    build_bar_fallback(bar_labels, bar_values),
                    size="medium",
                    title="Top 15 States by Bachelor's+ %",
                )
                charts_js += chart_init_js("demoEduBar", bar_json)
                body += "</div>"

            if dd_summary.get("avg_bachelors_pct") and dd_summary.get("avg_graduate_pct"):
                other_pct = round(100 - dd_summary["avg_bachelors_pct"] - dd_summary["avg_graduate_pct"], 1)
                if other_pct < 0:
                    other_pct = 0
                edu_labels = ["Bachelor's Degree", "Graduate Degree", "Below Bachelor's"]
                edu_values = [
                    dd_summary["avg_bachelors_pct"],
                    dd_summary["avg_graduate_pct"],
                    other_pct,
                ]
                edu_colors = [BLUE, PURPLE, GRAY]
                donut_config = build_doughnut_config(edu_labels, edu_values, edu_colors)
                donut_json = json.dumps(donut_config)
                body += "<div>"
                body += chart_container(
                    "demoEduDonut", donut_json,
                    build_bar_fallback(edu_labels, edu_values),
                    size="medium",
                    title="Avg Education Level Distribution",
                )
                charts_js += chart_init_js("demoEduDonut", donut_json)
                body += build_chart_legend(edu_labels, edu_values, edu_colors, show_pct=True)
                body += "</div>"

            body += "</div>"

            # Data table: top 25 states by education-to-medspa gap
            if dd_states:
                gap_sorted = sorted(dd_states, key=lambda x: x.get("gap_score", 0), reverse=True)
                body += '<h3 style="font-size:15px;font-weight:600;color:var(--primary);margin:20px 0 8px">Education-to-Medspa Gap Analysis</h3>'
                dd_rows = []
                for s in gap_sorted[:25]:
                    dd_rows.append([
                        s["state_name"][:20],
                        f"{s['avg_bachelors']:.1f}%",
                        f"{s['avg_graduate']:.1f}%",
                        _fmt(s["prospect_count"]),
                        _fmt(s["a_grade_count"]),
                        f"{s['gap_score']:.1f}",
                    ])
                body += '<div class="highlight-table">'
                body += data_table(
                    headers=["State", "Bachelor's %", "Graduate %",
                             "Prospects", "A-Grade", "Gap Score"],
                    rows=dd_rows,
                )
                body += "</div>"

            body += callout(
                f"<strong>{dd_summary.get('underserved_educated', 0)} underserved educated states</strong> "
                "identified with bachelor's rate >30% but fewer than 50 medspa prospects. "
                "These represent the strongest demand-supply mismatch — educated, affluent "
                "populations without adequate aesthetics service coverage.",
                variant="tip",
            )

        body += "\n" + section_end()

        # ==================================================================
        # Section 24: PE Competitive Heat Map
        # ==================================================================
        body += "\n" + section_start(24, "PE Competitive Heat Map", "pe-heatmap")

        pec_summary = pe_competitive.get("summary", {})
        pec_deals = pe_competitive.get("deals", [])
        pec_deal_types = pe_competitive.get("deal_type_breakdown", {})
        pec_state_counts = pe_competitive.get("state_counts", {})

        if not pe_competitive:
            body += callout(
                "<strong>PE deal data not yet ingested.</strong> "
                "Run <code>POST /api/v1/pe/collect</code> to enable "
                "PE competitive heat map analysis. This shows which firms are actively "
                "acquiring aesthetics platforms, deal multiples, and geographic concentration.",
                variant="info",
            )
        else:
            pec_cards = ""
            pec_cards += kpi_card("PE-Backed Platforms", _fmt(pec_summary.get("pe_platforms")))
            pec_cards += kpi_card(
                "Avg EV/EBITDA",
                f"{pec_summary.get('avg_ev_ebitda', 0):.1f}x",
            )
            pec_cards += kpi_card(
                "Total Deal Value",
                _fmt_currency(pec_summary.get("total_deal_value", 0)),
            )
            pec_cards += kpi_card("Most Active Buyer", pec_summary.get("most_active_buyer", "-"))
            body += "\n" + kpi_strip(pec_cards)

            body += """<p style="font-size:14px;color:var(--gray-600);margin:12px 0">
            <strong>PE competitive intelligence</strong> maps active buyers in the aesthetics
            space. Understanding deal multiples, preferred geographies, and deal types (LBO vs
            add-on vs growth) helps position acquisitions competitively and identifies
            white-space markets where PE capital hasn't yet concentrated.</p>"""

            # Charts: bar (deals by state) + doughnut (deal type)
            body += '<div class="provider-split-grid">'

            if pec_state_counts:
                top_states = list(pec_state_counts.items())[:15]
                bar_labels = [s[0] for s in top_states]
                bar_values = [float(s[1]) for s in top_states]
                bar_config = build_horizontal_bar_config(
                    bar_labels, bar_values,
                    dataset_label="Deals",
                )
                bar_json = json.dumps(bar_config)
                body += "<div>"
                body += chart_container(
                    "peHeatBar", bar_json,
                    build_bar_fallback(bar_labels, bar_values),
                    size="medium",
                    title="PE Aesthetics Deals by State",
                )
                charts_js += chart_init_js("peHeatBar", bar_json)
                body += "</div>"

            if pec_deal_types:
                dt_labels = list(pec_deal_types.keys())
                dt_values = [float(v) for v in pec_deal_types.values()]
                dt_colors = [BLUE, GREEN, ORANGE, PURPLE, TEAL, PINK][:len(dt_labels)]
                donut_config = build_doughnut_config(dt_labels, dt_values, dt_colors)
                donut_json = json.dumps(donut_config)
                body += "<div>"
                body += chart_container(
                    "peDealTypeDonut", donut_json,
                    build_bar_fallback(dt_labels, dt_values),
                    size="medium",
                    title="Deal Type Breakdown",
                )
                charts_js += chart_init_js("peDealTypeDonut", donut_json)
                body += build_chart_legend(dt_labels, dt_values, dt_colors, show_pct=True)
                body += "</div>"

            body += "</div>"

            # Data table: recent deals
            if pec_deals:
                body += '<h3 style="font-size:15px;font-weight:600;color:var(--primary);margin:20px 0 8px">Recent Aesthetics PE Deals</h3>'
                deal_rows = []
                for d in pec_deals[:25]:
                    ev_str = _fmt_currency(d["ev_usd"]) if d["ev_usd"] else "-"
                    mult_str = f"{d['ev_ebitda']:.1f}x" if d["ev_ebitda"] else "-"
                    deal_rows.append([
                        d["company"],
                        d["buyer"],
                        d["deal_type"],
                        ev_str,
                        mult_str,
                        d["date"],
                        d["state"],
                    ])
                body += '<div class="highlight-table">'
                body += data_table(
                    headers=["Company", "Buyer", "Type", "EV",
                             "EV/EBITDA", "Date", "State"],
                    rows=deal_rows,
                )
                body += "</div>"

            body += callout(
                f"<strong>{pec_summary.get('total_deals', 0)} aesthetics deals tracked</strong> "
                f"with an average EV/EBITDA of {pec_summary.get('avg_ev_ebitda', 0):.1f}x. "
                "Markets with low PE activity but high medspa density present "
                "first-mover acquisition opportunities before competitive bidding intensifies.",
                variant="tip",
            )

        body += "\n" + section_end()

        # ==================================================================
        # Section 25: Construction Momentum Signal
        # ==================================================================
        body += "\n" + section_start(25, "Construction Momentum Signal", "construction-momentum")

        cm_summary = construction_momentum.get("summary", {})
        cm_states = construction_momentum.get("state_data", [])
        cm_comp = construction_momentum.get("permit_composition", {})

        if not construction_momentum:
            body += callout(
                "<strong>HUD building permit data not yet ingested.</strong> "
                "Run <code>POST /api/v1/realestate/hud-permits/ingest</code> to enable "
                "construction momentum analysis. Rising building permits are a 1-2 year "
                "leading indicator of population growth and future medspa demand.",
                variant="info",
            )
        else:
            cm_cards = ""
            cm_cards += kpi_card("States Analyzed", _fmt(cm_summary.get("states_analyzed")))
            cm_cards += kpi_card(
                "Avg YoY Permit Growth",
                f"{cm_summary.get('avg_yoy_growth', 0):+.1f}%",
            )
            cm_cards += kpi_card(
                "High-Growth States",
                _fmt(cm_summary.get("high_growth_states")),
                delta=">10% YoY growth",
            )
            cm_cards += kpi_card("Top State by Permits", cm_summary.get("top_state", "-"))
            body += "\n" + kpi_strip(cm_cards)

            body += f"""<p style="font-size:14px;color:var(--gray-600);margin:12px 0">
            <strong>Construction momentum</strong> uses HUD building permit data as a
            <strong>1-2 year leading indicator</strong> of population growth. States with
            surging permit activity will see new residents needing services — including medspas.
            Cross-referenced with current medspa density to identify early-mover markets.
            Year: <strong>{cm_summary.get('latest_year', '?')}</strong>.</p>"""

            # Charts: bar (top 15 by YoY growth) + doughnut (permit type)
            body += '<div class="provider-split-grid">'

            if cm_states:
                growth_states = [s for s in cm_states if s.get("yoy_growth") is not None][:15]
                if growth_states:
                    bar_labels = [s["state_name"][:20] for s in growth_states]
                    bar_values = [s["yoy_growth"] for s in growth_states]
                    bar_colors = ["#48bb78" if v >= 0 else "#fc8181" for v in bar_values]
                    bar_config = build_horizontal_bar_config(
                        bar_labels, bar_values, bar_colors,
                        dataset_label="YoY Permit Growth %",
                    )
                    bar_json = json.dumps(bar_config)
                    body += "<div>"
                    body += chart_container(
                        "constructionGrowthBar", bar_json,
                        build_bar_fallback(bar_labels, bar_values),
                        size="medium",
                        title="Top States by YoY Permit Growth",
                    )
                    charts_js += chart_init_js("constructionGrowthBar", bar_json)
                    body += "</div>"

            if cm_comp:
                comp_labels = ["Single-Family", "2-4 Units", "5+ Units"]
                comp_values = [
                    cm_comp.get("single_family_pct", 0),
                    cm_comp.get("two_to_four_pct", 0),
                    cm_comp.get("five_plus_pct", 0),
                ]
                comp_colors = [GREEN, ORANGE, PURPLE]
                donut_config = build_doughnut_config(comp_labels, comp_values, comp_colors)
                donut_json = json.dumps(donut_config)
                body += "<div>"
                body += chart_container(
                    "permitTypeDonut", donut_json,
                    build_bar_fallback(comp_labels, comp_values),
                    size="medium",
                    title="Permit Type Composition",
                )
                charts_js += chart_init_js("permitTypeDonut", donut_json)
                body += build_chart_legend(comp_labels, comp_values, comp_colors, show_pct=True)
                body += "</div>"

            body += "</div>"

            # Data table: top 25 states
            if cm_states:
                body += '<h3 style="font-size:15px;font-weight:600;color:var(--primary);margin:20px 0 8px">State Permit Growth vs Medspa Density</h3>'
                cm_rows = []
                for s in cm_states[:25]:
                    yoy_str = f"{s['yoy_growth']:+.1f}%" if s.get("yoy_growth") is not None else "-"
                    yoy_color = "color:#48bb78" if (s.get("yoy_growth") or 0) >= 0 else "color:#fc8181"
                    yoy_html = f'<span style="{yoy_color};font-weight:600">{yoy_str}</span>'
                    # Growth signal grade
                    yoy_val = s.get("yoy_growth") or 0
                    if yoy_val > 15:
                        grade = "A"
                    elif yoy_val > 5:
                        grade = "B"
                    elif yoy_val > 0:
                        grade = "C"
                    else:
                        grade = "D"
                    cm_rows.append([
                        s["state_name"][:20],
                        _fmt(s["total_permits"]),
                        yoy_html,
                        _fmt(s["medspa_count"]),
                        _fmt(s["a_grade_count"]),
                        _grade_badge(grade),
                    ])
                body += '<div class="highlight-table">'
                body += data_table(
                    headers=["State", "Total Permits", "YoY Growth",
                             "Medspas", "A-Grade", "Signal"],
                    rows=cm_rows,
                )
                body += "</div>"

            body += callout(
                f"<strong>{cm_summary.get('high_growth_states', 0)} high-growth states</strong> "
                "with >10% YoY permit increases. Rising construction = incoming population = "
                "future medspa demand. Target acquisitions in these markets now, before "
                "competition follows population growth.",
                variant="tip",
            )

        body += "\n" + section_end()

        # ==================================================================
        # Section 26: Medical CPI Pricing Power
        # ==================================================================
        body += "\n" + section_start(26, "Medical CPI Pricing Power", "medical-cpi")

        mcpi_summary = medical_cpi.get("summary", {})
        mcpi_annual = medical_cpi.get("annual_data", [])

        if not medical_cpi:
            body += callout(
                "<strong>BLS CPI data not yet ingested.</strong> "
                "Run <code>POST /api/v1/bls/dataset/ingest</code> with medical care CPI series "
                "(CUSR0000SAM, CUSR0000SA0) to enable pricing power analysis. When medical "
                "CPI outpaces general CPI, medspa operators have pricing power — they can "
                "raise prices faster than general inflation.",
                variant="info",
            )
        else:
            mcpi_cards = ""
            mcpi_cards += kpi_card(
                "Current Medical CPI",
                f"{mcpi_summary.get('current_medical_cpi', 0):.1f}",
            )
            mcpi_cards += kpi_card(
                "YoY Medical CPI Change",
                f"{mcpi_summary.get('yoy_medical_change', 0):+.2f}%",
            )
            spread = mcpi_summary.get("medical_vs_general_spread", 0) or 0
            spread_label = "above" if spread >= 0 else "below"
            mcpi_cards += kpi_card(
                "Medical vs General Spread",
                f"{spread:+.2f} pp",
                delta=f"{spread_label} general inflation",
            )
            cagr = mcpi_summary.get("cagr_5yr")
            mcpi_cards += kpi_card(
                "5yr Medical CAGR",
                f"{cagr:.2f}%" if cagr is not None else "-",
            )
            body += "\n" + kpi_strip(mcpi_cards)

            body += """<p style="font-size:14px;color:var(--gray-600);margin:12px 0">
            <strong>Medical CPI vs general CPI</strong> measures medspa pricing power.
            When medical care inflation exceeds general inflation, operators can raise prices
            without losing customers — a structural advantage for the aesthetics sector.
            This is a <strong>national macro signal</strong> supporting the roll-up thesis.</p>"""

            # Chart: horizontal bar (annual medical CPI YoY %)
            if mcpi_annual:
                yoy_data = [d for d in reversed(mcpi_annual) if d.get("medical_yoy") is not None]
                if yoy_data:
                    bar_labels = [str(d["year"]) for d in yoy_data]
                    bar_values = [d["medical_yoy"] for d in yoy_data]
                    bar_colors = [GREEN if v >= 0 else RED for v in bar_values]
                    bar_config = build_horizontal_bar_config(
                        bar_labels, bar_values, bar_colors,
                        dataset_label="Medical CPI YoY %",
                    )
                    bar_json = json.dumps(bar_config)
                    body += chart_container(
                        "medCpiBar", bar_json,
                        build_bar_fallback(bar_labels, bar_values),
                        title="Annual Medical CPI YoY % Change",
                    )
                    charts_js += chart_init_js("medCpiBar", bar_json)

            # Data table: annual comparison
            if mcpi_annual:
                body += '<h3 style="font-size:15px;font-weight:600;color:var(--primary);margin:20px 0 8px">Medical vs General CPI Comparison</h3>'
                cpi_rows = []
                for d in mcpi_annual:
                    med_yoy_str = f"{d['medical_yoy']:+.2f}%" if d.get("medical_yoy") is not None else "-"
                    gen_yoy_str = f"{d['general_yoy']:+.2f}%" if d.get("general_yoy") is not None else "-"
                    spread_str = f"{d['yoy_spread']:+.2f} pp" if d.get("yoy_spread") is not None else "-"
                    spread_color = "color:#48bb78" if (d.get("yoy_spread") or 0) >= 0 else "color:#fc8181"
                    cpi_rows.append([
                        str(d["year"]),
                        f"{d['medical_cpi']:.1f}",
                        f"{d['general_cpi']:.1f}",
                        med_yoy_str,
                        gen_yoy_str,
                        f'<span style="{spread_color};font-weight:600">{spread_str}</span>',
                        f"{d['cumulative_divergence']:+.2f} pp",
                    ])
                body += data_table(
                    headers=["Year", "Medical CPI", "General CPI",
                             "Medical YoY", "General YoY", "Spread", "Cumulative"],
                    rows=cpi_rows,
                )

            signal = "positive" if spread >= 0 else "negative"
            pricing_msg = (
                "Medspa operators have structural pricing power \u2014 revenue growth can outpace cost inflation."
                if spread >= 0
                else "Monitor for reversal; current spread suggests pricing headwinds."
            )
            body += callout(
                f"<strong>Pricing power signal: {signal}.</strong> "
                f"Medical CPI is {'outpacing' if spread >= 0 else 'trailing'} general CPI "
                f"by {abs(spread):.2f} percentage points. {pricing_msg}",
                variant="tip" if spread >= 0 else "info",
            )

        body += "\n" + section_end()

        # ==================================================================
        # Section 27: Talent Pipeline Pressure
        # ==================================================================
        body += "\n" + section_start(27, "Talent Pipeline Pressure", "talent-pipeline")

        tp_summary = talent_pipeline.get("summary", {})
        tp_quarterly = talent_pipeline.get("quarterly_data", [])

        if not talent_pipeline:
            body += callout(
                "<strong>BLS JOLTS data not yet ingested.</strong> "
                "Run <code>POST /api/v1/bls/dataset/ingest</code> with healthcare JOLTS series "
                "(JTU6200000000000JOL, JTU6200000000000HIR) to enable talent pipeline analysis. "
                "The openings-to-hires ratio measures healthcare talent scarcity — a higher "
                "ratio means harder hiring, making well-staffed platforms more valuable.",
                variant="info",
            )
        else:
            tp_cards = ""
            tp_cards += kpi_card(
                "Latest HC Openings (000s)",
                f"{tp_summary.get('latest_openings', 0):.0f}",
            )
            tp_cards += kpi_card(
                "Openings/Hires Ratio",
                f"{tp_summary.get('openings_to_hires_ratio', 0):.2f}",
            )
            yoy_ch = tp_summary.get("yoy_change")
            tp_cards += kpi_card(
                "YoY Change",
                f"{yoy_ch:+.1f}%" if yoy_ch is not None else "-",
            )
            trend = tp_summary.get("trend", "stable")
            trend_display = {"tightening": "Tightening", "easing": "Easing", "stable": "Stable"}.get(trend, trend.title())
            tp_cards += kpi_card("Talent Scarcity Trend", trend_display)
            body += "\n" + kpi_strip(tp_cards)

            body += """<p style="font-size:14px;color:var(--gray-600);margin:12px 0">
            <strong>Talent pipeline pressure</strong> uses BLS JOLTS (Job Openings and Labor
            Turnover Survey) for healthcare. A high openings-to-hires ratio signals talent
            scarcity — well-staffed medspa platforms become <strong>competitive moats</strong>.
            PE acquirers should target platforms with strong teams; M&A to acquire talent,
            not just locations. This is a <strong>national macro signal</strong>.</p>"""

            # Chart: horizontal bar (quarterly openings trend)
            if tp_quarterly:
                recent_q = list(reversed(tp_quarterly[:8]))
                bar_labels = [q["period"] for q in recent_q]
                bar_values = [q["openings"] for q in recent_q]
                bar_config = build_horizontal_bar_config(
                    bar_labels, bar_values,
                    dataset_label="Healthcare Openings (000s)",
                )
                bar_json = json.dumps(bar_config)
                body += chart_container(
                    "talentOpeningsBar", bar_json,
                    build_bar_fallback(bar_labels, bar_values),
                    title="Healthcare Job Openings Trend (Quarterly)",
                )
                charts_js += chart_init_js("talentOpeningsBar", bar_json)

            # Data table: quarterly data
            if tp_quarterly:
                body += '<h3 style="font-size:15px;font-weight:600;color:var(--primary);margin:20px 0 8px">Quarterly Healthcare Labor Market Data</h3>'
                tp_rows = []
                for i, q in enumerate(tp_quarterly[:12]):
                    # Compute QoQ change for ratio
                    prev_ratio = tp_quarterly[i + 1]["ratio"] if i + 1 < len(tp_quarterly) else None
                    ratio_change = ""
                    if prev_ratio and prev_ratio > 0:
                        rc = (q["ratio"] - prev_ratio) / prev_ratio * 100
                        rc_color = "color:#fc8181" if rc > 0 else "color:#48bb78"
                        ratio_change = f'<span style="{rc_color}">{rc:+.1f}%</span>'

                    tp_rows.append([
                        q["period"],
                        f"{q['openings']:.0f}",
                        f"{q['hires']:.0f}" if q["hires"] else "-",
                        f"{q['ratio']:.2f}",
                        ratio_change or "-",
                    ])
                body += data_table(
                    headers=["Period", "Openings (000s)", "Hires (000s)",
                             "Open/Hire Ratio", "Ratio Change"],
                    rows=tp_rows,
                )

            talent_msg = (
                "A tightening labor market makes staffed platforms premium assets \u2014 "
                "acquirers should prioritize retention and team quality."
                if trend == "tightening"
                else "Current labor conditions are manageable, but monitor for shifts "
                "that could impact staffing costs."
            )
            body += callout(
                f"<strong>Talent scarcity trend: {trend_display.lower()}.</strong> "
                f"Healthcare openings-to-hires ratio is {tp_summary.get('openings_to_hires_ratio', 0):.2f}. "
                f"{talent_msg}",
                variant="tip" if trend == "tightening" else "info",
            )

        body += "\n" + section_end()

        # ---- Close container ----
        body += "\n</div>"

        # ---- Footer ----
        notes = [
            "Prospect data sourced from Yelp Business Search API; scores reflect publicly available ratings and review counts.",
            "ZIP affluence model uses IRS Statistics of Income (SOI) ZIP-level data, tax year 2021.",
            "PE competitive landscape data from SEC EDGAR filings and company websites.",
            "Acquisition scores are model-generated estimates and should be validated with on-the-ground diligence.",
            "Deal model uses industry benchmark economics (AmSpa, IBISWorld) — actual financials require location-level diligence.",
            "Stealth Wealth Signal uses IRS SOI ZIP-level income composition; dollar amounts reported in thousands.",
            "Migration Alpha uses IRS SOI county-to-county migration flows as a 2-3 year leading indicator.",
            "Provider Density Signal uses CMS Medicare Provider Utilization data for aesthetics-adjacent specialties.",
            "Real Estate Alpha uses Redfin ZIP-level data (preferred) or FHFA House Price Index (ZIP3 fallback).",
            "Deposit Wealth uses FDIC Summary of Deposits; deposit amounts are actual dollars.",
            "Business Formation uses IRS SOI Business Income; dollar amounts multiplied by 1,000 from source.",
            "Opportunity Zone data from CDFI Fund; 10-year capital gains deferral applies to Qualified Opportunity Fund investments.",
            "Demographic Demand uses Census ACS educational attainment as medspa demand proxy.",
            "PE Competitive Heat Map filters PE portfolio/deal databases for aesthetics-related keywords.",
            "Construction Momentum uses HUD SOCDS building permit data as 1-2 year population growth leading indicator.",
            "Medical CPI uses BLS Consumer Price Index series CUSR0000SAM (medical) vs CUSR0000SA0 (all items).",
            "Talent Pipeline uses BLS JOLTS healthcare series; openings-to-hires ratio measures labor scarcity.",
            "This report does not constitute investment advice. All data is from public sources.",
        ]
        body += "\n" + page_footer(
            notes=notes,
            generated_line=f"Report generated {data.get('generated_at', 'N/A')} | Nexdata Investment Intelligence",
        )

        return html_document(
            title=f"Med-Spa Market Analysis{scope_label}",
            body_content=body,
            charts_js=charts_js,
            extra_css=MEDSPA_EXTRA_CSS,
        )

    # ------------------------------------------------------------------
    # Excel Rendering
    # ------------------------------------------------------------------

    def render_excel(self, data: Dict[str, Any]) -> bytes:
        """Render report as Excel workbook."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()

        header_fill = PatternFill(
            start_color="1A365D", end_color="1A365D", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=16)
        section_font = Font(bold=True, size=13)

        summary = data.get("summary", {})

        # ---- Sheet 1: Executive Summary ----
        ws = wb.active
        ws.title = "Executive Summary"

        ws["A1"] = "Med-Spa Market Analysis"
        ws["A1"].font = title_font
        ws.merge_cells("A1:D1")

        ws["A3"] = "Total Prospects"
        ws["B3"] = summary.get("total_prospects", 0)
        ws["A4"] = "A-Grade Targets"
        ws["B4"] = summary.get("a_grade", 0)
        ws["A5"] = "A/B-Grade Targets"
        ws["B5"] = summary.get("ab_grade", 0)
        ws["A6"] = "States Covered"
        ws["B6"] = summary.get("states_covered", 0)
        ws["A7"] = "ZIPs Covered"
        ws["B7"] = summary.get("zips_covered", 0)
        ws["A8"] = "Avg Acquisition Score"
        ws["B8"] = summary.get("avg_score", 0)
        ws["A9"] = "Avg Yelp Rating"
        ws["B9"] = summary.get("avg_rating", 0)
        ws["A10"] = "Avg Reviews"
        ws["B10"] = summary.get("avg_reviews", 0)

        for row in range(3, 11):
            ws[f"A{row}"].font = Font(bold=True)

        ws["A12"] = "Grade Distribution"
        ws["A12"].font = section_font
        row_num = 13
        for gd in data.get("grade_distribution", []):
            ws[f"A{row_num}"] = f"Grade {gd['grade']}"
            ws[f"B{row_num}"] = gd["count"]
            ws[f"C{row_num}"] = f"{gd['pct']}%"
            row_num += 1

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 10

        # ---- Sheet 2: Top Targets ----
        ws_targets = wb.create_sheet("Top Targets")
        top_targets = data.get("top_targets", [])

        headers = [
            "Rank", "Name", "City", "State", "ZIP", "Score", "Grade",
            "Rating", "Reviews", "Price", "ZIP Grade", "ZIP Score",
            "Competitors in ZIP", "Phone",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws_targets.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for i, t in enumerate(top_targets, 1):
            row_num = i + 1
            ws_targets.cell(row=row_num, column=1, value=i)
            ws_targets.cell(row=row_num, column=2, value=t.get("name"))
            ws_targets.cell(row=row_num, column=3, value=t.get("city"))
            ws_targets.cell(row=row_num, column=4, value=t.get("state"))
            ws_targets.cell(row=row_num, column=5, value=t.get("zip_code"))
            ws_targets.cell(row=row_num, column=6, value=t.get("score"))
            ws_targets.cell(row=row_num, column=7, value=t.get("grade"))
            ws_targets.cell(row=row_num, column=8, value=t.get("rating"))
            ws_targets.cell(row=row_num, column=9, value=t.get("reviews"))
            ws_targets.cell(row=row_num, column=10, value=t.get("price"))
            ws_targets.cell(row=row_num, column=11, value=t.get("zip_grade"))
            ws_targets.cell(row=row_num, column=12, value=t.get("zip_score"))
            ws_targets.cell(row=row_num, column=13, value=t.get("competitors"))
            ws_targets.cell(row=row_num, column=14, value=t.get("phone"))

        col_widths = {
            "A": 6, "B": 35, "C": 18, "D": 8, "E": 10, "F": 8,
            "G": 8, "H": 8, "I": 10, "J": 8, "K": 10, "L": 10,
            "M": 18, "N": 15,
        }
        for col_letter, width in col_widths.items():
            ws_targets.column_dimensions[col_letter].width = width

        # ---- Sheet 3: By State ----
        ws_state = wb.create_sheet("By State")
        state_avgs = data.get("state_avg_scores", [])

        headers = ["State", "Prospects", "Avg Score", "Max Score", "A-Grade Count"]
        for col, header in enumerate(headers, 1):
            cell = ws_state.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for i, s in enumerate(state_avgs, 2):
            ws_state.cell(row=i, column=1, value=s.get("state"))
            ws_state.cell(row=i, column=2, value=s.get("prospect_count"))
            ws_state.cell(row=i, column=3, value=s.get("avg_score"))
            ws_state.cell(row=i, column=4, value=s.get("max_score"))
            ws_state.cell(row=i, column=5, value=s.get("a_count"))

        ws_state.column_dimensions["A"].width = 8
        ws_state.column_dimensions["B"].width = 12
        ws_state.column_dimensions["C"].width = 12
        ws_state.column_dimensions["D"].width = 12
        ws_state.column_dimensions["E"].width = 15

        # ---- Sheet 4: ZIP Concentration ----
        ws_zip = wb.create_sheet("ZIP Concentration")
        zip_conc = data.get("zip_concentration", [])

        headers = ["ZIP", "City", "State", "Prospects", "Avg Score", "ZIP Score"]
        for col, header in enumerate(headers, 1):
            cell = ws_zip.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for i, z in enumerate(zip_conc, 2):
            ws_zip.cell(row=i, column=1, value=z.get("zip_code"))
            ws_zip.cell(row=i, column=2, value=z.get("city"))
            ws_zip.cell(row=i, column=3, value=z.get("state"))
            ws_zip.cell(row=i, column=4, value=z.get("prospect_count"))
            ws_zip.cell(row=i, column=5, value=z.get("avg_score"))
            ws_zip.cell(row=i, column=6, value=z.get("avg_zip_score"))

        for col_letter in ["A", "B", "C", "D", "E", "F"]:
            ws_zip.column_dimensions[col_letter].width = 14

        # ---- Sheet 5: ZIP Affluence ----
        ws_aff = wb.create_sheet("ZIP Affluence")
        zip_affluence = data.get("zip_affluence_by_state", [])

        headers = ["State", "ZIPs Scored", "Avg Score", "Affluence Score", "Avg AGI", "A-Grade ZIPs"]
        for col, header in enumerate(headers, 1):
            cell = ws_aff.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for i, z in enumerate(zip_affluence, 2):
            ws_aff.cell(row=i, column=1, value=z.get("state"))
            ws_aff.cell(row=i, column=2, value=z.get("zip_count"))
            ws_aff.cell(row=i, column=3, value=z.get("avg_score"))
            ws_aff.cell(row=i, column=4, value=z.get("avg_affluence"))
            ws_aff.cell(row=i, column=5, value=z.get("avg_agi"))
            ws_aff.cell(row=i, column=6, value=z.get("a_zips"))

        for col_letter in ["A", "B", "C", "D", "E", "F"]:
            ws_aff.column_dimensions[col_letter].width = 16

        # ---- Sheet 6: PE Comps ----
        ws_comps = wb.create_sheet("PE Comps")
        pe_comps = data.get("pe_comps", [])

        headers = ["Company", "Industry", "Sub-Industry", "Location", "PE Owner", "Employees", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws_comps.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for i, c in enumerate(pe_comps, 2):
            location = ", ".join(p for p in [c.get("city"), c.get("state")] if p) or "-"
            ws_comps.cell(row=i, column=1, value=c.get("name"))
            ws_comps.cell(row=i, column=2, value=c.get("industry"))
            ws_comps.cell(row=i, column=3, value=c.get("sub_industry"))
            ws_comps.cell(row=i, column=4, value=location)
            ws_comps.cell(row=i, column=5, value=c.get("pe_owner"))
            ws_comps.cell(row=i, column=6, value=c.get("employees"))
            ws_comps.cell(row=i, column=7, value=c.get("status"))

        col_widths = {"A": 30, "B": 20, "C": 20, "D": 20, "E": 25, "F": 12, "G": 10}
        for col_letter, width in col_widths.items():
            ws_comps.column_dimensions[col_letter].width = width

        # ---- Sheet 7: Whitespace ZIPs ----
        ws_ws = wb.create_sheet("Whitespace ZIPs")
        whitespace_zips = data.get("whitespace_zips", [])

        headers = ["ZIP", "State", "Score", "Avg AGI", "Tax Returns", "Affluence Score"]
        for col, header in enumerate(headers, 1):
            cell = ws_ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for i, z in enumerate(whitespace_zips, 2):
            ws_ws.cell(row=i, column=1, value=z.get("zip_code"))
            ws_ws.cell(row=i, column=2, value=z.get("state"))
            ws_ws.cell(row=i, column=3, value=z.get("score"))
            ws_ws.cell(row=i, column=4, value=z.get("avg_agi"))
            ws_ws.cell(row=i, column=5, value=z.get("total_returns"))
            ws_ws.cell(row=i, column=6, value=z.get("affluence_score"))

        for col_letter in ["A", "B", "C", "D", "E", "F"]:
            ws_ws.column_dimensions[col_letter].width = 16

        # ---- Sheet 8: Workforce Wages ----
        ws_wage = wb.create_sheet("Workforce Wages")
        bls_wages = data.get("bls_wages", {})

        if bls_wages:
            all_years = sorted(set(yr for occ_data in bls_wages.values() for yr in occ_data.keys()))
            wage_headers = ["Occupation"] + [str(y) for y in all_years]
            for col, header in enumerate(wage_headers, 1):
                cell = ws_wage.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill

            for i, occ in enumerate(sorted(bls_wages.keys()), 2):
                ws_wage.cell(row=i, column=1, value=occ)
                for j, yr in enumerate(all_years, 2):
                    val = bls_wages[occ].get(yr)
                    ws_wage.cell(row=i, column=j, value=val)

            ws_wage.column_dimensions["A"].width = 25
            for j in range(2, len(all_years) + 2):
                from openpyxl.utils import get_column_letter
                ws_wage.column_dimensions[get_column_letter(j)].width = 14

        # ---- Sheet 9: Growth Signals ----
        ws_gs = wb.create_sheet("Growth Signals")
        top_by_reviews = data.get("top_by_reviews", [])
        low_competition_gems = data.get("low_competition_gems", [])

        # Top by reviews
        ws_gs["A1"] = "Top Prospects by Review Volume"
        ws_gs["A1"].font = Font(bold=True, size=13)
        ws_gs.merge_cells("A1:H1")

        headers = ["Name", "City", "State", "Reviews", "Rating", "Score", "Grade", "Competitors"]
        for col, header in enumerate(headers, 1):
            cell = ws_gs.cell(row=2, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for i, t in enumerate(top_by_reviews, 3):
            ws_gs.cell(row=i, column=1, value=t.get("name"))
            ws_gs.cell(row=i, column=2, value=t.get("city"))
            ws_gs.cell(row=i, column=3, value=t.get("state"))
            ws_gs.cell(row=i, column=4, value=t.get("reviews"))
            ws_gs.cell(row=i, column=5, value=t.get("rating"))
            ws_gs.cell(row=i, column=6, value=t.get("score"))
            ws_gs.cell(row=i, column=7, value=t.get("grade"))
            ws_gs.cell(row=i, column=8, value=t.get("competitors"))

        # Low competition gems section
        gem_start = len(top_by_reviews) + 5
        ws_gs.cell(row=gem_start, column=1, value="Low Competition Gems (≤3 Competitors)")
        ws_gs.cell(row=gem_start, column=1).font = Font(bold=True, size=13)
        ws_gs.merge_cells(f"A{gem_start}:H{gem_start}")

        headers = ["Name", "City", "State", "Score", "Grade", "Rating", "Reviews", "Competitors"]
        for col, header in enumerate(headers, 1):
            cell = ws_gs.cell(row=gem_start + 1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for i, g in enumerate(low_competition_gems, gem_start + 2):
            ws_gs.cell(row=i, column=1, value=g.get("name"))
            ws_gs.cell(row=i, column=2, value=g.get("city"))
            ws_gs.cell(row=i, column=3, value=g.get("state"))
            ws_gs.cell(row=i, column=4, value=g.get("score"))
            ws_gs.cell(row=i, column=5, value=g.get("grade"))
            ws_gs.cell(row=i, column=6, value=g.get("rating"))
            ws_gs.cell(row=i, column=7, value=g.get("reviews"))
            ws_gs.cell(row=i, column=8, value=g.get("competitors"))

        col_widths = {"A": 35, "B": 18, "C": 8, "D": 10, "E": 8, "F": 8, "G": 10, "H": 14}
        for col_letter, width in col_widths.items():
            ws_gs.column_dimensions[col_letter].width = width

        # ---- Sheet 10: Deal Model ----
        ws_deal = wb.create_sheet("Deal Model")
        deal_model = data.get("deal_model", {})
        dm_tiers = deal_model.get("tier_economics", [])
        dm_scenarios = deal_model.get("scenarios", {})
        dm_cap = deal_model.get("capital_stack", {})

        ws_deal["A1"] = "Deal Model — Unit Economics"
        ws_deal["A1"].font = Font(bold=True, size=13)
        ws_deal.merge_cells("A1:H1")

        # Tier economics table
        deal_headers = [
            "Price Tier", "Locations", "Avg Revenue", "EBITDA Margin",
            "Avg EBITDA", "Entry Multiple", "Total Revenue", "Total EBITDA", "Total Acq Cost",
        ]
        for col, header in enumerate(deal_headers, 1):
            cell = ws_deal.cell(row=2, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for i, t in enumerate(dm_tiers, 3):
            ws_deal.cell(row=i, column=1, value=t.get("tier", "Unknown"))
            ws_deal.cell(row=i, column=2, value=t.get("count", 0))
            ws_deal.cell(row=i, column=3, value=t.get("avg_revenue", 0))
            ws_deal.cell(row=i, column=4, value=t.get("ebitda_margin", 0))
            ws_deal.cell(row=i, column=5, value=t.get("avg_ebitda", 0))
            ws_deal.cell(row=i, column=6, value=t.get("entry_multiple", 0))
            ws_deal.cell(row=i, column=7, value=t.get("total_revenue", 0))
            ws_deal.cell(row=i, column=8, value=t.get("total_ebitda", 0))
            ws_deal.cell(row=i, column=9, value=t.get("total_acq_cost", 0))

        # Totals row
        total_row = len(dm_tiers) + 3
        ws_deal.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws_deal.cell(row=total_row, column=2, value=deal_model.get("total_locations", 0)).font = Font(bold=True)
        ws_deal.cell(row=total_row, column=7, value=deal_model.get("total_revenue", 0)).font = Font(bold=True)
        ws_deal.cell(row=total_row, column=8, value=deal_model.get("total_ebitda", 0)).font = Font(bold=True)
        ws_deal.cell(row=total_row, column=9, value=deal_model.get("total_acquisition_cost", 0)).font = Font(bold=True)

        # Capital stack section
        cap_start = total_row + 2
        ws_deal.cell(row=cap_start, column=1, value="Capital Requirements").font = Font(bold=True, size=13)
        ws_deal.merge_cells(f"A{cap_start}:D{cap_start}")

        cap_items = [
            ("Equity Contribution (40%)", dm_cap.get("equity", 0)),
            ("Senior Debt (60%)", dm_cap.get("debt", 0)),
            ("Transaction Costs (5%)", dm_cap.get("transaction_costs", 0)),
            ("Working Capital Reserve", dm_cap.get("working_capital", 0)),
            ("Total Capital Required", dm_cap.get("total_capital_required", 0)),
        ]
        for j, (label, value) in enumerate(cap_items, cap_start + 1):
            ws_deal.cell(row=j, column=1, value=label)
            ws_deal.cell(row=j, column=2, value=value)
            if label.startswith("Total"):
                ws_deal.cell(row=j, column=1).font = Font(bold=True)
                ws_deal.cell(row=j, column=2).font = Font(bold=True)

        ws_deal.cell(row=cap_start + len(cap_items) + 1, column=1, value="Leverage (Debt/EBITDA)")
        ws_deal.cell(row=cap_start + len(cap_items) + 1, column=2, value=deal_model.get("leverage_ratio", 0))

        # Scenarios section
        sc_start = cap_start + len(cap_items) + 3
        ws_deal.cell(row=sc_start, column=1, value="Returns Analysis").font = Font(bold=True, size=13)
        ws_deal.merge_cells(f"A{sc_start}:H{sc_start}")

        sc_headers = ["Scenario", "Entry EV", "Exit Multiple", "Margin Improvement",
                      "Hold Years", "Exit EV", "Gross MOIC", "Net IRR"]
        for col, header in enumerate(sc_headers, 1):
            cell = ws_deal.cell(row=sc_start + 1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        sc_row = sc_start + 2
        for sc_name in ["conservative", "base", "aggressive"]:
            s = dm_scenarios.get(sc_name, {})
            if not s:
                continue
            ws_deal.cell(row=sc_row, column=1, value=sc_name.title())
            ws_deal.cell(row=sc_row, column=2, value=s.get("entry_ev", 0))
            ws_deal.cell(row=sc_row, column=3, value=s.get("exit_multiple", 0))
            ws_deal.cell(row=sc_row, column=4, value=s.get("margin_improvement", 0))
            ws_deal.cell(row=sc_row, column=5, value=s.get("hold_years", 0))
            ws_deal.cell(row=sc_row, column=6, value=s.get("exit_ev", 0))
            ws_deal.cell(row=sc_row, column=7, value=s.get("gross_moic", 0))
            ws_deal.cell(row=sc_row, column=8, value=s.get("net_irr", 0))
            sc_row += 1

        deal_col_widths = {"A": 28, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16, "G": 16, "H": 14, "I": 16}
        for col_letter, width in deal_col_widths.items():
            ws_deal.column_dimensions[col_letter].width = width

        # ---- Sheet 11: Stealth Wealth ----
        ws_sw = wb.create_sheet("Stealth Wealth")
        stealth_wealth = data.get("stealth_wealth", {})
        sw_comp = stealth_wealth.get("wealth_composition", {})
        sw_zips = stealth_wealth.get("stealth_zips", [])
        sw_summary = stealth_wealth.get("summary", {})

        ws_sw["A1"] = "Stealth Wealth Signal"
        ws_sw["A1"].font = Font(bold=True, size=13)
        ws_sw.merge_cells("A1:E1")

        # Summary stats
        ws_sw["A3"] = "Tax Year"
        ws_sw["B3"] = sw_summary.get("tax_year", "-")
        ws_sw["A4"] = "Stealth ZIPs Found"
        ws_sw["B4"] = sw_summary.get("total_stealth", 0)
        ws_sw["A5"] = "Validated (Score ≥70)"
        ws_sw["B5"] = sw_summary.get("validated", 0)
        ws_sw["A6"] = "Avg Non-Wage Income"
        ws_sw["B6"] = sw_summary.get("avg_non_wage_income", 0)
        for r in range(3, 7):
            ws_sw[f"A{r}"].font = Font(bold=True)

        # Wealth composition
        if sw_comp:
            ws_sw["A8"] = "Income Composition"
            ws_sw["A8"].font = Font(bold=True, size=11)
            ws_sw["A9"] = "W-2 Wages %"
            ws_sw["B9"] = sw_comp.get("wages_pct", 0)
            ws_sw["A10"] = "Capital Gains %"
            ws_sw["B10"] = sw_comp.get("cap_gains_pct", 0)
            ws_sw["A11"] = "Dividends %"
            ws_sw["B11"] = sw_comp.get("dividends_pct", 0)
            ws_sw["A12"] = "Business/Partnership %"
            ws_sw["B12"] = sw_comp.get("biz_income_pct", 0)
            ws_sw["A13"] = "Other %"
            ws_sw["B13"] = sw_comp.get("other_pct", 0)

        # Stealth ZIPs table
        sw_start = 15
        ws_sw.cell(row=sw_start, column=1, value="Top Stealth Wealth ZIPs").font = Font(bold=True, size=11)
        sw_headers = [
            "ZIP", "State", "Non-Wage/Return", "Non-Wage %",
            "Returns", "Avg AGI", "Medspa Score", "Grade", "A-Grade Medspas",
        ]
        for col, header in enumerate(sw_headers, 1):
            cell = ws_sw.cell(row=sw_start + 1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for i, z in enumerate(sw_zips, sw_start + 2):
            ws_sw.cell(row=i, column=1, value=z.get("zip_code"))
            ws_sw.cell(row=i, column=2, value=z.get("state"))
            ws_sw.cell(row=i, column=3, value=z.get("non_wage_per_return", 0))
            ws_sw.cell(row=i, column=4, value=z.get("non_wage_pct", 0))
            ws_sw.cell(row=i, column=5, value=z.get("num_returns", 0))
            ws_sw.cell(row=i, column=6, value=z.get("avg_agi", 0))
            ws_sw.cell(row=i, column=7, value=z.get("medspa_score", 0))
            ws_sw.cell(row=i, column=8, value=z.get("medspa_grade", "-"))
            ws_sw.cell(row=i, column=9, value=z.get("a_grade_medspas", 0))

        sw_col_widths = {"A": 12, "B": 8, "C": 16, "D": 12, "E": 10, "F": 12, "G": 14, "H": 8, "I": 16}
        for col_letter, width in sw_col_widths.items():
            ws_sw.column_dimensions[col_letter].width = width

        # ---- Sheet 12: Migration Alpha ----
        ws_ma = wb.create_sheet("Migration Alpha")
        migration_alpha = data.get("migration_alpha", {})
        ma_flows = migration_alpha.get("state_flows", [])
        ma_emerging = migration_alpha.get("emerging_markets", [])
        ma_summary = migration_alpha.get("summary", {})

        ws_ma["A1"] = "Migration Alpha"
        ws_ma["A1"].font = Font(bold=True, size=13)
        ws_ma.merge_cells("A1:E1")

        ws_ma["A3"] = "Tax Year"
        ws_ma["B3"] = ma_summary.get("tax_year", "-")
        ws_ma["A4"] = "Top Gainer"
        ws_ma["B4"] = ma_summary.get("top_gainer", "-")
        ws_ma["A5"] = "Top Loser"
        ws_ma["B5"] = ma_summary.get("top_loser", "-")
        ws_ma["A6"] = "Emerging Markets"
        ws_ma["B6"] = ma_summary.get("emerging_count", 0)
        for r in range(3, 7):
            ws_ma[f"A{r}"].font = Font(bold=True)

        # State flows table
        ma_start = 8
        ws_ma.cell(row=ma_start, column=1, value="State Wealth Flows").font = Font(bold=True, size=11)
        ma_headers = [
            "State", "Inflow AGI ($M)", "Outflow AGI ($M)", "Net AGI ($M)",
            "Inflow Returns", "Total Medspas", "A-Grade", "Migration Alpha",
        ]
        for col, header in enumerate(ma_headers, 1):
            cell = ws_ma.cell(row=ma_start + 1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for i, s in enumerate(ma_flows, ma_start + 2):
            ws_ma.cell(row=i, column=1, value=s.get("state"))
            ws_ma.cell(row=i, column=2, value=s.get("inflow_agi_m", 0))
            ws_ma.cell(row=i, column=3, value=s.get("outflow_agi_m", 0))
            ws_ma.cell(row=i, column=4, value=s.get("net_agi_m", 0))
            ws_ma.cell(row=i, column=5, value=s.get("inflow_returns", 0))
            ws_ma.cell(row=i, column=6, value=s.get("total_medspas", 0))
            ws_ma.cell(row=i, column=7, value=s.get("a_grade_count", 0))
            ws_ma.cell(row=i, column=8, value=s.get("migration_alpha", 0))

        # Emerging markets highlight
        em_start = ma_start + len(ma_flows) + 4
        ws_ma.cell(row=em_start, column=1, value="Emerging Markets (<20 A-Grade, Positive Inflow)").font = Font(bold=True, size=11)
        ws_ma.merge_cells(f"A{em_start}:F{em_start}")

        em_headers = ["State", "Net AGI ($M)", "Inflow Returns", "Total Medspas", "A-Grade", "Migration Alpha"]
        for col, header in enumerate(em_headers, 1):
            cell = ws_ma.cell(row=em_start + 1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        green_fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
        for i, s in enumerate(ma_emerging, em_start + 2):
            ws_ma.cell(row=i, column=1, value=s.get("state"))
            ws_ma.cell(row=i, column=2, value=s.get("net_agi_m", 0))
            ws_ma.cell(row=i, column=3, value=s.get("inflow_returns", 0))
            ws_ma.cell(row=i, column=4, value=s.get("total_medspas", 0))
            ws_ma.cell(row=i, column=5, value=s.get("a_grade_count", 0))
            ws_ma.cell(row=i, column=6, value=s.get("migration_alpha", 0))
            for c in range(1, 7):
                ws_ma.cell(row=i, column=c).fill = green_fill

        ma_col_widths = {"A": 12, "B": 16, "C": 16, "D": 14, "E": 14, "F": 14, "G": 10, "H": 16}
        for col_letter, width in ma_col_widths.items():
            ws_ma.column_dimensions[col_letter].width = width

        # ---- Sheet 13: Provider Density ----
        ws_pd = wb.create_sheet("Provider Density")
        provider_density = data.get("provider_density", {})
        pd_opp_zips = provider_density.get("opportunity_zips", [])
        pd_summary = provider_density.get("summary", {})
        pd_by_type = provider_density.get("provider_by_type", [])

        ws_pd["A1"] = "Medical Provider Density Signal"
        ws_pd["A1"].font = Font(bold=True, size=13)
        ws_pd.merge_cells("A1:E1")

        ws_pd["A3"] = "Provider-ZIP Pairs"
        ws_pd["B3"] = pd_summary.get("total_provider_zip_pairs", 0)
        ws_pd["A4"] = "Opportunity ZIPs (score < 60)"
        ws_pd["B4"] = pd_summary.get("opportunity_zips", 0)
        ws_pd["A5"] = "Avg Providers/ZIP"
        ws_pd["B5"] = pd_summary.get("avg_providers_per_zip", 0)
        ws_pd["A6"] = "Top Provider Type"
        ws_pd["B6"] = pd_summary.get("top_provider_type", "-")
        for r in range(3, 7):
            ws_pd[f"A{r}"].font = Font(bold=True)

        # Provider type breakdown
        if pd_by_type:
            ws_pd["A8"] = "Provider Counts by Specialty"
            ws_pd["A8"].font = Font(bold=True, size=11)
            for i, t in enumerate(pd_by_type, 9):
                ws_pd.cell(row=i, column=1, value=t.get("type"))
                ws_pd.cell(row=i, column=2, value=t.get("count", 0))
                ws_pd.cell(row=i, column=3, value=t.get("beneficiaries", 0))

        pd_start = 9 + len(pd_by_type) + 2
        ws_pd.cell(row=pd_start, column=1, value="Opportunity ZIPs").font = Font(bold=True, size=11)
        pd_headers = [
            "ZIP", "Providers", "Types", "Beneficiaries", "Medspas",
            "Imbalance Score", "Medspa Score", "Grade",
        ]
        for col, header in enumerate(pd_headers, 1):
            cell = ws_pd.cell(row=pd_start + 1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for i, z in enumerate(pd_opp_zips, pd_start + 2):
            ws_pd.cell(row=i, column=1, value=z.get("zip_code"))
            ws_pd.cell(row=i, column=2, value=z.get("provider_count", 0))
            ws_pd.cell(row=i, column=3, value=z.get("provider_types", ""))
            ws_pd.cell(row=i, column=4, value=z.get("beneficiaries", 0))
            ws_pd.cell(row=i, column=5, value=z.get("medspa_count", 0))
            ws_pd.cell(row=i, column=6, value=z.get("imbalance_score", 0))
            ws_pd.cell(row=i, column=7, value=z.get("medspa_score", 0))
            ws_pd.cell(row=i, column=8, value=z.get("medspa_grade", "-"))

        pd_col_widths = {"A": 12, "B": 12, "C": 35, "D": 14, "E": 10, "F": 16, "G": 14, "H": 8}
        for col_letter, width in pd_col_widths.items():
            ws_pd.column_dimensions[col_letter].width = width

        # ---- Sheet 14: RE Appreciation Alpha ----
        ws_re = wb.create_sheet("RE Appreciation Alpha")
        real_estate_alpha = data.get("real_estate_alpha", {})
        re_timing = real_estate_alpha.get("timing_zips", [])
        re_states = real_estate_alpha.get("state_summary", [])
        re_summary = real_estate_alpha.get("summary", {})

        ws_re["A1"] = "Real Estate Appreciation Alpha"
        ws_re["A1"].font = Font(bold=True, size=13)
        ws_re.merge_cells("A1:E1")

        ws_re["A3"] = "Source"
        ws_re["B3"] = real_estate_alpha.get("source", "-")
        ws_re["A4"] = "Avg Median Price"
        ws_re["B4"] = re_summary.get("avg_median_price", 0)
        ws_re["A5"] = "Avg YoY Appreciation %"
        ws_re["B5"] = re_summary.get("avg_yoy_change", 0)
        ws_re["A6"] = "Timing Opportunity ZIPs"
        ws_re["B6"] = re_summary.get("timing_opportunity_zips", 0)
        ws_re["A7"] = "Hottest State"
        ws_re["B7"] = re_summary.get("hottest_state", "-")
        for r in range(3, 8):
            ws_re[f"A{r}"].font = Font(bold=True)

        # State summary
        re_st_start = 9
        ws_re.cell(row=re_st_start, column=1, value="State Summary").font = Font(bold=True, size=11)
        st_headers = ["State", "Avg Median Price", "Avg YoY %", "ZIP Count"]
        for col, header in enumerate(st_headers, 1):
            cell = ws_re.cell(row=re_st_start + 1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        for i, s in enumerate(re_states, re_st_start + 2):
            ws_re.cell(row=i, column=1, value=s.get("state"))
            ws_re.cell(row=i, column=2, value=s.get("avg_median_price", 0))
            ws_re.cell(row=i, column=3, value=s.get("avg_yoy_change", 0))
            ws_re.cell(row=i, column=4, value=s.get("zip_count", 0))

        # Timing ZIPs
        re_tz_start = re_st_start + len(re_states) + 4
        ws_re.cell(row=re_tz_start, column=1, value="Timing Opportunity ZIPs").font = Font(bold=True, size=11)
        tz_headers = ["ZIP", "State", "Median Price", "YoY Change %", "Medspa Score", "Grade"]
        for col, header in enumerate(tz_headers, 1):
            cell = ws_re.cell(row=re_tz_start + 1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        for i, z in enumerate(re_timing, re_tz_start + 2):
            ws_re.cell(row=i, column=1, value=z.get("zip_code"))
            ws_re.cell(row=i, column=2, value=z.get("state", "-"))
            ws_re.cell(row=i, column=3, value=z.get("median_price", 0))
            ws_re.cell(row=i, column=4, value=z.get("yoy_change"))
            ws_re.cell(row=i, column=5, value=z.get("medspa_score", 0))
            ws_re.cell(row=i, column=6, value=z.get("medspa_grade", "-"))

        re_col_widths = {"A": 12, "B": 10, "C": 16, "D": 14, "E": 14, "F": 8}
        for col_letter, width in re_col_widths.items():
            ws_re.column_dimensions[col_letter].width = width

        # ---- Sheet 15: Deposit Wealth ----
        ws_dw = wb.create_sheet("Deposit Wealth")
        deposit_wealth = data.get("deposit_wealth", {})
        dw_underserved = deposit_wealth.get("underserved_zips", [])
        dw_summary = deposit_wealth.get("summary", {})

        ws_dw["A1"] = "Deposit Wealth Concentration"
        ws_dw["A1"].font = Font(bold=True, size=13)
        ws_dw.merge_cells("A1:E1")

        ws_dw["A3"] = "Year"
        ws_dw["B3"] = dw_summary.get("year", "-")
        ws_dw["A4"] = "ZIPs Analyzed"
        ws_dw["B4"] = dw_summary.get("zips_analyzed", 0)
        ws_dw["A5"] = "Total Deposits ($T)"
        ws_dw["B5"] = dw_summary.get("total_deposits_t", 0)
        ws_dw["A6"] = "Avg Deposits/Return"
        ws_dw["B6"] = dw_summary.get("avg_deposits_per_return", 0)
        ws_dw["A7"] = "Underserved ZIPs"
        ws_dw["B7"] = dw_summary.get("underserved_zips", 0)
        for r in range(3, 8):
            ws_dw[f"A{r}"].font = Font(bold=True)

        dw_start = 9
        ws_dw.cell(row=dw_start, column=1, value="Underserved ZIPs (High Deposits, Low Medspa Score)").font = Font(bold=True, size=11)
        ws_dw.merge_cells(f"A{dw_start}:F{dw_start}")
        dw_headers = [
            "ZIP", "State", "Total Deposits", "Deposits/Return",
            "Branches", "Medspa Score", "Grade",
        ]
        for col, header in enumerate(dw_headers, 1):
            cell = ws_dw.cell(row=dw_start + 1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for i, z in enumerate(dw_underserved, dw_start + 2):
            ws_dw.cell(row=i, column=1, value=z.get("zip_code"))
            ws_dw.cell(row=i, column=2, value=z.get("state", "-"))
            ws_dw.cell(row=i, column=3, value=z.get("total_deposits", 0))
            ws_dw.cell(row=i, column=4, value=z.get("deposits_per_return", 0))
            ws_dw.cell(row=i, column=5, value=z.get("branch_count", 0))
            ws_dw.cell(row=i, column=6, value=z.get("medspa_score", 0))
            ws_dw.cell(row=i, column=7, value=z.get("medspa_grade", "-"))

        dw_col_widths = {"A": 12, "B": 8, "C": 18, "D": 16, "E": 10, "F": 14, "G": 8}
        for col_letter, width in dw_col_widths.items():
            ws_dw.column_dimensions[col_letter].width = width

        # ---- Sheet 16: Business Formation ----
        ws_bf = wb.create_sheet("Business Formation")
        business_formation = data.get("business_formation", {})
        bf_entrep = business_formation.get("entrepreneurial_zips", [])
        bf_summary = business_formation.get("summary", {})
        bf_comp = business_formation.get("income_composition", {})

        ws_bf["A1"] = "Business Formation Velocity"
        ws_bf["A1"].font = Font(bold=True, size=13)
        ws_bf.merge_cells("A1:E1")

        ws_bf["A3"] = "Tax Year"
        ws_bf["B3"] = bf_summary.get("tax_year", "-")
        ws_bf["A4"] = "ZIPs Analyzed"
        ws_bf["B4"] = bf_summary.get("zips_analyzed", 0)
        ws_bf["A5"] = "Avg Business Density"
        ws_bf["B5"] = bf_summary.get("avg_biz_density", 0)
        ws_bf["A6"] = "High-Growth ZIPs (>10% YoY)"
        ws_bf["B6"] = bf_summary.get("high_growth_zips", 0)
        ws_bf["A7"] = "Avg Schedule C Income"
        ws_bf["B7"] = bf_summary.get("avg_schedule_c", 0)
        for r in range(3, 8):
            ws_bf[f"A{r}"].font = Font(bold=True)

        # Income composition
        if bf_comp:
            ws_bf["A9"] = "Business Income Composition"
            ws_bf["A9"].font = Font(bold=True, size=11)
            ws_bf["A10"] = "Schedule C (Sole Props) %"
            ws_bf["B10"] = bf_comp.get("schedule_c_pct", 0)
            ws_bf["A11"] = "Partnerships %"
            ws_bf["B11"] = bf_comp.get("partnership_pct", 0)
            ws_bf["A12"] = "S-Corps %"
            ws_bf["B12"] = bf_comp.get("scorp_pct", 0)

        bf_start = 14
        ws_bf.cell(row=bf_start, column=1, value="Entrepreneurial ZIPs (Biz Density >20%, Medspa Score <60)").font = Font(bold=True, size=11)
        ws_bf.merge_cells(f"A{bf_start}:G{bf_start}")
        bf_headers = [
            "ZIP", "State", "Biz Density %", "Schedule C", "Partnership",
            "YoY Growth %", "Medspa Score", "Grade",
        ]
        for col, header in enumerate(bf_headers, 1):
            cell = ws_bf.cell(row=bf_start + 1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for i, z in enumerate(bf_entrep, bf_start + 2):
            ws_bf.cell(row=i, column=1, value=z.get("zip_code"))
            ws_bf.cell(row=i, column=2, value=z.get("state", "-"))
            ws_bf.cell(row=i, column=3, value=z.get("biz_density", 0))
            ws_bf.cell(row=i, column=4, value=z.get("schedule_c", 0))
            ws_bf.cell(row=i, column=5, value=z.get("partnership", 0))
            ws_bf.cell(row=i, column=6, value=z.get("yoy_growth"))
            ws_bf.cell(row=i, column=7, value=z.get("medspa_score", 0))
            ws_bf.cell(row=i, column=8, value=z.get("medspa_grade", "-"))

        bf_col_widths = {"A": 12, "B": 8, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14, "H": 8}
        for col_letter, width in bf_col_widths.items():
            ws_bf.column_dimensions[col_letter].width = width

        # ---- Sheet 17: OZ Overlay ----
        ws_oz = wb.create_sheet("OZ Overlay")
        opportunity_zones = data.get("opportunity_zones", {})
        oz_states = opportunity_zones.get("state_data", [])
        oz_summary = opportunity_zones.get("summary", {})

        ws_oz["A1"] = "Opportunity Zone Overlay"
        ws_oz["A1"].font = Font(bold=True, size=13)
        ws_oz.merge_cells("A1:E1")

        ws_oz["A3"] = "Total OZ Tracts"
        ws_oz["B3"] = oz_summary.get("total_tracts", 0)
        ws_oz["A4"] = "States with OZs"
        ws_oz["B4"] = oz_summary.get("states_with_oz", 0)
        ws_oz["A5"] = "Tax-Advantaged States"
        ws_oz["B5"] = oz_summary.get("tax_advantaged_states", 0)
        for r in range(3, 6):
            ws_oz[f"A{r}"].font = Font(bold=True)

        oz_start = 7
        ws_oz.cell(row=oz_start, column=1, value="State OZ Data").font = Font(bold=True, size=11)
        oz_headers = ["State", "OZ Tracts", "Low-Income", "Medspa Prospects", "A-Grade", "OZ/Prospect Ratio"]
        for col, header in enumerate(oz_headers, 1):
            cell = ws_oz.cell(row=oz_start + 1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for i, s in enumerate(oz_states, oz_start + 2):
            ws_oz.cell(row=i, column=1, value=s.get("state"))
            ws_oz.cell(row=i, column=2, value=s.get("oz_tracts", 0))
            ws_oz.cell(row=i, column=3, value=s.get("low_income_tracts", 0))
            ws_oz.cell(row=i, column=4, value=s.get("prospect_count", 0))
            ws_oz.cell(row=i, column=5, value=s.get("a_grade_count", 0))
            ws_oz.cell(row=i, column=6, value=s.get("oz_per_prospect", 0))

        oz_col_widths = {"A": 8, "B": 12, "C": 12, "D": 16, "E": 10, "F": 18}
        for col_letter, width in oz_col_widths.items():
            ws_oz.column_dimensions[col_letter].width = width

        # ---- Sheet 18: Demographic Demand ----
        ws_dd = wb.create_sheet("Demographic Demand")
        demographic_demand = data.get("demographic_demand", {})
        dd_states = demographic_demand.get("state_data", [])
        dd_summary = demographic_demand.get("summary", {})

        ws_dd["A1"] = "Demographic Demand Model"
        ws_dd["A1"].font = Font(bold=True, size=13)
        ws_dd.merge_cells("A1:E1")

        ws_dd["A3"] = "Period Year"
        ws_dd["B3"] = dd_summary.get("period_year", "-")
        ws_dd["A4"] = "Avg Bachelor's+ %"
        ws_dd["B4"] = dd_summary.get("avg_bachelors_pct", 0)
        ws_dd["A5"] = "Avg Graduate %"
        ws_dd["B5"] = dd_summary.get("avg_graduate_pct", 0)
        ws_dd["A6"] = "Underserved Educated States"
        ws_dd["B6"] = dd_summary.get("underserved_educated", 0)
        for r in range(3, 7):
            ws_dd[f"A{r}"].font = Font(bold=True)

        dd_start = 8
        ws_dd.cell(row=dd_start, column=1, value="State Education vs Medspa Gap").font = Font(bold=True, size=11)
        dd_headers = ["State", "Bachelor's %", "Graduate %", "Prospects", "A-Grade", "Gap Score"]
        for col, header in enumerate(dd_headers, 1):
            cell = ws_dd.cell(row=dd_start + 1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        dd_sorted = sorted(dd_states, key=lambda x: x.get("gap_score", 0), reverse=True)
        for i, s in enumerate(dd_sorted, dd_start + 2):
            ws_dd.cell(row=i, column=1, value=s.get("state_name", "-"))
            ws_dd.cell(row=i, column=2, value=s.get("avg_bachelors", 0))
            ws_dd.cell(row=i, column=3, value=s.get("avg_graduate", 0))
            ws_dd.cell(row=i, column=4, value=s.get("prospect_count", 0))
            ws_dd.cell(row=i, column=5, value=s.get("a_grade_count", 0))
            ws_dd.cell(row=i, column=6, value=s.get("gap_score", 0))

        dd_col_widths = {"A": 22, "B": 14, "C": 12, "D": 12, "E": 10, "F": 12}
        for col_letter, width in dd_col_widths.items():
            ws_dd.column_dimensions[col_letter].width = width

        # ---- Sheet 19: PE Heat Map ----
        ws_pe = wb.create_sheet("PE Heat Map")
        pe_competitive = data.get("pe_competitive", {})
        pec_deals = pe_competitive.get("deals", [])
        pec_summary = pe_competitive.get("summary", {})

        ws_pe["A1"] = "PE Competitive Heat Map"
        ws_pe["A1"].font = Font(bold=True, size=13)
        ws_pe.merge_cells("A1:E1")

        ws_pe["A3"] = "PE-Backed Platforms"
        ws_pe["B3"] = pec_summary.get("pe_platforms", 0)
        ws_pe["A4"] = "Avg EV/EBITDA"
        ws_pe["B4"] = pec_summary.get("avg_ev_ebitda", 0)
        ws_pe["A5"] = "Total Deal Value"
        ws_pe["B5"] = pec_summary.get("total_deal_value", 0)
        ws_pe["A6"] = "Most Active Buyer"
        ws_pe["B6"] = pec_summary.get("most_active_buyer", "-")
        for r in range(3, 7):
            ws_pe[f"A{r}"].font = Font(bold=True)

        pe_start = 8
        ws_pe.cell(row=pe_start, column=1, value="Aesthetics PE Deals").font = Font(bold=True, size=11)
        pe_headers = ["Company", "Buyer", "Deal Type", "EV ($)", "EV/EBITDA", "Date", "State"]
        for col, header in enumerate(pe_headers, 1):
            cell = ws_pe.cell(row=pe_start + 1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for i, d in enumerate(pec_deals, pe_start + 2):
            ws_pe.cell(row=i, column=1, value=d.get("company", "-"))
            ws_pe.cell(row=i, column=2, value=d.get("buyer", "-"))
            ws_pe.cell(row=i, column=3, value=d.get("deal_type", "-"))
            ws_pe.cell(row=i, column=4, value=d.get("ev_usd", 0))
            ws_pe.cell(row=i, column=5, value=d.get("ev_ebitda", 0))
            ws_pe.cell(row=i, column=6, value=d.get("date", "-"))
            ws_pe.cell(row=i, column=7, value=d.get("state", "-"))

        pe_col_widths = {"A": 30, "B": 25, "C": 12, "D": 16, "E": 12, "F": 12, "G": 8}
        for col_letter, width in pe_col_widths.items():
            ws_pe.column_dimensions[col_letter].width = width

        # ---- Sheet 20: Construction Momentum ----
        ws_cm = wb.create_sheet("Construction Momentum")
        construction_momentum = data.get("construction_momentum", {})
        cm_states = construction_momentum.get("state_data", [])
        cm_summary = construction_momentum.get("summary", {})

        ws_cm["A1"] = "Construction Momentum Signal"
        ws_cm["A1"].font = Font(bold=True, size=13)
        ws_cm.merge_cells("A1:E1")

        ws_cm["A3"] = "Year"
        ws_cm["B3"] = cm_summary.get("latest_year", "-")
        ws_cm["A4"] = "States Analyzed"
        ws_cm["B4"] = cm_summary.get("states_analyzed", 0)
        ws_cm["A5"] = "Avg YoY Growth"
        ws_cm["B5"] = cm_summary.get("avg_yoy_growth", 0)
        ws_cm["A6"] = "High-Growth States (>10%)"
        ws_cm["B6"] = cm_summary.get("high_growth_states", 0)
        for r in range(3, 7):
            ws_cm[f"A{r}"].font = Font(bold=True)

        cm_start = 8
        ws_cm.cell(row=cm_start, column=1, value="State Permit Data").font = Font(bold=True, size=11)
        cm_headers = ["State", "Total Permits", "1-Unit", "5+ Units", "YoY Growth %", "Medspas", "A-Grade"]
        for col, header in enumerate(cm_headers, 1):
            cell = ws_cm.cell(row=cm_start + 1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for i, s in enumerate(cm_states, cm_start + 2):
            ws_cm.cell(row=i, column=1, value=s.get("state_name", "-"))
            ws_cm.cell(row=i, column=2, value=s.get("total_permits", 0))
            ws_cm.cell(row=i, column=3, value=s.get("permits_1unit", 0))
            ws_cm.cell(row=i, column=4, value=s.get("permits_5plus", 0))
            ws_cm.cell(row=i, column=5, value=s.get("yoy_growth"))
            ws_cm.cell(row=i, column=6, value=s.get("medspa_count", 0))
            ws_cm.cell(row=i, column=7, value=s.get("a_grade_count", 0))

        cm_col_widths = {"A": 22, "B": 14, "C": 10, "D": 10, "E": 14, "F": 10, "G": 10}
        for col_letter, width in cm_col_widths.items():
            ws_cm.column_dimensions[col_letter].width = width

        # ---- Sheet 21: Medical CPI ----
        ws_cpi = wb.create_sheet("Medical CPI")
        medical_cpi = data.get("medical_cpi", {})
        mcpi_annual = medical_cpi.get("annual_data", [])
        mcpi_summary = medical_cpi.get("summary", {})

        ws_cpi["A1"] = "Medical CPI Pricing Power"
        ws_cpi["A1"].font = Font(bold=True, size=13)
        ws_cpi.merge_cells("A1:E1")

        ws_cpi["A3"] = "Current Medical CPI"
        ws_cpi["B3"] = mcpi_summary.get("current_medical_cpi", 0)
        ws_cpi["A4"] = "YoY Medical Change %"
        ws_cpi["B4"] = mcpi_summary.get("yoy_medical_change", 0)
        ws_cpi["A5"] = "Medical vs General Spread"
        ws_cpi["B5"] = mcpi_summary.get("medical_vs_general_spread", 0)
        ws_cpi["A6"] = "5yr CAGR %"
        ws_cpi["B6"] = mcpi_summary.get("cagr_5yr", "-")
        for r in range(3, 7):
            ws_cpi[f"A{r}"].font = Font(bold=True)

        cpi_start = 8
        ws_cpi.cell(row=cpi_start, column=1, value="Annual CPI Comparison").font = Font(bold=True, size=11)
        cpi_headers = ["Year", "Medical CPI", "General CPI", "Medical YoY %", "General YoY %", "Spread (pp)", "Cumulative (pp)"]
        for col, header in enumerate(cpi_headers, 1):
            cell = ws_cpi.cell(row=cpi_start + 1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for i, d in enumerate(mcpi_annual, cpi_start + 2):
            ws_cpi.cell(row=i, column=1, value=d.get("year"))
            ws_cpi.cell(row=i, column=2, value=d.get("medical_cpi"))
            ws_cpi.cell(row=i, column=3, value=d.get("general_cpi"))
            ws_cpi.cell(row=i, column=4, value=d.get("medical_yoy"))
            ws_cpi.cell(row=i, column=5, value=d.get("general_yoy"))
            ws_cpi.cell(row=i, column=6, value=d.get("yoy_spread"))
            ws_cpi.cell(row=i, column=7, value=d.get("cumulative_divergence"))

        cpi_col_widths = {"A": 8, "B": 14, "C": 14, "D": 14, "E": 14, "F": 12, "G": 16}
        for col_letter, width in cpi_col_widths.items():
            ws_cpi.column_dimensions[col_letter].width = width

        # ---- Sheet 22: Talent Pipeline ----
        ws_tp = wb.create_sheet("Talent Pipeline")
        talent_pipeline = data.get("talent_pipeline", {})
        tp_quarterly = talent_pipeline.get("quarterly_data", [])
        tp_summary = talent_pipeline.get("summary", {})

        ws_tp["A1"] = "Talent Pipeline Pressure"
        ws_tp["A1"].font = Font(bold=True, size=13)
        ws_tp.merge_cells("A1:E1")

        ws_tp["A3"] = "Latest HC Openings (000s)"
        ws_tp["B3"] = tp_summary.get("latest_openings", 0)
        ws_tp["A4"] = "Openings/Hires Ratio"
        ws_tp["B4"] = tp_summary.get("openings_to_hires_ratio", 0)
        ws_tp["A5"] = "YoY Change %"
        ws_tp["B5"] = tp_summary.get("yoy_change", "-")
        ws_tp["A6"] = "Talent Scarcity Trend"
        ws_tp["B6"] = tp_summary.get("trend", "-")
        for r in range(3, 7):
            ws_tp[f"A{r}"].font = Font(bold=True)

        tp_start = 8
        ws_tp.cell(row=tp_start, column=1, value="Quarterly Healthcare Labor Data").font = Font(bold=True, size=11)
        tp_headers = ["Period", "Openings (000s)", "Hires (000s)", "Open/Hire Ratio"]
        for col, header in enumerate(tp_headers, 1):
            cell = ws_tp.cell(row=tp_start + 1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for i, q in enumerate(tp_quarterly, tp_start + 2):
            ws_tp.cell(row=i, column=1, value=q.get("period", "-"))
            ws_tp.cell(row=i, column=2, value=q.get("openings", 0))
            ws_tp.cell(row=i, column=3, value=q.get("hires", 0))
            ws_tp.cell(row=i, column=4, value=q.get("ratio", 0))

        tp_col_widths = {"A": 12, "B": 18, "C": 16, "D": 16}
        for col_letter, width in tp_col_widths.items():
            ws_tp.column_dimensions[col_letter].width = width

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        return output.getvalue()
