"""
PE Deal Memo Report Template.

Generates an IC-ready multi-section deal memo with:
- Executive Summary with KPI strip
- Company Overview (profile, location, sector)
- Financial Analysis with benchmark radar chart
- Market Position (competitive landscape, fragmentation)
- Management Assessment (leadership team)
- Valuation & Comparables (comparable transactions table)
- Exit Readiness (composite score, sub-scores)
- Recommended Next Steps

Uses the shared design_system.py patterns.
"""

import json
import logging
from datetime import datetime
from typing import Dict, Any, List, Optional
from io import BytesIO

from sqlalchemy import text
from sqlalchemy.orm import Session

from app.reports.design_system import (
    html_document, page_header, kpi_strip, kpi_card,
    toc, section_start, section_end,
    data_table, callout, pill_badge, profile_card,
    chart_container, chart_init_js, page_footer,
    build_bar_fallback, CHART_COLORS,
    BLUE, GREEN, ORANGE, RED, GRAY, TEAL, PURPLE,
)

logger = logging.getLogger(__name__)


def _fmt_usd(val: Optional[float], unit: str = "M") -> str:
    if val is None:
        return "\u2014"
    if unit == "M":
        if abs(val) >= 1_000:
            return f"${val / 1_000:,.1f}B"
        return f"${val:,.0f}M"
    return f"${val:,.0f}"


def _fmt_pct(val: Optional[float]) -> str:
    if val is None:
        return "\u2014"
    return f"{val:.1f}%"


def _fmt_x(val: Optional[float]) -> str:
    if val is None:
        return "\u2014"
    return f"{val:.2f}x"


def _radar_chart_config(labels, values, label="Company", max_val=None):
    """Build a Chart.js radar config for financial benchmarking."""
    if max_val is None:
        max_val = max(values) * 1.3 if values and max(values) > 0 else 100
    return {
        "type": "radar",
        "data": {
            "labels": labels,
            "datasets": [{
                "label": label,
                "data": values,
                "backgroundColor": f"{BLUE}30",
                "borderColor": BLUE,
                "borderWidth": 2,
                "pointBackgroundColor": BLUE,
            }],
        },
        "options": {
            "responsive": True,
            "plugins": {"legend": {"display": True}},
            "scales": {
                "r": {
                    "beginAtZero": True,
                    "max": round(max_val, 1),
                    "ticks": {"stepSize": round(max_val / 5, 1)},
                },
            },
        },
    }


class PEDealMemoTemplate:
    """PE Deal Memo report template."""

    name = "pe_deal_memo"
    description = "IC-ready deal memo with financial benchmarks, exit readiness, and comparable transactions"

    # ── Data Gathering ─────────────────────────────────────────────

    def gather_data(self, db: Session, params: Dict[str, Any]) -> Dict[str, Any]:
        company_id = params.get("company_id")
        if not company_id:
            raise ValueError("company_id is required")

        company = self._get_company(db, company_id)
        if not company:
            return {
                "generated_at": datetime.utcnow().isoformat(),
                "company": None,
                "financials": [],
                "benchmarks": {},
                "comparables": [],
                "exit_readiness": {},
                "leadership": [],
                "competitors": [],
                "investment": {},
            }

        return {
            "generated_at": datetime.utcnow().isoformat(),
            "company": company,
            "financials": self._get_financials(db, company_id),
            "benchmarks": self._get_benchmarks(db, company_id),
            "comparables": self._get_comparables(db, company_id),
            "exit_readiness": self._get_exit_readiness(db, company_id),
            "leadership": self._get_leadership(db, company_id),
            "competitors": self._get_competitors(db, company_id),
            "investment": self._get_investment(db, company_id),
        }

    def _get_company(self, db: Session, company_id: int) -> Optional[Dict]:
        row = db.execute(text("""
            SELECT id, name, description, industry, sub_industry,
                   naics_code, headquarters_city, headquarters_state,
                   employee_count, founded_year, website,
                   ownership_status, current_pe_owner, sector
            FROM pe_portfolio_companies
            WHERE id = :id
        """), {"id": company_id}).fetchone()
        if not row:
            return None
        return {
            "id": row[0], "name": row[1], "description": row[2],
            "industry": row[3] or "Unknown", "sub_industry": row[4],
            "naics": row[5], "city": row[6], "state": row[7],
            "employees": row[8], "founded": row[9], "website": row[10],
            "ownership": row[11], "pe_owner": row[12], "sector": row[13],
        }

    def _get_financials(self, db: Session, company_id: int) -> List[Dict]:
        rows = db.execute(text("""
            SELECT fiscal_year, revenue_usd, revenue_growth_pct,
                   ebitda_usd, ebitda_margin_pct, gross_margin_pct,
                   net_income_usd, total_debt_usd, cash_usd,
                   free_cash_flow_usd, debt_to_ebitda
            FROM pe_company_financials
            WHERE company_id = :cid
            ORDER BY fiscal_year DESC
            LIMIT 5
        """), {"cid": company_id}).fetchall()
        return [
            {
                "year": r[0], "revenue": float(r[1]) if r[1] else None,
                "revenue_growth": float(r[2]) if r[2] else None,
                "ebitda": float(r[3]) if r[3] else None,
                "ebitda_margin": float(r[4]) if r[4] else None,
                "gross_margin": float(r[5]) if r[5] else None,
                "net_income": float(r[6]) if r[6] else None,
                "total_debt": float(r[7]) if r[7] else None,
                "cash": float(r[8]) if r[8] else None,
                "fcf": float(r[9]) if r[9] else None,
                "leverage": float(r[10]) if r[10] else None,
            }
            for r in rows
        ]

    def _get_benchmarks(self, db: Session, company_id: int) -> Dict:
        """Get benchmark data — company vs peer median for radar chart."""
        try:
            from app.core.pe_valuation_comps import ValuationCompsService
            svc = ValuationCompsService(db)
            return svc.get_comps(company_id)
        except Exception:
            logger.debug("Valuation comps unavailable for company %s", company_id)
            return {}

    def _get_comparables(self, db: Session, company_id: int) -> List[Dict]:
        """Get comparable transactions."""
        try:
            from app.core.pe_comparable_transactions import ComparableTransactionService
            svc = ComparableTransactionService(db)
            result = svc.get_comps(company_id)
            return result.get("deals", [])
        except Exception:
            logger.debug("Comparable transactions unavailable for company %s", company_id)
            return []

    def _get_exit_readiness(self, db: Session, company_id: int) -> Dict:
        try:
            from app.core.pe_exit_scoring import ExitReadinessScorer
            scorer = ExitReadinessScorer(db)
            return scorer.score(company_id)
        except Exception:
            logger.debug("Exit readiness unavailable for company %s", company_id)
            return {}

    def _get_leadership(self, db: Session, company_id: int) -> List[Dict]:
        rows = db.execute(text("""
            SELECT p.full_name, cl.title, cl.role_category,
                   cl.is_ceo, cl.is_cfo, cl.is_board_member,
                   cl.start_date, p.linkedin_url
            FROM pe_company_leadership cl
            JOIN pe_people p ON cl.person_id = p.id
            WHERE cl.company_id = :cid AND cl.is_current = true
            ORDER BY cl.is_ceo DESC, cl.is_cfo DESC, cl.is_board_member DESC, p.full_name
        """), {"cid": company_id}).fetchall()
        return [
            {
                "name": r[0], "title": r[1], "role": r[2],
                "is_ceo": r[3], "is_cfo": r[4], "is_board": r[5],
                "start_date": r[6].isoformat() if r[6] else None,
                "linkedin": r[7],
            }
            for r in rows
        ]

    def _get_competitors(self, db: Session, company_id: int) -> List[Dict]:
        rows = db.execute(text("""
            SELECT competitor_name, is_public, ticker, is_pe_backed,
                   pe_owner, competitor_type, relative_size, market_position
            FROM pe_competitor_mappings
            WHERE company_id = :cid
            ORDER BY market_position, competitor_name
        """), {"cid": company_id}).fetchall()
        return [
            {
                "name": r[0], "public": r[1], "ticker": r[2],
                "pe_backed": r[3], "pe_owner": r[4], "type": r[5],
                "size": r[6], "position": r[7],
            }
            for r in rows
        ]

    def _get_investment(self, db: Session, company_id: int) -> Dict:
        """Get the most recent investment record for this company."""
        row = db.execute(text("""
            SELECT fi.invested_amount_usd, fi.entry_ev_usd,
                   fi.entry_ev_ebitda_multiple, fi.entry_ev_revenue_multiple,
                   fi.ownership_pct, fi.investment_date, fi.status,
                   f.name as fund_name, fm.name as firm_name
            FROM pe_fund_investments fi
            JOIN pe_funds f ON fi.fund_id = f.id
            JOIN pe_firms fm ON f.firm_id = fm.id
            WHERE fi.company_id = :cid
            ORDER BY fi.investment_date DESC
            LIMIT 1
        """), {"cid": company_id}).fetchone()
        if not row:
            return {}
        return {
            "invested": float(row[0]) if row[0] else None,
            "entry_ev": float(row[1]) if row[1] else None,
            "ev_ebitda": float(row[2]) if row[2] else None,
            "ev_revenue": float(row[3]) if row[3] else None,
            "ownership": float(row[4]) if row[4] else None,
            "date": row[5].isoformat() if row[5] else None,
            "status": row[6],
            "fund": row[7], "firm": row[8],
        }

    # ── HTML Rendering ─────────────────────────────────────────────

    def render_html(self, data: Dict[str, Any]) -> str:
        company = data.get("company")
        if not company:
            return html_document(
                title="Company Not Found",
                body_content=callout("<strong>Company not found.</strong> Check the company_id parameter.", "warn"),
            )

        financials = data.get("financials", [])
        benchmarks = data.get("benchmarks", {})
        comparables = data.get("comparables", [])
        exit_data = data.get("exit_readiness", {})
        leadership = data.get("leadership", [])
        competitors = data.get("competitors", [])
        investment = data.get("investment", {})
        title = data.get("report_title", f"{company['name']} \u2014 Deal Memo")

        charts_js = ""
        body = ""

        # ── Header ─────────────────────────────────────────────────
        subtitle_parts = [p for p in [
            company.get("industry"),
            f"{company.get('city')}, {company.get('state')}" if company.get("city") else None,
        ] if p]
        body += page_header(
            title=title,
            subtitle=" \u00b7 ".join(subtitle_parts) if subtitle_parts else None,
            badge=f"Generated {datetime.utcnow().strftime('%b %d, %Y')}",
        )

        # ── KPI Strip ──────────────────────────────────────────────
        latest = financials[0] if financials else {}
        cards = ""
        cards += kpi_card("Revenue", _fmt_usd(latest.get("revenue") / 1_000_000 if latest.get("revenue") else None))
        cards += kpi_card("EBITDA", _fmt_usd(latest.get("ebitda") / 1_000_000 if latest.get("ebitda") else None))
        cards += kpi_card("EBITDA Margin", _fmt_pct(latest.get("ebitda_margin")))
        cards += kpi_card("Revenue Growth", _fmt_pct(latest.get("revenue_growth")))

        exit_score = exit_data.get("composite_score") or exit_data.get("score")
        cards += kpi_card("Exit Readiness", f'{exit_score:.0f}/100' if exit_score else "\u2014")
        body += "\n" + kpi_strip(cards)

        # ── TOC ────────────────────────────────────────────────────
        toc_items = [
            {"number": 1, "id": "exec-summary", "title": "Executive Summary"},
            {"number": 2, "id": "company-overview", "title": "Company Overview"},
            {"number": 3, "id": "financial-analysis", "title": "Financial Analysis"},
            {"number": 4, "id": "market-position", "title": "Market Position"},
            {"number": 5, "id": "management", "title": "Management Assessment"},
            {"number": 6, "id": "valuation", "title": "Valuation & Comparables"},
            {"number": 7, "id": "exit-readiness", "title": "Exit Readiness"},
            {"number": 8, "id": "next-steps", "title": "Recommended Next Steps"},
        ]
        body += "\n" + toc(toc_items)

        # ── Section 1: Executive Summary ───────────────────────────
        body += "\n" + section_start(1, "Executive Summary", "exec-summary")

        summary_parts = [f"<strong>{company['name']}</strong>"]
        if company.get("industry"):
            summary_parts.append(f"operates in the <strong>{company['industry']}</strong> sector")
        if company.get("city") and company.get("state"):
            summary_parts.append(f"headquartered in {company['city']}, {company['state']}")
        if company.get("employees"):
            summary_parts.append(f"with approximately {company['employees']:,} employees")
        if company.get("founded"):
            summary_parts.append(f"(founded {company['founded']})")

        body += f'<p>{" ".join(summary_parts)}.</p>'

        if company.get("description"):
            body += f'<p style="color:var(--gray-600);margin-top:8px">{company["description"]}</p>'

        if investment:
            body += '<h3 style="margin:16px 0 8px;font-size:1rem">Current Investment</h3>'
            inv_rows = [
                ["Sponsor", investment.get("firm") or "\u2014"],
                ["Fund", investment.get("fund") or "\u2014"],
                ["Investment Date", investment.get("date") or "\u2014"],
                ["Entry EV", _fmt_usd(investment.get("entry_ev") / 1_000_000 if investment.get("entry_ev") else None)],
                ["Entry EV/EBITDA", _fmt_x(investment.get("ev_ebitda"))],
                ["Ownership", _fmt_pct(investment.get("ownership"))],
            ]
            body += data_table(["Parameter", "Value"], inv_rows)

        body += "\n" + section_end()

        # ── Section 2: Company Overview ────────────────────────────
        body += "\n" + section_start(2, "Company Overview", "company-overview")

        overview_rows = [
            ["Company Name", company["name"]],
            ["Industry", company.get("industry") or "\u2014"],
            ["Sub-Industry", company.get("sub_industry") or "\u2014"],
            ["NAICS Code", str(company.get("naics") or "\u2014")],
            ["Headquarters", f"{company.get('city', '')}, {company.get('state', '')}".strip(", ")],
            ["Employees", f"{company['employees']:,}" if company.get("employees") else "\u2014"],
            ["Founded", str(company.get("founded") or "\u2014")],
            ["Website", company.get("website") or "\u2014"],
            ["Ownership", pill_badge((company.get("ownership") or "unknown").title(),
                                     "pe" if company.get("pe_owner") else "default")],
            ["PE Owner", company.get("pe_owner") or "\u2014"],
        ]
        body += data_table(["Field", "Value"], overview_rows)
        body += "\n" + section_end()

        # ── Section 3: Financial Analysis ──────────────────────────
        body += "\n" + section_start(3, "Financial Analysis", "financial-analysis")

        if financials:
            # Financial history table
            fin_headers = ["Year", "Revenue ($M)", "Growth", "EBITDA ($M)", "EBITDA Margin",
                           "Gross Margin", "FCF ($M)", "Leverage"]
            fin_rows = []
            for f in financials:
                fin_rows.append([
                    str(f.get("year") or "\u2014"),
                    f'${f["revenue"] / 1_000_000:,.0f}' if f.get("revenue") else "\u2014",
                    _fmt_pct(f.get("revenue_growth")),
                    f'${f["ebitda"] / 1_000_000:,.0f}' if f.get("ebitda") else "\u2014",
                    _fmt_pct(f.get("ebitda_margin")),
                    _fmt_pct(f.get("gross_margin")),
                    f'${f["fcf"] / 1_000_000:,.0f}' if f.get("fcf") else "\u2014",
                    f'{f["leverage"]:.1f}x' if f.get("leverage") else "\u2014",
                ])
            body += data_table(fin_headers, fin_rows, numeric_columns={1, 2, 3, 4, 5, 6, 7})

            # Radar chart for benchmarks
            if benchmarks.get("company") and benchmarks.get("peer_stats"):
                radar_labels = []
                radar_values = []
                co = benchmarks["company"]
                for metric_key, label in [
                    ("ev_revenue", "EV/Revenue"),
                    ("ev_ebitda", "EV/EBITDA"),
                ]:
                    val = co.get(metric_key)
                    if val is not None:
                        radar_labels.append(label)
                        radar_values.append(round(val, 2))

                # Add margin metrics from latest financials
                if latest.get("ebitda_margin") is not None:
                    radar_labels.append("EBITDA Margin %")
                    radar_values.append(round(latest["ebitda_margin"], 1))
                if latest.get("revenue_growth") is not None:
                    radar_labels.append("Revenue Growth %")
                    radar_values.append(round(latest["revenue_growth"], 1))
                if latest.get("gross_margin") is not None:
                    radar_labels.append("Gross Margin %")
                    radar_values.append(round(latest["gross_margin"], 1))

                if radar_labels:
                    radar_cfg = _radar_chart_config(radar_labels, radar_values, label=company["name"])
                    radar_json = json.dumps(radar_cfg)
                    body += '<div style="max-width:500px;margin:16px auto">'
                    body += chart_container(
                        "benchmarkRadar", radar_json,
                        build_bar_fallback(radar_labels, radar_values, BLUE),
                        size="medium", title="Financial Benchmark Profile",
                    )
                    charts_js += chart_init_js("benchmarkRadar", radar_json)
                    body += '</div>'
        else:
            body += callout("<strong>No financial data available.</strong>", "warn")

        body += "\n" + section_end()

        # ── Section 4: Market Position ─────────────────────────────
        body += "\n" + section_start(4, "Market Position", "market-position")

        if competitors:
            comp_headers = ["Competitor", "Type", "Size", "Position", "PE-Backed", "PE Owner"]
            comp_rows = []
            for c in competitors:
                pe_badge = pill_badge("PE", "pe") if c.get("pe_backed") else pill_badge("Independent", "public")
                comp_rows.append([
                    c["name"],
                    c.get("type") or "\u2014",
                    c.get("size") or "\u2014",
                    c.get("position") or "\u2014",
                    pe_badge,
                    c.get("pe_owner") or "\u2014",
                ])
            body += '<p>Known competitive landscape:</p>'
            body += data_table(comp_headers, comp_rows)
        else:
            body += callout("<strong>No competitor data available.</strong> Run competitive landscape analysis to populate.", "warn")

        body += "\n" + section_end()

        # ── Section 5: Management Assessment ───────────────────────
        body += "\n" + section_start(5, "Management Assessment", "management")

        if leadership:
            body += '<div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(280px,1fr));gap:16px">'
            for ldr in leadership:
                badges = []
                if ldr.get("is_ceo"):
                    badges.append("CEO")
                if ldr.get("is_cfo"):
                    badges.append("CFO")
                if ldr.get("is_board"):
                    badges.append("Board")
                initials = "".join(w[0].upper() for w in ldr["name"].split()[:2]) if ldr.get("name") else "?"
                body += profile_card(
                    name=ldr["name"],
                    title=ldr.get("title") or "\u2014",
                    initials=initials,
                    badges=badges,
                    linkedin=ldr.get("linkedin"),
                )
            body += '</div>'

            mgmt_headers = ["Name", "Title", "Role", "Since"]
            mgmt_rows = [[l["name"], l.get("title") or "\u2014",
                          l.get("role") or "\u2014",
                          l.get("start_date") or "\u2014"] for l in leadership]
            body += data_table(mgmt_headers, mgmt_rows)
        else:
            body += callout("<strong>No leadership data available.</strong> Run people collection to populate.", "warn")

        body += "\n" + section_end()

        # ── Section 6: Valuation & Comparables ─────────────────────
        body += "\n" + section_start(6, "Valuation & Comparables", "valuation")

        if comparables:
            body += f'<p>{len(comparables)} comparable transactions identified in the same industry.</p>'
            txn_headers = ["Deal", "Buyer", "EV ($M)", "EV/EBITDA", "EV/Revenue",
                           "Date", "Type"]
            txn_rows = []
            for d in comparables[:15]:
                ev = d.get("enterprise_value_usd")
                txn_rows.append([
                    d.get("deal_name") or d.get("company_name", "\u2014"),
                    d.get("buyer_name") or "\u2014",
                    _fmt_usd(ev / 1_000_000 if ev else None),
                    _fmt_x(d.get("ev_ebitda_multiple")),
                    _fmt_x(d.get("ev_revenue_multiple")),
                    str(d.get("closed_date") or d.get("announced_date") or "\u2014"),
                    d.get("deal_type") or "\u2014",
                ])
            body += data_table(txn_headers, txn_rows, numeric_columns={2, 3, 4})
        else:
            body += callout("<strong>No comparable transactions found.</strong>", "warn")

        # Valuation multiples from benchmarks
        if benchmarks.get("company"):
            co = benchmarks["company"]
            body += '<h3 style="margin:16px 0 8px;font-size:1rem">Current Valuation Multiples</h3>'
            val_rows = []
            if co.get("ev_revenue") is not None:
                peer_rev = benchmarks.get("peer_stats", {}).get("ev_revenue", {})
                val_rows.append([
                    "EV/Revenue", _fmt_x(co["ev_revenue"]),
                    _fmt_x(peer_rev.get("median")),
                    _fmt_x(peer_rev.get("p25")),
                    _fmt_x(peer_rev.get("p75")),
                ])
            if co.get("ev_ebitda") is not None:
                peer_eb = benchmarks.get("peer_stats", {}).get("ev_ebitda", {})
                val_rows.append([
                    "EV/EBITDA", _fmt_x(co["ev_ebitda"]),
                    _fmt_x(peer_eb.get("median")),
                    _fmt_x(peer_eb.get("p25")),
                    _fmt_x(peer_eb.get("p75")),
                ])
            if val_rows:
                body += data_table(
                    ["Multiple", "Company", "Peer Median", "P25", "P75"],
                    val_rows, numeric_columns={1, 2, 3, 4},
                )

        body += "\n" + section_end()

        # ── Section 7: Exit Readiness ──────────────────────────────
        body += "\n" + section_start(7, "Exit Readiness", "exit-readiness")

        if exit_data:
            score = exit_data.get("composite_score") or exit_data.get("score")
            grade = exit_data.get("grade", "N/A")

            if score is not None:
                color = GREEN if score >= 70 else ORANGE if score >= 50 else RED
                body += f'<div style="text-align:center;margin:16px 0">'
                body += f'<span style="font-size:3rem;font-weight:700;color:{color}">{score:.0f}</span>'
                body += f'<span style="font-size:1.5rem;color:var(--gray-500)">/100</span>'
                body += f'<br><span style="font-size:1.2rem;font-weight:600;color:{color}">{grade}</span>'
                body += '</div>'

            sub_scores = exit_data.get("sub_scores", [])
            if sub_scores:
                sub_headers = ["Dimension", "Score", "Grade", "Explanation"]
                sub_rows = []
                for ss in sub_scores:
                    sub_rows.append([
                        ss.get("label", ss.get("dimension", "")),
                        f'{ss.get("raw_score", 0):.0f}/100',
                        ss.get("grade", "N/A"),
                        ss.get("explanation", ""),
                    ])
                body += data_table(sub_headers, sub_rows, numeric_columns={1})

            recs = exit_data.get("recommendations", [])
            if recs:
                body += callout(
                    "<strong>Recommendations:</strong><ul>" +
                    "".join(f"<li>{r}</li>" for r in recs[:5]) +
                    "</ul>", "info",
                )
        else:
            body += callout("<strong>Exit readiness score not available.</strong> Run the scorer to populate.", "warn")

        body += "\n" + section_end()

        # ── Section 8: Recommended Next Steps ──────────────────────
        body += "\n" + section_start(8, "Recommended Next Steps", "next-steps")

        steps = []
        if not financials:
            steps.append("Obtain and upload financial statements for the last 3-5 years")
        if not leadership:
            steps.append("Complete management team assessment and background checks")
        if not competitors:
            steps.append("Commission competitive landscape analysis")
        if not comparables:
            steps.append("Identify and analyze comparable transactions")
        if not exit_data:
            steps.append("Run exit readiness scoring to identify improvement areas")
        if exit_data and (exit_data.get("composite_score") or 0) < 70:
            steps.append("Address exit readiness gaps before marketing process")
        if financials and len(financials) < 3:
            steps.append("Collect additional years of financial history for trend analysis")
        if not steps:
            steps.append("All key data collected \u2014 proceed to IC presentation preparation")

        body += '<ol style="font-size:1rem;line-height:1.8">'
        for step in steps:
            body += f'<li>{step}</li>'
        body += '</ol>'

        body += "\n" + section_end()

        # ── Footer ─────────────────────────────────────────────────
        body += "\n" + page_footer(
            notes=[
                "Financial data sourced from SEC filings, company disclosures, and public records.",
                "Comparable transactions from public deal databases and press releases.",
                "Exit readiness scores are model-generated estimates, not investment advice.",
                "This document is prepared for internal IC discussion purposes only.",
            ],
            generated_line=f"Report generated {data.get('generated_at', 'N/A')} | Nexdata PE Intelligence",
        )

        return html_document(
            title=title,
            body_content=body,
            charts_js=charts_js,
        )

    # ── Excel Rendering ────────────────────────────────────────────

    def render_excel(self, data: Dict[str, Any]) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        hdr_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
        hdr_font = Font(bold=True, color="FFFFFF")

        company = data.get("company") or {}
        financials = data.get("financials", [])

        # ── Sheet 1: Summary ───────────────────────────────────────
        ws = wb.active
        ws.title = "Summary"

        summary = [
            ["Company", company.get("name", "")],
            ["Industry", company.get("industry", "")],
            ["Location", f"{company.get('city', '')}, {company.get('state', '')}"],
            ["Employees", company.get("employees")],
            ["Founded", company.get("founded")],
            ["Ownership", company.get("ownership", "")],
            ["PE Owner", company.get("pe_owner", "")],
        ]
        for i, (label, val) in enumerate(summary, 1):
            ws.cell(row=i, column=1, value=label).font = Font(bold=True)
            ws.cell(row=i, column=2, value=val)
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 30

        # ── Sheet 2: Financials ────────────────────────────────────
        ws_f = wb.create_sheet("Financials")
        f_headers = ["Year", "Revenue ($)", "Rev Growth %", "EBITDA ($)",
                     "EBITDA Margin %", "Gross Margin %", "FCF ($)", "Leverage"]
        for col, h in enumerate(f_headers, 1):
            cell = ws_f.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
        for row_idx, f in enumerate(financials, 2):
            ws_f.cell(row=row_idx, column=1, value=f.get("year"))
            ws_f.cell(row=row_idx, column=2, value=f.get("revenue"))
            ws_f.cell(row=row_idx, column=3, value=f.get("revenue_growth"))
            ws_f.cell(row=row_idx, column=4, value=f.get("ebitda"))
            ws_f.cell(row=row_idx, column=5, value=f.get("ebitda_margin"))
            ws_f.cell(row=row_idx, column=6, value=f.get("gross_margin"))
            ws_f.cell(row=row_idx, column=7, value=f.get("fcf"))
            ws_f.cell(row=row_idx, column=8, value=f.get("leverage"))
        for letter in "ABCDEFGH":
            ws_f.column_dimensions[letter].width = 16

        # ── Sheet 3: Comparables ───────────────────────────────────
        ws_c = wb.create_sheet("Comparables")
        c_headers = ["Deal", "Buyer", "EV ($)", "EV/EBITDA", "EV/Revenue", "Date", "Type"]
        for col, h in enumerate(c_headers, 1):
            cell = ws_c.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
        for row_idx, d in enumerate(data.get("comparables", []), 2):
            ws_c.cell(row=row_idx, column=1, value=d.get("deal_name") or d.get("company_name"))
            ws_c.cell(row=row_idx, column=2, value=d.get("buyer_name"))
            ws_c.cell(row=row_idx, column=3, value=d.get("enterprise_value_usd"))
            ws_c.cell(row=row_idx, column=4, value=d.get("ev_ebitda_multiple"))
            ws_c.cell(row=row_idx, column=5, value=d.get("ev_revenue_multiple"))
            ws_c.cell(row=row_idx, column=6, value=str(d.get("closed_date") or d.get("announced_date") or ""))
            ws_c.cell(row=row_idx, column=7, value=d.get("deal_type"))
        for letter in "ABCDEFG":
            ws_c.column_dimensions[letter].width = 18

        # ── Sheet 4: Leadership ────────────────────────────────────
        ws_l = wb.create_sheet("Leadership")
        l_headers = ["Name", "Title", "Role", "CEO", "CFO", "Board", "Since"]
        for col, h in enumerate(l_headers, 1):
            cell = ws_l.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
        for row_idx, l in enumerate(data.get("leadership", []), 2):
            ws_l.cell(row=row_idx, column=1, value=l.get("name"))
            ws_l.cell(row=row_idx, column=2, value=l.get("title"))
            ws_l.cell(row=row_idx, column=3, value=l.get("role"))
            ws_l.cell(row=row_idx, column=4, value="Yes" if l.get("is_ceo") else "")
            ws_l.cell(row=row_idx, column=5, value="Yes" if l.get("is_cfo") else "")
            ws_l.cell(row=row_idx, column=6, value="Yes" if l.get("is_board") else "")
            ws_l.cell(row=row_idx, column=7, value=l.get("start_date"))
        for letter in "ABCDEFG":
            ws_l.column_dimensions[letter].width = 18

        output = BytesIO()
        wb.save(output)
        return output.getvalue()
