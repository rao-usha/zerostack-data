"""
PE Fund Tearsheet Report Template.

Generates a professional single-page fund tearsheet with:
- Fund metadata and KPI strip (Size, Net IRR, TVPI, DPI, MOIC)
- J-curve line chart from cash flow data
- Portfolio company valuation bar chart
- Sector allocation doughnut chart
- Portfolio company table with investment details
- Top/bottom performers
- Cash flow summary

Uses the shared design_system.py patterns.
"""

import json
import logging
from datetime import datetime
from typing import Dict, Any, List, Optional
from io import BytesIO

from sqlalchemy import text
from sqlalchemy.orm import Session

from app.reports.design_system import (
    html_document, page_header, kpi_strip, kpi_card,
    toc, section_start, section_end,
    data_table, callout, pill_badge,
    chart_container, chart_init_js, page_footer,
    build_doughnut_config, build_horizontal_bar_config,
    build_line_chart_config,
    build_bar_fallback, build_chart_legend, CHART_COLORS,
    BLUE, GREEN, ORANGE, RED, GRAY, TEAL, PURPLE,
)

logger = logging.getLogger(__name__)


def _fmt_usd(val: Optional[float], unit: str = "M") -> str:
    """Format USD values. Input assumed in millions unless specified."""
    if val is None:
        return "\u2014"
    if unit == "M":
        if abs(val) >= 1_000:
            return f"${val / 1_000:,.1f}B"
        return f"${val:,.0f}M"
    return f"${val:,.0f}"


def _fmt_pct(val: Optional[float]) -> str:
    if val is None:
        return "\u2014"
    return f"{val:.1f}%"


def _fmt_x(val: Optional[float]) -> str:
    if val is None:
        return "\u2014"
    return f"{val:.2f}x"


class PEFundTearsheetTemplate:
    """PE Fund Tearsheet report template."""

    name = "pe_fund_tearsheet"
    description = "Single-page fund tearsheet with performance metrics, J-curve, and portfolio summary"

    # ── Data Gathering ─────────────────────────────────────────────

    def gather_data(self, db: Session, params: Dict[str, Any]) -> Dict[str, Any]:
        """Gather all data for the fund tearsheet."""
        fund_id = params.get("fund_id")
        if not fund_id:
            raise ValueError("fund_id is required")

        fund = self._get_fund(db, fund_id)
        if not fund:
            return {
                "generated_at": datetime.utcnow().isoformat(),
                "fund": None,
                "performance": {},
                "portfolio": [],
                "cash_flows": [],
                "sector_allocation": [],
                "top_performers": [],
                "bottom_performers": [],
            }

        return {
            "generated_at": datetime.utcnow().isoformat(),
            "fund": fund,
            "performance": self._get_performance(db, fund_id),
            "portfolio": self._get_portfolio(db, fund_id),
            "cash_flows": self._get_cash_flows(db, fund_id),
            "sector_allocation": self._get_sector_allocation(db, fund_id),
            "top_performers": self._get_top_performers(db, fund_id),
            "bottom_performers": self._get_bottom_performers(db, fund_id),
        }

    def _get_fund(self, db: Session, fund_id: int) -> Optional[Dict]:
        row = db.execute(text("""
            SELECT f.id, f.name, f.vintage_year, f.target_size_usd_millions,
                   f.final_close_usd_millions, f.strategy, f.status,
                   f.management_fee_pct, f.carried_interest_pct,
                   f.preferred_return_pct, f.fund_life_years,
                   f.investment_period_years, f.first_close_date,
                   f.final_close_date, fi.name as firm_name
            FROM pe_funds f
            LEFT JOIN pe_firms fi ON f.firm_id = fi.id
            WHERE f.id = :id
        """), {"id": fund_id}).fetchone()
        if not row:
            return None
        return {
            "id": row[0], "name": row[1], "vintage": row[2],
            "target_size": float(row[3]) if row[3] else None,
            "final_close": float(row[4]) if row[4] else None,
            "strategy": row[5], "status": row[6],
            "mgmt_fee": float(row[7]) if row[7] else None,
            "carry": float(row[8]) if row[8] else None,
            "pref_return": float(row[9]) if row[9] else None,
            "fund_life": row[10], "invest_period": row[11],
            "first_close": row[12].isoformat() if row[12] else None,
            "final_close_date": row[13].isoformat() if row[13] else None,
            "firm_name": row[14],
        }

    def _get_performance(self, db: Session, fund_id: int) -> Dict:
        row = db.execute(text("""
            SELECT net_irr_pct, gross_irr_pct, tvpi, dpi, rvpi,
                   committed_capital, called_capital, distributed_capital,
                   remaining_value, active_investments, realized_investments,
                   written_off_investments, as_of_date
            FROM pe_fund_performance
            WHERE fund_id = :fid
            ORDER BY as_of_date DESC
            LIMIT 1
        """), {"fid": fund_id}).fetchone()
        if not row:
            return {}
        called = float(row[6]) if row[6] else 0
        distributed = float(row[7]) if row[7] else 0
        remaining = float(row[8]) if row[8] else 0
        moic = (distributed + remaining) / called if called > 0 else None
        return {
            "net_irr": float(row[0]) if row[0] else None,
            "gross_irr": float(row[1]) if row[1] else None,
            "tvpi": float(row[2]) if row[2] else None,
            "dpi": float(row[3]) if row[3] else None,
            "rvpi": float(row[4]) if row[4] else None,
            "committed": float(row[5]) if row[5] else None,
            "called": called,
            "distributed": distributed,
            "remaining": remaining,
            "active": row[9] or 0,
            "realized": row[10] or 0,
            "written_off": row[11] or 0,
            "as_of_date": row[12].isoformat() if row[12] else None,
            "moic": moic,
        }

    def _get_portfolio(self, db: Session, fund_id: int) -> List[Dict]:
        rows = db.execute(text("""
            SELECT pc.id, pc.name, pc.industry, pc.status,
                   fi.investment_date, fi.invested_amount_usd,
                   fi.entry_ev_usd, fi.entry_ev_ebitda_multiple,
                   fi.ownership_pct, fi.exit_date, fi.exit_type,
                   fi.exit_amount_usd, fi.exit_multiple, fi.exit_irr_pct,
                   fi.status as inv_status
            FROM pe_fund_investments fi
            JOIN pe_portfolio_companies pc ON fi.company_id = pc.id
            WHERE fi.fund_id = :fid
            ORDER BY fi.investment_date
        """), {"fid": fund_id}).fetchall()
        return [
            {
                "id": r[0], "name": r[1], "industry": r[2] or "Unknown",
                "status": r[3] or "active",
                "invested_date": r[4].isoformat() if r[4] else None,
                "invested_amount": float(r[5]) if r[5] else None,
                "entry_ev": float(r[6]) if r[6] else None,
                "entry_multiple": float(r[7]) if r[7] else None,
                "ownership_pct": float(r[8]) if r[8] else None,
                "exit_date": r[9].isoformat() if r[9] else None,
                "exit_type": r[10],
                "exit_amount": float(r[11]) if r[11] else None,
                "exit_multiple": float(r[12]) if r[12] else None,
                "exit_irr": float(r[13]) if r[13] else None,
                "inv_status": r[14] or "active",
            }
            for r in rows
        ]

    def _get_cash_flows(self, db: Session, fund_id: int) -> List[Dict]:
        rows = db.execute(text("""
            SELECT flow_date, amount, cash_flow_type, description
            FROM pe_cash_flows
            WHERE fund_id = :fid
            ORDER BY flow_date
        """), {"fid": fund_id}).fetchall()
        return [
            {
                "date": r[0].isoformat() if r[0] else None,
                "amount": float(r[1]) if r[1] else 0,
                "type": r[2], "description": r[3],
            }
            for r in rows
        ]

    def _get_sector_allocation(self, db: Session, fund_id: int) -> List[Dict]:
        rows = db.execute(text("""
            SELECT COALESCE(pc.industry, 'Unknown') as ind,
                   COUNT(*) as cnt,
                   SUM(COALESCE(fi.invested_amount_usd, 0)) as total_invested
            FROM pe_fund_investments fi
            JOIN pe_portfolio_companies pc ON fi.company_id = pc.id
            WHERE fi.fund_id = :fid
            GROUP BY pc.industry
            ORDER BY total_invested DESC
        """), {"fid": fund_id}).fetchall()
        total = sum(float(r[2]) for r in rows) or 1
        return [
            {
                "sector": r[0], "count": r[1],
                "invested": float(r[2]) if r[2] else 0,
                "pct": round(float(r[2]) / total * 100, 1) if r[2] else 0,
            }
            for r in rows
        ]

    def _get_top_performers(self, db: Session, fund_id: int) -> List[Dict]:
        rows = db.execute(text("""
            SELECT pc.name, fi.exit_multiple, fi.exit_irr_pct,
                   fi.invested_amount_usd, fi.exit_amount_usd, fi.exit_type
            FROM pe_fund_investments fi
            JOIN pe_portfolio_companies pc ON fi.company_id = pc.id
            WHERE fi.fund_id = :fid
              AND fi.exit_multiple IS NOT NULL
            ORDER BY fi.exit_multiple DESC
            LIMIT 5
        """), {"fid": fund_id}).fetchall()
        return [
            {
                "name": r[0], "multiple": float(r[1]) if r[1] else None,
                "irr": float(r[2]) if r[2] else None,
                "invested": float(r[3]) if r[3] else None,
                "exit_value": float(r[4]) if r[4] else None,
                "exit_type": r[5],
            }
            for r in rows
        ]

    def _get_bottom_performers(self, db: Session, fund_id: int) -> List[Dict]:
        rows = db.execute(text("""
            SELECT pc.name, fi.exit_multiple, fi.exit_irr_pct,
                   fi.invested_amount_usd, fi.exit_amount_usd, fi.exit_type
            FROM pe_fund_investments fi
            JOIN pe_portfolio_companies pc ON fi.company_id = pc.id
            WHERE fi.fund_id = :fid
              AND fi.exit_multiple IS NOT NULL
            ORDER BY fi.exit_multiple ASC
            LIMIT 3
        """), {"fid": fund_id}).fetchall()
        return [
            {
                "name": r[0], "multiple": float(r[1]) if r[1] else None,
                "irr": float(r[2]) if r[2] else None,
                "invested": float(r[3]) if r[3] else None,
                "exit_value": float(r[4]) if r[4] else None,
                "exit_type": r[5],
            }
            for r in rows
        ]

    # ── HTML Rendering ─────────────────────────────────────────────

    def render_html(self, data: Dict[str, Any]) -> str:
        fund = data.get("fund")
        if not fund:
            return html_document(
                title="Fund Not Found",
                body_content=callout("<strong>Fund not found.</strong> Check the fund_id parameter.", "warn"),
            )

        perf = data.get("performance", {})
        portfolio = data.get("portfolio", [])
        cash_flows = data.get("cash_flows", [])
        sectors = data.get("sector_allocation", [])
        top = data.get("top_performers", [])
        bottom = data.get("bottom_performers", [])
        title = data.get("report_title", f"{fund['name']} \u2014 Fund Tearsheet")

        charts_js = ""
        body = ""

        # ── Header ─────────────────────────────────────────────────
        subtitle_parts = [p for p in [
            fund.get("firm_name"),
            f"Vintage {fund['vintage']}" if fund.get("vintage") else None,
            fund.get("strategy"),
        ] if p]
        body += page_header(
            title=title,
            subtitle=" \u00b7 ".join(subtitle_parts) if subtitle_parts else None,
            badge=f"Generated {datetime.utcnow().strftime('%b %d, %Y')}",
        )

        # ── KPI Strip ──────────────────────────────────────────────
        fund_size = fund.get("final_close") or fund.get("target_size")
        cards = ""
        cards += kpi_card("Fund Size", _fmt_usd(fund_size))
        cards += kpi_card("Net IRR", _fmt_pct(perf.get("net_irr")),
                          delta=_fmt_pct(perf.get("gross_irr")) + " gross" if perf.get("gross_irr") else None)
        cards += kpi_card("TVPI", _fmt_x(perf.get("tvpi")))
        cards += kpi_card("DPI", _fmt_x(perf.get("dpi")))
        cards += kpi_card("MOIC", _fmt_x(perf.get("moic")))
        body += "\n" + kpi_strip(cards)

        # ── TOC ────────────────────────────────────────────────────
        toc_items = [
            {"number": 1, "id": "fund-overview", "title": "Fund Overview"},
            {"number": 2, "id": "performance", "title": "Performance & J-Curve"},
            {"number": 3, "id": "portfolio", "title": "Portfolio Companies"},
            {"number": 4, "id": "sectors", "title": "Sector Allocation"},
            {"number": 5, "id": "performers", "title": "Top & Bottom Performers"},
            {"number": 6, "id": "cashflows", "title": "Cash Flow Summary"},
        ]
        body += "\n" + toc(toc_items)

        # ── Section 1: Fund Overview ───────────────────────────────
        body += "\n" + section_start(1, "Fund Overview", "fund-overview")

        overview_rows = [
            ["Fund Name", fund["name"]],
            ["Firm", fund.get("firm_name") or "\u2014"],
            ["Vintage Year", str(fund.get("vintage") or "\u2014")],
            ["Strategy", fund.get("strategy") or "\u2014"],
            ["Status", pill_badge((fund.get("status") or "unknown").title(),
                                  "public" if fund.get("status") == "active" else "default")],
            ["Target Size", _fmt_usd(fund.get("target_size"))],
            ["Final Close", _fmt_usd(fund.get("final_close"))],
            ["Management Fee", _fmt_pct(fund.get("mgmt_fee"))],
            ["Carried Interest", _fmt_pct(fund.get("carry"))],
            ["Preferred Return", _fmt_pct(fund.get("pref_return"))],
            ["Fund Life", f"{fund['fund_life']} years" if fund.get("fund_life") else "\u2014"],
            ["Investment Period", f"{fund['invest_period']} years" if fund.get("invest_period") else "\u2014"],
        ]
        body += data_table(["Parameter", "Value"], overview_rows)

        if perf:
            body += '<h3 style="margin:16px 0 8px;font-size:1rem">Capital Summary</h3>'
            cap_rows = [
                ["Committed Capital", _fmt_usd(perf.get("committed"))],
                ["Called Capital", _fmt_usd(perf.get("called"))],
                ["Distributed Capital", _fmt_usd(perf.get("distributed"))],
                ["Remaining Value", _fmt_usd(perf.get("remaining"))],
                ["Active Investments", str(perf.get("active", 0))],
                ["Realized Investments", str(perf.get("realized", 0))],
                ["Written Off", str(perf.get("written_off", 0))],
            ]
            body += data_table(["Metric", "Value"], cap_rows, numeric_columns={1})
            if perf.get("as_of_date"):
                body += f'<p style="color:var(--gray-500);font-size:0.8rem;margin-top:4px">As of {perf["as_of_date"]}</p>'

        body += "\n" + section_end()

        # ── Section 2: Performance & J-Curve ───────────────────────
        body += "\n" + section_start(2, "Performance & J-Curve", "performance")

        if perf:
            perf_rows = [
                ["Net IRR", _fmt_pct(perf.get("net_irr"))],
                ["Gross IRR", _fmt_pct(perf.get("gross_irr"))],
                ["TVPI", _fmt_x(perf.get("tvpi"))],
                ["DPI", _fmt_x(perf.get("dpi"))],
                ["RVPI", _fmt_x(perf.get("rvpi"))],
                ["MOIC", _fmt_x(perf.get("moic"))],
            ]
            body += data_table(["Metric", "Value"], perf_rows, numeric_columns={1})

        # J-curve chart from cash flows
        if cash_flows:
            cumulative = []
            running = 0.0
            labels = []
            for cf in cash_flows:
                running += cf["amount"]
                cumulative.append(round(running, 2))
                labels.append(cf["date"][:7] if cf.get("date") else "")

            j_config = build_line_chart_config(
                labels=labels,
                datasets=[{
                    "label": "Cumulative Net Cash Flow ($M)",
                    "data": cumulative,
                    "borderColor": BLUE,
                    "backgroundColor": f"{BLUE}20",
                    "fill": True,
                    "tension": 0.3,
                }],
                y_label="Cumulative ($M)",
            )
            j_json = json.dumps(j_config)

            body += '<div style="margin:16px 0">'
            body += chart_container(
                "jCurveChart", j_json,
                build_bar_fallback(labels[-10:], cumulative[-10:], BLUE),
                size="large", title="J-Curve \u2014 Cumulative Net Cash Flow",
            )
            charts_js += chart_init_js("jCurveChart", j_json)
            body += '</div>'
        else:
            body += callout("<strong>No cash flow data available</strong> for J-curve chart.", "warn")

        body += "\n" + section_end()

        # ── Section 3: Portfolio Companies ─────────────────────────
        body += "\n" + section_start(3, "Portfolio Companies", "portfolio")

        if portfolio:
            # Valuation bar chart
            co_names = [p["name"][:20] for p in portfolio if p.get("entry_ev")]
            co_evs = [p["entry_ev"] / 1_000_000 for p in portfolio if p.get("entry_ev")]
            if co_names:
                bar_config = build_horizontal_bar_config(
                    co_names, co_evs,
                    [CHART_COLORS[i % len(CHART_COLORS)] for i in range(len(co_names))],
                    dataset_label="Entry EV ($M)",
                )
                bar_json = json.dumps(bar_config)
                body += chart_container(
                    "portfolioBarChart", bar_json,
                    build_bar_fallback(co_names, co_evs, BLUE),
                    size="medium", title="Portfolio Entry Valuations",
                )
                charts_js += chart_init_js("portfolioBarChart", bar_json)

            # Portfolio table
            headers = ["Company", "Industry", "Status", "Invested", "Entry EV",
                       "Entry Multiple", "Ownership", "Exit Multiple", "Exit IRR"]
            rows = []
            for p in portfolio:
                inv_status = p.get("inv_status", "active")
                variant = "public" if inv_status == "active" else "pe" if inv_status == "realized" else "default"
                rows.append([
                    p["name"],
                    p["industry"],
                    pill_badge(inv_status.title(), variant),
                    _fmt_usd(p.get("invested_amount") / 1_000_000 if p.get("invested_amount") else None),
                    _fmt_usd(p.get("entry_ev") / 1_000_000 if p.get("entry_ev") else None),
                    _fmt_x(p.get("entry_multiple")),
                    _fmt_pct(p.get("ownership_pct")),
                    _fmt_x(p.get("exit_multiple")),
                    _fmt_pct(p.get("exit_irr")),
                ])
            body += data_table(headers, rows, numeric_columns={3, 4, 5, 6, 7, 8})
        else:
            body += callout("<strong>No portfolio company data available.</strong>", "warn")

        body += "\n" + section_end()

        # ── Section 4: Sector Allocation ───────────────────────────
        body += "\n" + section_start(4, "Sector Allocation", "sectors")

        if sectors:
            sec_labels = [s["sector"] for s in sectors]
            sec_values = [float(s["invested"]) / 1_000_000 for s in sectors]
            sec_colors = [CHART_COLORS[i % len(CHART_COLORS)] for i in range(len(sec_labels))]
            donut_cfg = build_doughnut_config(sec_labels, sec_values, sec_colors)
            donut_json = json.dumps(donut_cfg)

            body += '<div class="chart-row"><div>'
            body += chart_container(
                "sectorChart", donut_json,
                build_bar_fallback(sec_labels, sec_values),
                size="medium", title="Capital Deployed by Sector",
            )
            charts_js += chart_init_js("sectorChart", donut_json)
            body += '</div><div>'
            body += build_chart_legend(sec_labels, sec_values, sec_colors,
                                       value_suffix="M", show_pct=True)

            sec_rows = [[s["sector"], str(s["count"]),
                         _fmt_usd(s["invested"] / 1_000_000 if s["invested"] else 0),
                         f'{s["pct"]}%'] for s in sectors]
            body += data_table(["Sector", "Companies", "Invested", "%"], sec_rows, numeric_columns={1, 2, 3})
            body += '</div></div>'
        else:
            body += callout("<strong>No sector allocation data available.</strong>", "warn")

        body += "\n" + section_end()

        # ── Section 5: Top & Bottom Performers ─────────────────────
        body += "\n" + section_start(5, "Top & Bottom Performers", "performers")

        if top:
            body += '<h3 style="margin:0 0 8px;font-size:1rem;color:var(--gray-700)">Top Performers</h3>'
            top_rows = []
            for t in top:
                top_rows.append([
                    t["name"],
                    _fmt_x(t.get("multiple")),
                    _fmt_pct(t.get("irr")),
                    _fmt_usd(t.get("invested") / 1_000_000 if t.get("invested") else None),
                    _fmt_usd(t.get("exit_value") / 1_000_000 if t.get("exit_value") else None),
                    t.get("exit_type") or "\u2014",
                ])
            body += data_table(
                ["Company", "Multiple", "IRR", "Invested", "Exit Value", "Exit Type"],
                top_rows, numeric_columns={1, 2, 3, 4},
            )

        if bottom:
            body += '<h3 style="margin:16px 0 8px;font-size:1rem;color:var(--gray-700)">Bottom Performers</h3>'
            bot_rows = []
            for b in bottom:
                bot_rows.append([
                    b["name"],
                    _fmt_x(b.get("multiple")),
                    _fmt_pct(b.get("irr")),
                    _fmt_usd(b.get("invested") / 1_000_000 if b.get("invested") else None),
                    _fmt_usd(b.get("exit_value") / 1_000_000 if b.get("exit_value") else None),
                    b.get("exit_type") or "\u2014",
                ])
            body += data_table(
                ["Company", "Multiple", "IRR", "Invested", "Exit Value", "Exit Type"],
                bot_rows, numeric_columns={1, 2, 3, 4},
            )

        if not top and not bottom:
            body += callout("<strong>No realized exits with return data available.</strong>", "warn")

        body += "\n" + section_end()

        # ── Section 6: Cash Flow Summary ───────────────────────────
        body += "\n" + section_start(6, "Cash Flow Summary", "cashflows")

        if cash_flows:
            calls = sum(cf["amount"] for cf in cash_flows if cf["amount"] < 0)
            distribs = sum(cf["amount"] for cf in cash_flows if cf["amount"] > 0)

            body += f'<p><strong>Total Capital Calls:</strong> ${abs(calls):,.0f}M &nbsp;|&nbsp; '
            body += f'<strong>Total Distributions:</strong> ${distribs:,.0f}M &nbsp;|&nbsp; '
            body += f'<strong>Net:</strong> ${calls + distribs:,.0f}M</p>'

            cf_rows = []
            for cf in cash_flows:
                color = GREEN if cf["amount"] > 0 else RED
                formatted = f'<span style="color:{color}">${cf["amount"]:,.1f}M</span>'
                cf_rows.append([
                    cf.get("date", "\u2014"),
                    cf.get("type", "\u2014"),
                    formatted,
                    cf.get("description") or "\u2014",
                ])
            body += data_table(
                ["Date", "Type", "Amount", "Description"],
                cf_rows, numeric_columns={2},
            )
        else:
            body += callout("<strong>No cash flow records available.</strong>", "warn")

        body += "\n" + section_end()

        # ── Footer ─────────────────────────────────────────────────
        body += "\n" + page_footer(
            notes=[
                "Performance data sourced from fund reports and SEC filings.",
                "MOIC calculated as (Distributions + Remaining Value) / Called Capital.",
                "J-curve shows cumulative net cash flow (distributions minus calls).",
            ],
            generated_line=f"Report generated {data.get('generated_at', 'N/A')} | Nexdata PE Intelligence",
        )

        extra_css = """
        .chart-row { display: grid; grid-template-columns: 1fr 1fr; gap: 24px; margin: 16px 0; }
        @media (max-width: 768px) { .chart-row { grid-template-columns: 1fr; } }
        """

        return html_document(
            title=title,
            body_content=body,
            charts_js=charts_js,
            extra_css=extra_css,
        )

    # ── Excel Rendering ────────────────────────────────────────────

    def render_excel(self, data: Dict[str, Any]) -> bytes:
        """Render report as multi-sheet Excel workbook."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        hdr_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
        hdr_font = Font(bold=True, color="FFFFFF")

        fund = data.get("fund") or {}
        perf = data.get("performance", {})

        # ── Sheet 1: Summary ───────────────────────────────────────
        ws = wb.active
        ws.title = "Summary"

        summary_data = [
            ["Fund Name", fund.get("name", "")],
            ["Firm", fund.get("firm_name", "")],
            ["Vintage", fund.get("vintage")],
            ["Strategy", fund.get("strategy", "")],
            ["Status", fund.get("status", "")],
            ["Final Close ($M)", fund.get("final_close")],
            ["Management Fee (%)", fund.get("mgmt_fee")],
            ["Carried Interest (%)", fund.get("carry")],
            [""],
            ["PERFORMANCE METRICS"],
            ["Net IRR (%)", perf.get("net_irr")],
            ["Gross IRR (%)", perf.get("gross_irr")],
            ["TVPI", perf.get("tvpi")],
            ["DPI", perf.get("dpi")],
            ["RVPI", perf.get("rvpi")],
            ["MOIC", perf.get("moic")],
            [""],
            ["CAPITAL"],
            ["Committed ($M)", perf.get("committed")],
            ["Called ($M)", perf.get("called")],
            ["Distributed ($M)", perf.get("distributed")],
            ["Remaining Value ($M)", perf.get("remaining")],
        ]
        for row_idx, row_data in enumerate(summary_data, 1):
            if len(row_data) == 1:
                continue
            ws.cell(row=row_idx, column=1, value=row_data[0]).font = Font(bold=True)
            if len(row_data) > 1:
                ws.cell(row=row_idx, column=2, value=row_data[1])
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20

        # ── Sheet 2: Portfolio ─────────────────────────────────────
        ws_p = wb.create_sheet("Portfolio")
        p_headers = ["Company", "Industry", "Status", "Invested ($)",
                     "Entry EV ($)", "Entry Multiple", "Ownership %",
                     "Exit Multiple", "Exit IRR %", "Exit Type"]
        for col, h in enumerate(p_headers, 1):
            cell = ws_p.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill

        for row_idx, p in enumerate(data.get("portfolio", []), 2):
            ws_p.cell(row=row_idx, column=1, value=p.get("name"))
            ws_p.cell(row=row_idx, column=2, value=p.get("industry"))
            ws_p.cell(row=row_idx, column=3, value=p.get("inv_status"))
            ws_p.cell(row=row_idx, column=4, value=p.get("invested_amount"))
            ws_p.cell(row=row_idx, column=5, value=p.get("entry_ev"))
            ws_p.cell(row=row_idx, column=6, value=p.get("entry_multiple"))
            ws_p.cell(row=row_idx, column=7, value=p.get("ownership_pct"))
            ws_p.cell(row=row_idx, column=8, value=p.get("exit_multiple"))
            ws_p.cell(row=row_idx, column=9, value=p.get("exit_irr"))
            ws_p.cell(row=row_idx, column=10, value=p.get("exit_type"))
        for col_letter in "ABCDEFGHIJ":
            ws_p.column_dimensions[col_letter].width = 16

        # ── Sheet 3: Cash Flows ────────────────────────────────────
        ws_cf = wb.create_sheet("Cash Flows")
        cf_headers = ["Date", "Type", "Amount ($M)", "Description"]
        for col, h in enumerate(cf_headers, 1):
            cell = ws_cf.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill

        for row_idx, cf in enumerate(data.get("cash_flows", []), 2):
            ws_cf.cell(row=row_idx, column=1, value=cf.get("date"))
            ws_cf.cell(row=row_idx, column=2, value=cf.get("type"))
            ws_cf.cell(row=row_idx, column=3, value=cf.get("amount"))
            ws_cf.cell(row=row_idx, column=4, value=cf.get("description"))
        ws_cf.column_dimensions["A"].width = 14
        ws_cf.column_dimensions["B"].width = 16
        ws_cf.column_dimensions["C"].width = 14
        ws_cf.column_dimensions["D"].width = 30

        # ── Sheet 4: Performance ───────────────────────────────────
        ws_perf = wb.create_sheet("Performance")
        perf_headers = ["Metric", "Value"]
        for col, h in enumerate(perf_headers, 1):
            cell = ws_perf.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill

        perf_data = [
            ("Net IRR (%)", perf.get("net_irr")),
            ("Gross IRR (%)", perf.get("gross_irr")),
            ("TVPI", perf.get("tvpi")),
            ("DPI", perf.get("dpi")),
            ("RVPI", perf.get("rvpi")),
            ("MOIC", perf.get("moic")),
            ("Committed ($M)", perf.get("committed")),
            ("Called ($M)", perf.get("called")),
            ("Distributed ($M)", perf.get("distributed")),
            ("Remaining Value ($M)", perf.get("remaining")),
            ("Active Investments", perf.get("active")),
            ("Realized Investments", perf.get("realized")),
            ("Written Off", perf.get("written_off")),
        ]
        for row_idx, (metric, value) in enumerate(perf_data, 2):
            ws_perf.cell(row=row_idx, column=1, value=metric)
            ws_perf.cell(row=row_idx, column=2, value=value)
        ws_perf.column_dimensions["A"].width = 25
        ws_perf.column_dimensions["B"].width = 16

        output = BytesIO()
        wb.save(output)
        return output.getvalue()
