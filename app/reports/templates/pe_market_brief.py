"""
PE Market Brief Report Template.

Generates a 2-3 page market intelligence brief with:
- Sector Overview (momentum scores, deal activity)
- Deal Activity Trends (volume, multiples over time)
- Fragmentation Opportunities (NAICS-based market concentration)
- Top Rollup Targets (high-fragmentation sectors with active PE interest)
- Market Timing Assessment (buy/sell signals)

Uses the shared design_system.py patterns.
"""

import json
import logging
from datetime import datetime
from typing import Dict, Any, List, Optional

from io import BytesIO
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.reports.design_system import (
    html_document, page_header, kpi_strip, kpi_card,
    toc, section_start, section_end,
    data_table, callout, pill_badge,
    chart_container, chart_init_js, page_footer,
    build_doughnut_config, build_horizontal_bar_config,
    build_line_chart_config,
    build_bar_fallback, build_chart_legend, CHART_COLORS,
    BLUE, GREEN, ORANGE, RED, GRAY, TEAL, PURPLE,
)

logger = logging.getLogger(__name__)


def _fmt_usd(val: Optional[float]) -> str:
    if val is None:
        return "\u2014"
    if abs(val) >= 1_000:
        return f"${val / 1_000:,.1f}B"
    return f"${val:,.0f}M"


def _fmt_pct(val: Optional[float]) -> str:
    if val is None:
        return "\u2014"
    return f"{val:.1f}%"


def _fmt_x(val: Optional[float]) -> str:
    if val is None:
        return "\u2014"
    return f"{val:.1f}x"


def _momentum_color(score: Optional[float]) -> str:
    if score is None:
        return GRAY
    if score >= 65:
        return GREEN
    if score >= 40:
        return ORANGE
    return RED


def _momentum_label(score: Optional[float]) -> str:
    if score is None:
        return "Unknown"
    if score >= 65:
        return "Bullish"
    if score >= 40:
        return "Neutral"
    return "Bearish"


class PEMarketBriefTemplate:
    """PE Market Brief report template."""

    name = "pe_market_brief"
    description = "Market intelligence brief with sector momentum, deal trends, and rollup opportunities"

    # ── Data Gathering ─────────────────────────────────────────────

    def gather_data(self, db: Session, params: Dict[str, Any]) -> Dict[str, Any]:
        industry = params.get("industry")  # Optional filter

        return {
            "generated_at": datetime.utcnow().isoformat(),
            "industry_filter": industry,
            "sector_signals": self._get_sector_signals(db),
            "deal_activity": self._get_deal_activity(db, industry),
            "deal_type_breakdown": self._get_deal_type_breakdown(db, industry),
            "top_buyers": self._get_top_buyers(db, industry),
            "fragmentation": self._get_fragmentation(db),
            "rollup_targets": self._get_rollup_targets(db),
            "timing_assessment": self._build_timing_assessment(db, industry),
        }

    def _get_sector_signals(self, db: Session) -> List[Dict]:
        """Get momentum signals per sector from market scanner."""
        try:
            from app.core.pe_market_scanner import MarketScannerService
            svc = MarketScannerService(db)
            return svc.get_sectors_ranked()
        except Exception:
            logger.debug("Market scanner unavailable, falling back to direct query")
            # Fallback: query deal counts by industry
            rows = db.execute(text("""
                SELECT COALESCE(d.deal_type, 'Unknown') as industry,
                       COUNT(*) as deal_count,
                       AVG(d.ev_ebitda_multiple) as avg_multiple
                FROM pe_deals d
                WHERE d.status = 'closed'
                  AND d.closed_date >= NOW() - INTERVAL '2 years'
                GROUP BY d.deal_type
                ORDER BY deal_count DESC
                LIMIT 10
            """)).fetchall()
            return [
                {
                    "industry": r[0], "deal_count": r[1],
                    "median_ev_ebitda": float(r[2]) if r[2] else None,
                    "momentum_score": None, "momentum": "unknown",
                }
                for r in rows
            ]

    def _get_deal_activity(self, db: Session, industry: Optional[str]) -> List[Dict]:
        """Get quarterly deal counts and avg multiples."""
        where = "AND d.deal_type = :ind" if industry else ""
        params = {"ind": industry} if industry else {}

        rows = db.execute(text(f"""
            SELECT DATE_TRUNC('quarter', d.closed_date) as quarter,
                   COUNT(*) as deal_count,
                   AVG(d.ev_ebitda_multiple) as avg_ev_ebitda,
                   SUM(d.enterprise_value_usd) as total_value
            FROM pe_deals d
            WHERE d.status = 'closed'
              AND d.closed_date IS NOT NULL
              {where}
            GROUP BY quarter
            ORDER BY quarter DESC
            LIMIT 12
        """), params).fetchall()

        return [
            {
                "quarter": f"{r[0].year}-Q{(r[0].month - 1) // 3 + 1}" if hasattr(r[0], 'year') else str(r[0])[:7],
                "quarter_date": r[0].isoformat() if r[0] else None,
                "deal_count": r[1],
                "avg_ev_ebitda": float(r[2]) if r[2] else None,
                "total_value": float(r[3]) / 1_000_000 if r[3] else 0,
            }
            for r in rows
        ]

    def _get_deal_type_breakdown(self, db: Session, industry: Optional[str]) -> List[Dict]:
        where = "AND d.deal_type = :ind" if industry else ""
        params = {"ind": industry} if industry else {}

        rows = db.execute(text(f"""
            SELECT COALESCE(d.deal_sub_type, d.deal_type, 'Unknown') as dtype,
                   COUNT(*) as cnt,
                   AVG(d.ev_ebitda_multiple) as avg_mult
            FROM pe_deals d
            WHERE d.status = 'closed'
              {where}
            GROUP BY dtype
            ORDER BY cnt DESC
        """), params).fetchall()

        total = sum(r[1] for r in rows) or 1
        return [
            {
                "type": r[0], "count": r[1],
                "pct": round(r[1] / total * 100, 1),
                "avg_multiple": float(r[2]) if r[2] else None,
            }
            for r in rows
        ]

    def _get_top_buyers(self, db: Session, industry: Optional[str]) -> List[Dict]:
        where = "AND d.deal_type = :ind" if industry else ""
        params = {"ind": industry} if industry else {}

        rows = db.execute(text(f"""
            SELECT d.buyer_name, COUNT(*) as deal_count,
                   SUM(d.enterprise_value_usd) as total_value
            FROM pe_deals d
            WHERE d.status = 'closed'
              AND d.buyer_name IS NOT NULL
              {where}
            GROUP BY d.buyer_name
            ORDER BY deal_count DESC
            LIMIT 10
        """), params).fetchall()

        return [
            {
                "buyer": r[0], "deals": r[1],
                "total_value": float(r[2]) / 1_000_000 if r[2] else 0,
            }
            for r in rows
        ]

    def _get_fragmentation(self, db: Session) -> List[Dict]:
        """Get market fragmentation data from census/CBP tables if available."""
        try:
            rows = db.execute(text("""
                SELECT naics_code, naics_title, establishments, score,
                       top_states, avg_employees_per_firm
                FROM market_fragmentation_scores
                ORDER BY score DESC
                LIMIT 15
            """)).fetchall()
            return [
                {
                    "naics": r[0], "title": r[1], "establishments": r[2],
                    "score": float(r[3]) if r[3] else None,
                    "top_states": r[4], "avg_employees": float(r[5]) if r[5] else None,
                }
                for r in rows
            ]
        except Exception:
            logger.debug("Fragmentation scores table not available")
            return []

    def _get_rollup_targets(self, db: Session) -> List[Dict]:
        """Get potential rollup targets — high-fragmentation industries with PE activity."""
        try:
            rows = db.execute(text("""
                SELECT pc.industry, COUNT(DISTINCT pc.id) as company_count,
                       COUNT(DISTINCT d.id) as deal_count,
                       AVG(d.ev_ebitda_multiple) as avg_multiple
                FROM pe_portfolio_companies pc
                LEFT JOIN pe_fund_investments fi ON fi.company_id = pc.id
                LEFT JOIN pe_deals d ON d.company_id = pc.id AND d.status = 'closed'
                WHERE pc.industry IS NOT NULL
                GROUP BY pc.industry
                HAVING COUNT(DISTINCT pc.id) >= 2
                ORDER BY company_count DESC
                LIMIT 10
            """)).fetchall()
            return [
                {
                    "industry": r[0], "companies": r[1],
                    "deals": r[2],
                    "avg_multiple": float(r[3]) if r[3] else None,
                }
                for r in rows
            ]
        except Exception:
            logger.debug("Rollup target query failed")
            return []

    def _build_timing_assessment(self, db: Session, industry: Optional[str]) -> Dict:
        """Build market timing signals from recent deal activity."""
        try:
            from app.core.pe_market_scanner import MarketScannerService
            svc = MarketScannerService(db)
            if industry:
                brief = svc.get_intelligence_brief(industry)
                return {
                    "headline": brief.get("headline", ""),
                    "findings": brief.get("key_findings", []),
                    "recommendations": brief.get("recommendations", []),
                    "market_condition": brief.get("market_condition", "unknown"),
                }
            else:
                signals = svc.get_market_signals()
                bullish = sum(1 for s in signals if s.get("momentum") == "bullish")
                bearish = sum(1 for s in signals if s.get("momentum") == "bearish")
                total = len(signals) or 1
                return {
                    "headline": f"{bullish}/{total} sectors bullish, {bearish}/{total} bearish",
                    "findings": [
                        f"{bullish} sectors showing positive momentum",
                        f"{bearish} sectors showing negative momentum",
                        f"{total - bullish - bearish} sectors neutral",
                    ],
                    "recommendations": [],
                    "market_condition": "bullish" if bullish > bearish else "bearish" if bearish > bullish else "mixed",
                }
        except Exception:
            logger.debug("Timing assessment unavailable")
            return {"headline": "", "findings": [], "recommendations": [], "market_condition": "unknown"}

    # ── HTML Rendering ─────────────────────────────────────────────

    def render_html(self, data: Dict[str, Any]) -> str:
        signals = data.get("sector_signals", [])
        deals = data.get("deal_activity", [])
        type_breakdown = data.get("deal_type_breakdown", [])
        buyers = data.get("top_buyers", [])
        frag = data.get("fragmentation", [])
        rollups = data.get("rollup_targets", [])
        timing = data.get("timing_assessment", {})
        industry_filter = data.get("industry_filter")

        title_suffix = f" \u2014 {industry_filter}" if industry_filter else ""
        title = data.get("report_title", f"PE Market Intelligence Brief{title_suffix}")

        charts_js = ""
        body = ""

        # ── Header ─────────────────────────────────────────────────
        body += page_header(
            title=title,
            subtitle=timing.get("headline") or "Cross-sector PE market analysis",
            badge=f"Generated {datetime.utcnow().strftime('%b %d, %Y')}",
        )

        # ── KPI Strip ──────────────────────────────────────────────
        total_deals = sum(d.get("deal_count", 0) for d in deals)
        total_value = sum(d.get("total_value", 0) for d in deals)
        bullish_count = sum(1 for s in signals if s.get("momentum") == "bullish")
        avg_mult = None
        mults = [d["avg_ev_ebitda"] for d in deals if d.get("avg_ev_ebitda")]
        if mults:
            avg_mult = sum(mults) / len(mults)

        cards = ""
        cards += kpi_card("Total Deals", str(total_deals))
        cards += kpi_card("Deal Volume", _fmt_usd(total_value))
        cards += kpi_card("Avg EV/EBITDA", _fmt_x(avg_mult))
        cards += kpi_card("Bullish Sectors", f"{bullish_count}/{len(signals)}")
        cards += kpi_card("Sectors Tracked", str(len(signals)))
        body += "\n" + kpi_strip(cards)

        # ── TOC ────────────────────────────────────────────────────
        toc_items = [
            {"number": 1, "id": "sector-overview", "title": "Sector Overview"},
            {"number": 2, "id": "deal-activity", "title": "Deal Activity Trends"},
            {"number": 3, "id": "fragmentation", "title": "Fragmentation Opportunities"},
            {"number": 4, "id": "rollup-targets", "title": "Top Rollup Targets"},
            {"number": 5, "id": "timing", "title": "Market Timing Assessment"},
        ]
        body += "\n" + toc(toc_items)

        # ── Section 1: Sector Overview ─────────────────────────────
        body += "\n" + section_start(1, "Sector Overview", "sector-overview")

        if signals:
            # Momentum bar chart
            sec_labels = [s.get("industry", "?")[:25] for s in signals[:10]]
            sec_scores = [s.get("momentum_score") or 50 for s in signals[:10]]
            sec_colors = [_momentum_color(s.get("momentum_score")) for s in signals[:10]]
            bar_cfg = build_horizontal_bar_config(sec_labels, sec_scores, sec_colors, "Momentum Score")
            bar_json = json.dumps(bar_cfg)

            body += chart_container(
                "momentumChart", bar_json,
                build_bar_fallback(sec_labels, sec_scores, BLUE),
                size="medium", title="Sector Momentum Rankings",
            )
            charts_js += chart_init_js("momentumChart", bar_json)

            # Sector table
            sig_headers = ["Sector", "Momentum", "Score", "Deals", "Median EV/EBITDA"]
            sig_rows = []
            for s in signals:
                momentum = s.get("momentum", "unknown")
                score = s.get("momentum_score")
                color = _momentum_color(score)
                badge = f'<span style="padding:2px 8px;border-radius:4px;background:{color}20;color:{color};font-weight:600">{momentum.title()}</span>'
                sig_rows.append([
                    s.get("industry", "\u2014"),
                    badge,
                    f'{score:.0f}' if score else "\u2014",
                    str(s.get("deal_count", 0)),
                    _fmt_x(s.get("median_ev_ebitda")),
                ])
            body += data_table(sig_headers, sig_rows, numeric_columns={2, 3, 4})
        else:
            body += callout("<strong>No sector signals available.</strong> Seed deal data to populate.", "warn")

        body += "\n" + section_end()

        # ── Section 2: Deal Activity Trends ────────────────────────
        body += "\n" + section_start(2, "Deal Activity Trends", "deal-activity")

        if deals:
            # Reverse for chronological order
            sorted_deals = list(reversed(deals))
            q_labels = [d["quarter"] for d in sorted_deals]
            q_counts = [d["deal_count"] for d in sorted_deals]
            q_values = [d["total_value"] for d in sorted_deals]

            line_cfg = build_line_chart_config(
                labels=q_labels,
                datasets=[
                    {
                        "label": "Deal Count",
                        "data": q_counts,
                        "borderColor": BLUE,
                        "backgroundColor": f"{BLUE}20",
                        "yAxisID": "y",
                        "tension": 0.3,
                    },
                    {
                        "label": "Total Value ($M)",
                        "data": q_values,
                        "borderColor": GREEN,
                        "backgroundColor": f"{GREEN}20",
                        "yAxisID": "y1",
                        "tension": 0.3,
                    },
                ],
                y_label="Count / Value",
            )
            line_json = json.dumps(line_cfg)
            body += chart_container(
                "dealTrendChart", line_json,
                build_bar_fallback(q_labels[-6:], q_counts[-6:], BLUE),
                size="large", title="Quarterly Deal Activity",
            )
            charts_js += chart_init_js("dealTrendChart", line_json)

        # Deal type breakdown
        if type_breakdown:
            body += '<h3 style="margin:16px 0 8px;font-size:1rem">Deal Type Breakdown</h3>'

            dt_labels = [d["type"] for d in type_breakdown]
            dt_values = [float(d["count"]) for d in type_breakdown]
            dt_colors = [CHART_COLORS[i % len(CHART_COLORS)] for i in range(len(dt_labels))]
            donut_cfg = build_doughnut_config(dt_labels, dt_values, dt_colors)
            donut_json = json.dumps(donut_cfg)

            body += '<div class="chart-row"><div>'
            body += chart_container(
                "dealTypeChart", donut_json,
                build_bar_fallback(dt_labels, dt_values),
                size="medium", title="Deals by Type",
            )
            charts_js += chart_init_js("dealTypeChart", donut_json)
            body += '</div><div>'
            body += build_chart_legend(dt_labels, dt_values, dt_colors, show_pct=True)

            dt_rows = [[d["type"], str(d["count"]), f'{d["pct"]}%',
                         _fmt_x(d.get("avg_multiple"))] for d in type_breakdown]
            body += data_table(["Type", "Count", "%", "Avg Multiple"], dt_rows, numeric_columns={1, 2, 3})
            body += '</div></div>'

        # Top buyers
        if buyers:
            body += '<h3 style="margin:16px 0 8px;font-size:1rem">Most Active Buyers</h3>'
            buyer_rows = [[b["buyer"], str(b["deals"]), _fmt_usd(b["total_value"])] for b in buyers]
            body += data_table(["Buyer", "Deals", "Total Value"], buyer_rows, numeric_columns={1, 2})

        if not deals and not type_breakdown:
            body += callout("<strong>No deal activity data available.</strong>", "warn")

        body += "\n" + section_end()

        # ── Section 3: Fragmentation Opportunities ─────────────────
        body += "\n" + section_start(3, "Fragmentation Opportunities", "fragmentation")

        if frag:
            body += '<p>Highly fragmented industries with many small establishments present rollup opportunities.</p>'
            frag_headers = ["Industry", "NAICS", "Establishments", "Frag. Score",
                            "Avg Employees", "Top States"]
            frag_rows = []
            for f in frag:
                score = f.get("score")
                color = GREEN if score and score >= 70 else ORANGE if score and score >= 40 else GRAY
                score_badge = f'<span style="padding:2px 8px;border-radius:4px;background:{color}20;color:{color};font-weight:600">{score:.0f}</span>' if score else "\u2014"
                frag_rows.append([
                    f.get("title") or "\u2014",
                    str(f.get("naics") or "\u2014"),
                    f'{f["establishments"]:,}' if f.get("establishments") else "\u2014",
                    score_badge,
                    f'{f["avg_employees"]:.0f}' if f.get("avg_employees") else "\u2014",
                    f.get("top_states") or "\u2014",
                ])
            body += data_table(frag_headers, frag_rows, numeric_columns={2, 3, 4})
        else:
            body += callout(
                "<strong>No fragmentation data available.</strong> Run Census CBP ingestion to populate market fragmentation scores.",
                "warn",
            )

        body += "\n" + section_end()

        # ── Section 4: Top Rollup Targets ──────────────────────────
        body += "\n" + section_start(4, "Top Rollup Targets", "rollup-targets")

        if rollups:
            body += '<p>Industries with multiple portfolio companies and active deal flow suggest rollup potential.</p>'
            rollup_headers = ["Industry", "Portfolio Companies", "Closed Deals", "Avg Multiple"]
            rollup_rows = [[r["industry"], str(r["companies"]), str(r["deals"]),
                            _fmt_x(r.get("avg_multiple"))] for r in rollups]
            body += data_table(rollup_headers, rollup_rows, numeric_columns={1, 2, 3})
        else:
            body += callout("<strong>No rollup target data available.</strong>", "warn")

        body += "\n" + section_end()

        # ── Section 5: Market Timing Assessment ────────────────────
        body += "\n" + section_start(5, "Market Timing Assessment", "timing")

        if timing.get("headline"):
            condition = timing.get("market_condition", "unknown")
            cond_color = GREEN if condition == "bullish" else RED if condition == "bearish" else ORANGE
            body += f'<div style="text-align:center;margin:16px 0">'
            body += f'<span style="font-size:1.5rem;font-weight:700;color:{cond_color}">{timing["headline"]}</span>'
            body += '</div>'

        findings = timing.get("findings", [])
        if findings:
            body += '<h3 style="margin:16px 0 8px;font-size:1rem">Key Findings</h3>'
            body += '<ul style="font-size:1rem;line-height:1.8">'
            for f in findings:
                body += f'<li>{f}</li>'
            body += '</ul>'

        recs = timing.get("recommendations", [])
        if recs:
            body += callout(
                "<strong>Recommendations:</strong><ul>" +
                "".join(f"<li>{r}</li>" for r in recs[:5]) +
                "</ul>", "info",
            )

        if not timing.get("headline") and not findings:
            body += callout("<strong>Timing assessment unavailable.</strong> Seed deal data to populate.", "warn")

        body += "\n" + section_end()

        # ── Footer ─────────────────────────────────────────────────
        body += "\n" + page_footer(
            notes=[
                "Deal data sourced from public deal databases, SEC filings, and press releases.",
                "Momentum scores derived from deal volume trends, multiple expansion, and sentiment.",
                "Fragmentation scores from Census Bureau County Business Patterns data.",
                "This brief is for informational purposes and does not constitute investment advice.",
            ],
            generated_line=f"Report generated {data.get('generated_at', 'N/A')} | Nexdata PE Intelligence",
        )

        extra_css = """
        .chart-row { display: grid; grid-template-columns: 1fr 1fr; gap: 24px; margin: 16px 0; }
        @media (max-width: 768px) { .chart-row { grid-template-columns: 1fr; } }
        """

        return html_document(
            title=title,
            body_content=body,
            charts_js=charts_js,
            extra_css=extra_css,
        )

    # ── Excel Rendering ────────────────────────────────────────────

    def render_excel(self, data: Dict[str, Any]) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        hdr_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
        hdr_font = Font(bold=True, color="FFFFFF")

        # ── Sheet 1: Sector Signals ────────────────────────────────
        ws = wb.active
        ws.title = "Sector Signals"
        headers = ["Industry", "Momentum", "Score", "Deals", "Median EV/EBITDA"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
        for idx, s in enumerate(data.get("sector_signals", []), 2):
            ws.cell(row=idx, column=1, value=s.get("industry"))
            ws.cell(row=idx, column=2, value=s.get("momentum"))
            ws.cell(row=idx, column=3, value=s.get("momentum_score"))
            ws.cell(row=idx, column=4, value=s.get("deal_count"))
            ws.cell(row=idx, column=5, value=s.get("median_ev_ebitda"))
        for letter in "ABCDE":
            ws.column_dimensions[letter].width = 20

        # ── Sheet 2: Deal Activity ─────────────────────────────────
        ws_d = wb.create_sheet("Deal Activity")
        d_headers = ["Quarter", "Deal Count", "Avg EV/EBITDA", "Total Value ($M)"]
        for col, h in enumerate(d_headers, 1):
            cell = ws_d.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
        for idx, d in enumerate(data.get("deal_activity", []), 2):
            ws_d.cell(row=idx, column=1, value=d.get("quarter"))
            ws_d.cell(row=idx, column=2, value=d.get("deal_count"))
            ws_d.cell(row=idx, column=3, value=d.get("avg_ev_ebitda"))
            ws_d.cell(row=idx, column=4, value=d.get("total_value"))
        for letter in "ABCD":
            ws_d.column_dimensions[letter].width = 18

        # ── Sheet 3: Top Buyers ────────────────────────────────────
        ws_b = wb.create_sheet("Top Buyers")
        b_headers = ["Buyer", "Deals", "Total Value ($M)"]
        for col, h in enumerate(b_headers, 1):
            cell = ws_b.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
        for idx, b in enumerate(data.get("top_buyers", []), 2):
            ws_b.cell(row=idx, column=1, value=b.get("buyer"))
            ws_b.cell(row=idx, column=2, value=b.get("deals"))
            ws_b.cell(row=idx, column=3, value=b.get("total_value"))
        ws_b.column_dimensions["A"].width = 30
        ws_b.column_dimensions["B"].width = 12
        ws_b.column_dimensions["C"].width = 18

        # ── Sheet 4: Rollup Targets ───────────────────────────────
        ws_r = wb.create_sheet("Rollup Targets")
        r_headers = ["Industry", "Companies", "Deals", "Avg Multiple"]
        for col, h in enumerate(r_headers, 1):
            cell = ws_r.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
        for idx, r in enumerate(data.get("rollup_targets", []), 2):
            ws_r.cell(row=idx, column=1, value=r.get("industry"))
            ws_r.cell(row=idx, column=2, value=r.get("companies"))
            ws_r.cell(row=idx, column=3, value=r.get("deals"))
            ws_r.cell(row=idx, column=4, value=r.get("avg_multiple"))
        for letter in "ABCD":
            ws_r.column_dimensions[letter].width = 18

        output = BytesIO()
        wb.save(output)
        return output.getvalue()
