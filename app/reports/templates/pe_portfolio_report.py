"""
PE Portfolio Report Template.

Generates a self-contained HTML report with:
- Executive Summary (firm overview + KPIs)
- Portfolio Heatmap (color-coded benchmarking grid)
- Per-Company Sections (financials, exit readiness, recommendations)

Uses the shared design_system.py patterns.
"""

import json
import logging
from datetime import datetime
from typing import Dict, Any, List, Optional
from io import BytesIO

from sqlalchemy import text
from sqlalchemy.orm import Session

from app.reports.design_system import (
    html_document, page_header, kpi_strip, kpi_card,
    toc, section_start, section_end,
    data_table, pill_badge, callout,
    chart_container, chart_init_js, page_footer,
    build_doughnut_config, build_horizontal_bar_config,
    build_bar_fallback, build_chart_legend, CHART_COLORS,
    BLUE, GREEN, ORANGE, RED, GRAY, TEAL, PURPLE,
)

logger = logging.getLogger(__name__)

# Percentile color thresholds
_PCT_COLORS = {
    "top": "#22c55e",      # green — >=75th
    "above": "#86efac",    # light green — 50-74
    "mid": "#fbbf24",      # amber — 25-49
    "bottom": "#ef4444",   # red — <25
}


def _pct_color(p: Optional[int]) -> str:
    if p is None:
        return GRAY
    if p >= 75:
        return _PCT_COLORS["top"]
    if p >= 50:
        return _PCT_COLORS["above"]
    if p >= 25:
        return _PCT_COLORS["mid"]
    return _PCT_COLORS["bottom"]


def _pct_cell(p: Optional[int]) -> str:
    """Colored table cell for a percentile value."""
    if p is None:
        return '<td style="text-align:right;color:#94a3b8">—</td>'
    color = _pct_color(p)
    return (
        f'<td style="text-align:right">'
        f'<span style="display:inline-block;padding:2px 10px;border-radius:4px;'
        f'background:{color}20;color:{color};font-weight:600">{p}th</span></td>'
    )


def _grade_pill(grade: str) -> str:
    """Colored pill for a letter grade."""
    colors = {"A": GREEN, "B": BLUE, "C": ORANGE, "D": RED, "F": RED}
    c = colors.get(grade[0].upper(), GRAY) if grade else GRAY
    return (
        f'<span style="display:inline-block;padding:2px 12px;border-radius:4px;'
        f'background:{c}20;color:{c};font-weight:700;font-size:0.9rem">{grade}</span>'
    )


def _fmt_usd(val: Optional[float]) -> str:
    if val is None:
        return "—"
    if abs(val) >= 1_000:
        return f"${val / 1_000:,.1f}B"
    return f"${val:,.0f}M"


def _fmt_pct(val: Optional[float]) -> str:
    if val is None:
        return "—"
    return f"{val:.1f}%"


class PEPortfolioReportTemplate:
    """PE Portfolio report template."""

    name = "pe_portfolio"
    description = "PE portfolio report with benchmarking heatmap and exit readiness"

    # ── Data Gathering ─────────────────────────────────────────────

    def gather_data(self, db: Session, params: Dict[str, Any]) -> Dict[str, Any]:
        """Gather all data for the report."""
        firm_id = params.get("firm_id")
        if not firm_id:
            raise ValueError("firm_id is required")

        return {
            "generated_at": datetime.utcnow().isoformat(),
            "firm": self._get_firm(db, firm_id),
            "funds": self._get_funds(db, firm_id),
            "companies": self._get_companies(db, firm_id),
            "heatmap": self._get_heatmap(db, firm_id),
            "exit_scores": self._get_exit_scores(db, firm_id),
            "industry_breakdown": self._get_industry_breakdown(db, firm_id),
        }

    def _get_firm(self, db: Session, firm_id: int) -> Dict:
        row = db.execute(text("""
            SELECT id, name, firm_type, primary_strategy,
                   headquarters_city, headquarters_state, aum_usd_millions,
                   total_funds_raised_usd_millions, active_portfolio_count
            FROM pe_firms WHERE id = :id
        """), {"id": firm_id}).fetchone()
        if not row:
            return {"id": firm_id, "name": "Unknown Firm"}
        return {
            "id": row[0], "name": row[1], "type": row[2],
            "strategy": row[3], "city": row[4], "state": row[5],
            "aum": float(row[6]) if row[6] else None,
            "total_raised": float(row[7]) if row[7] else None,
            "active_count": row[8] or 0,
        }

    def _get_funds(self, db: Session, firm_id: int) -> List[Dict]:
        rows = db.execute(text("""
            SELECT name, vintage_year, final_close_usd_millions, strategy, status
            FROM pe_funds WHERE firm_id = :fid ORDER BY vintage_year DESC
        """), {"fid": firm_id}).fetchall()
        return [
            {"name": r[0], "vintage": r[1],
             "size": float(r[2]) if r[2] else None,
             "strategy": r[3], "status": r[4]}
            for r in rows
        ]

    def _get_companies(self, db: Session, firm_id: int) -> List[Dict]:
        rows = db.execute(text("""
            SELECT pc.id, pc.name, pc.industry, pc.status,
                   pc.headquarters_city, pc.headquarters_state,
                   fi.investment_date, fi.exit_date
            FROM pe_portfolio_companies pc
            LEFT JOIN pe_fund_investments fi ON fi.company_id = pc.id
            LEFT JOIN pe_funds f ON fi.fund_id = f.id
            WHERE f.firm_id = :fid OR pc.id IN (
                SELECT fi2.company_id FROM pe_fund_investments fi2
                JOIN pe_funds f2 ON fi2.fund_id = f2.id
                WHERE f2.firm_id = :fid
            )
            ORDER BY pc.name
        """), {"fid": firm_id}).fetchall()

        seen = set()
        companies = []
        for r in rows:
            if r[0] in seen:
                continue
            seen.add(r[0])
            companies.append({
                "id": r[0], "name": r[1], "industry": r[2] or "Unknown",
                "status": r[3] or "active",
                "city": r[4], "state": r[5],
                "invested": r[6].isoformat() if r[6] else None,
                "exited": r[7].isoformat() if r[7] else None,
            })
        return companies

    def _get_heatmap(self, db: Session, firm_id: int) -> List[Dict]:
        """Get latest-year financials + percentiles for each portfolio company."""
        rows = db.execute(text("""
            SELECT pc.id, pc.name, pc.industry, pc.status,
                   f.fiscal_year,
                   f.revenue_usd, f.revenue_growth_pct,
                   f.ebitda_margin_pct, f.gross_margin_pct,
                   f.net_debt_to_ebitda
            FROM pe_portfolio_companies pc
            JOIN pe_fund_investments fi ON fi.company_id = pc.id
            JOIN pe_funds fu ON fi.fund_id = fu.id
            LEFT JOIN LATERAL (
                SELECT * FROM pe_company_financials cf
                WHERE cf.company_id = pc.id
                ORDER BY cf.fiscal_year DESC LIMIT 1
            ) f ON TRUE
            WHERE fu.firm_id = :fid
            ORDER BY pc.name
        """), {"fid": firm_id}).fetchall()

        seen = set()
        results = []
        for r in rows:
            if r[0] in seen:
                continue
            seen.add(r[0])
            results.append({
                "id": r[0], "name": r[1], "industry": r[2],
                "status": r[3] or "active", "year": r[4],
                "revenue": float(r[5]) if r[5] else None,
                "revenue_growth": float(r[6]) if r[6] else None,
                "ebitda_margin": float(r[7]) if r[7] else None,
                "gross_margin": float(r[8]) if r[8] else None,
                "leverage": float(r[9]) if r[9] else None,
            })

        # Compute simple percentiles within the portfolio
        metrics = ["revenue_growth", "ebitda_margin", "gross_margin"]
        for m in metrics:
            vals = sorted([c[m] for c in results if c[m] is not None])
            n = len(vals)
            for c in results:
                if c[m] is not None and n > 1:
                    rank = vals.index(c[m])
                    c[f"{m}_pct"] = int(rank / (n - 1) * 100)
                else:
                    c[f"{m}_pct"] = None
        # Leverage: lower is better → invert
        lev_vals = sorted([c["leverage"] for c in results if c["leverage"] is not None])
        n_lev = len(lev_vals)
        for c in results:
            if c["leverage"] is not None and n_lev > 1:
                rank = lev_vals.index(c["leverage"])
                c["leverage_pct"] = int((1 - rank / (n_lev - 1)) * 100)
            else:
                c["leverage_pct"] = None

        return results

    def _get_exit_scores(self, db: Session, firm_id: int) -> Dict[int, Dict]:
        """Get pre-computed exit readiness scores from the scoring engine."""
        from app.core.pe_exit_scoring import ExitReadinessScorer
        scorer = ExitReadinessScorer(db)

        company_ids = [r[0] for r in db.execute(text("""
            SELECT DISTINCT fi.company_id
            FROM pe_fund_investments fi
            JOIN pe_funds f ON fi.fund_id = f.id
            WHERE f.firm_id = :fid
        """), {"fid": firm_id}).fetchall()]

        scores = {}
        for cid in company_ids:
            try:
                result = scorer.score(cid)
                scores[cid] = {
                    "composite": result.get("composite_score", 0),
                    "grade": result.get("grade", "N/A"),
                    "sub_scores": result.get("sub_scores", []),
                    "recommendations": result.get("recommendations", []),
                    "data_gaps": result.get("data_gaps", []),
                }
            except Exception:
                logger.debug("Exit score unavailable for company %s", cid)
        return scores

    def _get_industry_breakdown(self, db: Session, firm_id: int) -> List[Dict]:
        rows = db.execute(text("""
            SELECT COALESCE(pc.industry, 'Unknown') as ind, COUNT(*) as cnt
            FROM pe_portfolio_companies pc
            JOIN pe_fund_investments fi ON fi.company_id = pc.id
            JOIN pe_funds f ON fi.fund_id = f.id
            WHERE f.firm_id = :fid
            GROUP BY pc.industry
            ORDER BY cnt DESC
        """), {"fid": firm_id}).fetchall()
        total = sum(r[1] for r in rows) or 1
        return [{"industry": r[0], "count": r[1],
                 "pct": round(r[1] / total * 100, 1)} for r in rows]

    # ── HTML Rendering ─────────────────────────────────────────────

    def render_html(self, data: Dict[str, Any]) -> str:
        firm = data.get("firm", {})
        funds = data.get("funds", [])
        companies = data.get("companies", [])
        heatmap = data.get("heatmap", [])
        exit_scores = data.get("exit_scores", {})
        industries = data.get("industry_breakdown", [])
        title = data.get("report_title", f"{firm.get('name', 'PE')} Portfolio Report")

        active = [c for c in companies if c.get("status", "").lower() in ("active", "")]
        exited = [c for c in companies if c.get("status", "").lower() not in ("active", "")]
        charts_js = ""
        body = ""

        # ── Header ─────────────────────────────────────────────────
        parts = [p for p in [firm.get("type"), firm.get("strategy")] if p]
        subtitle = " · ".join(parts) if parts else None
        body += page_header(title=title, subtitle=subtitle,
                            badge=f"Generated {datetime.utcnow().strftime('%b %d, %Y')}")

        # ── KPIs ───────────────────────────────────────────────────
        cards = ""
        cards += kpi_card("AUM", _fmt_usd(firm.get("aum")))
        cards += kpi_card("Total Raised", _fmt_usd(firm.get("total_raised")))
        cards += kpi_card("Active Companies", str(len(active)))
        cards += kpi_card("Exited Companies", str(len(exited)))
        cards += kpi_card("Funds", str(len(funds)))
        body += "\n" + kpi_strip(cards)

        # ── TOC ────────────────────────────────────────────────────
        toc_items = [
            {"number": 1, "id": "heatmap", "title": "Portfolio Heatmap"},
            {"number": 2, "id": "exit", "title": "Exit Readiness Summary"},
            {"number": 3, "id": "funds", "title": "Fund Overview"},
            {"number": 4, "id": "industries", "title": "Industry Breakdown"},
        ]
        sec_num = 4
        for c in companies[:20]:
            sec_num += 1
            toc_items.append({
                "number": sec_num,
                "id": f"co-{c['id']}",
                "title": c["name"],
            })
        body += "\n" + toc(toc_items)

        # ── Section 1: Heatmap ─────────────────────────────────────
        body += "\n" + section_start(1, "Portfolio Heatmap", "heatmap")
        body += '<p>Color-coded percentile rankings across key financial metrics. '
        body += 'Green indicates top quartile; red indicates bottom quartile.</p>'

        heatmap_headers = [
            "Company", "Industry", "Status", "Revenue",
            "Rev. Growth", "EBITDA Margin", "Gross Margin", "Leverage",
        ]
        heatmap_rows = []
        for h in heatmap:
            status = pill_badge(h["status"].title(), "public" if h["status"] == "active" else "default")
            heatmap_rows.append([
                f'<a href="#co-{h["id"]}" style="color:var(--primary-light);font-weight:600">{h["name"]}</a>',
                h.get("industry") or "—",
                status,
                _fmt_usd(h.get("revenue")),
                _pct_cell_inline(h.get("revenue_growth"), h.get("revenue_growth_pct")),
                _pct_cell_inline(h.get("ebitda_margin"), h.get("ebitda_margin_pct")),
                _pct_cell_inline(h.get("gross_margin"), h.get("gross_margin_pct")),
                _pct_cell_inline(h.get("leverage"), h.get("leverage_pct"), fmt="x", lower_better=True),
            ])
        body += data_table(heatmap_headers, heatmap_rows, numeric_columns={3, 4, 5, 6, 7})
        body += "\n" + section_end()

        # ── Section 2: Exit Readiness Summary ──────────────────────
        body += "\n" + section_start(2, "Exit Readiness Summary", "exit")
        if exit_scores:
            body += '<p>Composite exit readiness scores for portfolio companies.</p>'
            exit_headers = ["Company", "Score", "Grade", "Top Recommendations"]
            exit_rows = []
            for c in companies:
                es = exit_scores.get(c["id"])
                if not es:
                    continue
                recs = "; ".join(es.get("recommendations", [])[:2]) or "—"
                exit_rows.append([
                    f'<a href="#co-{c["id"]}" style="color:var(--primary-light);font-weight:600">{c["name"]}</a>',
                    f'{es["composite"]:.0f}/100',
                    _grade_pill(es["grade"]),
                    recs,
                ])
            body += data_table(exit_headers, exit_rows, numeric_columns={1})
        else:
            body += callout(
                "<strong>No exit scores available.</strong> Run the exit readiness scorer to populate this section.",
                "warn",
            )
        body += "\n" + section_end()

        # ── Section 3: Fund Overview ───────────────────────────────
        body += "\n" + section_start(3, "Fund Overview", "funds")
        if funds:
            fund_headers = ["Fund", "Vintage", "Size", "Strategy", "Status"]
            fund_rows = []
            for f in funds:
                status_variant = "public" if f.get("status") == "active" else "default"
                fund_rows.append([
                    f["name"],
                    str(f.get("vintage") or "—"),
                    _fmt_usd(f.get("size")),
                    f.get("strategy") or "—",
                    pill_badge((f.get("status") or "unknown").title(), status_variant),
                ])
            body += data_table(fund_headers, fund_rows, numeric_columns={1, 2})
        else:
            body += callout("<strong>No fund data available.</strong>", "warn")
        body += "\n" + section_end()

        # ── Section 4: Industry Breakdown ──────────────────────────
        body += "\n" + section_start(4, "Industry Breakdown", "industries")
        if industries:
            ind_labels = [i["industry"] for i in industries]
            ind_values = [float(i["count"]) for i in industries]
            ind_colors = [CHART_COLORS[idx % len(CHART_COLORS)] for idx in range(len(ind_labels))]
            donut_cfg = build_doughnut_config(ind_labels, ind_values, ind_colors)
            donut_json = json.dumps(donut_cfg)

            body += '<div class="chart-row"><div>'
            body += chart_container(
                "industryChart", donut_json,
                build_bar_fallback(ind_labels, ind_values),
                size="medium", title="Portfolio by Industry",
            )
            charts_js += chart_init_js("industryChart", donut_json)
            body += '</div><div>'
            body += build_chart_legend(ind_labels, ind_values, ind_colors, show_pct=True)

            ind_table_rows = [[i["industry"], str(i["count"]), f'{i["pct"]}%'] for i in industries]
            body += data_table(["Industry", "Companies", "%"], ind_table_rows, numeric_columns={1, 2})
            body += '</div></div>'
        else:
            body += callout("<strong>No industry data available.</strong>", "warn")
        body += "\n" + section_end()

        # ── Per-Company Sections ───────────────────────────────────
        sec_num = 4
        heatmap_by_id = {h["id"]: h for h in heatmap}
        for c in companies[:20]:
            sec_num += 1
            body += "\n" + section_start(sec_num, c["name"], f"co-{c['id']}")

            # Company info line
            info_parts = [p for p in [c.get("industry"), c.get("city"), c.get("state")] if p]
            if info_parts:
                body += f'<p style="color:var(--gray-500);margin-bottom:12px">{" · ".join(info_parts)}</p>'

            status_variant = "public" if c.get("status", "").lower() == "active" else "default"
            body += f'<p>Status: {pill_badge(c.get("status", "unknown").title(), status_variant)}'
            if c.get("invested"):
                body += f' &nbsp;|&nbsp; Invested: {c["invested"]}'
            if c.get("exited"):
                body += f' &nbsp;|&nbsp; Exited: {c["exited"]}'
            body += '</p>'

            # Financials from heatmap
            hm = heatmap_by_id.get(c["id"])
            if hm:
                body += '<h3 style="margin:16px 0 8px;font-size:1rem">Key Financials</h3>'
                fin_rows = []
                fin_rows.append(["Revenue", _fmt_usd(hm.get("revenue")), "—"])
                fin_rows.append([
                    "Revenue Growth",
                    _fmt_pct(hm.get("revenue_growth")),
                    _pct_label(hm.get("revenue_growth_pct")),
                ])
                fin_rows.append([
                    "EBITDA Margin",
                    _fmt_pct(hm.get("ebitda_margin")),
                    _pct_label(hm.get("ebitda_margin_pct")),
                ])
                fin_rows.append([
                    "Gross Margin",
                    _fmt_pct(hm.get("gross_margin")),
                    _pct_label(hm.get("gross_margin_pct")),
                ])
                lev_str = f"{hm['leverage']:.1f}x" if hm.get("leverage") is not None else "—"
                fin_rows.append([
                    "Net Debt / EBITDA",
                    lev_str,
                    _pct_label(hm.get("leverage_pct")),
                ])
                body += data_table(["Metric", "Value", "Percentile"], fin_rows, numeric_columns={1, 2})

            # Exit readiness
            es = exit_scores.get(c["id"])
            if es:
                body += '<h3 style="margin:16px 0 8px;font-size:1rem">Exit Readiness</h3>'
                body += f'<p>Composite Score: <strong>{es["composite"]:.0f}/100</strong> '
                body += f'{_grade_pill(es["grade"])}</p>'

                if es.get("sub_scores"):
                    sub_headers = ["Dimension", "Score", "Grade", "Explanation"]
                    sub_rows = []
                    for ss in es["sub_scores"]:
                        sub_rows.append([
                            ss.get("label", ss.get("dimension", "")),
                            f'{ss.get("raw_score", 0):.0f}',
                            _grade_pill(ss.get("grade", "N/A")),
                            ss.get("explanation", ""),
                        ])
                    body += data_table(sub_headers, sub_rows, numeric_columns={1})

                if es.get("recommendations"):
                    body += '<div style="margin-top:8px">'
                    body += callout(
                        "<strong>Recommendations:</strong><ul>" +
                        "".join(f"<li>{r}</li>" for r in es["recommendations"][:5]) +
                        "</ul>",
                        "info",
                    )
                    body += '</div>'

                if es.get("data_gaps"):
                    body += callout(
                        "<strong>Data Gaps:</strong> " + ", ".join(es["data_gaps"]),
                        "warn",
                    )

            body += "\n" + section_end()

        # ── Footer ─────────────────────────────────────────────────
        body += "\n" + page_footer(
            notes=[
                "Financial data sourced from SEC filings, company disclosures, and public records.",
                "Percentiles are computed within the portfolio peer set, not industry-wide.",
                "Exit readiness scores are model-generated estimates, not investment advice.",
            ],
            generated_line=f"Report generated {data.get('generated_at', 'N/A')} | Nexdata PE Intelligence",
        )

        extra_css = """
        .chart-row { display: grid; grid-template-columns: 1fr 1fr; gap: 24px; margin: 16px 0; }
        @media (max-width: 768px) { .chart-row { grid-template-columns: 1fr; } }
        """

        return html_document(
            title=title,
            body_content=body,
            charts_js=charts_js,
            extra_css=extra_css,
        )

    def render_excel(self, data: Dict[str, Any]) -> bytes:
        """Render report as Excel workbook."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        hdr_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
        hdr_font = Font(bold=True, color="FFFFFF")

        # Portfolio sheet
        ws = wb.active
        ws.title = "Portfolio"
        headers = ["Company", "Industry", "Status", "Revenue ($M)",
                    "Rev Growth %", "EBITDA Margin %", "Gross Margin %",
                    "Leverage (x)", "Exit Score", "Exit Grade"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill

        heatmap = data.get("heatmap", [])
        exit_scores = data.get("exit_scores", {})
        for row_idx, hm in enumerate(heatmap, 2):
            ws.cell(row=row_idx, column=1, value=hm.get("name"))
            ws.cell(row=row_idx, column=2, value=hm.get("industry"))
            ws.cell(row=row_idx, column=3, value=hm.get("status"))
            ws.cell(row=row_idx, column=4, value=hm.get("revenue"))
            ws.cell(row=row_idx, column=5, value=hm.get("revenue_growth"))
            ws.cell(row=row_idx, column=6, value=hm.get("ebitda_margin"))
            ws.cell(row=row_idx, column=7, value=hm.get("gross_margin"))
            ws.cell(row=row_idx, column=8, value=hm.get("leverage"))
            es = exit_scores.get(hm["id"], {})
            ws.cell(row=row_idx, column=9, value=es.get("composite"))
            ws.cell(row=row_idx, column=10, value=es.get("grade"))

        for col_letter in "ABCDEFGHIJ":
            ws.column_dimensions[col_letter].width = 18

        # Funds sheet
        ws_funds = wb.create_sheet("Funds")
        fund_headers = ["Fund", "Vintage", "Size ($M)", "Strategy", "Status"]
        for col, h in enumerate(fund_headers, 1):
            cell = ws_funds.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
        for row_idx, f in enumerate(data.get("funds", []), 2):
            ws_funds.cell(row=row_idx, column=1, value=f.get("name"))
            ws_funds.cell(row=row_idx, column=2, value=f.get("vintage"))
            ws_funds.cell(row=row_idx, column=3, value=f.get("size"))
            ws_funds.cell(row=row_idx, column=4, value=f.get("strategy"))
            ws_funds.cell(row=row_idx, column=5, value=f.get("status"))
        ws_funds.column_dimensions["A"].width = 30

        output = BytesIO()
        wb.save(output)
        return output.getvalue()


# ── Module-level helpers (used in render_html) ─────────────────────

def _pct_cell_inline(
    value: Optional[float],
    percentile: Optional[int],
    fmt: str = "pct",
    lower_better: bool = False,
) -> str:
    """Format a value + percentile for inline table cell content."""
    if value is None:
        return "—"
    if fmt == "pct":
        val_str = f"{value:.1f}%"
    elif fmt == "x":
        val_str = f"{value:.1f}x"
    else:
        val_str = f"{value:,.0f}"

    if percentile is None:
        return val_str

    color = _pct_color(percentile)
    return (
        f'{val_str} <span style="display:inline-block;padding:1px 6px;border-radius:3px;'
        f'font-size:0.75rem;background:{color}20;color:{color}">{percentile}th</span>'
    )


def _pct_label(percentile: Optional[int]) -> str:
    """Colored percentile label for per-company tables."""
    if percentile is None:
        return "—"
    color = _pct_color(percentile)
    return (
        f'<span style="padding:2px 8px;border-radius:4px;'
        f'background:{color}20;color:{color};font-weight:600">{percentile}th</span>'
    )
