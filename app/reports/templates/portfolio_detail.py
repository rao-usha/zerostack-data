"""
Portfolio Detail Report Template.

Generates a detailed portfolio breakdown with all holdings,
sector distribution, stage breakdown, and recent changes.
"""

import json
import logging
from datetime import datetime
from typing import Dict, Any, List
from io import BytesIO

from sqlalchemy import text
from sqlalchemy.orm import Session

from app.reports.design_system import (
    html_document, hero_header, kpi_card, kpi_grid,
    section_heading, data_table, pill_badge,
    chart_container, chart_init_js, footer,
    build_doughnut_config, build_horizontal_bar_config,
    build_bar_fallback, CHART_COLORS,
)

logger = logging.getLogger(__name__)


class PortfolioDetailTemplate:
    """Portfolio detail report template."""

    name = "portfolio_detail"
    description = "Detailed portfolio breakdown with holdings and analytics"

    def gather_data(self, db: Session, params: Dict[str, Any]) -> Dict[str, Any]:
        """Gather all data needed for the report."""
        investor_id = params.get("investor_id")
        investor_type = params.get("investor_type", "lp")

        data = {
            "generated_at": datetime.utcnow().isoformat(),
            "investor": self._get_investor(db, investor_id, investor_type),
            "holdings": self._get_all_holdings(db, investor_id, investor_type),
            "sector_breakdown": self._get_sector_breakdown(
                db, investor_id, investor_type
            ),
            "stage_breakdown": self._get_stage_breakdown(
                db, investor_id, investor_type
            ),
            "location_breakdown": self._get_location_breakdown(
                db, investor_id, investor_type
            ),
        }

        return data

    def _get_investor(self, db: Session, investor_id: int, investor_type: str) -> Dict:
        """Get investor details."""
        if investor_type == "lp":
            result = db.execute(
                text("""
                SELECT id, name, lp_type, jurisdiction
                FROM lp_fund WHERE id = :id
            """),
                {"id": investor_id},
            )
            row = result.fetchone()
            if row:
                return {
                    "id": row[0],
                    "name": row[1],
                    "type": row[2],
                    "jurisdiction": row[3],
                }
        else:
            result = db.execute(
                text("""
                SELECT id, name, 'family_office' as lp_type, location as jurisdiction
                FROM family_offices WHERE id = :id
            """),
                {"id": investor_id},
            )
            row = result.fetchone()
            if row:
                return {
                    "id": row[0],
                    "name": row[1],
                    "type": row[2],
                    "jurisdiction": row[3],
                }

        return {"id": investor_id, "name": "Unknown", "type": investor_type}

    def _get_all_holdings(
        self, db: Session, investor_id: int, investor_type: str
    ) -> List[Dict]:
        """Get all portfolio holdings."""
        result = db.execute(
            text("""
            SELECT
                company_name,
                company_industry,
                company_location,
                company_stage,
                current_holding,
                collected_date
            FROM portfolio_companies
            WHERE investor_id = :investor_id AND investor_type = :investor_type
            ORDER BY current_holding DESC, company_name
        """),
            {"investor_id": investor_id, "investor_type": investor_type},
        )

        return [
            {
                "name": row[0],
                "industry": row[1] or "Unknown",
                "location": row[2] or "Unknown",
                "stage": row[3] or "Unknown",
                "current": row[4] == 1,
                "collected_date": row[5].isoformat() if row[5] else None,
            }
            for row in result.fetchall()
        ]

    def _get_sector_breakdown(
        self, db: Session, investor_id: int, investor_type: str
    ) -> List[Dict]:
        """Get sector breakdown."""
        result = db.execute(
            text("""
            SELECT
                COALESCE(company_industry, 'Unknown') as sector,
                COUNT(*) as total,
                COUNT(CASE WHEN current_holding = 1 THEN 1 END) as current
            FROM portfolio_companies
            WHERE investor_id = :investor_id AND investor_type = :investor_type
            GROUP BY company_industry
            ORDER BY total DESC
        """),
            {"investor_id": investor_id, "investor_type": investor_type},
        )

        rows = result.fetchall()
        total = sum(r[1] for r in rows)

        return [
            {
                "sector": row[0],
                "total": row[1],
                "current": row[2],
                "pct": round(row[1] / total * 100, 1) if total > 0 else 0,
            }
            for row in rows
        ]

    def _get_stage_breakdown(
        self, db: Session, investor_id: int, investor_type: str
    ) -> List[Dict]:
        """Get stage breakdown."""
        result = db.execute(
            text("""
            SELECT
                COALESCE(company_stage, 'Unknown') as stage,
                COUNT(*) as total,
                COUNT(CASE WHEN current_holding = 1 THEN 1 END) as current
            FROM portfolio_companies
            WHERE investor_id = :investor_id AND investor_type = :investor_type
            GROUP BY company_stage
            ORDER BY total DESC
        """),
            {"investor_id": investor_id, "investor_type": investor_type},
        )

        rows = result.fetchall()
        total = sum(r[1] for r in rows)

        return [
            {
                "stage": row[0],
                "total": row[1],
                "current": row[2],
                "pct": round(row[1] / total * 100, 1) if total > 0 else 0,
            }
            for row in rows
        ]

    def _get_location_breakdown(
        self, db: Session, investor_id: int, investor_type: str
    ) -> List[Dict]:
        """Get location breakdown."""
        result = db.execute(
            text("""
            SELECT
                COALESCE(company_location, 'Unknown') as location,
                COUNT(*) as total
            FROM portfolio_companies
            WHERE investor_id = :investor_id AND investor_type = :investor_type
            GROUP BY company_location
            ORDER BY total DESC
            LIMIT 15
        """),
            {"investor_id": investor_id, "investor_type": investor_type},
        )

        rows = result.fetchall()
        total = sum(r[1] for r in rows)

        return [
            {
                "location": row[0],
                "total": row[1],
                "pct": round(row[1] / total * 100, 1) if total > 0 else 0,
            }
            for row in rows
        ]

    def render_html(self, data: Dict[str, Any]) -> str:
        """Render report as HTML using the shared design system."""
        investor = data.get("investor", {})
        holdings = data.get("holdings", [])
        sectors = data.get("sector_breakdown", [])
        stages = data.get("stage_breakdown", [])
        locations = data.get("location_breakdown", [])

        charts_js = ""
        body = ""
        current_count = len([h for h in holdings if h.get("current")])

        # ── Hero Header ──────────────────────────────────────────
        pills = []
        if investor.get("type"):
            pills.append({"label": "Type", "value": str(investor["type"])})
        if investor.get("jurisdiction"):
            pills.append({"label": "Location", "value": str(investor["jurisdiction"])})

        body += hero_header(
            title=f"{investor.get('name', 'Unknown')} — Portfolio Detail",
            pills=pills if pills else None,
        )

        # ── KPI Cards ────────────────────────────────────────────
        cards = ""
        cards += kpi_card(str(current_count), "Current Holdings", "blue")
        cards += kpi_card(str(len(holdings)), "Total Holdings", "emerald")
        cards += kpi_card(str(len(sectors)), "Sectors", "slate")

        body += '\n    <main class="container">'
        body += "\n" + kpi_grid(cards)

        # ── Holdings Table ───────────────────────────────────────
        body += "\n" + section_heading("All Holdings", count=len(holdings))

        table_rows = []
        for h in holdings:
            status = pill_badge("Current", "public") if h.get("current") else pill_badge("Exited", "default")
            table_rows.append([
                f'<span class="company-name">{h.get("name", "N/A")}</span>',
                h.get("industry", "N/A"),
                h.get("location", "N/A"),
                h.get("stage", "N/A"),
                status,
            ])

        body += "\n" + data_table(
            headers=["Company", "Industry", "Location", "Stage", "Status"],
            rows=table_rows,
        )

        # ── Sector Breakdown (doughnut chart + table) ────────────
        if sectors:
            body += "\n" + section_heading("Sector Breakdown", count=len(sectors))

            sector_labels = [s.get("sector", "N/A") for s in sectors]
            sector_values = [float(s.get("total", 0)) for s in sectors]
            doughnut_config = build_doughnut_config(sector_labels, sector_values)
            doughnut_json = json.dumps(doughnut_config)
            doughnut_fallback = build_bar_fallback(sector_labels, sector_values)

            body += '\n<div class="charts-row">'
            body += "\n" + chart_container("sectorChart", doughnut_json, doughnut_fallback)
            charts_js += chart_init_js("sectorChart", doughnut_json)

            # Sector table alongside chart
            sector_table_rows = []
            for s in sectors:
                sector_table_rows.append([
                    s.get("sector", "N/A"),
                    str(s.get("current", 0)),
                    str(s.get("total", 0)),
                    f'{s.get("pct", 0)}%',
                ])
            body += "\n" + data_table(
                headers=["Sector", "Current", "Total", "% of Portfolio"],
                rows=sector_table_rows,
                numeric_columns={1, 2, 3},
            )
            body += "\n</div>"

        # ── Stage Breakdown (horizontal bar chart) ───────────────
        if stages:
            body += "\n" + section_heading("Stage Breakdown", count=len(stages))

            stage_labels = [s.get("stage", "N/A") for s in stages]
            stage_values = [float(s.get("total", 0)) for s in stages]
            bar_config = build_horizontal_bar_config(stage_labels, stage_values, dataset_label="Companies")
            bar_json = json.dumps(bar_config)
            bar_fallback = build_bar_fallback(stage_labels, stage_values)

            body += '\n<div class="charts-row">'
            body += "\n" + chart_container("stageChart", bar_json, bar_fallback)
            charts_js += chart_init_js("stageChart", bar_json)

            # Stage table alongside chart
            stage_table_rows = []
            for s in stages:
                stage_table_rows.append([
                    s.get("stage", "N/A"),
                    str(s.get("current", 0)),
                    str(s.get("total", 0)),
                    f'{s.get("pct", 0)}%',
                ])
            body += "\n" + data_table(
                headers=["Stage", "Current", "Total", "% of Portfolio"],
                rows=stage_table_rows,
                numeric_columns={1, 2, 3},
            )
            body += "\n</div>"

        # ── Location Breakdown (horizontal bar chart) ────────────
        if locations:
            body += "\n" + section_heading("Location Breakdown", count=len(locations))

            loc_labels = [loc.get("location", "N/A") for loc in locations]
            loc_values = [float(loc.get("total", 0)) for loc in locations]
            loc_config = build_horizontal_bar_config(loc_labels, loc_values, dataset_label="Companies")
            loc_json = json.dumps(loc_config)
            loc_fallback = build_bar_fallback(loc_labels, loc_values)

            body += "\n" + chart_container("locationChart", loc_json, loc_fallback)
            charts_js += chart_init_js("locationChart", loc_json)

        body += "\n    </main>"

        # ── Footer ───────────────────────────────────────────────
        body += "\n" + footer(data.get("generated_at", "N/A"))

        return html_document(
            title=f"{investor.get('name', 'Portfolio')} - Detail Report",
            body_content=body,
            charts_js=charts_js,
        )

    def render_excel(self, data: Dict[str, Any]) -> bytes:
        """Render report as Excel workbook."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()

        header_fill = PatternFill(
            start_color="27AE60", end_color="27AE60", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        # Holdings sheet
        ws = wb.active
        ws.title = "Holdings"

        holdings = data.get("holdings", [])
        headers = ["Company", "Industry", "Location", "Stage", "Status"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for row, h in enumerate(holdings, 2):
            ws.cell(row=row, column=1, value=h.get("name"))
            ws.cell(row=row, column=2, value=h.get("industry"))
            ws.cell(row=row, column=3, value=h.get("location"))
            ws.cell(row=row, column=4, value=h.get("stage"))
            ws.cell(
                row=row, column=5, value="Current" if h.get("current") else "Exited"
            )

        for col in ["A", "B", "C", "D", "E"]:
            ws.column_dimensions[col].width = 20

        # Sectors sheet
        ws_sectors = wb.create_sheet("Sectors")
        sectors = data.get("sector_breakdown", [])

        headers = ["Sector", "Current", "Total", "Percentage"]
        for col, header in enumerate(headers, 1):
            cell = ws_sectors.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for row, s in enumerate(sectors, 2):
            ws_sectors.cell(row=row, column=1, value=s.get("sector"))
            ws_sectors.cell(row=row, column=2, value=s.get("current"))
            ws_sectors.cell(row=row, column=3, value=s.get("total"))
            ws_sectors.cell(row=row, column=4, value=f"{s.get('pct', 0)}%")

        ws_sectors.column_dimensions["A"].width = 25

        # Stages sheet
        ws_stages = wb.create_sheet("Stages")
        stages = data.get("stage_breakdown", [])

        for col, header in enumerate(headers, 1):
            cell = ws_stages.cell(
                row=1, column=col, value=header.replace("Sector", "Stage")
            )
            cell.font = header_font
            cell.fill = header_fill

        for row, s in enumerate(stages, 2):
            ws_stages.cell(row=row, column=1, value=s.get("stage"))
            ws_stages.cell(row=row, column=2, value=s.get("current"))
            ws_stages.cell(row=row, column=3, value=s.get("total"))
            ws_stages.cell(row=row, column=4, value=f"{s.get('pct', 0)}%")

        ws_stages.column_dimensions["A"].width = 20

        # Save
        output = BytesIO()
        wb.save(output)
        return output.getvalue()
