"""
Census of Governments Collector.

Fetches jurisdiction counts per county from the 2022 Census of Governments.
Fewer overlapping government layers = faster permitting = better for
datacenter site selection.

Data source: https://www.census.gov/programs-surveys/cog.html
No API key required — downloads ZIP file containing Excel data.
"""

import io
import logging
import zipfile
from datetime import datetime
from typing import Optional, Dict, Any

from sqlalchemy.orm import Session

from app.core.models_site_intel import GovernmentUnit
from app.sources.site_intel.base_collector import BaseCollector
from app.sources.site_intel.types import (
    SiteIntelDomain,
    SiteIntelSource,
    CollectionConfig,
    CollectionResult,
    CollectionStatus,
)
from app.sources.site_intel.runner import register_collector

logger = logging.getLogger(__name__)

# Census of Governments ZIP file URL (2022) — contains Excel with govt unit data
COG_2022_URL = (
    "https://www2.census.gov/programs-surveys/gus/datasets/2022/"
    "govt_units_2022.ZIP"
)

# Unit type prefix mapping from the UNIT_TYPE column (e.g. "1 - COUNTY")
UNIT_TYPE_PREFIX_MAP = {
    "1": "county",
    "2": "municipal",
    "3": "township",
    "4": "special_district",
    "5": "school_district",
}


@register_collector(SiteIntelSource.CENSUS_GOV)
class CensusGovCollector(BaseCollector):
    """
    Collector for Census of Governments data.

    Downloads the 2022 individual government units file and aggregates
    counts by county FIPS. Counties with fewer overlapping jurisdictions
    tend to have simpler permitting processes.
    """

    domain = SiteIntelDomain.LABOR
    source = SiteIntelSource.CENSUS_GOV

    default_timeout = 120.0
    rate_limit_delay = 0.5

    def __init__(self, db: Session, api_key: Optional[str] = None, **kwargs):
        super().__init__(db, api_key, **kwargs)

    def get_default_base_url(self) -> str:
        return "https://www2.census.gov/programs-surveys/gus/datasets"

    def get_default_headers(self) -> Dict[str, str]:
        return {
            "User-Agent": "Nexdata-SiteIntel/1.0",
            "Accept": "text/csv, */*",
        }

    async def collect(self, config: CollectionConfig) -> CollectionResult:
        """Collect Census of Governments data."""
        try:
            logger.info("Collecting Census of Governments 2022 data...")

            client = await self.get_client()
            await self.apply_rate_limit()

            response = await client.get(COG_2022_URL, timeout=180.0)
            response.raise_for_status()

            zip_bytes = response.content
            if not zip_bytes or len(zip_bytes) < 1000:
                return self.create_result(
                    status=CollectionStatus.FAILED,
                    error_message="Empty ZIP from Census of Governments",
                )

            # Extract Excel from ZIP and parse
            county_data = self._parse_zip(zip_bytes, config.states)

            records = list(county_data.values())
            logger.info(f"Census Gov: aggregated {len(records)} county records")

            if records:
                inserted, updated = self.bulk_upsert(
                    GovernmentUnit,
                    records,
                    unique_columns=["county_fips", "census_year"],
                    update_columns=[
                        "county_name", "state", "total_governments",
                        "county_govts", "municipal_govts", "township_govts",
                        "special_district_govts", "school_district_govts",
                        "population", "govts_per_10k_pop",
                        "source", "collected_at",
                    ],
                )
                return self.create_result(
                    status=CollectionStatus.SUCCESS,
                    total=len(records),
                    processed=len(records),
                    inserted=inserted + updated,
                )

            return self.create_result(
                status=CollectionStatus.SUCCESS,
                total=0,
                processed=0,
                inserted=0,
            )

        except Exception as e:
            logger.error(f"Census Gov collection failed: {e}", exc_info=True)
            return self.create_result(
                status=CollectionStatus.FAILED,
                error_message=str(e),
            )

    def _parse_zip(
        self, zip_bytes: bytes, states: Optional[list] = None
    ) -> Dict[str, Dict[str, Any]]:
        """Extract Excel from ZIP and aggregate government units by county."""
        import openpyxl

        county_data: Dict[str, Dict[str, Any]] = {}

        with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
            # Find the Excel file
            xlsx_name = None
            for name in zf.namelist():
                if name.endswith(".xlsx"):
                    xlsx_name = name
                    break
            if not xlsx_name:
                logger.warning("No xlsx file found in Census Gov ZIP")
                return county_data

            with zf.open(xlsx_name) as f:
                wb = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True)

            from app.sources.site_intel.labor.census_bps_collector import FIPS_TO_STATE

            # Process all sheets (General Purpose, Special District, School District)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows_iter = ws.iter_rows(values_only=True)

                # First row is header
                try:
                    header = next(rows_iter)
                except StopIteration:
                    continue

                header = [str(h or "").strip() for h in header]

                # Build column index
                col_idx = {name: i for i, name in enumerate(header)}
                fips_st = col_idx.get("FIPS_STATE")
                fips_co = col_idx.get("FIPS_COUNTY")
                unit_type_idx = col_idx.get("UNIT_TYPE")
                county_area_idx = col_idx.get("COUNTY_AREA_NAME")
                pop_idx = col_idx.get("POPULATION")
                active_idx = col_idx.get("IS_ACTIVE")

                if fips_st is None or fips_co is None:
                    continue

                for row in rows_iter:
                    if not row or len(row) <= max(fips_st, fips_co):
                        continue

                    # Skip inactive governments
                    if active_idx is not None and str(row[active_idx] or "").strip() != "Y":
                        continue

                    state_code = str(row[fips_st] or "").strip().zfill(2)
                    county_code = str(row[fips_co] or "").strip().zfill(3)

                    if not state_code or not county_code or county_code == "000":
                        continue

                    state_abbr = FIPS_TO_STATE.get(state_code)
                    if not state_abbr:
                        continue
                    if states and state_abbr not in states:
                        continue

                    county_fips = f"{state_code}{county_code}"

                    if county_fips not in county_data:
                        county_name = str(row[county_area_idx] or "").strip() if county_area_idx is not None else ""
                        county_data[county_fips] = {
                            "county_fips": county_fips,
                            "county_name": county_name[:255] if county_name else None,
                            "state": state_abbr,
                            "census_year": 2022,
                            "total_governments": 0,
                            "county_govts": 0,
                            "municipal_govts": 0,
                            "township_govts": 0,
                            "special_district_govts": 0,
                            "school_district_govts": 0,
                            "population": None,
                            "govts_per_10k_pop": None,
                            "source": "census_gov",
                            "collected_at": datetime.utcnow(),
                        }

                    entry = county_data[county_fips]
                    entry["total_governments"] += 1

                    # Parse unit type (e.g., "1 - COUNTY" → prefix "1")
                    if unit_type_idx is not None:
                        unit_type_str = str(row[unit_type_idx] or "").strip()
                        type_prefix = unit_type_str.split(" ")[0] if unit_type_str else ""
                        gov_category = UNIT_TYPE_PREFIX_MAP.get(type_prefix)
                        if gov_category:
                            key = f"{gov_category}_govts"
                            entry[key] = entry.get(key, 0) + 1

                    # Grab population from county-level records
                    if pop_idx is not None:
                        pop = self._safe_int(row[pop_idx])
                        if pop and pop > 0 and (entry["population"] is None or pop > entry["population"]):
                            entry["population"] = pop

                logger.info(f"Census Gov: processed sheet '{sheet_name}'")

            wb.close()

        # Compute govts_per_10k_pop where population is available
        for entry in county_data.values():
            pop = entry.get("population")
            total = entry.get("total_governments", 0)
            if pop and pop > 0 and total > 0:
                entry["govts_per_10k_pop"] = round(total / pop * 10000, 2)

        return county_data

    def _safe_int(self, value: Any) -> Optional[int]:
        if value is None or value == "" or value == "-":
            return None
        try:
            return int(float(str(value).strip()))
        except (ValueError, TypeError):
            return None
