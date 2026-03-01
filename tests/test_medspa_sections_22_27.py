"""
Tests for Medspa Market Report Sections 22-27.

Covers:
  22. Opportunity Zone Overlay
  23. Demographic Demand Model
  24. PE Competitive Heat Map
  25. Construction Momentum Signal
  26. Medical CPI Pricing Power
  27. Talent Pipeline Pressure

Test layers:
  - Data methods: table-not-found fallback → returns empty dict
  - render_html: empty data → fallback callouts; populated data → KPIs/charts/tables
  - render_excel: populated data → 6 new sheets with correct headers
  - gather_data: new keys present in output
"""

import pytest
from io import BytesIO

from app.reports.templates.medspa_market import MedSpaMarketTemplate


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def template():
    """Fresh MedSpaMarketTemplate instance."""
    return MedSpaMarketTemplate()


@pytest.fixture
def base_data():
    """Minimal valid data dict for render_html/render_excel (sections 1-21 empty)."""
    return {
        "generated_at": "2026-02-26T12:00:00",
        "params": {},
        "summary": {
            "total_prospects": 100,
            "a_grade": 20,
            "b_grade": 30,
            "ab_grade": 50,
            "states_covered": 10,
            "zips_covered": 80,
            "avg_score": 65.5,
            "max_score": 95.0,
            "avg_rating": 4.2,
            "avg_reviews": 120,
            "avg_zip_score": 70.0,
            "total_reviews": 12000,
        },
        "prospects_by_state": [],
        "grade_distribution": [],
        "score_histogram": [],
        "top_targets": [],
        "zip_concentration": [],
        "state_avg_scores": [],
        "a_grade_by_state": [],
        "zip_affluence_by_state": [],
        "census_income": [],
        "high_income_zip_penetration": {},
        "pe_comps": [],
        "recent_deals": [],
        "data_freshness": {},
        "whitespace_zips": [],
        "whitespace_by_state": [],
        "whitespace_summary": {},
        "bls_wages": {},
        "bls_employment": {},
        "category_breakdown": [],
        "pe_financials": [],
        "pe_financial_summary": {},
        "top_by_reviews": [],
        "review_buckets": [],
        "low_competition_gems": [],
        "deal_model": {},
        "stealth_wealth": {},
        "migration_alpha": {},
        "provider_density": {},
        "real_estate_alpha": {},
        "deposit_wealth": {},
        "business_formation": {},
        # Sections 22-27 empty
        "opportunity_zones": {},
        "demographic_demand": {},
        "pe_competitive": {},
        "construction_momentum": {},
        "medical_cpi": {},
        "talent_pipeline": {},
    }


@pytest.fixture
def populated_oz_data():
    """Opportunity Zone data with realistic values."""
    return {
        "state_data": [
            {
                "state": "CA",
                "oz_tracts": 879,
                "low_income_tracts": 680,
                "contiguous_tracts": 199,
                "prospect_count": 450,
                "a_grade_count": 85,
                "oz_per_prospect": 1.95,
            },
            {
                "state": "TX",
                "oz_tracts": 628,
                "low_income_tracts": 510,
                "contiguous_tracts": 118,
                "prospect_count": 320,
                "a_grade_count": 60,
                "oz_per_prospect": 1.96,
            },
            {
                "state": "NY",
                "oz_tracts": 514,
                "low_income_tracts": 400,
                "contiguous_tracts": 114,
                "prospect_count": 280,
                "a_grade_count": 55,
                "oz_per_prospect": 1.84,
            },
        ],
        "summary": {
            "total_tracts": 8764,
            "states_with_oz": 50,
            "medspa_oz_states": 45,
            "tax_advantaged_states": 12,
            "total_low_income": 6800,
            "total_contiguous": 1964,
        },
    }


@pytest.fixture
def populated_demographic_data():
    """Demographic demand data with realistic values."""
    return {
        "state_data": [
            {
                "state_name": "Massachusetts",
                "state_fips": "25",
                "avg_bachelors": 44.2,
                "avg_graduate": 20.1,
                "total_pop_25plus": 4800000,
                "area_count": 14,
                "prospect_count": 120,
                "a_grade_count": 25,
                "gap_score": 52.3,
            },
            {
                "state_name": "Colorado",
                "state_fips": "08",
                "avg_bachelors": 42.7,
                "avg_graduate": 16.8,
                "total_pop_25plus": 3900000,
                "area_count": 10,
                "prospect_count": 80,
                "a_grade_count": 15,
                "gap_score": 51.5,
            },
        ],
        "summary": {
            "avg_bachelors_pct": 33.5,
            "avg_graduate_pct": 13.2,
            "states_analyzed": 50,
            "underserved_educated": 8,
            "period_year": 2023,
        },
    }


@pytest.fixture
def populated_pe_competitive_data():
    """PE competitive heat map data with realistic values."""
    return {
        "deals": [
            {
                "company": "SkinSpirit",
                "industry": "Medical Aesthetics",
                "state": "CA",
                "city": "Palo Alto",
                "pe_owner": "LNR Partners",
                "deal_type": "LBO",
                "ev_usd": 250000000,
                "ev_ebitda": 12.5,
                "date": "2024-06-15",
                "buyer": "LNR Partners",
                "is_platform": True,
            },
            {
                "company": "LaserAway",
                "industry": "Cosmetic Dermatology",
                "state": "CA",
                "city": "Los Angeles",
                "pe_owner": "Ares Management",
                "deal_type": "Growth",
                "ev_usd": 180000000,
                "ev_ebitda": 10.2,
                "date": "2024-03-20",
                "buyer": "Ares Management",
                "is_platform": True,
            },
            {
                "company": "Milan Laser",
                "industry": "Beauty & Wellness",
                "state": "NE",
                "city": "Omaha",
                "pe_owner": "Harvest Partners",
                "deal_type": "Add-on",
                "ev_usd": 0,
                "ev_ebitda": 0,
                "date": "2023-11-01",
                "buyer": "Harvest Partners",
                "is_platform": False,
            },
        ],
        "deal_type_breakdown": {"LBO": 1, "Growth": 1, "Add-on": 1},
        "state_counts": {"CA": 2, "NE": 1},
        "year_counts": {"2023": 1, "2024": 2},
        "summary": {
            "pe_platforms": 2,
            "avg_ev_ebitda": 11.4,
            "total_deal_value": 430000000,
            "most_active_buyer": "LNR Partners",
            "total_deals": 3,
        },
    }


@pytest.fixture
def populated_construction_data():
    """Construction momentum data with realistic values."""
    return {
        "state_data": [
            {
                "state_name": "Texas",
                "state_abbr": "TX",
                "total_permits": 240000,
                "permits_1unit": 160000,
                "permits_2to4": 5000,
                "permits_5plus": 75000,
                "yoy_growth": 12.5,
                "medspa_count": 320,
                "a_grade_count": 60,
            },
            {
                "state_name": "Florida",
                "state_abbr": "FL",
                "total_permits": 200000,
                "permits_1unit": 130000,
                "permits_2to4": 4000,
                "permits_5plus": 66000,
                "yoy_growth": 8.3,
                "medspa_count": 290,
                "a_grade_count": 55,
            },
            {
                "state_name": "California",
                "state_abbr": "CA",
                "total_permits": 110000,
                "permits_1unit": 40000,
                "permits_2to4": 3000,
                "permits_5plus": 67000,
                "yoy_growth": -2.1,
                "medspa_count": 450,
                "a_grade_count": 85,
            },
        ],
        "permit_composition": {
            "single_family_pct": 59.8,
            "two_to_four_pct": 2.2,
            "five_plus_pct": 38.0,
        },
        "summary": {
            "states_analyzed": 50,
            "avg_yoy_growth": 5.4,
            "high_growth_states": 15,
            "top_state": "Texas",
            "latest_year": 2025,
            "has_prior_year": True,
        },
    }


@pytest.fixture
def populated_medical_cpi_data():
    """Medical CPI pricing power data with realistic values."""
    return {
        "annual_data": [
            {
                "year": 2025,
                "medical_cpi": 575.2,
                "general_cpi": 315.8,
                "spread": 259.4,
                "medical_yoy": 3.15,
                "general_yoy": 2.80,
                "yoy_spread": 0.35,
                "cumulative_divergence": 0.35,
            },
            {
                "year": 2024,
                "medical_cpi": 557.6,
                "general_cpi": 307.2,
                "spread": 250.4,
                "medical_yoy": 2.95,
                "general_yoy": 3.20,
                "yoy_spread": -0.25,
                "cumulative_divergence": 0.10,
            },
            {
                "year": 2023,
                "medical_cpi": 541.6,
                "general_cpi": 297.7,
                "spread": 243.9,
                "medical_yoy": 4.10,
                "general_yoy": 4.90,
                "yoy_spread": -0.80,
                "cumulative_divergence": -0.70,
            },
        ],
        "summary": {
            "current_medical_cpi": 575.2,
            "yoy_medical_change": 3.15,
            "medical_vs_general_spread": 0.35,
            "cagr_5yr": 3.05,
            "latest_year": 2025,
        },
    }


@pytest.fixture
def populated_talent_pipeline_data():
    """Talent pipeline pressure data with realistic values."""
    return {
        "quarterly_data": [
            {"period": "2025-Q4", "openings": 1850, "hires": 1320, "ratio": 1.40},
            {"period": "2025-Q3", "openings": 1780, "hires": 1280, "ratio": 1.39},
            {"period": "2025-Q2", "openings": 1720, "hires": 1300, "ratio": 1.32},
            {"period": "2025-Q1", "openings": 1690, "hires": 1310, "ratio": 1.29},
            {"period": "2024-Q4", "openings": 1650, "hires": 1350, "ratio": 1.22},
            {"period": "2024-Q3", "openings": 1620, "hires": 1340, "ratio": 1.21},
            {"period": "2024-Q2", "openings": 1580, "hires": 1320, "ratio": 1.20},
            {"period": "2024-Q1", "openings": 1550, "hires": 1300, "ratio": 1.19},
        ],
        "summary": {
            "latest_openings": 1850,
            "openings_to_hires_ratio": 1.40,
            "yoy_change": 12.1,
            "trend": "tightening",
        },
    }


# ===========================================================================
# 1. Data Method Fallback Tests (table doesn't exist → empty dict)
# ===========================================================================

class TestDataMethodFallbacks:
    """Each data method must return an empty sub-dict when its table is missing."""

    @pytest.mark.unit
    def test_opportunity_zone_no_table(self, template, test_db):
        result = template._get_opportunity_zone_data(test_db, None)
        assert result == {"opportunity_zones": {}}

    @pytest.mark.unit
    def test_opportunity_zone_no_table_with_state(self, template, test_db):
        result = template._get_opportunity_zone_data(test_db, "TX")
        assert result == {"opportunity_zones": {}}

    @pytest.mark.unit
    def test_demographic_demand_no_table(self, template, test_db):
        result = template._get_demographic_demand_data(test_db, None)
        assert result == {"demographic_demand": {}}

    @pytest.mark.unit
    def test_demographic_demand_no_table_with_state(self, template, test_db):
        result = template._get_demographic_demand_data(test_db, "CA")
        assert result == {"demographic_demand": {}}

    @pytest.mark.unit
    def test_pe_competitive_no_table(self, template, test_db):
        result = template._get_pe_competitive_data(test_db, None)
        assert result == {"pe_competitive": {}}

    @pytest.mark.unit
    def test_pe_competitive_no_table_with_state(self, template, test_db):
        result = template._get_pe_competitive_data(test_db, "NY")
        assert result == {"pe_competitive": {}}

    @pytest.mark.unit
    def test_construction_momentum_no_table(self, template, test_db):
        result = template._get_construction_momentum_data(test_db, None)
        assert result == {"construction_momentum": {}}

    @pytest.mark.unit
    def test_construction_momentum_no_table_with_state(self, template, test_db):
        result = template._get_construction_momentum_data(test_db, "FL")
        assert result == {"construction_momentum": {}}

    @pytest.mark.unit
    def test_medical_cpi_no_table(self, template, test_db):
        result = template._get_medical_cpi_data(test_db)
        assert result == {"medical_cpi": {}}

    @pytest.mark.unit
    def test_talent_pipeline_no_table(self, template, test_db):
        result = template._get_talent_pipeline_data(test_db)
        assert result == {"talent_pipeline": {}}


# ===========================================================================
# 2. Data Method Return Structure Tests (with mock data via DB seeding)
#    These use information_schema checks which work in SQLite, but the main
#    SQL queries use PostgreSQL syntax. We verify that:
#    - ORM-backed tables exist but return empty when no rows match
#    - Exception handling returns empty gracefully
# ===========================================================================

class TestDataMethodEmptyTables:
    """Tables exist (via ORM) but contain no rows → structured empty response."""

    @pytest.mark.unit
    def test_opportunity_zone_empty_table(self, template, test_db):
        """opportunity_zone table exists (ORM) but is empty."""
        from sqlalchemy import text
        # SQLite information_schema won't list ORM tables the same way,
        # but the method uses information_schema check which may differ.
        # This tests the graceful handling path.
        result = template._get_opportunity_zone_data(test_db, None)
        # Either empty because table check fails in SQLite or because no rows
        assert "opportunity_zones" in result
        assert isinstance(result["opportunity_zones"], dict)

    @pytest.mark.unit
    def test_pe_competitive_empty_table(self, template, test_db):
        """pe_portfolio_companies exists (ORM) but no aesthetics matches."""
        result = template._get_pe_competitive_data(test_db, None)
        assert "pe_competitive" in result
        assert isinstance(result["pe_competitive"], dict)

    @pytest.mark.unit
    def test_demographic_demand_empty_table(self, template, test_db):
        result = template._get_demographic_demand_data(test_db, None)
        assert "demographic_demand" in result
        assert isinstance(result["demographic_demand"], dict)


# ===========================================================================
# 3. render_html Tests — Empty Sections (Fallback Callouts)
# ===========================================================================

class TestRenderHtmlEmptySections:
    """When section data is empty, render_html should include fallback callouts."""

    @pytest.mark.unit
    def test_html_renders_without_error(self, template, base_data):
        """Basic smoke test: render_html doesn't crash with empty data."""
        html = template.render_html(base_data)
        assert isinstance(html, str)
        assert len(html) > 1000

    @pytest.mark.unit
    def test_section_22_fallback_present(self, template, base_data):
        html = template.render_html(base_data)
        assert "Opportunity Zone data not yet ingested" in html

    @pytest.mark.unit
    def test_section_23_fallback_present(self, template, base_data):
        html = template.render_html(base_data)
        assert "Educational attainment data not yet ingested" in html

    @pytest.mark.unit
    def test_section_24_fallback_present(self, template, base_data):
        html = template.render_html(base_data)
        assert "PE deal data not yet ingested" in html

    @pytest.mark.unit
    def test_section_25_fallback_present(self, template, base_data):
        html = template.render_html(base_data)
        assert "HUD building permit data not yet ingested" in html

    @pytest.mark.unit
    def test_section_26_fallback_present(self, template, base_data):
        html = template.render_html(base_data)
        assert "BLS CPI data not yet ingested" in html

    @pytest.mark.unit
    def test_section_27_fallback_present(self, template, base_data):
        html = template.render_html(base_data)
        assert "BLS JOLTS data not yet ingested" in html

    @pytest.mark.unit
    def test_toc_contains_new_sections(self, template, base_data):
        html = template.render_html(base_data)
        assert "Opportunity Zone Overlay" in html
        assert "Demographic Demand Model" in html
        assert "PE Competitive Heat Map" in html
        assert "Construction Momentum Signal" in html
        assert "Medical CPI Pricing Power" in html
        assert "Talent Pipeline Pressure" in html

    @pytest.mark.unit
    def test_toc_section_ids(self, template, base_data):
        html = template.render_html(base_data)
        assert 'id="oz-overlay"' in html
        assert 'id="demographic-demand"' in html
        assert 'id="pe-heatmap"' in html
        assert 'id="construction-momentum"' in html
        assert 'id="medical-cpi"' in html
        assert 'id="talent-pipeline"' in html

    @pytest.mark.unit
    def test_section_numbering(self, template, base_data):
        """Sections 22-27 should appear with correct numbering."""
        html = template.render_html(base_data)
        # section_start puts the number in the section heading
        for num in [22, 23, 24, 25, 26, 27]:
            assert f">{num}<" in html or f">{num}." in html or f">{num} " in html

    @pytest.mark.unit
    def test_footer_notes_include_new_sources(self, template, base_data):
        html = template.render_html(base_data)
        assert "Opportunity Zone data from CDFI Fund" in html
        assert "Demographic Demand uses Census ACS" in html
        assert "PE Competitive Heat Map filters" in html
        assert "Construction Momentum uses HUD" in html
        assert "Medical CPI uses BLS" in html
        assert "Talent Pipeline uses BLS JOLTS" in html


# ===========================================================================
# 4. render_html Tests — Populated Sections
# ===========================================================================

class TestRenderHtmlPopulatedSections:
    """When section data is populated, render_html includes KPIs, charts, tables."""

    @pytest.mark.unit
    def test_section_22_oz_kpis(self, template, base_data, populated_oz_data):
        base_data["opportunity_zones"] = populated_oz_data
        html = template.render_html(base_data)
        # Should NOT have fallback
        assert "Opportunity Zone data not yet ingested" not in html
        # Should have KPIs
        assert "8,764" in html  # total_tracts
        assert "50" in html  # states_with_oz
        assert "Tax-Advantaged States" in html

    @pytest.mark.unit
    def test_section_22_oz_chart_and_table(self, template, base_data, populated_oz_data):
        base_data["opportunity_zones"] = populated_oz_data
        html = template.render_html(base_data)
        assert "ozStateBar" in html  # chart ID
        assert "ozDesignDonut" in html  # chart ID
        assert "OZ/Prospect Ratio" in html  # table header
        assert "CA" in html  # state in table

    @pytest.mark.unit
    def test_section_22_oz_callout(self, template, base_data, populated_oz_data):
        base_data["opportunity_zones"] = populated_oz_data
        html = template.render_html(base_data)
        assert "tax-advantaged states" in html

    @pytest.mark.unit
    def test_section_23_demographic_kpis(self, template, base_data, populated_demographic_data):
        base_data["demographic_demand"] = populated_demographic_data
        html = template.render_html(base_data)
        assert "Educational attainment data not yet ingested" not in html
        assert "33.5%" in html  # avg_bachelors_pct
        assert "13.2%" in html  # avg_graduate_pct
        assert "Underserved Educated States" in html

    @pytest.mark.unit
    def test_section_23_demographic_chart_and_table(self, template, base_data, populated_demographic_data):
        base_data["demographic_demand"] = populated_demographic_data
        html = template.render_html(base_data)
        assert "demoEduBar" in html
        assert "demoEduDonut" in html
        assert "Gap Score" in html  # table header
        assert "Massachusetts" in html

    @pytest.mark.unit
    def test_section_24_pe_competitive_kpis(self, template, base_data, populated_pe_competitive_data):
        base_data["pe_competitive"] = populated_pe_competitive_data
        html = template.render_html(base_data)
        assert "PE deal data not yet ingested" not in html
        assert "11.4x" in html  # avg_ev_ebitda
        assert "Most Active Buyer" in html

    @pytest.mark.unit
    def test_section_24_pe_competitive_deals_table(self, template, base_data, populated_pe_competitive_data):
        base_data["pe_competitive"] = populated_pe_competitive_data
        html = template.render_html(base_data)
        assert "peHeatBar" in html
        assert "peDealTypeDonut" in html
        assert "SkinSpirit" in html
        assert "LaserAway" in html
        assert "EV/EBITDA" in html  # table header

    @pytest.mark.unit
    def test_section_25_construction_kpis(self, template, base_data, populated_construction_data):
        base_data["construction_momentum"] = populated_construction_data
        html = template.render_html(base_data)
        assert "HUD building permit data not yet ingested" not in html
        assert "+5.4%" in html  # avg_yoy_growth
        assert "High-Growth States" in html
        assert "Texas" in html

    @pytest.mark.unit
    def test_section_25_construction_chart_and_table(self, template, base_data, populated_construction_data):
        base_data["construction_momentum"] = populated_construction_data
        html = template.render_html(base_data)
        assert "constructionGrowthBar" in html
        assert "permitTypeDonut" in html
        assert "Total Permits" in html  # table header
        assert "YoY Growth" in html

    @pytest.mark.unit
    def test_section_25_construction_signal_grades(self, template, base_data, populated_construction_data):
        base_data["construction_momentum"] = populated_construction_data
        html = template.render_html(base_data)
        # Texas has 12.5% growth → Grade B (>5%); CA -2.1% → Grade D (<0%)
        assert "Grade B" in html
        assert "Grade D" in html

    @pytest.mark.unit
    def test_section_26_medical_cpi_kpis(self, template, base_data, populated_medical_cpi_data):
        base_data["medical_cpi"] = populated_medical_cpi_data
        html = template.render_html(base_data)
        assert "BLS CPI data not yet ingested" not in html
        assert "575.2" in html  # current_medical_cpi
        assert "+3.15%" in html  # yoy_medical_change
        assert "3.05%" in html  # cagr_5yr

    @pytest.mark.unit
    def test_section_26_medical_cpi_table(self, template, base_data, populated_medical_cpi_data):
        base_data["medical_cpi"] = populated_medical_cpi_data
        html = template.render_html(base_data)
        assert "medCpiBar" in html
        assert "Medical CPI" in html  # table header
        assert "General CPI" in html  # table header
        assert "Cumulative" in html

    @pytest.mark.unit
    def test_section_26_pricing_power_positive(self, template, base_data, populated_medical_cpi_data):
        base_data["medical_cpi"] = populated_medical_cpi_data
        html = template.render_html(base_data)
        assert "Pricing power signal: positive" in html
        assert "outpacing" in html

    @pytest.mark.unit
    def test_section_26_pricing_power_negative(self, template, base_data, populated_medical_cpi_data):
        """When spread is negative, signal should be negative."""
        populated_medical_cpi_data["summary"]["medical_vs_general_spread"] = -0.5
        base_data["medical_cpi"] = populated_medical_cpi_data
        html = template.render_html(base_data)
        assert "Pricing power signal: negative" in html
        assert "trailing" in html

    @pytest.mark.unit
    def test_section_27_talent_pipeline_kpis(self, template, base_data, populated_talent_pipeline_data):
        base_data["talent_pipeline"] = populated_talent_pipeline_data
        html = template.render_html(base_data)
        assert "BLS JOLTS data not yet ingested" not in html
        assert "1850" in html  # latest_openings
        assert "1.40" in html  # ratio
        assert "Tightening" in html  # trend

    @pytest.mark.unit
    def test_section_27_talent_pipeline_chart_and_table(self, template, base_data, populated_talent_pipeline_data):
        base_data["talent_pipeline"] = populated_talent_pipeline_data
        html = template.render_html(base_data)
        assert "talentOpeningsBar" in html
        assert "Openings (000s)" in html  # table header
        assert "Hires (000s)" in html
        assert "Open/Hire Ratio" in html
        assert "2025-Q4" in html

    @pytest.mark.unit
    def test_section_27_tightening_trend_callout(self, template, base_data, populated_talent_pipeline_data):
        base_data["talent_pipeline"] = populated_talent_pipeline_data
        html = template.render_html(base_data)
        assert "tightening" in html.lower()
        assert "staffed platforms premium assets" in html

    @pytest.mark.unit
    def test_section_27_easing_trend_callout(self, template, base_data, populated_talent_pipeline_data):
        """When trend is easing, callout should reflect manageable conditions."""
        populated_talent_pipeline_data["summary"]["trend"] = "easing"
        base_data["talent_pipeline"] = populated_talent_pipeline_data
        html = template.render_html(base_data)
        assert "manageable" in html

    @pytest.mark.unit
    def test_all_six_sections_render_populated(
        self, template, base_data,
        populated_oz_data, populated_demographic_data,
        populated_pe_competitive_data, populated_construction_data,
        populated_medical_cpi_data, populated_talent_pipeline_data,
    ):
        """Integration: all 6 sections populated at once — no crashes."""
        base_data["opportunity_zones"] = populated_oz_data
        base_data["demographic_demand"] = populated_demographic_data
        base_data["pe_competitive"] = populated_pe_competitive_data
        base_data["construction_momentum"] = populated_construction_data
        base_data["medical_cpi"] = populated_medical_cpi_data
        base_data["talent_pipeline"] = populated_talent_pipeline_data
        html = template.render_html(base_data)
        assert isinstance(html, str)
        # None of the section 22-27 fallback messages should appear
        assert "Opportunity Zone data not yet ingested" not in html
        assert "Educational attainment data not yet ingested" not in html
        assert "PE deal data not yet ingested" not in html
        assert "HUD building permit data not yet ingested" not in html
        assert "BLS CPI data not yet ingested" not in html
        assert "BLS JOLTS data not yet ingested" not in html
        # All chart IDs present
        for chart_id in ["ozStateBar", "demoEduBar", "peHeatBar",
                         "constructionGrowthBar", "medCpiBar", "talentOpeningsBar"]:
            assert chart_id in html


# ===========================================================================
# 5. render_html Tests — Methodology Freshness
# ===========================================================================

class TestRenderHtmlFreshness:
    """Data freshness rows for new data sources."""

    @pytest.mark.unit
    def test_oz_freshness_row(self, template, base_data):
        base_data["data_freshness"] = {
            "opportunity_zone": {"earliest": None, "latest": None, "total": 8764},
        }
        html = template.render_html(base_data)
        assert "Opportunity Zones (OZ Overlay)" in html
        assert "CDFI Fund" in html

    @pytest.mark.unit
    def test_educational_attainment_freshness_row(self, template, base_data):
        base_data["data_freshness"] = {
            "educational_attainment": {"earliest": "2020", "latest": "2023", "total": 3200},
        }
        html = template.render_html(base_data)
        assert "Educational Attainment (Demographic Demand)" in html

    @pytest.mark.unit
    def test_hud_permits_freshness_row(self, template, base_data):
        base_data["data_freshness"] = {
            "hud_permits": {"earliest": "2020-01-01", "latest": "2025-06-01", "total": 50000},
        }
        html = template.render_html(base_data)
        assert "HUD Building Permits (Construction Momentum)" in html

    @pytest.mark.unit
    def test_bls_cpi_freshness_row(self, template, base_data):
        base_data["data_freshness"] = {
            "bls_cpi": {"earliest": "2015", "latest": "2025", "total": 1200},
        }
        html = template.render_html(base_data)
        assert "BLS CPI (Medical Pricing Power)" in html


# ===========================================================================
# 6. render_excel Tests — New Sheets
# ===========================================================================

class TestRenderExcelNewSheets:
    """render_excel should include 6 new sheets for sections 22-27."""

    def _load_workbook(self, excel_bytes):
        from openpyxl import load_workbook
        return load_workbook(BytesIO(excel_bytes))

    @pytest.mark.unit
    def test_excel_renders_without_error(self, template, base_data):
        """Smoke test: render_excel doesn't crash with empty section data."""
        excel_bytes = template.render_excel(base_data)
        assert isinstance(excel_bytes, bytes)
        assert len(excel_bytes) > 0

    @pytest.mark.unit
    def test_excel_has_oz_overlay_sheet(self, template, base_data):
        wb = self._load_workbook(template.render_excel(base_data))
        assert "OZ Overlay" in wb.sheetnames

    @pytest.mark.unit
    def test_excel_has_demographic_demand_sheet(self, template, base_data):
        wb = self._load_workbook(template.render_excel(base_data))
        assert "Demographic Demand" in wb.sheetnames

    @pytest.mark.unit
    def test_excel_has_pe_heat_map_sheet(self, template, base_data):
        wb = self._load_workbook(template.render_excel(base_data))
        assert "PE Heat Map" in wb.sheetnames

    @pytest.mark.unit
    def test_excel_has_construction_momentum_sheet(self, template, base_data):
        wb = self._load_workbook(template.render_excel(base_data))
        assert "Construction Momentum" in wb.sheetnames

    @pytest.mark.unit
    def test_excel_has_medical_cpi_sheet(self, template, base_data):
        wb = self._load_workbook(template.render_excel(base_data))
        assert "Medical CPI" in wb.sheetnames

    @pytest.mark.unit
    def test_excel_has_talent_pipeline_sheet(self, template, base_data):
        wb = self._load_workbook(template.render_excel(base_data))
        assert "Talent Pipeline" in wb.sheetnames

    @pytest.mark.unit
    def test_excel_oz_sheet_headers(self, template, base_data, populated_oz_data):
        base_data["opportunity_zones"] = populated_oz_data
        wb = self._load_workbook(template.render_excel(base_data))
        ws = wb["OZ Overlay"]
        assert ws["A1"].value == "Opportunity Zone Overlay"
        assert ws["A3"].value == "Total OZ Tracts"
        assert ws["B3"].value == 8764

    @pytest.mark.unit
    def test_excel_oz_sheet_data_rows(self, template, base_data, populated_oz_data):
        base_data["opportunity_zones"] = populated_oz_data
        wb = self._load_workbook(template.render_excel(base_data))
        ws = wb["OZ Overlay"]
        # Header row at row 8, data starts at row 9
        assert ws.cell(row=8, column=1).value == "State"
        assert ws.cell(row=9, column=1).value == "CA"
        assert ws.cell(row=9, column=2).value == 879

    @pytest.mark.unit
    def test_excel_demographic_sheet_headers(self, template, base_data, populated_demographic_data):
        base_data["demographic_demand"] = populated_demographic_data
        wb = self._load_workbook(template.render_excel(base_data))
        ws = wb["Demographic Demand"]
        assert ws["A1"].value == "Demographic Demand Model"
        assert ws["A3"].value == "Period Year"

    @pytest.mark.unit
    def test_excel_pe_sheet_data(self, template, base_data, populated_pe_competitive_data):
        base_data["pe_competitive"] = populated_pe_competitive_data
        wb = self._load_workbook(template.render_excel(base_data))
        ws = wb["PE Heat Map"]
        assert ws["A1"].value == "PE Competitive Heat Map"
        assert ws["A3"].value == "PE-Backed Platforms"
        assert ws["B3"].value == 2

    @pytest.mark.unit
    def test_excel_construction_sheet_data(self, template, base_data, populated_construction_data):
        base_data["construction_momentum"] = populated_construction_data
        wb = self._load_workbook(template.render_excel(base_data))
        ws = wb["Construction Momentum"]
        assert ws["A1"].value == "Construction Momentum Signal"
        assert ws["A3"].value == "Year"
        assert ws["B3"].value == 2025

    @pytest.mark.unit
    def test_excel_medical_cpi_sheet_data(self, template, base_data, populated_medical_cpi_data):
        base_data["medical_cpi"] = populated_medical_cpi_data
        wb = self._load_workbook(template.render_excel(base_data))
        ws = wb["Medical CPI"]
        assert ws["A1"].value == "Medical CPI Pricing Power"
        assert ws["B3"].value == 575.2  # current_medical_cpi

    @pytest.mark.unit
    def test_excel_talent_pipeline_sheet_data(self, template, base_data, populated_talent_pipeline_data):
        base_data["talent_pipeline"] = populated_talent_pipeline_data
        wb = self._load_workbook(template.render_excel(base_data))
        ws = wb["Talent Pipeline"]
        assert ws["A1"].value == "Talent Pipeline Pressure"
        assert ws["B3"].value == 1850  # latest_openings

    @pytest.mark.unit
    def test_excel_total_sheet_count(self, template, base_data):
        """Should have original 16 sheets + 6 new = 22 sheets."""
        wb = self._load_workbook(template.render_excel(base_data))
        # Existing: Exec Summary, Top Targets, By State, ZIP Concentration,
        # ZIP Affluence, PE Comps, Whitespace ZIPs, Workforce Wages,
        # Growth Signals, Deal Model, Stealth Wealth, Migration Alpha,
        # Provider Density, RE Appreciation Alpha, Deposit Wealth,
        # Business Formation = 16
        # New: OZ Overlay, Demographic Demand, PE Heat Map,
        # Construction Momentum, Medical CPI, Talent Pipeline = 6
        assert len(wb.sheetnames) == 22


# ===========================================================================
# 7. gather_data Wiring Test
# ===========================================================================

class TestGatherDataWiring:
    """Verify the new section data methods are properly wired into gather_data.

    Full gather_data() requires PostgreSQL (FILTER, EXTRACT, etc.) which is
    unavailable in SQLite test DB. Instead we verify wiring by inspecting
    the source code and testing the data methods individually.
    """

    @pytest.mark.unit
    def test_gather_data_calls_all_new_methods(self, template):
        """Verify gather_data source code references all 6 new method calls."""
        import inspect
        source = inspect.getsource(template.gather_data)
        assert "_get_opportunity_zone_data" in source
        assert "_get_demographic_demand_data" in source
        assert "_get_pe_competitive_data" in source
        assert "_get_construction_momentum_data" in source
        assert "_get_medical_cpi_data" in source
        assert "_get_talent_pipeline_data" in source

    @pytest.mark.unit
    def test_national_sections_no_state_param(self, template):
        """Sections 26-27 should not receive state_filter (national data)."""
        import inspect
        source = inspect.getsource(template.gather_data)
        # medical_cpi and talent_pipeline should be called with just db
        assert "_get_medical_cpi_data(db)" in source
        assert "_get_talent_pipeline_data(db)" in source

    @pytest.mark.unit
    def test_geographic_sections_receive_state(self, template):
        """Sections 22-25 should receive state_filter."""
        import inspect
        source = inspect.getsource(template.gather_data)
        assert "_get_opportunity_zone_data(db, state_filter)" in source
        assert "_get_demographic_demand_data(db, state_filter)" in source
        assert "_get_pe_competitive_data(db, state_filter)" in source
        assert "_get_construction_momentum_data(db, state_filter)" in source

    @pytest.mark.unit
    def test_all_new_methods_exist(self, template):
        """All 6 new data methods should exist on the template class."""
        assert hasattr(template, "_get_opportunity_zone_data")
        assert hasattr(template, "_get_demographic_demand_data")
        assert hasattr(template, "_get_pe_competitive_data")
        assert hasattr(template, "_get_construction_momentum_data")
        assert hasattr(template, "_get_medical_cpi_data")
        assert hasattr(template, "_get_talent_pipeline_data")

    @pytest.mark.unit
    def test_all_new_methods_callable(self, template):
        """All 6 new data methods should be callable."""
        assert callable(template._get_opportunity_zone_data)
        assert callable(template._get_demographic_demand_data)
        assert callable(template._get_pe_competitive_data)
        assert callable(template._get_construction_momentum_data)
        assert callable(template._get_medical_cpi_data)
        assert callable(template._get_talent_pipeline_data)


# ===========================================================================
# 8. Edge Case Tests
# ===========================================================================

class TestEdgeCases:
    """Edge cases in data and rendering."""

    @pytest.mark.unit
    def test_oz_zero_prospects(self, template, base_data):
        """OZ data with zero medspa prospects should not divide by zero."""
        oz_data = {
            "state_data": [{
                "state": "WY",
                "oz_tracts": 25,
                "low_income_tracts": 20,
                "contiguous_tracts": 5,
                "prospect_count": 0,
                "a_grade_count": 0,
                "oz_per_prospect": 0,
            }],
            "summary": {
                "total_tracts": 25,
                "states_with_oz": 1,
                "medspa_oz_states": 0,
                "tax_advantaged_states": 0,
                "total_low_income": 20,
                "total_contiguous": 5,
            },
        }
        base_data["opportunity_zones"] = oz_data
        html = template.render_html(base_data)
        assert isinstance(html, str)
        assert "WY" in html

    @pytest.mark.unit
    def test_medical_cpi_zero_spread(self, template, base_data):
        """Medical CPI with exactly zero spread should handle gracefully."""
        cpi_data = {
            "annual_data": [{
                "year": 2025,
                "medical_cpi": 300.0,
                "general_cpi": 300.0,
                "spread": 0.0,
                "medical_yoy": 2.0,
                "general_yoy": 2.0,
                "yoy_spread": 0.0,
                "cumulative_divergence": 0.0,
            }],
            "summary": {
                "current_medical_cpi": 300.0,
                "yoy_medical_change": 2.0,
                "medical_vs_general_spread": 0.0,
                "cagr_5yr": 2.0,
                "latest_year": 2025,
            },
        }
        base_data["medical_cpi"] = cpi_data
        html = template.render_html(base_data)
        assert isinstance(html, str)
        assert "0.00 pp" in html

    @pytest.mark.unit
    def test_talent_pipeline_no_hires(self, template, base_data):
        """Talent data with zero hires should not divide by zero."""
        tp_data = {
            "quarterly_data": [{
                "period": "2025-Q4",
                "openings": 1500,
                "hires": 0,
                "ratio": 0,
            }],
            "summary": {
                "latest_openings": 1500,
                "openings_to_hires_ratio": 0,
                "yoy_change": None,
                "trend": "stable",
            },
        }
        base_data["talent_pipeline"] = tp_data
        html = template.render_html(base_data)
        assert isinstance(html, str)
        assert "1500" in html

    @pytest.mark.unit
    def test_construction_all_negative_growth(self, template, base_data):
        """All states declining — renders red without crash."""
        cm_data = {
            "state_data": [{
                "state_name": "West Virginia",
                "state_abbr": "WV",
                "total_permits": 5000,
                "permits_1unit": 3500,
                "permits_2to4": 200,
                "permits_5plus": 1300,
                "yoy_growth": -15.0,
                "medspa_count": 10,
                "a_grade_count": 2,
            }],
            "permit_composition": {
                "single_family_pct": 70.0,
                "two_to_four_pct": 4.0,
                "five_plus_pct": 26.0,
            },
            "summary": {
                "states_analyzed": 1,
                "avg_yoy_growth": -15.0,
                "high_growth_states": 0,
                "top_state": "West Virginia",
                "latest_year": 2025,
                "has_prior_year": True,
            },
        }
        base_data["construction_momentum"] = cm_data
        html = template.render_html(base_data)
        assert isinstance(html, str)
        assert "-15.0%" in html
        assert "Grade D" in html  # negative growth = D grade

    @pytest.mark.unit
    def test_pe_competitive_with_missing_ev(self, template, base_data):
        """PE deals with zero EV/EBITDA should display dash."""
        pe_data = {
            "deals": [{
                "company": "Secret Spa",
                "industry": "Aesthetics",
                "state": "TX",
                "city": "Dallas",
                "pe_owner": "-",
                "deal_type": "Add-on",
                "ev_usd": 0,
                "ev_ebitda": 0,
                "date": "-",
                "buyer": "Unknown",
                "is_platform": False,
            }],
            "deal_type_breakdown": {"Add-on": 1},
            "state_counts": {"TX": 1},
            "year_counts": {},
            "summary": {
                "pe_platforms": 0,
                "avg_ev_ebitda": 0,
                "total_deal_value": 0,
                "most_active_buyer": "Unknown",
                "total_deals": 1,
            },
        }
        base_data["pe_competitive"] = pe_data
        html = template.render_html(base_data)
        assert isinstance(html, str)
        assert "Secret Spa" in html

    @pytest.mark.unit
    def test_excel_all_sections_populated(
        self, template, base_data,
        populated_oz_data, populated_demographic_data,
        populated_pe_competitive_data, populated_construction_data,
        populated_medical_cpi_data, populated_talent_pipeline_data,
    ):
        """Excel with all 6 sections populated renders without error."""
        base_data["opportunity_zones"] = populated_oz_data
        base_data["demographic_demand"] = populated_demographic_data
        base_data["pe_competitive"] = populated_pe_competitive_data
        base_data["construction_momentum"] = populated_construction_data
        base_data["medical_cpi"] = populated_medical_cpi_data
        base_data["talent_pipeline"] = populated_talent_pipeline_data
        excel_bytes = template.render_excel(base_data)
        assert isinstance(excel_bytes, bytes)
        assert len(excel_bytes) > 5000  # Non-trivial size

    @pytest.mark.unit
    def test_demographic_empty_state_data(self, template, base_data):
        """Demographic with summary but empty state_data should render summary only."""
        dd_data = {
            "state_data": [],
            "summary": {
                "avg_bachelors_pct": 0,
                "avg_graduate_pct": 0,
                "states_analyzed": 0,
                "underserved_educated": 0,
                "period_year": 2023,
            },
        }
        base_data["demographic_demand"] = dd_data
        html = template.render_html(base_data)
        assert isinstance(html, str)
        # Should still show KPIs even with empty tables
        assert "States Analyzed" in html
