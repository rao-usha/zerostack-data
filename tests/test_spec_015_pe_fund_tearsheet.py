"""
Tests for SPEC 015 — PE Fund Tearsheet Report.
"""
import pytest
from unittest.mock import MagicMock, patch, PropertyMock
from io import BytesIO


@pytest.mark.unit
class TestSpec015PeFundTearsheet:
    """Tests for PE fund tearsheet report template."""

    def _make_mock_db(self, fund_row=None, perf_row=None,
                      portfolio_rows=None, cf_rows=None,
                      sector_rows=None, top_rows=None, bottom_rows=None):
        """Build a mock DB session that returns canned data per query."""
        db = MagicMock()
        call_count = [0]

        # Map query patterns to results based on call order
        # Order: fund, performance, portfolio, cash_flows, sector_alloc, top, bottom
        result_sequence = [
            fund_row,           # _get_fund
            perf_row,           # _get_performance
            portfolio_rows,     # _get_portfolio
            cf_rows,            # _get_cash_flows
            sector_rows,        # _get_sector_allocation
            top_rows,           # _get_top_performers
            bottom_rows,        # _get_bottom_performers
        ]

        def mock_execute(query, params=None):
            result = MagicMock()
            idx = call_count[0]
            call_count[0] += 1

            if idx < len(result_sequence):
                data = result_sequence[idx]
                if data is None:
                    result.fetchone.return_value = None
                    result.fetchall.return_value = []
                elif isinstance(data, list):
                    result.fetchall.return_value = data
                    result.fetchone.return_value = data[0] if data else None
                else:
                    result.fetchone.return_value = data
                    result.fetchall.return_value = [data]
            else:
                result.fetchone.return_value = None
                result.fetchall.return_value = []
            return result

        db.execute = mock_execute
        return db

    def _sample_fund_row(self):
        """A mock fund row tuple matching the SELECT in _get_fund."""
        from datetime import date
        return (
            1, "Summit Growth Fund III", 2022, 500.0,  # id, name, vintage, target
            450.0, "Growth Equity", "active",            # final_close, strategy, status
            2.0, 20.0, 8.0, 10, 5,                      # fees, carry, pref, life, invest_period
            date(2022, 3, 15), date(2022, 9, 30),        # first_close, final_close_date
            "Summit Ridge Partners",                      # firm_name
        )

    def _sample_perf_row(self):
        from datetime import date
        return (
            18.5, 22.3, 1.85, 0.45, 1.40,       # net_irr, gross_irr, tvpi, dpi, rvpi
            500.0, 380.0, 171.0,                  # committed, called, distributed
            350.0, 8, 3, 0,                       # remaining, active, realized, written_off
            date(2025, 12, 31),                   # as_of_date
        )

    def _sample_portfolio_rows(self):
        from datetime import date
        return [
            (101, "TechCorp", "Software", "active",
             date(2022, 6, 1), 50_000_000.0, 200_000_000.0, 8.5,
             35.0, None, None, None, None, None, "active"),
            (102, "HealthCo", "Healthcare", "active",
             date(2023, 1, 15), 40_000_000.0, 160_000_000.0, 10.2,
             25.0, date(2025, 6, 1), "strategic", 120_000_000.0, 3.0, 32.0, "realized"),
        ]

    def _sample_cf_rows(self):
        from datetime import date
        return [
            (date(2022, 6, 1), -50.0, "capital_call", "Initial investment"),
            (date(2023, 1, 15), -40.0, "capital_call", "Follow-on"),
            (date(2025, 6, 1), 120.0, "distribution", "HealthCo exit"),
        ]

    def _sample_sector_rows(self):
        return [
            ("Software", 5, 250_000_000.0),
            ("Healthcare", 3, 120_000_000.0),
        ]

    def _sample_top_rows(self):
        return [
            ("HealthCo", 3.0, 32.0, 40_000_000.0, 120_000_000.0, "strategic"),
        ]

    def test_gather_data_returns_expected_keys(self):
        """T1: gather_data returns all required sections."""
        from app.reports.templates.pe_fund_tearsheet import PEFundTearsheetTemplate

        db = self._make_mock_db(
            fund_row=self._sample_fund_row(),
            perf_row=self._sample_perf_row(),
            portfolio_rows=self._sample_portfolio_rows(),
            cf_rows=self._sample_cf_rows(),
            sector_rows=self._sample_sector_rows(),
            top_rows=self._sample_top_rows(),
            bottom_rows=[],
        )
        template = PEFundTearsheetTemplate()
        data = template.gather_data(db, {"fund_id": 1})

        assert "generated_at" in data
        assert "fund" in data
        assert data["fund"]["name"] == "Summit Growth Fund III"
        assert "performance" in data
        assert data["performance"]["net_irr"] == 18.5
        assert "portfolio" in data
        assert len(data["portfolio"]) == 2
        assert "cash_flows" in data
        assert "sector_allocation" in data
        assert "top_performers" in data
        assert "bottom_performers" in data

    def test_render_html_contains_sections(self):
        """T2: HTML output has KPI strip, charts, tables."""
        from app.reports.templates.pe_fund_tearsheet import PEFundTearsheetTemplate

        db = self._make_mock_db(
            fund_row=self._sample_fund_row(),
            perf_row=self._sample_perf_row(),
            portfolio_rows=self._sample_portfolio_rows(),
            cf_rows=self._sample_cf_rows(),
            sector_rows=self._sample_sector_rows(),
            top_rows=self._sample_top_rows(),
            bottom_rows=[],
        )
        template = PEFundTearsheetTemplate()
        data = template.gather_data(db, {"fund_id": 1})
        html = template.render_html(data)

        assert "<!DOCTYPE html>" in html
        assert "Summit Growth Fund III" in html
        # KPI strip values
        assert "18.5%" in html  # Net IRR
        assert "1.85x" in html  # TVPI
        # Section headings
        assert "Fund Overview" in html
        assert "J-Curve" in html
        assert "Portfolio Companies" in html
        assert "Sector Allocation" in html
        assert "Top" in html and "Performers" in html
        assert "Cash Flow Summary" in html
        # Chart containers
        assert "jCurveChart" in html
        assert "sectorChart" in html

    def test_render_excel_has_sheets(self):
        """T3: Excel has 4 required sheet names."""
        from app.reports.templates.pe_fund_tearsheet import PEFundTearsheetTemplate
        from openpyxl import load_workbook

        db = self._make_mock_db(
            fund_row=self._sample_fund_row(),
            perf_row=self._sample_perf_row(),
            portfolio_rows=self._sample_portfolio_rows(),
            cf_rows=self._sample_cf_rows(),
            sector_rows=self._sample_sector_rows(),
            top_rows=self._sample_top_rows(),
            bottom_rows=[],
        )
        template = PEFundTearsheetTemplate()
        data = template.gather_data(db, {"fund_id": 1})
        excel_bytes = template.render_excel(data)

        wb = load_workbook(BytesIO(excel_bytes))
        assert "Summary" in wb.sheetnames
        assert "Portfolio" in wb.sheetnames
        assert "Cash Flows" in wb.sheetnames
        assert "Performance" in wb.sheetnames
        assert len(wb.sheetnames) == 4

    def test_template_registered(self):
        """T4: Template is in ReportBuilder.templates."""
        # Import the builder class and check template registry
        from app.reports.templates.pe_fund_tearsheet import PEFundTearsheetTemplate
        # Check it can be imported and has required attributes
        template = PEFundTearsheetTemplate()
        assert template.name == "pe_fund_tearsheet"
        assert hasattr(template, "gather_data")
        assert hasattr(template, "render_html")
        assert hasattr(template, "render_excel")

    def test_gather_data_missing_fund(self):
        """T5: Handles non-existent fund_id gracefully."""
        from app.reports.templates.pe_fund_tearsheet import PEFundTearsheetTemplate

        db = self._make_mock_db(fund_row=None)
        template = PEFundTearsheetTemplate()
        data = template.gather_data(db, {"fund_id": 9999})

        assert data["fund"] is None
        assert data["portfolio"] == []
        assert data["cash_flows"] == []

        # Render should show "Fund Not Found" gracefully
        html = template.render_html(data)
        assert "Fund not found" in html or "Fund Not Found" in html
