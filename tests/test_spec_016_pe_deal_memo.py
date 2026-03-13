"""
Tests for SPEC 016 — PE Deal Memo Report.
"""
import pytest
from unittest.mock import MagicMock
from io import BytesIO


@pytest.mark.unit
class TestSpec016PeDealMemo:
    """Tests for PE deal memo report template."""

    def _make_mock_db(self, company_row=None, financials_rows=None,
                      leadership_rows=None, competitor_rows=None,
                      investment_row=None):
        """Build a mock DB that returns canned data per query order."""
        db = MagicMock()
        call_count = [0]

        # Query order in gather_data:
        # 0: company, 1: financials, 2: benchmarks (service call), 3: comparables (service call),
        # 4: exit_readiness (service call), 5: leadership, 6: competitors, 7: investment
        # Service calls use their own DB queries so we just need the direct ones
        result_sequence = [
            company_row,        # _get_company
            financials_rows,    # _get_financials
            # benchmarks and comparables and exit_readiness use service classes — will fail gracefully
            leadership_rows,    # _get_leadership (after service calls fail)
            competitor_rows,    # _get_competitors
            investment_row,     # _get_investment
        ]

        def mock_execute(query, params=None):
            result = MagicMock()
            idx = call_count[0]
            call_count[0] += 1

            if idx < len(result_sequence):
                data = result_sequence[idx]
                if data is None:
                    result.fetchone.return_value = None
                    result.fetchall.return_value = []
                elif isinstance(data, list):
                    result.fetchall.return_value = data
                    result.fetchone.return_value = data[0] if data else None
                else:
                    result.fetchone.return_value = data
                    result.fetchall.return_value = [data]
            else:
                result.fetchone.return_value = None
                result.fetchall.return_value = []
            return result

        db.execute = mock_execute
        return db

    def _sample_company_row(self):
        return (
            101, "DataBridge Analytics", "AI-powered data integration platform",
            "Software", "Enterprise Software", "511210",
            "Austin", "TX", 250, 2018, "https://databridge.io",
            "pe_backed", "Summit Ridge Partners", "Technology",
        )

    def _sample_financials_rows(self):
        return [
            (2025, 45_000_000.0, 28.5, 12_000_000.0, 26.7, 72.3,
             8_500_000.0, 15_000_000.0, 5_000_000.0, 6_000_000.0, 1.25),
            (2024, 35_000_000.0, 22.0, 8_500_000.0, 24.3, 70.1,
             5_000_000.0, 12_000_000.0, 3_000_000.0, 4_000_000.0, 1.41),
        ]

    def _sample_leadership_rows(self):
        from datetime import date
        return [
            ("John Smith", "CEO", "executive", True, False, True,
             date(2018, 1, 1), "https://linkedin.com/in/jsmith"),
            ("Jane Doe", "CFO", "finance", False, True, False,
             date(2020, 3, 15), None),
        ]

    def _sample_competitor_rows(self):
        return [
            ("Competitor A", True, "CMP", False, None, "direct", "larger", "leader"),
            ("Competitor B", False, None, True, "KKR", "direct", "similar", "challenger"),
        ]

    def _sample_investment_row(self):
        from datetime import date
        return (
            30_000_000.0, 180_000_000.0, 15.0, 4.0,  # invested, ev, ev_ebitda, ev_rev
            65.0, date(2022, 6, 15), "active",         # ownership, date, status
            "Summit Growth Fund III", "Summit Ridge Partners",
        )

    def test_gather_data_returns_expected_keys(self):
        """T1: gather_data returns all required sections."""
        from app.reports.templates.pe_deal_memo import PEDealMemoTemplate

        db = self._make_mock_db(
            company_row=self._sample_company_row(),
            financials_rows=self._sample_financials_rows(),
            leadership_rows=self._sample_leadership_rows(),
            competitor_rows=self._sample_competitor_rows(),
            investment_row=self._sample_investment_row(),
        )
        template = PEDealMemoTemplate()
        data = template.gather_data(db, {"company_id": 101})

        assert "generated_at" in data
        assert "company" in data
        assert data["company"]["name"] == "DataBridge Analytics"
        assert "financials" in data
        assert len(data["financials"]) == 2
        assert "benchmarks" in data
        assert "comparables" in data
        assert "exit_readiness" in data
        assert "leadership" in data
        assert "competitors" in data
        assert "investment" in data

    def test_render_html_contains_sections(self):
        """T2: HTML output has all 8 IC memo sections."""
        from app.reports.templates.pe_deal_memo import PEDealMemoTemplate

        db = self._make_mock_db(
            company_row=self._sample_company_row(),
            financials_rows=self._sample_financials_rows(),
            leadership_rows=self._sample_leadership_rows(),
            competitor_rows=self._sample_competitor_rows(),
            investment_row=self._sample_investment_row(),
        )
        template = PEDealMemoTemplate()
        data = template.gather_data(db, {"company_id": 101})
        html = template.render_html(data)

        assert "<!DOCTYPE html>" in html
        assert "DataBridge Analytics" in html
        # All 8 sections
        assert "Executive Summary" in html
        assert "Company Overview" in html
        assert "Financial Analysis" in html
        assert "Market Position" in html
        assert "Management Assessment" in html
        assert "Valuation" in html and "Comparables" in html
        assert "Exit Readiness" in html
        assert "Recommended Next Steps" in html
        # KPI values
        assert "26.7%" in html  # EBITDA margin
        assert "28.5%" in html  # Revenue growth

    def test_render_excel_has_sheets(self):
        """T3: Excel has required sheet names."""
        from app.reports.templates.pe_deal_memo import PEDealMemoTemplate
        from openpyxl import load_workbook

        db = self._make_mock_db(
            company_row=self._sample_company_row(),
            financials_rows=self._sample_financials_rows(),
            leadership_rows=self._sample_leadership_rows(),
            competitor_rows=self._sample_competitor_rows(),
            investment_row=self._sample_investment_row(),
        )
        template = PEDealMemoTemplate()
        data = template.gather_data(db, {"company_id": 101})
        excel_bytes = template.render_excel(data)

        wb = load_workbook(BytesIO(excel_bytes))
        assert "Summary" in wb.sheetnames
        assert "Financials" in wb.sheetnames
        assert "Comparables" in wb.sheetnames
        assert "Leadership" in wb.sheetnames
        assert len(wb.sheetnames) == 4

    def test_template_registered(self):
        """T4: Template name and methods exist."""
        from app.reports.templates.pe_deal_memo import PEDealMemoTemplate

        template = PEDealMemoTemplate()
        assert template.name == "pe_deal_memo"
        assert hasattr(template, "gather_data")
        assert hasattr(template, "render_html")
        assert hasattr(template, "render_excel")

    def test_gather_data_missing_company(self):
        """T5: Handles non-existent company_id gracefully."""
        from app.reports.templates.pe_deal_memo import PEDealMemoTemplate

        db = self._make_mock_db(company_row=None)
        template = PEDealMemoTemplate()
        data = template.gather_data(db, {"company_id": 9999})

        assert data["company"] is None
        assert data["financials"] == []
        assert data["leadership"] == []

        html = template.render_html(data)
        assert "not found" in html.lower() or "Not Found" in html
