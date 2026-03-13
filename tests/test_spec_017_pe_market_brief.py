"""
Tests for SPEC 017 — PE Market Brief Report.
"""
import pytest
from unittest.mock import MagicMock, patch
from io import BytesIO
from datetime import date


@pytest.mark.unit
class TestSpec017PeMarketBrief:
    """Tests for PE market brief report template."""

    def _make_mock_db(self, sector_fallback_rows=None, deal_rows=None,
                      type_rows=None, buyer_rows=None, frag_rows=None,
                      rollup_rows=None):
        """Build a mock DB with canned data.

        Since gather_data calls service classes first (which will fail in
        unit tests), we mock the fallback direct queries.
        """
        db = MagicMock()
        call_count = [0]

        # Call order after service failures:
        # 0: sector fallback (from _get_sector_signals fallback)
        # 1: deal_activity
        # 2: deal_type_breakdown
        # 3: top_buyers
        # 4: fragmentation
        # 5: rollup_targets
        # 6: timing (service call — will fail gracefully)
        result_sequence = [
            sector_fallback_rows,
            deal_rows,
            type_rows,
            buyer_rows,
            frag_rows,
            rollup_rows,
        ]

        def mock_execute(query, params=None):
            result = MagicMock()
            idx = call_count[0]
            call_count[0] += 1

            if idx < len(result_sequence):
                data = result_sequence[idx]
                if data is None:
                    result.fetchone.return_value = None
                    result.fetchall.return_value = []
                elif isinstance(data, list):
                    result.fetchall.return_value = data
                    result.fetchone.return_value = data[0] if data else None
                else:
                    result.fetchone.return_value = data
                    result.fetchall.return_value = [data]
            else:
                result.fetchone.return_value = None
                result.fetchall.return_value = []
            return result

        db.execute = mock_execute
        return db

    def _sample_sector_rows(self):
        return [
            ("Software", 12, 8.5),
            ("Healthcare", 8, 10.2),
            ("Industrials", 6, 6.8),
        ]

    def _sample_deal_rows(self):
        return [
            (date(2025, 10, 1), 15, 9.2, 850_000_000.0),
            (date(2025, 7, 1), 12, 8.8, 620_000_000.0),
            (date(2025, 4, 1), 10, 7.5, 450_000_000.0),
        ]

    def _sample_type_rows(self):
        return [
            ("LBO", 18, 9.0),
            ("Strategic", 12, 10.5),
            ("Add-on", 8, 7.2),
        ]

    def _sample_buyer_rows(self):
        return [
            ("Thoma Bravo", 5, 2_500_000_000.0),
            ("Vista Equity", 4, 1_800_000_000.0),
        ]

    def _sample_rollup_rows(self):
        return [
            ("Healthcare", 8, 5, 12.3),
            ("Software", 6, 4, 9.1),
        ]

    def test_gather_data_returns_expected_keys(self):
        """T1: gather_data returns all required sections."""
        from app.reports.templates.pe_market_brief import PEMarketBriefTemplate

        db = self._make_mock_db(
            sector_fallback_rows=self._sample_sector_rows(),
            deal_rows=self._sample_deal_rows(),
            type_rows=self._sample_type_rows(),
            buyer_rows=self._sample_buyer_rows(),
            frag_rows=[],
            rollup_rows=self._sample_rollup_rows(),
        )
        template = PEMarketBriefTemplate()
        data = template.gather_data(db, {})

        assert "generated_at" in data
        assert "sector_signals" in data
        assert "deal_activity" in data
        assert "deal_type_breakdown" in data
        assert "top_buyers" in data
        assert "fragmentation" in data
        assert "rollup_targets" in data
        assert "timing_assessment" in data

    def test_render_html_contains_sections(self):
        """T2: HTML output has all 5 sections."""
        from app.reports.templates.pe_market_brief import PEMarketBriefTemplate

        db = self._make_mock_db(
            sector_fallback_rows=self._sample_sector_rows(),
            deal_rows=self._sample_deal_rows(),
            type_rows=self._sample_type_rows(),
            buyer_rows=self._sample_buyer_rows(),
            frag_rows=[],
            rollup_rows=self._sample_rollup_rows(),
        )
        template = PEMarketBriefTemplate()
        data = template.gather_data(db, {})
        html = template.render_html(data)

        assert "<!DOCTYPE html>" in html
        assert "Market Intelligence Brief" in html
        # All 5 sections
        assert "Sector Overview" in html
        assert "Deal Activity Trends" in html
        assert "Fragmentation Opportunities" in html
        assert "Rollup Targets" in html
        assert "Market Timing Assessment" in html
        # Deal type chart (always present with deal data)
        assert "dealTypeChart" in html

    def test_render_excel_has_sheets(self):
        """T3: Excel has required sheets."""
        from app.reports.templates.pe_market_brief import PEMarketBriefTemplate
        from openpyxl import load_workbook

        db = self._make_mock_db(
            sector_fallback_rows=self._sample_sector_rows(),
            deal_rows=self._sample_deal_rows(),
            type_rows=self._sample_type_rows(),
            buyer_rows=self._sample_buyer_rows(),
            frag_rows=[],
            rollup_rows=self._sample_rollup_rows(),
        )
        template = PEMarketBriefTemplate()
        data = template.gather_data(db, {})
        excel_bytes = template.render_excel(data)

        wb = load_workbook(BytesIO(excel_bytes))
        assert "Sector Signals" in wb.sheetnames
        assert "Deal Activity" in wb.sheetnames
        assert "Top Buyers" in wb.sheetnames
        assert "Rollup Targets" in wb.sheetnames
        assert len(wb.sheetnames) == 4

    def test_template_registered(self):
        """T4: Template name and methods exist."""
        from app.reports.templates.pe_market_brief import PEMarketBriefTemplate

        template = PEMarketBriefTemplate()
        assert template.name == "pe_market_brief"
        assert hasattr(template, "gather_data")
        assert hasattr(template, "render_html")
        assert hasattr(template, "render_excel")

    def test_gather_data_no_industry(self):
        """T5: Works without industry filter (all sectors)."""
        from app.reports.templates.pe_market_brief import PEMarketBriefTemplate

        db = self._make_mock_db(
            sector_fallback_rows=self._sample_sector_rows(),
            deal_rows=self._sample_deal_rows(),
            type_rows=self._sample_type_rows(),
            buyer_rows=self._sample_buyer_rows(),
            frag_rows=[],
            rollup_rows=self._sample_rollup_rows(),
        )
        template = PEMarketBriefTemplate()
        data = template.gather_data(db, {})

        assert data["industry_filter"] is None
        # Should still have data
        assert len(data["sector_signals"]) > 0 or len(data["deal_activity"]) > 0
